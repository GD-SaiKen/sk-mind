"""SRM 数据梳理 Excel 生成 — 简洁版
从 SRM API 文档中直接提取字段定义表，一个 API 对应一个 Sheet。
格式：英文字段名 | 中文描述 | 类型 | 枚举说明（可选）
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式 ──────────────────────────────────────────
HEADER_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
BODY_FONT = Font(name="Microsoft YaHei", size=9)
SECTION_FONT = Font(name="Microsoft YaHei", size=10, bold=True, color="2F5496")
SECTION_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)


def style_header(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER


def style_section(ws, row, ncols, text):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    cell = ws.cell(row=row, column=1)
    cell.value = text
    cell.font = SECTION_FONT
    cell.fill = SECTION_FILL
    cell.alignment = LEFT
    cell.border = THIN_BORDER


def style_body(ws, start_row, end_row, ncols):
    for r in range(start_row, end_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            cell.alignment = LEFT if c != 1 else LEFT


def write_info_row(ws, row, label, value, ncols=5):
    """写标签-值行"""
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=1)
    ws.cell(row=row, column=1, value=label).font = Font(name="Microsoft YaHei", size=9, bold=True)
    ws.cell(row=row, column=1).border = THIN_BORDER
    ws.merge_cells(start_row=row, start_column=2, end_row=row, end_column=ncols)
    ws.cell(row=row, column=2, value=value).font = BODY_FONT
    ws.cell(row=row, column=2).border = THIN_BORDER
    ws.cell(row=row, column=2).alignment = LEFT


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def make_api_sheet(wb, sheet_name, info: dict, request_params: list, response_fields: list):
    """创建一个 API Sheet
    info: {label: value} 头部信息
    request_params: [(field, desc, type, required, enum), ...]
    response_fields: [(field, desc, type, enum), ...]
    """
    ws = wb.create_sheet(title=sheet_name)
    ncols = 4  # 字段名 | 中文描述 | 类型 | 枚举/说明
    set_col_widths(ws, [32, 36, 12, 40])

    row = 1
    # ── 标题 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=ncols)
    title_cell = ws.cell(row=row, column=1, value=f"📋 {info.get('名称', sheet_name)}")
    title_cell.font = Font(name="Microsoft YaHei", size=13, bold=True, color="1F3864")
    title_cell.alignment = LEFT
    row += 2

    # ── 基本信息 ──
    style_section(ws, row, ncols, "▎基本信息")
    row += 1
    base_keys = ["数据源", "API 路径", "请求方式", "所属模块", "调用场景", "访问频率限制"]
    for k in base_keys:
        if k in info:
            write_info_row(ws, row, k, info[k], ncols)
            row += 1
    row += 1

    # ── 请求参数 ──
    if request_params:
        style_section(ws, row, ncols, "▎请求参数（Body）")
        row += 1
        headers = ["英文字段名", "中文描述", "类型", "必填 / 枚举说明"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, ncols)
        row += 1
        req_start = row
        for p in request_params:
            for i, v in enumerate(p, 1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        style_body(ws, req_start, row - 1, ncols)
        row += 1

    # ── 响应字段 ──
    if response_fields:
        style_section(ws, row, ncols, "▎响应字段")
        row += 1
        headers = ["英文字段名", "中文描述", "类型", "枚举值 / 说明"]
        for i, h in enumerate(headers, 1):
            ws.cell(row=row, column=i, value=h)
        style_header(ws, row, ncols)
        row += 1
        resp_start = row
        for f in response_fields:
            for i, v in enumerate(f, 1):
                ws.cell(row=row, column=i, value=v)
            row += 1
        style_body(ws, resp_start, row - 1, ncols)

    # 冻结首行
    ws.freeze_panes = "A2"
    return ws


# ═══════════════════════════════════════════════════════════
# 📦 API 数据 — 直接从文档中提取的字段定义表
# ═══════════════════════════════════════════════════════════


def build_sample(wb):
    # ── Sheet 1: 企业物料供应商交货方式查询 ──
    make_api_sheet(
        wb,
        "01-物料供应商交货方式",
        info={
            "名称": "企业物料供应商交货方式查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/product/vendor/receiveTypes.json",
            "请求方式": "POST (JSON)",
            "所属模块": "物料供应商关系",
            "调用场景": "第三方系统(ERP)按物料和供应商从平台读取交货方式",
            "访问频率限制": "同一供应商 ≥ 1 分钟",
        },
        request_params=[
            ("prodCodes", "物料编码列表", "List<String>", "选填", "最多 500 条；至少 1 条"),
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "用于数据鉴权"),
            ("innerVendorCode", "内部供应商编码", "String", "必填", "供应商档案中的编码"),
        ],
        response_fields=[
            ("productCode", "物料编码", "String", ""),
            ("innerVendorCode", "内部供应商编码", "String", ""),
            ("receiveType", "收货方式", "Integer", "1：按订单；2：按排程"),
        ],
    )

    # ── Sheet 2: 根据物料编码查询核定供应商 ──
    make_api_sheet(
        wb,
        "02-核定供应商查询",
        info={
            "名称": "根据物料编码查询核定供应商",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/compProd/pvRelation/queryList.json",
            "请求方式": "POST (JSON)",
            "所属模块": "物料供应商关系",
            "调用场景": "第三方系统按物料获取核定供应商列表",
            "访问频率限制": "≥ 1 分钟",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "用于数据鉴权"),
            ("prodCodeList", "物料编码列表", "JsonArray", "与时间范围二选一必填", "最多 500 条"),
            ("updateStart", "最新修改开始时间", "Number", "与物料列表二选一必填", "毫秒时间戳；>= updateTime"),
            ("updateEnd", "最新修改结束时间", "Number", "与物料列表二选一必填", "毫秒时间戳；<= updateTime；区间 ≤ 24h"),
            ("categoryCodeList", "分类编码列表", "JsonArray", "选填", ""),
            ("categoryType", "分类类型", "String", "选填", "2000:ERP分类; 2001:分类1; 2002:分类2; 2003:分类3"),
        ],
        response_fields=[
            # 主表字段
            ("innerVendorCode", "内部供应商编码（ERP）", "String", ""),
            ("innerVendorName", "内部供应商名称", "String", ""),
            ("innerVendorAbbr", "内部供应商简称", "String", ""),
            ("vendorCompanyOrgNo", "供应商企业组织号", "String", ""),
            ("vendorStatus", "供应商状态", "String", "0:正式；1:潜在；2:停用；3:黑名单"),
            ("stage", "供应商阶段", "String", "准入评估/试样/小批量采购/正式采购/退出"),
            ("prodCode", "物料编码", "String", ""),
            ("prodName", "物料名称", "String", ""),
            ("prodScale", "物料规格", "String", ""),
            ("plantCode", "工厂编码", "String", ""),
            ("plantName", "工厂名称", "String", ""),
            ("deliveryMethodCode", "交货方式编码", "String", ""),
            ("deliveryMethodName", "交货方式名称", "String", ""),
            ("deliveryStockDays", "备货天数", "Number", ""),
            ("purchaseQuota", "采购配额", "Number", ""),
            ("minPurchaseQty", "最小采购量", "Number", ""),
            ("overDeliveryRate", "排程交货超交率", "Number", ""),
            ("packageQty", "包装数量", "Number", ""),
            ("minPackageQty", "最小包装量", "Number", ""),
            ("unitNetWeight", "单位净重", "Number", ""),
            ("customsRecordStatus", "关务备案状态", "String", ""),
            ("inspectionFlag", "检验否", "String", ""),
            ("brandName", "品牌名称", "String", ""),
            ("manufacturerPartNo", "制造商料号", "String", ""),
            ("bidType", "中标类型", "String", ""),
            ("outsourceLossRate", "委外损耗率", "Number", ""),
            ("updateTime", "最后更新时间", "Number", "毫秒时间戳"),
        ],
    )

    # ── Sheet 3: 内部商城引入单列表查询 ──
    make_api_sheet(
        wb,
        "03-内部商城引入单列表",
        info={
            "名称": "内部商城引入单列表查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/introduceOrder/getList.json",
            "请求方式": "POST (JSON)",
            "所属模块": "内部商城协同",
            "调用场景": "查询内部商城引入单列表数据",
            "访问频率限制": "≥ 5 分钟",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("startDate", "引入时间-开始", "Number", "必填", "毫秒时间戳；>= updateTime"),
            ("endDate", "引入时间-结束", "Number", "必填", "毫秒时间戳；<= updateTime；区间 ≤ 24h"),
        ],
        response_fields=[
            ("inductXkNo", "内部商城引入单平台单据号", "String", ""),
            ("status", "引入单状态", "Number", "0:草稿; 1:待选品; 2:待发布; 3:已完成; 4:已撤回; 5:已作废"),
        ],
    )

    # ── Sheet 4: 内部商城引入单详情查询 ──
    make_api_sheet(
        wb,
        "04-内部商城引入单详情",
        info={
            "名称": "内部商城引入单详情查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/introduceOrder/getDetail.json",
            "请求方式": "POST (JSON)",
            "所属模块": "内部商城协同",
            "调用场景": "查询内部商城引入单详情数据",
            "访问频率限制": "≥ 6 秒",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("inductXkNo", "内部商城引入单平台单据号", "String", "必填", ""),
        ],
        response_fields=[
            # data 公共字段
            ("inductXkNo", "内部商城引入单平台单据号", "String", ""),
            ("status", "引入单状态", "Number", "0:草稿; 1:待选品; 2:待发布; 3:已完成; 4:已撤回; 5:已作废"),
            ("remark", "备注", "String", ""),
            ("createTime", "创建时间", "Number", "毫秒时间戳"),
            ("publishTime", "发布时间", "Number", "毫秒时间戳"),
            ("creatorErpCode", "创建人 ERP 帐号", "String", ""),
            ("creatorName", "创建人姓名", "String", ""),
            # lineList 明细 (响应中的嵌套数组)
            ("── lineList[].lineNo", "行项次", "String", ""),
            ("── lineList[].prodCode", "物料编码", "String", ""),
            ("── lineList[].prodName", "物料名称", "String", ""),
            ("── lineList[].prodScale", "物料规格", "String", ""),
            ("── lineList[].prodScaleCode", "物料规格编号", "String", ""),
            ("── lineList[].quantity", "数量", "Number", ""),
            ("── lineList[].unit", "单位", "String", ""),
            ("── lineList[].price", "单价", "Number", ""),
            ("── lineList[].amount", "金额", "Number", ""),
            ("── lineList[].deliveryDate", "交货日期", "Number", "毫秒时间戳"),
            ("── lineList[].remark", "备注", "String", ""),
            # userInfo 嵌套对象
            ("── userInfo.name", "引入用户姓名", "String", ""),
            ("── userInfo.erpCode", "引入用户 ERP 帐号", "String", ""),
        ],
    )

    # ── Sheet 5: 潜在供应商列表查询 ──
    make_api_sheet(
        wb,
        "05-潜在供应商列表",
        info={
            "名称": "潜在供应商列表查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/potentialVendor/getList.json",
            "请求方式": "POST (JSON)",
            "所属模块": "生命周期管理",
            "调用场景": "查询企业潜在供应商列表数据",
            "访问频率限制": "≥ 5 分钟",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("dateStart", "资质审核完成时间-从", "Number", "四组时间范围必填一项", "毫秒时间戳；区间 ≤ 24h"),
            ("dateEnd", "资质审核完成时间-至", "Number", "", ""),
            ("createStart", "创建时间-从", "Number", "", "毫秒时间戳"),
            ("createEnd", "创建时间-至", "Number", "", ""),
            ("checkTimeStart", "审核时间-从", "Number", "", "毫秒时间戳"),
            ("checkTimeEnd", "审核时间-至", "Number", "", ""),
            ("turnFormalTimeStart", "转正时间-从", "Number", "", "毫秒时间戳"),
            ("turnFormalTimeEnd", "转正时间-至", "Number", "", ""),
            ("applyStatus", "申请状态", "Number[]", "选填", "0:草稿; 1:申请中; 2:拒绝; 3:已确认; 4:已转正式; 5:转正中; 6:不合格"),
            ("inviteStatus", "邀约状态", "Number[]", "选填", "0:未发起; 1:已发起; 2:对方已确认; 3:邀约码失败; 4:待采购审核; 5:采购退回"),
        ],
        response_fields=[
            ("innerVendorCode", "内部供应商编码（ERP）", "String", ""),
            ("innerVendorName", "内部供应商名称", "String", ""),
            ("certificateAuditFinishTime", "资质审核完成时间", "Number", "毫秒时间戳"),
            ("admissionReportFlag", "完成准入评估标识", "Number", "0:未完成; 1:已完成"),
            ("applyStatus", "申请状态", "Number", "0:草稿; 1:申请中; 2:拒绝; 3:已确认; 4:已转正式; 5:转正中; 6:不合格"),
            ("vendorStatus", "供应商状态", "Number", "0:无效; 1:有效; 2:停用; 3:黑名单"),
            ("inviteStatus", "邀约状态", "Number", "0:未发起; 1:已发起; 2:对方已确认; 3:邀约码失败; 4:待采购审核; 5:采购退回"),
            ("checkTime", "审核时间", "Number", "毫秒时间戳"),
            ("updateTime", "最后更新时间", "Number", "毫秒时间戳"),
        ],
    )

    # ── Sheet 6: 潜在供应商详情查询 ──
    make_api_sheet(
        wb,
        "06-潜在供应商详情",
        info={
            "名称": "潜在供应商详情查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/potentialVendor/getDetail.json",
            "请求方式": "POST (JSON)",
            "所属模块": "生命周期管理",
            "调用场景": "查询单个潜在供应商的完整档案信息",
            "访问频率限制": "≥ 6 秒",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("innerVendorCode", "内部供应商编码", "String", "必填", ""),
        ],
        response_fields=[
            # 基本信息
            ("innerVendorCode", "内部供应商编码（ERP）", "String", ""),
            ("innerVendorName", "内部供应商名称", "String", ""),
            ("vendorCompanyOrgNo", "供应商企业组织号", "String", ""),
            ("companyName", "企业名称", "String", ""),
            ("companyNameEn", "公司英文名称", "String", ""),
            ("groupCompanyName", "集团公司名称", "String", ""),
            ("creditCode", "统一社会信用代码", "String", ""),
            ("legalPerson", "注册法人", "String", ""),
            ("registeredCapital", "注册资本", "String", ""),
            ("establishDate", "成立日期", "String", ""),
            ("companyAddress", "企业地址", "String", ""),
            ("businessScope", "经营范围", "String", ""),
            # 资质与阶段
            ("applyStatus", "申请状态", "Number", "0:草稿; 1:申请中; 2:拒绝; 3:已确认; 4:已转正式; 5:转正中; 6:不合格"),
            ("approvalStatus", "审批状态", "String", ""),
            ("vendorStatus", "供应商状态", "Number", "0:无效; 1:有效; 2:停用; 3:黑名单"),
            ("vendorLifecycle", "生命周期阶段", "String", "准入评估/试样/小批量采购/正式采购/退出"),
            ("turnFormalTime", "转正时间", "String", ""),
            ("stopReason", "停用原因", "String", ""),
            # 分类与品类
            ("vendorCategory", "潜在供应商分类", "String", ""),
            ("vendorCategoryList", "准入品类类别列表", "JsonArray", "含 status 状态字段"),
            # 税务与财务
            ("einCategoryCode", "税号类别编码", "String", ""),
            ("einCategoryName", "税号类别名称", "String", ""),
            ("taxRate", "税率", "String", ""),
            ("isTaxIncluded", "是否含税", "String", ""),
            ("paymentDaysCode", "账期编码", "String", ""),
            ("paymentDaysName", "账期名称", "String", ""),
            ("invoiceName", "发票名称", "String", ""),
            ("invoiceAddress", "发票地址", "String", ""),
            # 银行信息 (bankList 嵌套)
            ("── bankList[].bankName", "开户行名称", "String", ""),
            ("── bankList[].bankAccount", "银行账号", "String", ""),
            ("── bankList[].unionBankNo", "联行号", "String", ""),
            ("── bankList[].bankCountry", "开户地国家", "String", ""),
            ("── bankList[].bankProvince", "开户地省份", "String", ""),
            ("── bankList[].bankRegion", "开户地地区", "String", ""),
            # 产能与设备
            ("companyCapacity", "企业产能/产量", "String", ""),
            ("availableCapacity", "可提供产能", "String", ""),
            ("keyEquipment", "主要生产设备", "String", ""),
            ("keyManufacturingAbility", "关键制造能力", "String", ""),
            ("mainProducts", "主要供应产品", "String", ""),
            # 经营规模
            ("employeeCount", "企业员工数量", "Number", ""),
            ("techEmployeeCount", "技术人员数量", "Number", ""),
            ("annualTurnover", "年营业额", "String", ""),
            ("businessArea", "经营面积", "String", ""),
            ("productionArea", "生产面积", "String", ""),
            ("companyAdvantage", "企业优势", "String", ""),
            # 采购负责人
            ("buyerErpCode", "采购员编码", "String", ""),
            ("buyerName", "采购员姓名", "String", ""),
            ("buyerPhone", "采购员手机号", "String", ""),
            ("buyerEmail", "采购员邮箱", "String", ""),
            ("validExpireTime", "有效到期时间", "String", ""),
            # 关联信息
            ("assoCompany", "关联企业", "String", ""),
            ("mainCustomerList", "主要客户列表", "JsonArray", ""),
            ("customFields", "自定义字段", "String", "自定义字段 1-10"),
            # 时间
            ("createTime", "创建时间", "Number", "毫秒时间戳"),
            ("updateTime", "最后更新时间", "Number", "毫秒时间戳"),
        ],
    )

    # ── Sheet 7: 客户订单列表查询 ──
    make_api_sheet(
        wb,
        "07-客户订单列表",
        info={
            "名称": "客户订单列表查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/customerOrder/orderList.json",
            "请求方式": "POST (JSON)",
            "所属模块": "销售协同",
            "调用场景": "销售方获取客户订单列表数据",
            "访问频率限制": "≥ 5 分钟",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("publishStartTime", "发布开始时间", "Number", "必填（与确认时间二选一）", "毫秒时间戳；区间 ≤ 24h"),
            ("publishEndTime", "发布结束时间", "Number", "必填（与确认时间二选一）", "毫秒时间戳"),
            ("confirmStartTime", "确认时间起始", "Number", "必填（与发布时间二选一）", "毫秒时间戳"),
            ("confirmEndTime", "确认时间截止", "Number", "必填（与发布时间二选一）", "毫秒时间戳"),
            ("orderStatus", "订单状态", "Number", "选填", "1:待答交; 2:差异待确认; 3:退回待答交; 4:变更确认中; 5:已确认; 6:已结案; 7:已冻结; 8:已留置; 10:撤回答交; 11:采购方撤回; 12:作废; 13:供应商方拒绝"),
            ("customerXkUniqueCode", "客户档案平台唯一编码", "String", "选填", ""),
            ("innerCustomerCode", "内部客户编码", "String", "选填", ""),
        ],
        response_fields=[
            ("poXkNo", "采购单平台单据号", "String", ""),
            ("poErpNo", "采购单 ERP 单据号", "String", ""),
            ("innerCustomerCode", "内部客户编码", "String", ""),
            ("innerCustomerName", "内部客户名称", "String", ""),
            ("customerCode", "客户平台企业号", "String", ""),
            ("orderStatus", "订单状态", "Number", "1:待答交; 2:差异待确认; 3:退回待答交; 4:变更确认中; 5:已确认; 6:已结案; 7:已冻结; 8:已留置; 10:撤回答交; 11:采购方撤回; 12:作废; 13:供应商方拒绝"),
            ("erpPurchaseDate", "采购日期", "Number", "毫秒时间戳"),
            ("publishTime", "发布日期", "Number", "毫秒时间戳"),
        ],
    )

    # ── Sheet 8: 客户订单详情查询 ──
    make_api_sheet(
        wb,
        "08-客户订单详情",
        info={
            "名称": "客户订单详情查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/customerOrder/orderDetail.json",
            "请求方式": "POST (JSON)",
            "所属模块": "销售协同",
            "调用场景": "第三方系统(ERP)按平台单据号获取客户订单详情数据",
            "访问频率限制": "≥ 6 秒",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("poXkNo", "采购单平台单据号", "String", "必填", ""),
        ],
        response_fields=[
            ("poXkNo", "采购单平台单据号", "String", ""),
            ("poErpNo", "采购单 ERP 单据号", "String", ""),
            ("innerCustomerCode", "内部客户编码", "String", ""),
            ("innerCustomerName", "内部客户名称", "String", ""),
            ("customerCode", "客户平台企业号", "String", ""),
            ("customerName", "客户企业名称", "String", ""),
            ("orderStatus", "订单状态", "Number", "1:待答交; 2:差异待确认; 3:退回待答交; 4:变更确认中; 5:已确认; 6:已结案; 7:已冻结; 8:已留置; 10:撤回答交; 11:采购方撤回; 12:作废; 13:供应商方拒绝"),
            ("erpPurchaseDate", "采购日期", "Number", "毫秒时间戳"),
            ("publishTime", "发布日期", "Number", "毫秒时间戳"),
            ("confirmTime", "确认时间", "Number", "毫秒时间戳"),
            ("currencyCode", "币种编码", "String", ""),
            ("remark", "采购说明", "String", ""),
            ("overseasDirectCompanyCode", "境外直送企业编码", "String", ""),
            ("overseasDirectCompanyName", "境外直送企业名称", "String", ""),
            # 订单明细 lineList
            ("── lineList[].lineNo", "订单行号", "String", ""),
            ("── lineList[].poLineNoShow", "订单行显示项次", "String", ""),
            ("── lineList[].prodCode", "物料编码", "String", ""),
            ("── lineList[].prodName", "物料名称", "String", ""),
            ("── lineList[].prodScale", "物料规格", "String", ""),
            ("── lineList[].customerProdCode", "客户物料编码", "String", ""),
            ("── lineList[].customerProdName", "客户物料名称", "String", ""),
            ("── lineList[].customerProdScale", "客户物料规格", "String", ""),
            ("── lineList[].orderQty", "订单数量", "Number", ""),
            ("── lineList[].unit", "单位", "String", ""),
            ("── lineList[].unitPrice", "单价", "Number", ""),
            ("── lineList[].amount", "金额", "Number", ""),
            ("── lineList[].deliveryDate", "交货日期", "Number", "毫秒时间戳"),
            ("── lineList[].receivedQty", "已收货数量", "Number", ""),
            ("── lineList[].receivingDeptCode", "收货部门编码", "String", ""),
            ("── lineList[].receivingDeptName", "收货部门名称", "String", ""),
            ("── lineList[].remark", "行备注", "String", ""),
        ],
    )

    # ── Sheet 9: 委外发料单详情查询 ──
    make_api_sheet(
        wb,
        "09-委外发料单详情",
        info={
            "名称": "委外发料单详情查询",
            "数据源": "携客云 SRM 平台",
            "API 路径": "https://openapi.xiekeyun.com/outsourceIssuanceOrder/getDetail.json",
            "请求方式": "POST (JSON)",
            "所属模块": "物料委托管理",
            "调用场景": "查询委外发料单的详细信息，包括发料明细和签收状态",
            "访问频率限制": "≥ 18 秒",
        },
        request_params=[
            ("erpCode", "请求者用户 ERP 帐号", "String", "必填", "数据鉴权"),
            ("oioXkNo", "委外发料单平台单号", "String", "二选一必填", ""),
            ("oioErpNo", "委外发料单 ERP 单号", "String", "二选一必填", ""),
        ],
        response_fields=[
            ("oioXkNo", "平台发料单据号", "String", ""),
            ("oioErpNo", "ERP 单据号", "String", ""),
            ("erpOrgCode", "ERP 组织编码", "String", ""),
            ("erpOrgName", "ERP 组织名称", "String", ""),
            ("orderType", "单据类型", "Number", "1:一般发料; 2:退货发料"),
            ("innerVendorName", "内部供应商名称", "String", ""),
            ("innerVendorCode", "内部供应商编码", "String", ""),
            ("issueTypeCode", "发料类型编码", "String", "1:一般发料; 2:工序发料; 3:调拨发料; 4:超耗发料; 5:补料发料"),
            ("issueTypeName", "发料类型名称", "String", ""),
            ("issueDate", "发料时间", "Number", "毫秒时间戳"),
            ("publishName", "发布人姓名", "String", ""),
            ("attachCount", "附件数量", "Number", ""),
            ("backReason", "退回原因说明", "String", ""),
            ("status", "状态", "Number", "0:待签收; 1:已退回; 2:差异待确认; 3:已完成; 4:冻结"),
            # 单身明细 lineList
            ("── lineList[].lineNo", "发料单项次", "String", ""),
            ("── lineList[].oioLineNoShow", "发料单平台显示项次", "String", ""),
            ("── lineList[].poXkNo", "携客云平台采购单号", "String", ""),
            ("── lineList[].poErpNo", "采购单 ERP 单号", "String", ""),
            ("── lineList[].poLineNo", "采购单项次", "String", ""),
            ("── lineList[].productCode", "产品编码", "String", ""),
            ("── lineList[].productName", "产品名称", "String", ""),
            ("── lineList[].productScale", "产品规格", "String", ""),
            ("── lineList[].prodFeature", "产品特征", "String", ""),
            ("── lineList[].materialCode", "材料编码", "String", ""),
            ("── lineList[].materialName", "材料名称", "String", ""),
            ("── lineList[].materialScale", "材料规格", "String", ""),
            ("── lineList[].profitCenterCode", "工厂编码", "String", ""),
            ("── lineList[].profitCenterName", "工厂名称", "String", ""),
            ("── lineList[].storeCode", "仓库编码", "String", ""),
            ("── lineList[].storeName", "仓库名称", "String", ""),
            ("── lineList[].batchNumber", "批号", "String", ""),
            ("── lineList[].issueQty", "发料数量", "Number", ""),
            ("── lineList[].issueUnitCode", "发料单位编码", "String", ""),
            ("── lineList[].issueUnitName", "发料单位名称", "String", ""),
            ("── lineList[].waitIssueQty", "应发料量", "Number", ""),
            ("── lineList[].alreadyIssueQty", "已发料量", "Number", ""),
            ("── lineList[].remark", "备注", "String", ""),
            ("── lineList[].receiveQty", "签收数量", "Number", ""),
            ("── lineList[].receiveRemark", "签收说明", "String", ""),
            ("── lineList[].attachCount", "附件数量", "Number", ""),
        ],
    )


# ═══════════════════════════════════════════════════════════
# 🏁 主入口
# ═══════════════════════════════════════════════════════════

wb = openpyxl.Workbook()
# 删除默认 Sheet
del wb["Sheet"]

build_sample(wb)

# 把模块总览放在第一个 Sheet
ws_index = wb.create_sheet(title="00-模块总览", index=0)
set_col_widths(ws_index, [6, 26, 28, 50, 22, 38])

# 标题
ws_index.merge_cells("A1:F1")
ws_index.cell(row=1, column=1, value="SRM 数据盘点 — API 接口清单").font = Font(
    name="Microsoft YaHei", size=14, bold=True, color="1F3864"
)

row = 3
headers = ["序号", "模块", "API 名称", "URL 路径", "请求方式", "Sheet"]
for i, h in enumerate(headers, 1):
    ws_index.cell(row=row, column=i, value=h)
style_header(ws_index, row, 6)

apis_index = [
    (1, "物料供应商关系", "企业物料供应商交货方式查询", "product/vendor/receiveTypes.json", "POST", "01-物料供应商交货方式"),
    (2, "物料供应商关系", "根据物料编码查询核定供应商", "compProd/pvRelation/queryList.json", "POST", "02-核定供应商查询"),
    (3, "内部商城协同", "内部商城引入单列表查询", "introduceOrder/getList.json", "POST", "03-内部商城引入单列表"),
    (4, "内部商城协同", "内部商城引入单详情查询", "introduceOrder/getDetail.json", "POST", "04-内部商城引入单详情"),
    (5, "生命周期管理", "潜在供应商列表查询", "potentialVendor/getList.json", "POST", "05-潜在供应商列表"),
    (6, "生命周期管理", "潜在供应商详情查询", "potentialVendor/getDetail.json", "POST", "06-潜在供应商详情"),
    (7, "销售协同", "客户订单列表查询", "customerOrder/orderList.json", "POST", "07-客户订单列表"),
    (8, "销售协同", "客户订单详情查询", "customerOrder/orderDetail.json", "POST", "08-客户订单详情"),
    (9, "物料委托管理", "委外发料单详情查询", "outsourceIssuanceOrder/getDetail.json", "POST", "09-委外发料单详情"),
]

row = 4
for api in apis_index:
    for i, v in enumerate(api, 1):
        ws_index.cell(row=row, column=i, value=v)
    row += 1
style_body(ws_index, 4, row - 1, 6)

# 底部说明
row += 1
ws_index.merge_cells(start_row=row, start_column=1, end_row=row, end_column=6)
ws_index.cell(row=row, column=1, value="数据源：携客云 (XieKeyun) SRM 平台 OpenAPI — https://openapi.xiekeyun.com/").font = Font(
    name="Microsoft YaHei", size=8, color="808080"
)

ws_index.freeze_panes = "A4"

# 保存
out_path = "docs/数据梳理/SRM/SRM-字段字典-打样.xlsx"
wb.save(out_path)
print(f"✅ 打样 Excel 已生成: {out_path}")
print(f"   Sheets: {wb.sheetnames}")
