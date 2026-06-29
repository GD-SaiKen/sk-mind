#!/usr/bin/env python3
"""
SRM API 字段字典解析器
解析 携客云 OpenAPI 文档 → Excel 字段字典
每个 API = 一个 Sheet，嵌套对象作为子表堆叠
"""

import re
import sys
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 样式定义 ──────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
SUBHEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
SUBHEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="1F4E79")
NORMAL_FONT = Font(name="微软雅黑", size=10)
URL_FONT = Font(name="Consolas", size=9, color="0563C1")
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)
WRAP = Alignment(wrap_text=True, vertical="top")

# ── 列定义 ──────────────────────────────────────────
REQ_COLS = ["字段名称", "中文描述", "类型", "必填", "说明"]
RESP_COLS = ["字段名称", "中文描述", "必填", "类型"]

def clean_cell(text):
    """去掉 markdown 的 **bold** 标记和多余空白"""
    if not text:
        return ""
    return re.sub(r'\*\*(.+?)\*\*', r'\1', text.strip())

def parse_md_table(lines, start_idx):
    """从 start_idx 开始解析 markdown 表格，返回 (headers, rows, end_idx)"""
    # 找到表头行（以 | 开头）
    i = start_idx
    while i < len(lines) and not lines[i].strip().startswith("|"):
        i += 1
    if i >= len(lines):
        return None, [], i

    header_line = lines[i].strip()
    headers = [clean_cell(h) for h in header_line.split("|")[1:-1]]
    i += 1

    # 跳过分隔行 (|---|---|)
    if i < len(lines) and re.match(r'^\|[\s\-:|]+\|$', lines[i].strip()):
        i += 1

    # 解析数据行
    rows = []
    while i < len(lines):
        line = lines[i].strip()
        if not line.startswith("|"):
            break
        cells = [clean_cell(c) for c in line.split("|")[1:-1]]
        # 对齐列数
        while len(cells) < len(headers):
            cells.append("")
        cells = cells[:len(headers)]
        rows.append(cells)
        i += 1

    return headers, rows, i

def find_section_title(lines, start_idx, patterns):
    """找到下一个匹配 patterns 之一的章节标题，返回 (title, line_idx)"""
    for i in range(start_idx, len(lines)):
        line = lines[i].strip()
        for pat in patterns:
            m = re.match(pat, line)
            if m:
                return m.group(0), i
    return None, len(lines)

def normalize_header(h):
    """将中文列名映射到标准列名"""
    mapping = {
        "参数名称": "字段名称",
        "参数描述": "中文描述",
        "参数类型": "类型",
        "要求": "必填",
    }
    return mapping.get(h, h)

# ── 字段定义标题的匹配模式 ──
FIELD_DEF_PATTERNS = [
    re.compile(r'^#{2,5}\s*(.+字段定义.*)'),
    re.compile(r'^#{2,5}\s*(.+明细字段定义.*)'),
    re.compile(r'^#{2,5}\s*(.+字段定义说明.*)'),
    re.compile(r'^#{2,5}\s*(.+字段明细定义.*)'),
]

def is_field_def_title(line):
    """检查是否是字段定义章节标题"""
    for pat in FIELD_DEF_PATTERNS:
        if pat.match(line):
            return True
    return False

def extract_field_def_title(line):
    """从章节标题中提取字段定义的标题文本"""
    for pat in FIELD_DEF_PATTERNS:
        m = pat.match(line)
        if m:
            return m.group(1).strip()
    return None

def is_api_boundary(line, start_idx, i):
    """检查是否应该停止解析当前 API。
    只有 ## 顶级数字标题（如 ## 1 xxx, ## 2 xxx）才算边界，
    ## N.M 子章节（如 ## 1.1 请求, ## 2.2 结果返回）不算。"""
    if i <= start_idx + 3:
        return False
    # # 一级标题 → 边界
    if re.match(r'^#\s[^#]', line):
        return True
    # ## 二级标题：只有顶级编号（纯数字，如 "## 1 xxx"）是边界
    # "## 1.1 请求" "## 2.2.3 字段定义" 等是子章节，不是边界
    m = re.match(r'^##\s+(\d+)\s+(.+)', line)
    if m and not is_field_def_title(line):
        # 检查是否是真的 API 标题（只有一个数字），而不是子章节号（1.1, 2.2.3）
        # 子章节的标题中包含 "请求" "结果返回" "字段定义" 等关键词
        title = m.group(2).strip()
        subsection_keywords = ["请求", "结果返回", "描述", "说明", "示例", "Demo"]
        is_subsection = any(kw in title for kw in subsection_keywords)
        if not is_subsection:
            return True
    return False

def parse_api(lines, start_idx, file_name):
    """
    从 start_idx 开始解析一个 API 块
    """
    result = {
        "name": file_name,
        "url": "",
        "req_params": [],
        "resp_tables": [],
    }

    i = start_idx

    # ── 找请求地址 ──
    url_found = False
    while i < len(lines):
        line = lines[i].strip()
        if is_api_boundary(line, start_idx, i):
            return result, i
        if "请求地址" in line:
            for j in range(i+1, min(i+5, len(lines))):
                candidate = lines[j].strip()
                # 处理裸 URL 和 markdown 链接 [text](url)
                url_m = re.search(r'https?://openapi\.xiekeyun\.com/([^\s\)]+)', candidate)
                if url_m:
                    result["url"] = url_m.group(1).rstrip(".json").split(".json")[0]
                    url_found = True
                    break
            if url_found:
                break
        i += 1

    if not url_found:
        return result, i

    # ── 收集请求参数表 ──
    while i < len(lines):
        line = lines[i].strip()
        if is_api_boundary(line, start_idx, i) and "业务请求参数" not in line:
            break
        if "业务请求参数" in line:
            headers, rows, next_i = parse_md_table(lines, i + 1)
            if headers:
                mapped_headers = [normalize_header(h) for h in headers]
                for row in rows:
                    d = dict(zip(mapped_headers, row))
                    result["req_params"].append((
                        d.get("字段名称", ""),
                        d.get("中文描述", ""),
                        d.get("类型", ""),
                        d.get("必填", ""),
                        d.get("说明", "")
                    ))
            i = next_i
            break
        i += 1

    # ── 收集所有响应字段表 ──
    while i < len(lines):
        line = lines[i].strip()

        if is_api_boundary(line, start_idx, i):
            break

        if is_field_def_title(line):
            title = extract_field_def_title(line)
            sec_match = re.search(r'(\d[\d.]*\d)', title) if title else None
            section = sec_match.group(1) if sec_match else ""

            headers, rows, next_i = parse_md_table(lines, i + 1)
            if headers and rows:
                mapped_headers = [normalize_header(h) for h in headers]
                cleaned_rows = []
                for row in rows:
                    d = dict(zip(mapped_headers, row))
                    cleaned_rows.append((
                        d.get("字段名称", ""),
                        d.get("描述", d.get("中文描述", "")),
                        d.get("类型", ""),
                        ""
                    ))
                result["resp_tables"].append({
                    "title": title,
                    "section": section,
                    "rows": cleaned_rows,
                })
            i = next_i
            continue

        i += 1

    return result, i


def parse_file(filepath):
    """解析整个 SRM 文档，返回 API 列表"""
    with open(filepath, "r", encoding="utf-8") as f:
        content = f.read()
    lines = content.split("\n")

    # 提取文件名作为模块名
    module_name = Path(filepath).stem.replace("SRM-查询-接口-", "")

    apis = []
    i = 0
    while i < len(lines):
        line = lines[i].strip()
        # 匹配 ## N API名称 或 ## API名称（无编号）
        m = re.match(r'^##\s+(\d+)\s+(.+)', line)
        if not m:
            m = re.match(r'^##\s+([^#\d].+)', line)
        if m:
            groups = m.groups()
            if len(groups) == 2:
                api_name = f"{groups[0]}-{groups[1].strip()}"
            else:
                api_name = groups[0].strip()
            # 跳过非 API 标题：修订历史、概述、目录等
            skip_titles = ["修订历史", "概述", "目录", "文档说明", "版本"]
            if any(api_name.startswith(t) for t in skip_titles):
                i += 1
                continue
            api, next_i = parse_api(lines, i, api_name)
            if api["url"]:
                api["module"] = module_name
                apis.append(api)
            i = next_i
        else:
            i += 1

    return apis


# ── Excel 输出 ──────────────────────────────────────────

# ── 子对象名称提取 ──
def extract_sub_object_name(title):
    """从字段定义标题中提取子对象名称，如 'lineList' 'bankList' 'userInfo'"""
    # 匹配 xxx字段定义 / xxx明细字段定义 / xxx公共字段定义 中的 xxx 对象名
    m = re.search(r'([a-zA-Z_][a-zA-Z0-9_]*)\s*(?:字段|明细|公共|中)', title)
    if m:
        return m.group(1)
    # 匹配 JSON对象(xxx) 中的 xxx
    m = re.search(r'JSON对象[（(](\w+)[）)]', title)
    if m:
        return m.group(1)
    return None

def _normalize_type(raw_type):
    """归一化类型：去 bold、小写、Integer(int32) 等"""
    t = clean_cell(raw_type).strip()
    mapping = {
        "String": "string", "string": "string",
        "Number": "number", "number": "number",
        "Integer": "integer", "integer": "integer",
        "Boolean": "boolean", "boolean": "boolean",
        "List": "array", "list": "array",
    }
    # 带括号的如 Integer(int32)、Number（毫秒数）
    t = re.sub(r'[（(].+[）)]', '', t)
    for k, v in mapping.items():
        if t == k:
            return v
    return t.lower() if t else t

SUB_OBJ_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # 淡黄色
SUB_OBJ_FONT = Font(name="微软雅黑", size=11, bold=True, color="BF8F00")

def write_sheet(ws, api):
    """将一个 API 写入一个 Sheet，嵌套对象用独立行标出"""
    row = 1

    # ── 标题行 ──
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=f"{api['module']}  /  {api['name']}")
    cell.font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    row += 1

    # URL
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=4)
    cell = ws.cell(row=row, column=1, value=f"https://openapi.xiekeyun.com/{api['url']}")
    cell.font = URL_FONT
    row += 2

    # ── 请求参数 ──
    if api["req_params"]:
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        cell = ws.cell(row=row, column=1, value="📥 请求参数 (Request Body)")
        cell.font = SUBHEADER_FONT
        cell.fill = SUBHEADER_FILL
        for c in range(1, 6):
            ws.cell(row=row, column=c).fill = SUBHEADER_FILL
        row += 1

        for ci, h in enumerate(REQ_COLS, 1):
            cell = ws.cell(row=row, column=ci, value=h)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.border = THIN_BORDER
            cell.alignment = WRAP
        row += 1

        for param in api["req_params"]:
            for ci, val in enumerate(param, 1):
                cell = ws.cell(row=row, column=ci, value=val)
                cell.font = NORMAL_FONT
                cell.border = THIN_BORDER
                cell.alignment = WRAP
            row += 1

        row += 2

    # ── 响应字段：嵌套对象 = 父行(array/object Xxx) + --子字段 ──
    if api["resp_tables"]:
        # 第一遍：收集 wrapper(data/dataList) 中的父字段说明
        parent_info = {}  # { "lineList": ("xxx列表", "array"), ... }
        wrapper_fields = []  # [(name, desc, req, type_), ...]
        for tbl in api["resp_tables"]:
            title = tbl["title"]
            if "公共字段" in title:
                continue
            is_wrapper = title.strip().startswith("data") or "dataList" in title
            if is_wrapper:
                for r in tbl["rows"]:
                    name, desc, type_, _ = r[0], r[1], r[2], r[3]
                    clean_type = _normalize_type(type_)
                    # JSON数组/对象 → 父字段
                    if "JSON数组" in type_:
                        parent_info[name] = (desc, f"array {name}")
                    elif "JSON对象" in type_:
                        parent_info[name] = (desc, f"object {name}")
                    else:
                        wrapper_fields.append((name, desc, "", clean_type))

        # 第二遍：组装最终行
        all_resp_fields = list(wrapper_fields)

        for tbl in api["resp_tables"]:
            title = tbl["title"]
            if "公共字段" in title:
                continue
            is_wrapper = title.strip().startswith("data") or "dataList" in title
            if is_wrapper:
                continue

            obj_name = extract_sub_object_name(title)
            if obj_name:
                p_desc, p_type = parent_info.get(obj_name, (title, f"array {obj_name}"))
                all_resp_fields.append((obj_name, p_desc, "", p_type))
                for r in tbl["rows"]:
                    clean_type = _normalize_type(r[2])
                    all_resp_fields.append((
                        f"--{obj_name}.{r[0]}",
                        r[1], "", clean_type
                    ))

        if all_resp_fields:
            cell = ws.cell(row=row, column=1, value="📤 响应字段 (Response)")
            cell.font = SUBHEADER_FONT
            cell.fill = SUBHEADER_FILL
            for c in range(1, 5):
                ws.cell(row=row, column=c).fill = SUBHEADER_FILL
            row += 1

            for ci, h in enumerate(RESP_COLS, 1):
                cell = ws.cell(row=row, column=ci, value=h)
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.border = THIN_BORDER
                cell.alignment = WRAP
            row += 1

            for r in all_resp_fields:
                for ci, val in enumerate(r, 1):
                    cell = ws.cell(row=row, column=ci, value=val)
                    cell.font = NORMAL_FONT
                    cell.border = THIN_BORDER
                    cell.alignment = WRAP
                row += 1

    # 列宽：字段名 / 中文描述 / 必填 / 类型 / 说明(E)
    col_widths = [32, 52, 10, 20, 55]
    for ci, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def main():
    src_dir = Path("docs/数据梳理/SRM")
    output = src_dir / "SRM-字段字典-打样v2.xlsx"

    all_apis = []
    for md_file in sorted(src_dir.glob("SRM-查询-接口-*.md")):
        print(f"📄 解析: {md_file.name}")
        apis = parse_file(md_file)
        all_apis.extend(apis)
        for api in apis:
            n_resp = len(api["resp_tables"])
            print(f"   ✓ {api['name']} — {api['url']} (请求:{len(api['req_params'])}字段, 响应:{n_resp}表)")

    print(f"\n总计: {len(all_apis)} 个 API 端点")

    # ── 打样：选嵌套最多的 8 个 API ──
    all_apis.sort(key=lambda a: len(a["resp_tables"]), reverse=True)
    sample_apis = all_apis[:8]

    # ── 也输出完整版（若被 Excel 占用则跳过） ──
    output_full = src_dir / "SRM-字段字典-完整版.xlsx"
    try:
        wb_full = Workbook()
        wb_full.remove(wb_full.active)
        ws_fi = wb_full.create_sheet("00-API索引")
        for ci, h in enumerate(["模块", "API名称", "URL路径", "请求字段数", "响应子表数"], 1):
            cell = ws_fi.cell(row=1, column=ci, value=h)
            cell.font = HEADER_FONT; cell.fill = HEADER_FILL; cell.border = THIN_BORDER
        for ri, api in enumerate(all_apis, 2):
            for ci, v in enumerate([api["module"], api["name"], api["url"],
                                    len(api["req_params"]), len(api["resp_tables"])], 1):
                cell = ws_fi.cell(row=ri, column=ci, value=v)
                cell.font = NORMAL_FONT; cell.border = THIN_BORDER
        ws_fi.column_dimensions["A"].width = 18
        ws_fi.column_dimensions["B"].width = 40
        ws_fi.column_dimensions["C"].width = 55
        ws_fi.column_dimensions["D"].width = 12
        ws_fi.column_dimensions["E"].width = 12
        for api in all_apis:
            sheet_name = api["name"][:31]
            write_sheet(wb_full.create_sheet(sheet_name), api)
        wb_full.save(output_full)
        print(f"\n📦 完整版: {output_full} ({len(all_apis)} Sheets)")
    except PermissionError:
        print(f"\n⚠️ 完整版被 Excel 占用，跳过覆盖。请关闭后重跑。")

    wb = Workbook()
    # 删除默认 Sheet
    wb.remove(wb.active)

    # 先写总览
    ws_index = wb.create_sheet("00-API索引")
    index_cols = ["模块", "API名称", "URL路径", "请求参数字段数", "响应字段表数"]
    for ci, h in enumerate(index_cols, 1):
        cell = ws_index.cell(row=1, column=ci, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.border = THIN_BORDER

    for ri, api in enumerate(sample_apis, 2):
        vals = [api["module"], api["name"], api["url"],
                len(api["req_params"]), len(api["resp_tables"])]
        for ci, v in enumerate(vals, 1):
            cell = ws_index.cell(row=ri, column=ci, value=v)
            cell.font = NORMAL_FONT
            cell.border = THIN_BORDER

    ws_index.column_dimensions["A"].width = 18
    ws_index.column_dimensions["B"].width = 35
    ws_index.column_dimensions["C"].width = 55
    ws_index.column_dimensions["D"].width = 16
    ws_index.column_dimensions["E"].width = 16

    # 每个 API 一个 Sheet
    for api in sample_apis:
        # Sheet 名最多 31 字符
        sheet_name = api["name"][:31]
        ws = wb.create_sheet(sheet_name)
        write_sheet(ws, api)

    wb.save(output)
    print(f"\n✅ 打样 Excel 已生成: {output}")
    print(f"   {len(sample_apis)} 个 API + 1 个索引页")


if __name__ == "__main__":
    main()
