"""
生成 SRM 数据梳理 Excel 示例（打样版）
对齐 docs/ai-native-enterprise-data-platform-team-development-tasks.md 的 T00 交付物
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()

# ── Style definitions ──
HEADER_FONT = Font(name="Microsoft YaHei", bold=True, size=11, color="FFFFFF")
HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
SUB_HEADER_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
WARN_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
NORMAL_FONT = Font(name="Microsoft YaHei", size=10)
TITLE_FONT = Font(name="Microsoft YaHei", bold=True, size=14, color="1F4E79")
BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

def style_header(ws, row, cols):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = BORDER

def style_row(ws, row, cols, fill=None):
    for col in range(1, cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = NORMAL_FONT
        cell.alignment = LEFT if col > 1 else CENTER
        cell.border = BORDER
        if fill:
            cell.fill = fill

def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ═══════════════════════════════════════════
# Sheet 0: 数据源盘点总表
# ═══════════════════════════════════════════
ws0 = wb.active
ws0.title = "00-数据源盘点总表"

ws0.merge_cells("A1:L1")
ws0.cell(row=1, column=1, value="SRM 数据源盘点总表（携客云 OpenAPI）").font = TITLE_FONT

# Metadata rows
meta = [
    ("数据源名称", "携客云 SRM（Supplier Relationship Management）"),
    ("数据源类型", "SaaS 平台 OpenAPI（REST/JSON）"),
    ("API 基础地址", "https://openapi.xiekeyun.com/"),
    ("认证方式", "appKey + sign 签名（MD5），commonParam 公传参数"),
    ("业务负责人", "（待确认）"),
    ("技术负责人", "（待确认）"),
    ("系统负责人", "（待确认）"),
    ("文档版本", "基于携客云 OpenAPI 文档（2019-2026 修订）"),
    ("盘点日期", "2026-07"),
]
for i, (k, v) in enumerate(meta, 2):
    ws0.cell(row=i, column=1, value=k).font = Font(name="Microsoft YaHei", bold=True, size=10)
    ws0.cell(row=i, column=2, value=v).font = NORMAL_FONT
    ws0.merge_cells(start_row=i, start_column=2, end_row=i, end_column=8)

headers0 = [
    "序号", "API 端点路径", "接口名称", "所属模块", "方法",
    "查询类型", "主要业务对象", "阶段1纳入", "纳入理由/不纳入原因",
    "访问频率限制", "数据方向", "备注"
]
row_start = len(meta) + 4
ws0.cell(row=row_start, column=1, value="API 接口清单").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
hr = row_start + 1
for i, h in enumerate(headers0, 1):
    ws0.cell(row=hr, column=i, value=h)
style_header(ws0, hr, len(headers0))

apis = [
    # 内部商城协同
    ("introduceOrder/getList.json", "内部商城引入单列表查询", "内部商城协同", "POST", "列表", "RequisitionOrder", "✅ 是", "对接采购需求入口", "≥5 min", "SRM→平台", ""),
    ("introduceOrder/getDetail.json", "内部商城引入单详情查询", "内部商城协同", "POST", "详情", "RequisitionOrder", "✅ 是", "获取请购明细和行项目", "≥6 sec", "SRM→平台", ""),
    ("purchaseOrder/getList.json", "采购订单列表查询", "内部商城协同", "POST", "列表", "PurchaseOrder", "✅ 是", "采购订单是核心交易对象", "≥5 min", "SRM→平台", ""),
    ("purchaseOrder/getDetail.json", "采购订单详情查询", "内部商城协同", "POST", "详情", "PurchaseOrder", "✅ 是", "获取订单行、交期、收发货信息", "≥6 sec", "SRM→平台", ""),
    ("purchaseChangeOrder/getList.json", "采购变更单列表查询", "内部商城协同", "POST", "列表", "PurchaseOrder", "⚠️ 待定", "变更单数据量小,按需纳入", "≥5 min", "SRM→平台", ""),
    ("deliveryNotice/getList.json", "送货通知单列表查询", "内部商城协同", "POST", "列表", "Shipment", "✅ 是", "发货和交付追踪关键数据", "≥5 min", "SRM→平台", "含条码明细"),
    ("sampleRequest/getList.json", "样品申请单列表查询", "内部商城协同", "POST", "列表", "Sample", "❌ 阶段2", "非核心流程,后续按需补充", "≥5 min", "SRM→平台", ""),
    # 物料供应商关系
    ("product/vendor/receiveTypes.json", "物料供应商交货方式查询", "物料供应商关系", "POST", "详情", "Material,Supplier", "✅ 是", "物料-供应商关系核心数据", "≥1 min", "SRM→平台", ""),
    ("compProd/pvRelation/queryList.json", "根据物料编码查询核定供应商", "物料供应商关系", "POST", "列表", "Material,Supplier", "✅ 是", "物料准入和供应商资质关键", "≥1 min", "SRM→平台", "支持分类筛选"),
    ("compOrg/itgOrgPlants.json", "采购组织/工厂列表查询", "物料供应商关系", "POST", "列表", "Plant,Org", "✅ 是", "组织主数据基础", "≥1 min", "SRM→平台", ""),
    ("vendorArchive/getList.json", "供应商档案列表查询", "物料供应商关系", "POST", "列表", "Supplier", "✅ 是", "供应商主数据核心", "≥5 min", "SRM→平台", ""),
    ("empArchive/getList.json", "员工档案列表查询", "物料供应商关系", "POST", "列表", "Employee", "⚠️ 待定", "用于采购员权限映射", "≥5 min", "SRM→平台", ""),
    # 销售协同
    ("customerOrder/orderList.json", "客户订单列表查询", "销售协同", "POST", "列表", "SalesOrder", "✅ 是", "销售订单核心交易对象", "≥5 min", "SRM→平台", ""),
    ("customerOrder/orderDetail.json", "客户订单详情查询", "销售协同", "POST", "详情", "SalesOrder", "✅ 是", "获取订单行和客户信息", "≥6 sec", "SRM→平台", "含条码"),
    ("customerOrder/progressList.json", "客户订单进度列表查询", "销售协同", "POST", "列表", "SalesOrder", "✅ 是", "对应高价值问题:订单进度", "≥5 min", "SRM→平台", ""),
    ("customerStatementOrder/getList.json", "对账单列表查询", "销售协同", "POST", "列表", "Receivable", "⚠️ 待定", "应收场景纳入阶段2", "≥5 min", "SRM→平台", ""),
    # 生命周期管理
    ("potentialVendor/getList.json", "潜在供应商列表查询", "生命周期管理", "POST", "列表", "Supplier", "✅ 是", "供应商准入管理入口", "≥5 min", "SRM→平台", ""),
    ("potentialVendor/getDetail.json", "潜在供应商详情查询", "生命周期管理", "POST", "详情", "Supplier", "✅ 是", "供应商资质、银行、分类等", "≥6 sec", "SRM→平台", "含银行/资质子表"),
    ("vendorCompChanged/getList.json", "供应商信息变更列表查询", "生命周期管理", "POST", "列表", "Supplier", "⚠️ 待定", "变更记录按需纳入", "≥5 min", "SRM→平台", ""),
    ("admissionReport/getList.json", "供应商准入报告查询", "生命周期管理", "POST", "列表", "Supplier", "⚠️ 待定", "准入评估报告按需纳入", "≥5 min", "SRM→平台", ""),
    ("eliminateVendorNotice/getList.json", "供应商退出通知查询", "生命周期管理", "POST", "列表", "Supplier", "❌ 阶段2", "低频场景", "≥5 min", "SRM→平台", ""),
    # 物料委托管理
    ("outsourceIssuanceOrder/getDetail.json", "委外发料单详情查询", "物料委托管理", "POST", "详情", "ProductionOrder", "✅ 是", "委外加工发料和库存追踪", "≥18 sec", "SRM→平台", "含lineList行项目"),
    ("outsourceInventoryOrder/getList.json", "委外物料盘点单列表查询", "物料委托管理", "POST", "列表", "Inventory", "⚠️ 待定", "盘点数据按需纳入", "≥5 min", "SRM→平台", ""),
    ("outsourceInventoryOrder/getDetail.json", "委外物料盘点单详情查询", "物料委托管理", "POST", "详情", "Inventory", "⚠️ 待定", "", "≥6 sec", "SRM→平台", ""),
]

for i, row_data in enumerate(apis):
    r = hr + 1 + i
    for j, val in enumerate(row_data, 1):
        ws0.cell(row=r, column=j, value=val)
    fill = None
    if "✅" in row_data[7]:
        fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    elif "❌" in row_data[7]:
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "⚠️" in row_data[7]:
        fill = WARN_FILL
    style_row(ws0, r, len(headers0), fill)

last_r = hr + len(apis) + 1
ws0.cell(row=last_r, column=1, value="统计").font = Font(name="Microsoft YaHei", bold=True, size=10)
ws0.cell(row=last_r, column=2, value=f"共 {len(apis)} 个查询接口: ✅ 阶段1 纳入 {sum(1 for a in apis if '✅' in a[7])} 个, ⚠️ 待定 {sum(1 for a in apis if '⚠️' in a[7])} 个, ❌ 阶段2 {sum(1 for a in apis if '❌' in a[7])} 个")

set_col_widths(ws0, [6, 38, 28, 16, 8, 8, 24, 12, 28, 12, 14, 16])


# ═══════════════════════════════════════════
# Sheet 1: 核定供应商查询 (核心 API 示例)
# ═══════════════════════════════════════════
ws1 = wb.create_sheet("01-核定供应商查询")

ws1.merge_cells("A1:H1")
ws1.cell(row=1, column=1, value="API: 根据物料编码查询核定供应商").font = TITLE_FONT
ws1.cell(row=2, column=1, value="URL: https://openapi.xiekeyun.com/compProd/pvRelation/queryList.json").font = Font(name="Consolas", size=10, color="555555")
ws1.cell(row=3, column=1, value="模块: 物料供应商关系 | 方法: POST | 频率: ≥1 分钟 | 业务对象: Material, Supplier").font = NORMAL_FONT

# ── Request Params ──
ws1.cell(row=5, column=1, value="请求参数 (Request Parameters)").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
req_headers = ["序号", "英文字段名", "中文名称", "类型", "必填/选填", "说明", "敏感等级", "备注"]
for i, h in enumerate(req_headers, 1):
    ws1.cell(row=6, column=i, value=h)
style_header(ws1, 6, len(req_headers))

req_fields = [
    ("1", "erpCode", "请求者用户ERP帐号", "String", "必填", "对应员工档案中ERP帐户的值，做数据鉴权", "L2-内部", "鉴权字段"),
    ("2", "prodCodeList", "物料编码列表", "Json数组", "必填(二选一)", "最大500条；与时间范围二选一", "L1-公开", ""),
    ("3", "updateStart", "最新修改开始时间", "Number", "必填(二选一)", "毫秒时间戳; >= updateStart", "L1-公开", ""),
    ("4", "updateEnd", "最新修改结束时间", "Number", "必填(二选一)", "毫秒时间戳; <= updateEnd，最大区间24h", "L1-公开", ""),
    ("5", "categoryCodeList", "分类编码列表", "Json数组", "选填", "物料分类编码过滤", "L1-公开", ""),
    ("6", "categoryType", "分类类型", "String", "选填(categoryCodeList非空时必填)", "2000:ERP分类 2001:分类1 2002:分类2 2003:分类3", "L1-公开", ""),
]
for i, row_data in enumerate(req_fields):
    r = 7 + i
    for j, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=j, value=val)
    style_row(ws1, r, len(req_headers))

# ── Response Fields ──
resp_start = 7 + len(req_fields) + 2
ws1.cell(row=resp_start, column=1, value="返回字段 (Response Fields)").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
resp_headers = ["序号", "英文字段名", "中文名称", "类型", "所在层级", "枚举值/说明", "敏感等级", "业务对象映射", "备注"]
hr2 = resp_start + 1
for i, h in enumerate(resp_headers, 1):
    ws1.cell(row=hr2, column=i, value=h)
style_header(ws1, hr2, len(resp_headers))

resp_fields = [
    # 公共字段
    ("1", "result", "返回结果状态", "Number", "公共", "1:成功 9:失败", "L1-公开", "-", ""),
    ("2", "errorCode", "接口错误编码", "String", "公共", "失败时有值", "L1-公开", "-", ""),
    ("3", "errorMsg", "接口错误信息", "String", "公共", "失败时有值", "L1-公开", "-", ""),
    # dataList 字段
    ("4", "productCode", "物料编码", "String", "dataList", "", "L1-公开", "Material.code", "物料主数据标识"),
    ("5", "productName", "物料名称", "String", "dataList", "", "L1-公开", "Material.name", ""),
    ("6", "productScale", "物料规格", "String", "dataList", "", "L1-公开", "Material.spec", ""),
    ("7", "categoryCode", "物料分类编码", "String", "dataList", "", "L1-公开", "Material.category", ""),
    ("8", "categoryName", "物料分类名称", "String", "dataList", "", "L1-公开", "Material.categoryName", ""),
    ("9", "innerVendorCode", "内部供应商编码", "String", "dataList", "", "L1-公开", "Supplier.code", "供应商主数据标识"),
    ("10", "innerVendorName", "内部供应商名称", "String", "dataList", "", "L2-内部", "Supplier.name", "供应商名称可能有商业敏感"),
    ("11", "innerVendorAbbr", "内部供应商简称", "String", "dataList", "", "L2-内部", "Supplier.shortName", ""),
    ("12", "vendorCompanyOrgNo", "供应商企业组织号", "String", "dataList", "", "L1-公开", "Supplier.orgNo", ""),
    ("13", "vendorType", "供应商类型", "Number", "dataList", "0:正式 1:潜在", "L1-公开", "Supplier.type", ""),
    ("14", "receiveType", "收货方式", "Integer", "dataList", "1:按订单 2:按排程", "L1-公开", "Supplier.deliveryMethod", ""),
    ("15", "stockDays", "备货天数", "Number", "dataList", "", "L2-内部", "Supplier.stockDays", "商业敏感"),
    ("16", "purchaseQuota", "采购配额", "Number", "dataList", "", "L3-机密", "Supplier.quota", "价格/配额敏感"),
    ("17", "minPurchaseQty", "最小采购量", "Number", "dataList", "", "L2-内部", "Supplier.minQty", ""),
    ("18", "overDeliveryRate", "超交率", "Number", "dataList", "", "L2-内部", "Supplier.overDelivery", ""),
    ("19", "packQty", "包装数量", "Number", "dataList", "", "L1-公开", "Supplier.packQty", ""),
    ("20", "minPackQty", "最小包装量", "Number", "dataList", "", "L1-公开", "Supplier.minPackQty", ""),
    ("21", "unitNetWeight", "单位净重", "Number", "dataList", "", "L1-公开", "Supplier.unitWeight", ""),
    ("22", "customsStatus", "关务备案状态", "String", "dataList", "", "L2-内部", "Supplier.customsStatus", ""),
    ("23", "checkFlag", "检验否", "String", "dataList", "Y/N", "L1-公开", "Supplier.needCheck", ""),
    ("24", "brandName", "品牌名称", "String", "dataList", "", "L1-公开", "Supplier.brand", ""),
    ("25", "mfrPartNo", "制造商料号", "String", "dataList", "", "L1-公开", "Supplier.mfrPartNo", ""),
    ("26", "bidType", "中标类型", "String", "dataList", "", "L2-内部", "Supplier.bidType", ""),
    ("27", "outsourceLossRate", "委外损耗率", "Number", "dataList", "", "L2-内部", "Supplier.outsourceLoss", ""),
    ("28", "plantCode", "工厂编码", "String", "dataList", "", "L1-公开", "Plant.code", ""),
    ("29", "plantName", "工厂名称", "String", "dataList", "", "L1-公开", "Plant.name", ""),
    ("30", "status", "状态", "Number", "dataList", "0:无效 1:有效", "L1-公开", "Supplier.status", ""),
    ("31", "customFields", "自定义字段", "Json对象", "dataList", "平台自定义扩展字段", "L2-内部", "Supplier.custom", ""),
]
for i, row_data in enumerate(resp_fields):
    r = hr2 + 1 + i
    for j, val in enumerate(row_data, 1):
        ws1.cell(row=r, column=j, value=val)
    fill = None
    if "L3-机密" in str(row_data[6]):
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "L2-内部" in str(row_data[6]):
        fill = WARN_FILL
    style_row(ws1, r, len(resp_headers), fill)

set_col_widths(ws1, [6, 24, 22, 14, 12, 28, 12, 24, 22])


# ═══════════════════════════════════════════
# Sheet 2: 潜在供应商详情查询 (复杂嵌套示例)
# ═══════════════════════════════════════════
ws2 = wb.create_sheet("02-潜在供应商详情查询")

ws2.merge_cells("A1:H1")
ws2.cell(row=1, column=1, value="API: 潜在供应商详情查询").font = TITLE_FONT
ws2.cell(row=2, column=1, value="URL: https://openapi.xiekeyun.com/potentialVendor/getDetail.json").font = Font(name="Consolas", size=10, color="555555")
ws2.cell(row=3, column=1, value="模块: 生命周期管理 | 方法: POST | 频率: ≥6 秒 | 业务对象: Supplier").font = NORMAL_FONT

# Request
ws2.cell(row=5, column=1, value="请求参数 (Request Parameters)").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
req_h2 = ["序号", "英文字段名", "中文名称", "类型", "必填/选填", "说明", "敏感等级", "备注"]
for i, h in enumerate(req_h2, 1):
    ws2.cell(row=6, column=i, value=h)
style_header(ws2, 6, len(req_h2))
req2 = [
    ("1", "erpCode", "请求者用户ERP帐号", "String", "必填", "数据鉴权", "L2-内部", ""),
    ("2", "innerVendorCode", "内部供应商编码", "String", "必填", "供应商唯一标识", "L1-公开", ""),
]
for i, row_data in enumerate(req2):
    r = 7 + i
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=j, value=val)
    style_row(ws2, r, len(req_h2))

# Response - main data
rs = 7 + len(req2) + 2
ws2.cell(row=rs, column=1, value="返回字段 - data (主对象)").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
resp_h2 = ["序号", "英文字段名", "中文名称", "类型", "所在层级", "枚举值/说明", "敏感等级", "业务对象映射", "备注"]
hr3 = rs + 1
for i, h in enumerate(resp_h2, 1):
    ws2.cell(row=hr3, column=i, value=h)
style_header(ws2, hr3, len(resp_h2))

resp2 = [
    ("1", "result", "返回结果状态", "Number", "公共", "1:成功 9:失败", "L1-公开", "-", ""),
    ("2", "errorCode", "接口错误编码", "String", "公共", "", "L1-公开", "-", ""),
    ("3", "innerVendorCode", "内部供应商编码", "String", "data", "", "L1-公开", "Supplier.code", ""),
    ("4", "innerVendorName", "内部供应商名称", "String", "data", "", "L2-内部", "Supplier.name", ""),
    ("5", "innerVendorAbbr", "内部供应商简称", "String", "data", "", "L2-内部", "Supplier.shortName", ""),
    ("6", "creditCode", "统一社会信用代码", "String", "data", "", "L3-机密", "Supplier.creditCode", "企业敏感信息"),
    ("7", "einCategoryCode", "税号类别编码", "String", "data", "", "L3-机密", "Supplier.taxCode", ""),
    ("8", "einCategoryName", "税号类别名称", "String", "data", "", "L3-机密", "Supplier.taxName", ""),
    ("9", "companyAddress", "企业地址", "String", "data", "", "L2-内部", "Supplier.address", ""),
    ("10", "legalPerson", "注册法人", "String", "data", "", "L2-内部", "Supplier.legalPerson", ""),
    ("11", "registeredCapital", "注册资本", "String", "data", "", "L3-机密", "Supplier.capital", "财务敏感"),
    ("12", "businessScope", "经营范围", "String", "data", "", "L2-内部", "Supplier.bizScope", ""),
    ("13", "establishedDate", "成立日期", "Number", "data", "时间戳", "L1-公开", "Supplier.establishedDate", ""),
    ("14", "employeeCount", "企业员工数量", "Number", "data", "", "L2-内部", "Supplier.employeeCount", ""),
    ("15", "techStaffCount", "技术人员数量", "Number", "data", "", "L2-内部", "Supplier.techCount", ""),
    ("16", "annualRevenue", "年营业额", "String", "data", "", "L3-机密", "Supplier.annualRevenue", "财务敏感"),
    ("17", "productionArea", "生产面积", "String", "data", "", "L2-内部", "Supplier.prodArea", ""),
    ("18", "businessArea", "经营面积", "String", "data", "", "L2-内部", "Supplier.bizArea", ""),
    ("19", "capacity", "企业产能/产量", "String", "data", "", "L2-内部", "Supplier.capacity", ""),
    ("20", "availableCapacity", "可提供产能", "String", "data", "", "L2-内部", "Supplier.availCapacity", ""),
    ("21", "mainEquipment", "主要生产设备", "String", "data", "", "L2-内部", "Supplier.mainEquipment", ""),
    ("22", "mainSupplyProducts", "主要供应产品", "String", "data", "", "L1-公开", "Supplier.mainProducts", ""),
    ("23", "keyManufacturing", "关键制造能力", "String", "data", "", "L2-内部", "Supplier.keyMfg", ""),
    ("24", "applyStatus", "申请状态", "Number", "data", "0:草稿 1:申请中 2:拒绝 3:已确认 4:已转正式 5:转正中 6:不合格", "L1-公开", "Supplier.applyStatus", ""),
    ("25", "vendorStatus", "供应商状态", "Number", "data", "0:无效 1:有效 2:停用 3:黑名单", "L1-公开", "Supplier.status", ""),
    ("26", "lifecycleStage", "生命周期阶段", "String", "data", "准入评估/试样/小批量采购/正式采购/退出", "L1-公开", "Supplier.lifecycle", ""),
    ("27", "approvalStatus", "审批状态", "Number", "data", "", "L1-公开", "Supplier.approval", ""),
    ("28", "paymentDaysCode", "账期编码", "String", "data", "", "L3-机密", "Supplier.paymentDays", "商业条款敏感"),
    ("29", "turnFormalTime", "转正时间", "Number", "data", "时间戳", "L1-公开", "Supplier.turnFormalTime", ""),
    ("30", "buyerCode", "采购员编码", "String", "data", "", "L2-内部", "Employee.code", ""),
    ("31", "buyerName", "采购员姓名", "String", "data", "", "L2-内部", "Employee.name", ""),
    ("32", "buyerPhone", "采购员手机号", "String", "data", "", "L3-机密", "Employee.phone", "个人隐私"),
    ("33", "buyerEmail", "采购员邮箱", "String", "data", "", "L3-机密", "Employee.email", "个人隐私"),
    ("34", "detailUrl", "供应商详情URL", "String", "data", "", "L1-公开", "Supplier.url", ""),
    ("35", "qualificationUrl", "供应商资质动态查看URL", "String", "data", "", "L2-内部", "Supplier.qualUrl", ""),
]
for i, row_data in enumerate(resp2):
    r = hr3 + 1 + i
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=j, value=val)
    fill = None
    if "L3-机密" in str(row_data[6]):
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "L2-内部" in str(row_data[6]):
        fill = WARN_FILL
    style_row(ws2, r, len(resp_h2), fill)

# Nested: bankList
bnk_start = hr3 + len(resp2) + 2
ws2.cell(row=bnk_start, column=1, value="返回字段 - bankList (银行资料子表)").font = Font(name="Microsoft YaHei", bold=True, size=11, color="1F4E79")
bnk_h = ["序号", "英文字段名", "中文名称", "类型", "所在层级", "枚举值/说明", "敏感等级", "业务对象映射", "备注"]
for i, h in enumerate(bnk_h, 1):
    ws2.cell(row=bnk_start+1, column=i, value=h)
style_header(ws2, bnk_start+1, len(bnk_h))
bank_fields = [
    ("B1", "bankName", "开户银行名称", "String", "data.bankList[]", "", "L3-机密", "Supplier.bank.name", "银行信息敏感"),
    ("B2", "bankAccount", "银行账号", "String", "data.bankList[]", "", "L3-机密", "Supplier.bank.account", "银行账号高度敏感"),
    ("B3", "accountName", "开户名称", "String", "data.bankList[]", "", "L3-机密", "Supplier.bank.accountName", ""),
    ("B4", "unionBankNo", "联行号", "String", "data.bankList[]", "", "L3-机密", "Supplier.bank.unionNo", ""),
    ("B5", "bankCountry", "开户地国家", "String", "data.bankList[]", "", "L2-内部", "Supplier.bank.country", ""),
    ("B6", "bankProvince", "开户地省份", "String", "data.bankList[]", "", "L2-内部", "Supplier.bank.province", ""),
    ("B7", "bankRegion", "开户地地区", "String", "data.bankList[]", "", "L2-内部", "Supplier.bank.region", ""),
]
for i, row_data in enumerate(bank_fields):
    r = bnk_start + 2 + i
    for j, val in enumerate(row_data, 1):
        ws2.cell(row=r, column=j, value=val)
    style_row(ws2, r, len(bnk_h), PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"))

set_col_widths(ws2, [6, 24, 22, 14, 16, 40, 12, 24, 22])


# ═══════════════════════════════════════════
# Sheet 3: 客户订单详情查询 (销售协同示例)
# ═══════════════════════════════════════════
ws3 = wb.create_sheet("03-客户订单详情查询")

ws3.merge_cells("A1:H1")
ws3.cell(row=1, column=1, value="API: 客户订单详情查询").font = TITLE_FONT
ws3.cell(row=2, column=1, value="URL: https://openapi.xiekeyun.com/customerOrder/orderDetail.json").font = Font(name="Consolas", size=10, color="555555")
ws3.cell(row=3, column=1, value="模块: 销售协同 | 方法: POST | 频率: ≥6 秒 | 业务对象: SalesOrder, Customer, Product").font = NORMAL_FONT

# Request
ws3.cell(row=5, column=1, value="请求参数").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
for i, h in enumerate(req_h2, 1):
    ws3.cell(row=6, column=i, value=h)
style_header(ws3, 6, len(req_h2))
req3 = [
    ("1", "erpCode", "请求者用户ERP帐号", "String", "必填", "数据鉴权", "L2-内部", ""),
    ("2", "poXkNo", "采购单平台单据号", "String", "必填(二选一)", "平台内部唯一标识", "L1-公开", ""),
    ("3", "poErpNo", "采购单ERP单据号", "String", "必填(二选一)", "对接ERP系统单据号", "L1-公开", ""),
]
for i, row_data in enumerate(req3):
    r = 7 + i
    for j, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=j, value=val)
    style_row(ws3, r, len(req_h2))

# Response
rs3 = 7 + len(req3) + 2
ws3.cell(row=rs3, column=1, value="返回字段 - data + lineList[] (主单+行项目)").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
for i, h in enumerate(resp_h2, 1):
    ws3.cell(row=rs3+1, column=i, value=h)
style_header(ws3, rs3+1, len(resp_h2))

resp3 = [
    # 公共
    ("1", "result", "返回结果状态", "Number", "公共", "1:成功 9:失败", "L1-公开", "-", ""),
    ("2", "errorCode", "接口错误编码", "String", "公共", "", "L1-公开", "-", ""),
    # 主单
    ("3", "poXkNo", "采购单平台单据号", "String", "data", "", "L1-公开", "SalesOrder.id", "业务主键"),
    ("4", "poErpNo", "采购单ERP单据号", "String", "data", "", "L1-公开", "SalesOrder.erpId", ""),
    ("5", "innerCustomerCode", "内部客户编码", "String", "data", "", "L1-公开", "Customer.code", ""),
    ("6", "innerCustomerName", "内部客户名称", "String", "data", "", "L2-内部", "Customer.name", ""),
    ("7", "customerCode", "客户平台企业号", "String", "data", "", "L1-公开", "Customer.platformId", ""),
    ("8", "customerName", "客户企业名称", "String", "data", "", "L2-内部", "Customer.companyName", ""),
    ("9", "orderStatus", "订单状态", "Number", "data", "1:待答交 2:差异待确认 3:退回待答交 4:变更确认中 5:已确认 6:已结案 7:已冻结 8:已留置 10:撤回答交 11:采购方撤回 12:作废 13:供应商方拒绝", "L1-公开", "SalesOrder.status", "高价值:订单进度"),
    ("10", "erpPurchaseDate", "采购日期", "Number", "data", "时间戳", "L1-公开", "SalesOrder.purchaseDate", ""),
    ("11", "publishTime", "发布日期", "Number", "data", "时间戳", "L1-公开", "SalesOrder.publishTime", ""),
    ("12", "confirmTime", "确认时间", "Number", "data", "时间戳", "L1-公开", "SalesOrder.confirmTime", ""),
    ("13", "currencyCode", "币种编码", "String", "data", "", "L2-内部", "SalesOrder.currency", ""),
    ("14", "extendN01", "采购说明", "String", "data", "自定义扩展字段", "L2-内部", "SalesOrder.remark", ""),
    # 行项目 lineList
    ("15", "lineList[].lineNo", "订单行项次", "String", "data.lineList[]", "", "L1-公开", "SalesOrderItem.lineNo", ""),
    ("16", "lineList[].poLineNoShow", "订单行显示项次", "String", "data.lineList[]", "", "L1-公开", "SalesOrderItem.displayNo", ""),
    ("17", "lineList[].productCode", "产品编码", "String", "data.lineList[]", "", "L1-公开", "Product.code", ""),
    ("18", "lineList[].productName", "产品名称", "String", "data.lineList[]", "", "L1-公开", "Product.name", ""),
    ("19", "lineList[].productScale", "产品规格", "String", "data.lineList[]", "", "L1-公开", "Product.spec", ""),
    ("20", "lineList[].customerProductCode", "客户物料编码", "String", "data.lineList[]", "", "L2-内部", "Product.customerCode", ""),
    ("21", "lineList[].orderQty", "订单数量", "Number", "data.lineList[]", "", "L2-内部", "SalesOrderItem.qty", ""),
    ("22", "lineList[].unitCode", "单位编码", "String", "data.lineList[]", "", "L1-公开", "SalesOrderItem.unit", ""),
    ("23", "lineList[].deliveryDate", "交期", "Number", "data.lineList[]", "时间戳", "L2-内部", "SalesOrderItem.deliveryDate", "高价值:交付风险"),
    ("24", "lineList[].receivedQty", "已收货数量", "Number", "data.lineList[]", "", "L2-内部", "SalesOrderItem.receivedQty", ""),
    ("25", "lineList[].shippedQty", "已发货数量", "Number", "data.lineList[]", "", "L2-内部", "SalesOrderItem.shippedQty", ""),
    ("26", "lineList[].unitPrice", "单价", "Number", "data.lineList[]", "", "L3-机密", "SalesOrderItem.unitPrice", "价格高度敏感"),
    ("27", "lineList[].receiveDeptCode", "收货部门编码", "String", "data.lineList[]", "", "L2-内部", "SalesOrderItem.deptCode", ""),
    ("28", "lineList[].receiveDeptName", "收货部门名称", "String", "data.lineList[]", "", "L2-内部", "SalesOrderItem.deptName", ""),
    ("29", "lineList[].overseasCompanyCode", "境外直送企业编码", "String", "data.lineList[]", "", "L2-内部", "SalesOrderItem.overseasCode", ""),
]
for i, row_data in enumerate(resp3):
    r = rs3 + 2 + i
    for j, val in enumerate(row_data, 1):
        ws3.cell(row=r, column=j, value=val)
    fill = None
    if "L3-机密" in str(row_data[6]):
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "L2-内部" in str(row_data[6]):
        fill = WARN_FILL
    style_row(ws3, r, len(resp_h2), fill)

set_col_widths(ws3, [6, 28, 22, 14, 18, 45, 12, 26, 24])


# ═══════════════════════════════════════════
# Sheet 4: 敏感等级汇总
# ═══════════════════════════════════════════
ws4 = wb.create_sheet("97-敏感等级汇总")

ws4.merge_cells("A1:F1")
ws4.cell(row=1, column=1, value="字段敏感等级定义与汇总").font = TITLE_FONT

ws4.cell(row=3, column=1, value="敏感等级定义").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
def_headers = ["等级", "标识", "定义", "Agent 可用边界", "示例字段", "建议审批"]
for i, h in enumerate(def_headers, 1):
    ws4.cell(row=4, column=i, value=h)
style_header(ws4, 4, len(def_headers))

defs = [
    ("L1", "公开", "企业内部公开数据，无商业或个人敏感信息", "Agent 可直接查询和展示", "物料编码、订单状态、工厂名称、分类编码", "无需审批"),
    ("L2", "内部", "企业内部使用，外部泄露可能造成轻微商业影响", "Agent 可查询但需标注来源，展示需权限", "供应商名称、订单数量、交期、库存量、设备信息", "数据负责人审批"),
    ("L3", "机密", "涉及商业机密、个人隐私或财务敏感数据", "Agent 不可直接展示原始值，仅可表述结论", "银行账号、单价、注册资本、年营业额、手机号", "业务负责人+数据负责人双审批"),
]
for i, row_data in enumerate(defs):
    r = 5 + i
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=j, value=val)
    fill = None
    if "L3" in str(row_data[0]):
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "L2" in str(row_data[0]):
        fill = WARN_FILL
    style_row(ws4, r, len(def_headers), fill)

# Summary
ws4.cell(row=9, column=1, value="三个示例 API 字段敏感等级统计").font = Font(name="Microsoft YaHei", bold=True, size=12, color="1F4E79")
sum_headers = ["API 名称", "L1 公开", "L2 内部", "L3 机密", "合计字段", "备注"]
for i, h in enumerate(sum_headers, 1):
    ws4.cell(row=10, column=i, value=h)
style_header(ws4, 10, len(sum_headers))
sum_data = [
    ("01-核定供应商查询", 17, 10, 1, 31, "含请求参数6个字段"),
    ("02-潜在供应商详情查询", 14, 15, 10, 42, "含 bankList 子表7个字段"),
    ("03-客户订单详情查询", 14, 14, 2, 32, "含 lineList 行项目字段"),
]
for i, row_data in enumerate(sum_data):
    r = 11 + i
    for j, val in enumerate(row_data, 1):
        ws4.cell(row=r, column=j, value=val)
    style_row(ws4, r, len(sum_headers))

set_col_widths(ws4, [26, 12, 55, 40, 40, 30])


# ═══════════════════════════════════════════
# Sheet 5: 业务对象映射
# ═══════════════════════════════════════════
ws5 = wb.create_sheet("98-业务对象映射")

ws5.merge_cells("A1:G1")
ws5.cell(row=1, column=1, value="API 字段 → 平台语义 BusinessObject 映射表").font = TITLE_FONT

map_headers = ["BusinessObject", "平台对象类型", "来源 API", "来源字段 (示例)", "映射说明", "优先级", "备注"]
for i, h in enumerate(map_headers, 1):
    ws5.cell(row=3, column=i, value=h)
style_header(ws5, 3, len(map_headers))

mappings = [
    ("Supplier", "主数据对象", "核定供应商查询, 潜在供应商详情, 供应商档案", "innerVendorCode → Supplier.code", "供应商唯一编码映射", "P0", ""),
    ("Supplier", "主数据对象", "潜在供应商详情", "innerVendorName → Supplier.name", "供应商名称", "P0", ""),
    ("Supplier", "主数据对象", "潜在供应商详情", "creditCode → Supplier.creditCode", "统一社会信用代码 (L3)", "P0", "敏感字段"),
    ("Supplier", "主数据对象", "潜在供应商详情", "vendorStatus → Supplier.status", "供应商状态 0:无效 1:有效 2:停用 3:黑名单", "P0", ""),
    ("Supplier", "主数据对象", "潜在供应商详情", "lifecycleStage → Supplier.lifecycle", "生命周期阶段", "P1", ""),
    ("Material", "主数据对象", "核定供应商查询", "productCode → Material.code", "物料编码", "P0", ""),
    ("Material", "主数据对象", "核定供应商查询", "productName → Material.name", "物料名称", "P0", ""),
    ("Material", "主数据对象", "核定供应商查询", "productScale → Material.spec", "物料规格", "P0", ""),
    ("PurchaseOrder", "交易对象", "采购订单列表/详情", "poXkNo → PurchaseOrder.id", "采购单平台单号", "P0", ""),
    ("PurchaseOrder", "交易对象", "采购订单列表/详情", "orderStatus → PurchaseOrder.status", "采购单状态", "P0", "高价值:交付风险"),
    ("SalesOrder", "交易对象", "客户订单列表/详情", "poXkNo → SalesOrder.id", "销售订单平台单号", "P0", ""),
    ("SalesOrder", "交易对象", "客户订单详情", "orderStatus → SalesOrder.status", "订单状态 (13种)", "P0", "高价值:订单进度"),
    ("SalesOrderItem", "交易对象", "客户订单详情 lineList", "lineNo → SalesOrderItem.lineNo", "订单行项次", "P0", ""),
    ("SalesOrderItem", "交易对象", "客户订单详情 lineList", "deliveryDate → SalesOrderItem.deliveryDate", "交期", "P0", "高价值:交付风险"),
    ("SalesOrderItem", "交易对象", "客户订单详情 lineList", "orderQty → SalesOrderItem.qty", "订单数量", "P0", ""),
    ("Product", "主数据对象", "客户订单详情 lineList", "productCode → Product.code", "产品编码", "P1", ""),
    ("Customer", "主数据对象", "客户订单详情", "customerCode → Customer.code", "客户编码", "P0", ""),
    ("Customer", "主数据对象", "客户订单详情", "customerName → Customer.name", "客户名称", "P0", ""),
    ("Shipment", "事件/交易对象", "送货通知单", "deliveryNo → Shipment.id", "送货单号", "P1", ""),
    ("Employee", "主数据对象", "潜在供应商详情", "buyerCode → Employee.code", "采购员编码", "P1", ""),
    ("Plant", "主数据对象", "核定供应商查询", "plantCode → Plant.code", "工厂编码", "P1", ""),
    ("Material-Supplier", "关系(Relation)", "核定供应商查询", "productCode+innerVendorCode → BusinessGraphEdge", "物料-供应商供应关系", "P0", "核心关系边"),
    ("Supplier.bankList", "子对象(属性组)", "潜在供应商详情 bankList", "bankAccount → Supplier.bank.account", "银行账号 (L3)", "P0", "敏感子对象"),
]
for i, row_data in enumerate(mappings):
    r = 4 + i
    for j, val in enumerate(row_data, 1):
        ws5.cell(row=r, column=j, value=val)
    style_row(ws5, r, len(map_headers))

set_col_widths(ws5, [20, 16, 30, 36, 36, 8, 14])


# ═══════════════════════════════════════════
# Sheet 6: 试点问题清单
# ═══════════════════════════════════════════
ws6 = wb.create_sheet("99-试点问题-API覆盖")

ws6.merge_cells("A1:G1")
ws6.cell(row=1, column=1, value="试点高价值问题清单 & API 覆盖分析").font = TITLE_FONT

prob_headers = ["问题编号", "高价值问题", "问题描述", "所需业务对象", "覆盖 API", "数据可行性", "备注"]
for i, h in enumerate(prob_headers, 1):
    ws6.cell(row=3, column=i, value=h)
style_header(ws6, 3, len(prob_headers))

problems = [
    ("P-01", "订单进度查询", "当前某客户订单处于什么状态？是否已确认、已发货、已完结？",
     "SalesOrder, SalesOrderItem, Customer",
     "客户订单列表查询, 客户订单详情查询, 客户订单进度列表查询",
     "✅ 可行", "阶段1 核心场景"),
    ("P-02", "交付风险预警", "哪些订单即将到交期但未完成发货？延期天数多少？",
     "SalesOrder, SalesOrderItem, Shipment",
     "客户订单列表查询, 客户订单详情查询, 送货通知单查询",
     "✅ 可行", "需要结合当前日期计算延期"),
    ("P-03", "库存异常识别", "委外物料库存是否异常？发料和签收是否有差异？",
     "Inventory, ProductionOrder, Material",
     "委外发料单详情查询, 委外物料盘点单查询",
     "⚠️ 部分可行", "需 MES/WMS 补充在库数据"),
    ("P-04", "质量异常追踪", "哪些供应商出现质量问题？是否有准入报告和ENC报告？",
     "Supplier, ExceptionEvent",
     "潜在供应商详情, 供应商准入报告, ENC报告查询",
     "⚠️ 部分可行", "质量检测数据可能在 QMS/MES 中"),
    ("P-05", "设备异常关联", "设备异常是否影响关键供应商的物料供应？",
     "Equipment, Supplier, Material, ExceptionEvent",
     "核定供应商查询, 供应商档案查询",
     "❌ SRM不覆盖", "设备数据需从 MES/IoT 接入"),
    ("P-06", "供应商准入状态", "潜在供应商处于哪个准入阶段？资质是否齐全？",
     "Supplier",
     "潜在供应商列表/详情查询, 供应商准入报告查询",
     "✅ 可行", "阶段1 核心场景"),
    ("P-07", "物料供应稳定性", "某物料有多少核定供应商？是否有替代供应商？",
     "Material, Supplier",
     "核定供应商查询, 供应商档案查询",
     "✅ 可行", "需结合供应商状态判断"),
]
for i, row_data in enumerate(problems):
    r = 4 + i
    for j, val in enumerate(row_data, 1):
        ws6.cell(row=r, column=j, value=val)
    fill = None
    if "❌" in str(row_data[5]):
        fill = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
    elif "⚠️" in str(row_data[5]):
        fill = WARN_FILL
    elif "✅" in str(row_data[5]):
        fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    style_row(ws6, r, len(prob_headers), fill)

set_col_widths(ws6, [10, 10, 44, 36, 44, 14, 24])

# ── Save ──
output_path = "docs/数据梳理/SRM/SRM-数据梳理-打样示例.xlsx"
wb.save(output_path)
print(f"✅ Sample Excel saved to: {output_path}")
print(f"   Sheets: {wb.sheetnames}")
