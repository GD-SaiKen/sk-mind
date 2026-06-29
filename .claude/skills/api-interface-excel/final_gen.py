#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SRM采购协同 - 完整字段 最终版"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side

DEEP_BLUE=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
LIGHT_BLUE=PatternFill(start_color='D6E4F0',end_color='D6E4F0',fill_type='solid')
LIGHT_GRAY=PatternFill(start_color='F2F2F2',end_color='F2F2F2',fill_type='solid')
WHITE_B=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF')
BOLD_F=Font(name='微软雅黑',size=10,bold=True)
TITLE_F=Font(name='微软雅黑',size=11,bold=True)
NORMAL_F=Font(name='微软雅黑',size=10)
THIN_B=Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
WRAP=Alignment(wrap_text=True,vertical='center')
CENTER=Alignment(wrap_text=True,vertical='center',horizontal='center')
C_W={'A':28,'B':52,'C':14,'D':20}
OV_C_W={'A':10,'B':36,'C':60,'D':12,'E':50}

def ac(c,font=NORMAL_F,fill=None,align=WRAP,border=THIN_B):
    c.font=font;c.alignment=align;c.border=border
    if fill:c.fill=fill

def write_overview(ws,mods):
    for k,v in OV_C_W.items():ws.column_dimensions[k].width=v
    ws.merge_cells('A1:E1');c=ws['A1'];c.value='携客云SRM采购协同平台 - API接口总览';c.font=Font(name='微软雅黑',size=14,bold=True);c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        ac(ws.cell(row=2,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[2].height=22;row=3
    for mod in mods:
        ws.merge_cells(f'A{row}:E{row}')
        ac(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
        for c_ in 'BCDE':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
        ws.row_dimensions[row].height=22;row+=1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,LIGHT_GRAY,CENTER if i in(1,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    ws.auto_filter.ref=f'A2:E{row-1}'

def write_subs(ws,api):
    for k,v in C_W.items():ws.column_dimensions[k].width=v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        ac(ws.cell(row=idx,column=1,value=l),BOLD_F);ws.merge_cells(f'B{idx}:D{idx}')
        ac(ws.cell(row=idx,column=2,value=v))
        for c_ in 'CD':ac(ws[f'{c_}{idx}'])
        ws.row_dimensions[idx].height=22
    row=5;ws.row_dimensions[4].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='请求参数'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;row+=1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        ac(ws.cell(row=row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[row].height=22;row+=1
    if api.get('params'):
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,None,CENTER if i in(3,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    else:
        ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'),NORMAL_F)
        ws.row_dimensions[row].height=20;row+=1
    row+=2;ws.row_dimensions[row-1].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='响应数据字段'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;hdr_row=row+1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        ac(ws.cell(row=hdr_row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[hdr_row].height=22;row=hdr_row+1
    rl=[]
    for item in api['responses']:
        lvl=item[0];rname=('  '*lvl)+item[1]if lvl>0 else item[1];rdesc=item[2]if len(item)>2 else'';rtype=item[3]if len(item)>3 else''
        ac(ws.cell(row=row,column=1,value=rname),NORMAL_F);ac(ws.cell(row=row,column=2,value=rdesc),NORMAL_F)
        ac(ws.cell(row=row,column=3,value=''),NORMAL_F,None,CENTER)
        ac(ws.cell(row=row,column=4,value=rtype),NORMAL_F,None,CENTER)
        rl.append((row,lvl));ws.row_dimensions[row].height=20;row+=1
    if rl:
        max_lvl=max(l for _,l in rl)
        for lvl in range(max_lvl,0,-1):
            i=0
            while i<len(rl):
                r,l=rl[i]
                if l==lvl:
                    j=i+1
                    while j<len(rl)and rl[j][1]>=lvl:j+=1
                    if j-1>i:ws.row_dimensions.group(r,rl[j-1][0],outline_level=lvl)
                    i=j
                else:i+=1

# ===================== COMPLETE DATA =====================
# All 12 modules with full nested fields

# 1. 供应商管理
VM = {'name':'供应商管理','interfaces':[
{'id':'1','name':'查询已加入的供应商列表','url':'https://openapi.xiekeyun.com/compVendor/confirmList.json','method':'POST','desc':'查询邀约状态为已加入的供应商列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('handshakeTimeStart','握手查询开始时间','选填','Number'),('handshakeTimeEnd','握手查询结束时间','选填','Number'),('categoryCode01','供应商分类1编码','选填','String'),('categoryCode02','供应商分类2编码','选填','String'),('categoryCode03','供应商分类3编码','选填','String'),('performanceClassificationCode','绩效分类编码','选填','String')],
 'responses':[(0,'result','返回结果状态 1成功9失败','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商档案集合(Json数组)','Json数组'),(1,'vendorXkUniqueCode','供应商平台唯一编码','String'),(1,'innerVendorCode','供应商编码','String'),(1,'innerVendorName','供应商名称','String'),(1,'employeeName','邀约供应商的员工姓名','String'),(1,'handshakeTime','握手时间(时间戳)','Number'),(1,'categoryCode01','供应商分类1编码','String'),(1,'categoryName01','供应商分类1名称','String'),(1,'categoryCode02','供应商分类2编码','String'),(1,'categoryName02','供应商分类2名称','String'),(1,'categoryCode03','供应商分类3编码','String'),(1,'categoryName03','供应商分类3名称','String'),(1,'performanceClassificationCode','绩效分类编码','String'),(1,'performanceClassificationName','绩效分类名称','String')]},
{'id':'2','name':'查询供应商绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore.json','method':'POST','desc':'查询供应商月度绩效考评结果列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型 1月份2季度3年份','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份 01-12','选填','String'),('quarterTime','季度 1-4','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'categoryCode','分类编码','String'),(1,'categoryName','分类名称','String'),(1,'assessmentMonth','绩效月份 yyyyMM','String'),(1,'assessmentScore','绩效得分','Number'),(1,'performanceScore','绩效评分','Number'),(1,'deductionScore','绩效扣分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'vendorCategoryList','供应商分类列表','json数组'),(2,'categoryType','分类类型(0001-0004)','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String')]},
{'id':'3','name':'供应商档案详情','url':'https://openapi.xiekeyun.com/compVendor/getDetail.json','method':'POST','desc':'查询一个供应商详情明细数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码(ERP内)','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商详情','Json对象'),(1,'xkUniqueCode','供应商平台编码(32位)','String'),(1,'vendorCode','企业号','String'),(1,'vendorName','企业名称','String'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'inviteStatus','邀约状态 0未发起1已发起2已确认3失败','Number'),(1,'handshakeTime','握手成功时间(时间戳)','Number'),(1,'vendorStatus','供应商状态 0无效1有效3黑名单','Number'),(1,'buyerName','采购员名称','String'),(1,'lifeCycle','生命周期 0准入评估1试样2小批量3正式4退出','Number'),(1,'gradeInfo','供应商绩效等级','String'),(1,'contacts','联系人明细列表','Json对象数组'),(2,'fullName','姓名','String'),(2,'departmentName','部门名称','String'),(2,'phoneNo','联系电话','String'),(2,'mobile','手机号','String'),(2,'email','邮箱','String'),(2,'mainFlag','是否主要联系人 0否1是','Number'),(1,'vendorCategoryList','供应商分类/准入品类列表','Json对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(2,'status','状态 1有效2备用3暂停4退出','Number'),(1,'orgList','供应商组织关系列表','Json对象数组'),(2,'innerVendorCode','内部供应商编码(ERP)','String'),(2,'purchaseOrgCode','采购组织编码','String'),(2,'purchaseOrgName','采购组织名称','String'),(2,'currencyCode','交易币种编码','String'),(2,'currencyName','交易币种名称','String'),(2,'payWayCode','付款条件编码','String'),(2,'payWayName','付款条件名称','String'),(2,'status','有效否 0无效1有效','Number'),(2,'priceConditionList','价格条件列表','Json对象数组'),(3,'profitCenterCode','工厂编码','String'),(3,'erpPriceConditionCode','价格条件编码','String'),(3,'erpPriceConditionName','价格条件名称','String'),(1,'sqeList','SQE列表','Json对象数组'),(2,'sqeCode','SQE编码','String'),(2,'sqeName','SQE名称','String'),(1,'companyEnglishName','公司英文名称','String'),(1,'groupCompanyName','集团公司名称','String')]},
{'id':'4','name':'查询供应商应用度报表','url':'https://openapi.xiekeyun.com/compVendor/report.json','method':'POST','desc':'查询供应商应用报表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','Number'),('year','年份','必填','Number'),('month','月份','必填','Number'),('type','查询类型 1公司2组织','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商应用度报表','Json对象'),(1,'reportMonth','报表月份 yyyyMM','String'),(1,'vendorTotalSum','供应商数量','Number'),(1,'vendorHandShakeSum','已握手供应商数量','Number'),(1,'reportList','供应商应用度报表列表','Json对象数组'),(2,'companyCode','企业号','String'),(2,'innerVendorCode','内部供应商编码','String'),(2,'innerVendorName','内部供应商名称','String'),(2,'currOrderSum','本月订单数','Number'),(2,'orderHandRate','订单处理率','Number'),(2,'onlineDeliveryRate','在线送货率','Number'),(2,'enquiryJoinRate','询价参与率','Number'),(2,'biddingJoinRate','竞价参与率','Number')]},
{'id':'5','name':'查询集团法人绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/list.json','method':'POST','desc':'查询集团法人绩效考评结果列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'assessmentScore','绩效得分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'relationCompanyList','关联公司列表','json数组'),(2,'relationCompanyCode','关联企业编码','String'),(2,'relationCompanyName','关联企业名称','String'),(2,'assessmentScore','绩效总得分','Number'),(2,'gradeInfo','评级','String'),(2,'relAsDetailList','关联公司绩效明细','json数组'),(3,'curSrcScore','本项达成率(%)','Number'),(3,'curTargetScore','本项得分','Number'),(3,'assementStandardItemCode','绩效考核项编码','String'),(3,'assementStandardItemName','绩效考核项名称','String')]},
{'id':'6','name':'查询集团物料分类绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/prodCategory/list.json','method':'POST','desc':'查询集团物料分类绩效考评结果列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'prodCategoryCode','物料分类编码','String'),(1,'prodCategoryName','物料分类名称','String'),(1,'assessmentScore','绩效得分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'vendorAsList','供应商列表','json数组'),(2,'innerVendorCode','供应商编码','String'),(2,'innerVendorName','供应商名称','String'),(2,'assessmentScore','绩效总得分','Number'),(2,'gradeInfo','评级','String'),(2,'vendorDetailList','供应商绩效明细','json数组'),(3,'curSrcScore','本项达成率(%)','Number'),(3,'curTargetScore','本项得分','Number')]}]}

# 2. 生命周期管理
LC = {'name':'生命周期管理','interfaces':[
{'id':'1','name':'潜在供应商列表查询','url':'https://openapi.xiekeyun.com/potentialVendor/getList.json','method':'POST','desc':'查询企业潜在供应商列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('dateStart','资质审核完成时间-从','选填','Number'),('dateEnd','资质审核完成时间-至','选填','Number'),('createStart','创建时间-从','选填','Number'),('createEnd','创建时间-至','选填','Number'),('checkTimeStart','审核时间-从','选填','Number'),('checkTimeEnd','审核时间-至','选填','Number'),('applyStatus','申请状态','选填','Number数组'),('inviteStatus','邀约状态','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','潜在供应商数据列表','Json对象'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'admissionReportFlag','完成准入评估标识 0未完成1已完成','Number'),(1,'applyStatus','申请状态','Number'),(1,'vendorStatus','供应商状态 0无效1有效2停用3黑名单','Number'),(1,'inviteStatus','邀约状态','Number'),(1,'checkTime','审核时间','Number'),(1,'updateTime','最后更新时间','Number')]},
{'id':'2','name':'潜在供应商详情查询','url':'https://openapi.xiekeyun.com/potentialVendor/getDetail.json','method':'POST','desc':'查询一个潜在供应商明细数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','潜在供应商编码','必填','String'),('approvalStatusFlag','是否返回审批状态','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','潜在供应商详情','Json对象'),(1,'xkUniqueCode','内部供应商平台编码','String'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'applyStatus','审批状态','number'),(1,'vendorStatus','供应商状态','number'),(1,'inviteStatus','邀约状态','number'),(1,'contacts','联系人明细列表','Json对象数组'),(2,'fullName','姓名','String'),(2,'phoneNo','联系电话','String'),(2,'mobile','手机号','String'),(2,'email','邮箱','String'),(2,'mainFlag','是否主要联系人','Number'),(1,'purchaseOrgList','采购组织明细列表','Json对象数组'),(2,'purchaseOrgCode','采购组织编码','String'),(2,'purchaseOrgName','采购组织名称','String'),(1,'financeList','企业资料明细列表','Json对象数组'),(2,'erpCompanyCode','公司代码编码','String'),(2,'payWayCode','付款条件编码','String'),(1,'bankList','银行明细列表','Json对象数组'),(2,'bankCardNo','银行帐号','String'),(1,'vendorCategoryList','供应商分类/准入品类列表','Json对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(1,'companyEnglishName','公司英文名称','String'),(1,'groupCompanyName','集团公司名称','String')]},
{'id':'3','name':'样品管理列表查询','url':'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getList.json','method':'POST','desc':'查询样品管理列表信息','params':[
 ('startDate','创建时间开始','必填','Number'),('endDate','创建时间结束','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','状态过滤','选填','Integer[]')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','样品列表','Json数组'),(1,'stkXkNo','样品单号','String'),(1,'dnXkNo','送样单号','String'),(1,'sampleDate','样品日期','Number'),(1,'status','状态','Number'),(1,'cableSampleTypeCode','索样类型编码','String'),(1,'cableSampleTypeName','索样类型名称','String')]},
{'id':'4','name':'样品单详情查询','url':'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getDetailByStkXkNo.json','method':'POST','desc':'查询样品单详情信息','params':[
 ('stkXkNo','样品单号','必填','String'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','样品单数据','Json对象'),(1,'stkXkNo','样品单号','String'),(1,'dnXkNo','送样单号','String'),(1,'sampleType','样品类别 1一般物料2工装','Number'),(1,'sampleDate','样品日期','Number'),(1,'status','状态 0待签收1待检验2待试产3待判定4已完成','Number'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'dnLine','送货明细列表','Json对象数组'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'deliveryQty','出货数量','Number'),(1,'verifies','检验列表','Json对象数组'),(2,'standardName','检验指标名称','String'),(2,'resultDesc','检验结果','String'),(1,'produces','试产工艺列表','Json对象数组'),(2,'paramName','工艺参数名称','String'),(2,'paramValue','工艺参数取值','String')]},
{'id':'5','name':'PCN报告列表查询','url':'https://openapi.xiekeyun.com/ecn/getList.json','method':'POST','desc':'查询PCN报告列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','创建开始时间','必填','Number'),('endDate','创建结束时间','必填','Number'),('statusArray','状态过滤','选填','Number数组'),('changeTypeArray','变更类型过滤','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','PCN报告列表','JSON数组'),(1,'reportXkNo','PCN报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'reportTitle','PCN报告名称','String'),(1,'reportStatus','PCN报告状态','Number'),(1,'changeDate','变更日期','Number'),(1,'reportPublishTime','报告发布日期','Number'),(1,'changeType','变更类型','Number')]},
{'id':'6','name':'PCN报告详情查询','url':'https://openapi.xiekeyun.com/ecn/getDetail.json','method':'POST','desc':'查询PCN报告详情数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('reportXkNo','PCN报告平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','PCN报告详情','JSON对象'),(1,'reportXkNo','PCN报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'reportTitle','PCN报告名称','String'),(1,'reportStatus','PCN报告状态','Number'),(1,'changeType','变更类型','Number'),(1,'contentList','报告内容信息列表','JSON对象数组'),(2,'changeType','变更类型','String'),(2,'changeOption','变更项','String'),(2,'changeBefore','变更前','String'),(2,'changeAfter','变更后','String'),(2,'changeReason','变更原因','String'),(1,'productList','报告物料信息列表','JSON对象数组'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'brandName','品牌名称','String')]},
{'id':'7','name':'供应商准入报告列表查询','url':'https://openapi.xiekeyun.com/admissionReport/getList.json','method':'POST','desc':'查询供应商准入报告列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('statusArray','状态过滤','选填','Number数组'),('workflowStatus','审批状态','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商准入报告列表','JSON数组'),(1,'admissionReportXkNo','准入报告平台编码','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','供应商内部名称','String'),(1,'assessmentType','评估类型','Number'),(1,'assessmentDate','评估日期','Number'),(1,'status','报告状态','Number'),(1,'inspectionResult','评审结果','Number')]},
{'id':'8','name':'供应商准入报告详情查询','url':'https://openapi.xiekeyun.com/admissionReport/getDetail.json','method':'POST','desc':'查询供应商准入报告详情数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('admissionReportXkNo','准入报告平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商准入报告详情','JSON对象'),(1,'admissionReportXkNo','准入报告平台编码','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'assessmentType','评估类型','Number'),(1,'status','报告状态','Number'),(1,'inspectionResult','评审结果','Number'),(1,'grade','评分等级','String'),(1,'detailList','评审项明细信息列表','JSON对象数组'),(2,'itemName','评审项目名称','String'),(2,'ownScore','自评分','Number'),(2,'inspectionScore','评审得分','Number'),(1,'vendorCategoryList','供应商分类','JSON对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(1,'prodCategoryList','准入品类列表','JSON对象数组'),(2,'categoryType','分类类型(物料ERP分类)','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String')]}]}

# 3. 内部商城协同
IM = {'name':'内部商城协同','interfaces':[
{'id':'1','name':'内部商城引入单列表查询','url':'https://openapi.xiekeyun.com/introduceOrder/getList.json','method':'POST','desc':'查询内部商城引入单列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','引入时间开始','必填','Number'),('endDate','引入时间结束','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','内部商城引入单列表','JSON数组'),(1,'inductXkNo','平台单据号','String'),(1,'status','引入单状态','Number')]},
{'id':'2','name':'内部商城引入单详情查询','url':'https://openapi.xiekeyun.com/introduceOrder/getDetail.json','method':'POST','desc':'查询内部商城引入单详情数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('inductXkNo','引入单平台单据号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','引入单详情','JSON对象'),(1,'inductXkNo','平台单据号','String'),(1,'status','引入单状态','Number'),(1,'currencyCode','币别编码','String'),(1,'currencyName','币别名称','String'),(1,'sourceSystem','来源','Number'),(1,'applyCompList','适用公司列表','JSON对象数组'),(2,'applyCompanyName','适用公司名称','String'),(1,'lineList','明细信息列表','JSON对象数组'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'price','未税价','Number'),(2,'taxPrice','含税价','Number'),(1,'userInfo','引入用户信息','JSON对象'),(2,'employeeName','员工姓名','String'),(2,'erpAccount','ERP账号','String')]}]}

# 4. 采购订单协同
PO = {'name':'采购订单协同','interfaces':[
{'id':'1','name':'请购单列表查询','url':'https://openapi.xiekeyun.com/requisition/getList.json','method':'POST','desc':'查找请购单列表信息','params':[
 ('startDate','请购时间开始','必填','Number'),('endDate','请购时间结束','必填','Number'),('dateType','查询时间类型','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','请购单状态','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','请购单列表','JSON数组'),(1,'reqXkNo','请购单平台编号','String'),(1,'reqInnerNo','请购单ERP单据号','String'),(1,'status','单据状态','Number')]},
{'id':'2','name':'请购单详情查询','url':'https://openapi.xiekeyun.com/requisition/getDetail.json','method':'POST','desc':'查询请购单详情信息','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('reqXkNo','请购单平台编号','必填','String'),('reqErpNo','请购单ERP单号','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','请购单详情','JSON对象'),(1,'reqXkNo','平台单据号','String'),(1,'reqInnerNo','ERP单据号','String'),(1,'status','单据状态','Number'),(1,'departmentName','请购部门名称','String'),(1,'reqDate','请购日期','Number'),(1,'lineList','单身列表','Json对象数组'),(2,'lineNo','项次号','String'),(2,'productCode','产品编码','String'),(2,'productName','物料名称','String'),(2,'reqQty','请购数量','Number'),(2,'demandDate','需求日期','Number'),(2,'estimatePrice','预估单价','Number'),(2,'estimateAmount','预估金额','Number'),(2,'brand','品牌','String'),(2,'remark','备注','String'),(2,'buyerList','负责采购员列表','Json对象数组'),(3,'employeeName','员工名字','String'),(3,'erpEmployeeCode','ERP员工编码','String'),(1,'infoList','业务操作列表','Json对象数组'),(2,'poXkNo','平台采购单号','String'),(2,'purchaseQty','累计转采购数量','Number'),(2,'expectedDate','采购交期','Number'),(1,'empList','协作人列表','Json对象数组'),(2,'employeeName','员工名字','String')]},
{'id':'3','name':'查询请购未转采购预估金额','url':'https://openapi.xiekeyun.com/requisition/sumRequisitionAmount.json','method':'POST','desc':'根据WBS编码统计请购未转采购预估金额','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('wbsCode','WBS编码','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','汇总数据','JSON对象'),(1,'wbsCode','Wbs编码','String'),(1,'estimateAmount','请购预估金额合计','Number'),(1,'poAmount','未同步采购单金额汇总','Number'),(1,'syncPoAmount','已同步采购单金额汇总','Number')]},
{'id':'4','name':'采购订单列表查询','url':'https://openapi.xiekeyun.com/purchaseOrder/getList.json','method':'POST','desc':'查询已答交采购单号列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('status','订单状态','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','采购订单列表','JSON数组'),(1,'poErpNo','采购单ERP单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'orderStatus','订单状态','Number'),(1,'poCreateDate','订单推到平台时间','Number'),(1,'answerDate','答交时间','Number'),(1,'vStatus','答交状态 1无差异2有差异','Number')]},
{'id':'5','name':'采购订单详情查询','url':'https://openapi.xiekeyun.com/purchaseOrder/getDetail.json','method':'POST','desc':'查询采购单答交明细','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('poXkNo','采购订单平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','采购订单详情','JSON对象'),(1,'poXkNo','采购单平台单号','String'),(1,'poErpNo','采购单ERP单号','String'),(1,'purchaseType','采购类型','Number'),(1,'innerVendorCode','供应商编码','String'),(1,'innerVendorName','供应商名称','String'),(1,'orderStatus','订单状态','Number'),(1,'currencyCode','币种编码','String'),(1,'currencyName','币种名称','String'),(1,'payWayCode','付款条件编码','String'),(1,'payWayName','付款条件名称','String'),(1,'totalAmount','原币单据总合计','Number'),(1,'lineList','产品明细列表','Json对象数组'),(2,'lineNo','项次号','String'),(2,'productCode','物料编码','String'),(2,'productName','产品名称','String'),(2,'purchaseQty','采购数量','Number'),(2,'price','无税单价','Number'),(2,'taxPrice','含税单价','Number'),(2,'expectedDate','交货日期','Number'),(2,'remark','备注','String'),(2,'isFreebie','是否赠品','Number'),(2,'brandName','品牌','String'),(2,'warrantyPeriod','保修期(天)','Number'),(2,'manufacturerProdCode','制造商料号','String'),(2,'prcLineList','核价单明细列表','Json对象数组'),(3,'prcXkNo','平台单据号','String'),(3,'prcErpNo','ERP单据号','String'),(3,'lineNo','项次号','String'),(2,'reqLineList','请购单单身列表','Json对象数组'),(3,'sourceInnerNo','ERP请购单号','String'),(3,'sourceReqXkNo','请购单单号','String'),(3,'demandDate','需求日期','Number'),(2,'lowestPriceInfo','集团最低价信息','Json对象'),(3,'groupMinPrice','集团最低核价单价','Number'),(3,'historyMinPrice','历史低价','Number'),(3,'lastPrcXkNo','最近核价单号','String'),(1,'empList','协作人列表','Json对象数组'),(2,'employeeName','员工名字','String'),(1,'answer','答交数据','Json对象'),(2,'answerDate','答交时间','Number'),(2,'answerEmpName','答交人姓名','String'),(2,'vStatus','答交状态','Number'),(2,'lineList','答交明细列表','Json对象数组'),(3,'lineNo','项次号','String'),(3,'vPurchaseQty','答复的数量','Number'),(3,'vPrice','答复的单价','Number'),(1,'afterSaleOrderList','售后单列表','Json对象数组'),(2,'asoXkNo','售后单号','String'),(2,'status','售后状态','Number')]},
{'id':'6','name':'采购变更单列表查询','url':'https://openapi.xiekeyun.com/changeOrder/getList.json','method':'POST','desc':'查询采购变更单列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','采购变更单列表','JSON数组'),(1,'pocXkNo','变更单平台单据号','String'),(1,'poErpNo','采购单ERP单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'changeSeq','变更版序','String')]},
{'id':'7','name':'采购变更单详情查询','url':'https://openapi.xiekeyun.com/changeOrder/getDetail.json','method':'POST','desc':'查询采购变更单详情','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('changeXkNo','变更单平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','采购变更单详情','JSON对象'),(1,'pocXkNo','变更单平台单据号','String'),(1,'changeSeq','变更版序','Number'),(1,'status','变更单状态 1已确认2已结案3已冻结4差异待确认5退回','Number'),(1,'poErpNo','采购单ERP单据号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'lineList','产品明细列表','Json对象数组'),(2,'lineNo','变更单项次','String'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'purchaseQty','采购数量','Number'),(2,'price','未税单价','Number'),(2,'taxPrice','含税单价','Number'),(2,'changedPurchaseQty','变更后采购数量','Number'),(2,'changedPrice','变更后未税单价','Number'),(2,'changedFlag','是否有变更 0未变1有变','Number')]},
{'id':'8','name':'交料通知单查询','url':'https://openapi.xiekeyun.com/acNotifyOrder/getList.json','method':'POST','desc':'查询交料通知单答交列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','交料通知单列表','JSON数组'),(1,'acnXkNo','平台单号','String'),(1,'acnInnerNo','ERP单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','通知单状态','Number'),(1,'answerDate','答交时间','Number')]},
{'id':'9','name':'索样通知单查询','url':'https://openapi.xiekeyun.com/cableSampleOrder/getList.json','method':'POST','desc':'查询索样通知单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','索样通知单列表','JSON数组'),(1,'ncsoXkNo','平台单据号','String'),(1,'ncsoErpNo','ERP单据号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'status','单据状态','Number')]}]}

# 5. 采购收货协同
RC = {'name':'采购收货协同','interfaces':[
{'id':'1','name':'送货单号列表查询','url':'https://openapi.xiekeyun.com/delivery/getNoList.json','method':'POST','desc':'按时间范围拉取新送货单号列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('profitCenterCode','收货部门编码','选填','String'),('status','状态过滤','选填','number数组'),('logisticsStatus','物流状态','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','送货单号数据列表','JSON数组'),(1,'deliveryNo','送货单号','String'),(1,'deliveryType','送货类型','Number'),(1,'innerVendorCode','内部供应商编码','String'),(1,'purchaseType','采购类型','Number'),(1,'status','状态','String'),(1,'logisticsStatus','物流状态','String'),(1,'lastOperateTime','最后变更时间','String'),(1,'buyerName','采购员名称','String'),(1,'deliveryDate','送货日期','String')]},
{'id':'2','name':'送货单详情查询','url':'https://openapi.xiekeyun.com/delivery/getDetail.json','method':'POST','desc':'查询一个送货单明细数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号或物流单号','必填','String'),('returnVerifyFlag','是否返回送检统计数量','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','送货单数据','Json对象'),(1,'deliveryNo','送货单号','String'),(1,'innerVendorCode','供应商编码','String'),(1,'innerVendorName','供应商名称','String'),(1,'purchaseType','采购类型','Number'),(1,'deliveryType','送货类型','Number'),(1,'deliveryDate','送货日期','Number'),(1,'status','送货单状态','Number'),(1,'logisticsStatus','物流状态','Number'),(1,'lineList','送货单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'deliveryQty','出货数量','Number'),(2,'price','单价','Number'),(2,'taxPrice','含税单价','Number'),(2,'receiveQty','收货数量','Number'),(2,'storageQty','入库数量','Number'),(1,'customsData','报关资料','Json对象'),(2,'originalCountryName','原产国名称','String'),(2,'customsTypeName','报关类型名称','String')]},
{'id':'3','name':'条码查询','url':'https://openapi.xiekeyun.com/barcode/byDeliveryNo.json','method':'POST','desc':'查询送货单包含的所有条码数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','条码数据','JsoN对象数组'),(1,'productCode','产品编码','String'),(1,'smallBarcode','小包条码','String'),(1,'bigBarcode','大包条码','String'),(1,'outerBarcode','外箱条码','String'),(1,'includeQty','包含数量','Number'),(1,'packLevel','包装层级','Number'),(1,'dynamicData','动态字段值','MAP'),(0,'packageList','条码数量信息','JsoN对象数组'),(1,'productCode','物料编码','String'),(1,'smallPackageCount','小包条码数量','Number'),(0,'dynamicDescList','条码规则描述','Json对象数组'),(1,'field','字段名称','String'),(1,'text','字段文本描述','String')]},
{'id':'4','name':'送货预约看板查询','url':'https://openapi.xiekeyun.com/reserves/dayInfo.json','method':'POST','desc':'查询一天所有供应商送货预约数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('planArrivedDate','预计到厂时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','看板数据','Json对象'),(1,'reserveRule','规则模板','Json对象'),(2,'reserveStep','预约时距','String'),(2,'timeList','可预约时段','Json数据'),(3,'timeSlot','时段名称','String'),(2,'areaList','可预约收货区域','Json数据'),(3,'areaName','收货区名称','String'),(1,'reservesList','预约点亮数据','Json对象数组'),(2,'dnXkNo','送货通知单号','String'),(2,'planArrivedDate','预计到货日期','Number'),(2,'deliveryNotice','送货通知数据','Json数据'),(3,'innerVendorName','内部供应商名称','String')]},
{'id':'5','name':'送货单供应商填写的检验数据查询','url':'https://openapi.xiekeyun.com/delivery/getDeliveryVendorWriteStandard.json','method':'POST','desc':'查询送货单供应商填写的检验数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String'),('standardDesc','检测说明','选填','String'),('rangStart','合格范围开始值','选填','String'),('rangEnd','合格范围结束值','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商填写检验标准列表','JsoN数组'),(1,'dnXkNo','送货单号','String'),(1,'lineNo','送货单项次号','Number'),(1,'productCode','产品编码','String'),(1,'productName','产品名称','String'),(1,'standardName','检验指标名称','String'),(1,'standardValue','检测值','String')]},
{'id':'6','name':'客供料代收单条码查询','url':'https://openapi.xiekeyun.com/signforOrder/barcode/byDeliveryNo.json','method':'POST','desc':'查询客供料代收单包含的所有条码数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','条码数据','JsoN对象数组'),(1,'productCode','产品编码','String'),(1,'smallBarcode','小包条码','String'),(0,'packageList','条码数量信息','JsoN对象数组'),(0,'dynamicDescList','条码规则描述','Json对象数组')]},
{'id':'7','name':'送货单待签收列表查询','url':'https://openapi.xiekeyun.com/delivery/getWaitSignList.json','method':'POST','desc':'按送货时间范围拉取待签收送货单列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','待签收送货单数据列表','JSON数组'),(1,'dnXkNo','送货单号','String'),(1,'lineNo','送货单项次','Number'),(1,'poErpNo','订单ERP单号','String'),(1,'deliveryQty','送货数量','Number'),(1,'innerVendorCode','内部供应商编码','String'),(1,'productCode','物料编码','String')]},
{'id':'8','name':'送样通知单号列表查询','url':'https://openapi.xiekeyun.com/sampleDelivery/getList.json','method':'POST','desc':'按时间范围拉取新送样通知单号列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','送样通知单号数据列表','JSON数组'),(1,'deliveryNo','送样通知单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','状态','String'),(1,'logisticsStatus','物流状态','String')]},
{'id':'9','name':'送样通知单详情查询','url':'https://openapi.xiekeyun.com/sampleDelivery/getDetail.json','method':'POST','desc':'查询一个送样通知单明细数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送样通知单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','送样通知单数据','Json对象'),(1,'deliveryNo','送样通知单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','状态','String'),(1,'lineList','送样明细列表','Json对象数组'),(2,'productCode','产品编码','String'),(2,'deliveryQty','出货数量','Number')]}]}

# 6. 来料检验协同
IK = {'name':'来料检验协同','interfaces':[
{'id':'1','name':'检验单列表查询','url':'https://openapi.xiekeyun.com/verifyOrder/getList.json','method':'POST','desc':'查询检验单列表信息','params':[
 ('startDate','创建开始时间','必填','Number'),('endDate','创建结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','检验状态','选填','Integer[]'),('innerVendorCode','供应商编码','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','检验单列表','JSON数组'),(1,'vfoXkNo','检验单平台单号','String'),(1,'sourceOrderType','来源单据 1送货单2收货单','Number'),(1,'resultFlag','检验结果 1全部合格2部分合格3拒收','Number'),(1,'status','检验状态 0检验中1待确认2退回3已确认4已作废','Number')]},
{'id':'2','name':'检验单详情查询','url':'https://openapi.xiekeyun.com/verifyOrder/getDetail.json','method':'POST','desc':'查询检验单详情信息','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('verifyXkNo','检验单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','检验单数据','Json对象'),(1,'vfoXkNo','检验单据号','String'),(1,'vfoInnerNo','检验单内部单据号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'resultFlag','检验结果标识','Number'),(1,'status','检验状态','Number'),(1,'dnXkNo','送货单号','String'),(1,'lineList','检测项次列表','Json对象数组'),(2,'lineNo','检验单项次','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'verifyQty','送检数量','Number'),(2,'allowQty','允收数量','Number'),(2,'refusedQty','拒收数量','Number'),(2,'batchDetermine','批次判定 0合格1不合格2免检3让步接收','Number'),(2,'detailList','不良记录详情列表','Json对象数组'),(3,'reasonCode','不良原因编码','String'),(3,'curQty','不良数量','Number'),(3,'operateType','收货判定 1让步接收2拒收','Number'),(3,'processMode','处理方式编码','String'),(2,'standardRecordList','检测项目列表','Json对象数组'),(3,'standardName','检验项目','String'),(3,'rangStart','开始合格范围','String'),(3,'rangEnd','结束合格范围','String'),(3,'standardValue','检验值','String'),(3,'standardResult','检验判断 1合格2不合格','Number')]},
{'id':'3','name':'8D报告列表查询','url':'https://openapi.xiekeyun.com/improveReport/getList.json','method':'POST','desc':'查询8D报告列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','8D报告列表','JSON数组'),(1,'p8dXkNo','8D报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','8D报告状态','Number'),(1,'reportResult','报告结果 0不合格1合格','Number'),(1,'createTime','创建时间','Number'),(1,'productCode','物料编码','String')]},
{'id':'4','name':'8D报告详情查询','url':'https://openapi.xiekeyun.com/improveReport/getDetail.json','method':'POST','desc':'查询8D报告详情数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('irXkNo','8D报告编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','8D报告详情','JSON对象'),(1,'p8dXkNo','8D报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'status','8D报告状态','Number'),(1,'reportResult','报告结果','Number'),(1,'productCode','物料编码','String'),(1,'productName','物料名称','String'),(1,'badReason','不良原因','String'),(1,'detailInfoList','8D报告项明细列表','JSON对象数组'),(2,'itemCode','报告项编码','String'),(2,'itemName','报告项名称','String'),(2,'status','状态 0未开始1进行中2已完成','Number'),(2,'headName','负责人姓名','String'),(1,'reportTraceList','批次检验信息列表','JSON对象数组'),(2,'vfoXkNo','平台检验单据号','String'),(2,'verifyTime','检验时间','Number'),(2,'resultFlag','检验结果','Number'),(2,'badQty','不良数量','Number')]}]}

# 7. 供应计划协同
SP = {'name':'供应计划协同','interfaces':[
{'id':'1','name':'供应计划排程列表查询','url':'https://openapi.xiekeyun.com/schedule/getList.json','method':'POST','desc':'查询排程单号列表','params':[
 ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('statusList','状态过滤','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','排程单号数据列表','JSON数组'),(1,'batchNo','排程批次号','String'),(1,'oldBatchNo','父排程单号','String'),(1,'status','状态 1已发布2部分答交3有差异4全部答交5已失效','Number'),(1,'publishName','发布人员工姓名','String'),(1,'publishTime','发布时间','Number'),(1,'schedulePlanType','计划类型','Number')]},
{'id':'2','name':'供应计划排程详情查询','url':'https://openapi.xiekeyun.com/schedule/getDetail.json','method':'POST','desc':'查询排程详情信息','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('scheduleXkNo','排程单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','排程详情','Json对象'),(1,'batchNo','排程号','String'),(1,'status','排程状态','Number'),(1,'publishName','发布人','String'),(1,'publishTime','发布时间','Number'),(1,'itemList','排程明细列表','Json对象数组'),(2,'itemId','排程物料唯一编码','String'),(2,'prodCode','物料编码','String'),(2,'prodName','物料名称','String'),(2,'innerVendorCode','供应商编码','String'),(2,'status','物料状态','Number'),(2,'detailList','排程时间明细','Json数组'),(3,'scheduleDate','排程日期','Number'),(3,'scheduleQty','排程量','Number'),(3,'replyQty','答交量','Number')]},
{'id':'3','name':'生产计划排程查询','url':'https://openapi.xiekeyun.com/productSchedule/getList.json','method':'POST','desc':'查询生产计划排程数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','明细数据','Json对象'),(1,'batchNo','排程号','String'),(1,'status','状态','Number'),(1,'itemList','排程明细列表','Json对象数组'),(2,'prodCode','物料编码','String'),(2,'prodName','物料名称','String'),(2,'detailList','排程时间明细','Json数组'),(3,'scheduleDate','排程日期','Number'),(3,'scheduleQty','排程量','Number')]},
{'id':'4','name':'供应计划看板查询','url':'https://openapi.xiekeyun.com/receiveboard/getList.json','method':'POST','desc':'查询供应计划看板数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','数据列表','JSON数组'),(1,'productCode','物料编码','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'itemList','天维度数据明细','JSON数组'),(2,'boardDate','看板日期','Number'),(2,'planQty','计划数量','Number'),(2,'answerQty','答交数量','Number'),(2,'deliveriedQty','累计已出货量','Number'),(2,'receivedQty','累计签收数量','Number'),(2,'detailList','看板明细列表','JSON数组'),(3,'boardTime','看板日期含时分秒','Number'),(3,'curQty','答交数量','Number'),(3,'billType','单据类型 1采购单2排程单','Number'),(2,'deliveryList','送货明细列表','JSON数组'),(3,'dnXkNo','送货通知单号码','String'),(3,'curQty','送货数量','Number'),(2,'receiveList','收货入库明细列表','JSON数组'),(3,'billXkNo','平台单据号','String'),(3,'signQty','收货数量','Number')]},
{'id':'5','name':'直运看板查询','url':'https://openapi.xiekeyun.com/directDeliveryBoard/getList.json','method':'POST','desc':'查询直运看板数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','直运看板列表','JSON数组'),(1,'poErpNo','采购单ERP单号','String'),(1,'productCode','物料编码','String'),(1,'productName','物料名称','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'directType','直送类型','Number'),(1,'deliveryQty','送货数量','Number'),(1,'status','状态','Number'),(1,'deliveryAddress','送货详细地址','String')]},
{'id':'6','name':'供应商库存看板查询','url':'https://openapi.xiekeyun.com/prodStockBoard/getList.json','method':'POST','desc':'查询供应商库存看板数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','数据列表','JSON数组'),(1,'prodCode','物料编码','String'),(1,'prodName','物料名称','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'stockQty','库存数量','Number'),(1,'planReceiveQty','预计收货','Number'),(1,'planQty','预计产出数量','Number'),(1,'produceGoodQty','产出良品','Number'),(1,'itemList','天维度数据明细','JSON数组'),(2,'boardDate','看板日期','Number'),(2,'planQty','预计生产产出数量','Number'),(2,'planReceiveQty','预计收货','Number'),(2,'produceGoodQty','产出良品','Number')]}]}

# 8. 应付账款协同
AP = {'name':'应付账款协同','interfaces':[
{'id':'1','name':'对账单列表查询','url':'https://openapi.xiekeyun.com/vendorStatementOrder/getList.json','method':'POST','desc':'查询对账单列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','对账单列表','JSON数组'),(1,'recoXkNo','对账单平台单号','String'),(1,'status','单据状态','Number'),(1,'innerVendorCode','内部供应商编码','String'),(1,'esignStatus','签章状态','Number')]},
{'id':'2','name':'对账单详情查询','url':'https://openapi.xiekeyun.com/vendorStatementOrder/getDetail.json','method':'POST','desc':'查询对账单详情数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('statementXkNo','对账单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','对账单单数据','Json对象'),(1,'recoXkNo','对账单平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'totalAmount','对账总金额','Number'),(1,'totalTaxAmount','对账含税金额','Number'),(1,'status','单据状态','Number'),(1,'invoiceStatus','开票状态','Number'),(1,'deliveryList','出货明细列表','Json对象数组'),(2,'productCode','产品编码','String'),(2,'deliveryQty','出货数量','Number'),(2,'deliveryAmount','出货未税金额','Number'),(2,'curReconAmount','本次对账未税金额','Number'),(1,'returnList','仓退明细列表','Json对象数组'),(2,'rtoErpNo','退货单ERP单号','String'),(2,'returnQty','仓退数量','Number'),(2,'returnAmount','仓退未税金额','Number'),(1,'costList','费用明细列表','Json对象数组'),(2,'csoXkNo','费用单平台单号','String'),(2,'costAmount','费用未税金额','Number'),(1,'empList','协作人列表','Json对象数组'),(2,'employeeName','员工名字','String')]},
{'id':'3','name':'发票单列表查询','url':'https://openapi.xiekeyun.com/invoiceOrder/getList.json','method':'POST','desc':'查询发票单列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','发票单列表','JSON数组'),(1,'invoiceNo','发票单平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','发票状态','Number'),(1,'publishTime','发布时间','Number')]},
{'id':'4','name':'费用单列表查询','url':'https://openapi.xiekeyun.com/costOrder/getList.json','method':'POST','desc':'查询费用单列表数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','费用单列表','JSON数组'),(1,'costXkNo','费用单平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','状态','Number')]},
{'id':'5','name':'付款申请单查询','url':'https://openapi.xiekeyun.com/paymentApplyOrder/getList.json','method':'POST','desc':'查询付款申请单数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','付款申请单列表','JSON数组'),(1,'pmaXkNo','付款申请单平台单号','String'),(1,'status','单据状态','Number')]},
{'id':'6','name':'预付款单查询','url':'https://openapi.xiekeyun.com/prepaymentOrder/getList.json','method':'POST','desc':'查询预付款单数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','预付款单列表','JSON数组'),(1,'ppmXkNo','预付款单平台单号','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'prepayTotalAmount','预付总金额','Number'),(1,'status','状态','Number')]},
{'id':'7','name':'付款单查询','url':'https://openapi.xiekeyun.com/receipt/getList.json','method':'POST','desc':'查询付款单数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','付款单列表','JSON数组'),(1,'rcpXkNo','付款单平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'status','状态','Number'),(1,'payStatus','支付状态','Number')]},
{'id':'8','name':'折让单查询','url':'https://openapi.xiekeyun.com/vendorStatementDiscountOrder/getList.json','method':'POST','desc':'查询折让单数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','折让单列表','JSON数组'),(1,'disXkNo','折让单平台单号','String'),(1,'status','单据状态','Number')]},
{'id':'9','name':'付款计划查询','url':'https://openapi.xiekeyun.com/paymentPlanOrder/getList.json','method':'POST','desc':'查询付款计划数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','付款计划列表','JSON数组'),(1,'pmpXkNo','付款计划平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'planTaxAmount','排定付款金额','Number'),(1,'planDate','排定付款日期','Number'),(1,'payStatus','付款状态','Number')]}]}

# 9. 询价招标竞价
IQ = {'name':'询价招标竞价','interfaces':[
{'id':'1','name':'询价需求单列表查询','url':'https://openapi.xiekeyun.com/inquiryInnerOrder/getList.json','method':'POST','desc':'查询询价需求单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','询价需求单列表','JSON数组'),(1,'inqXkNo','询价需求单平台单号','String'),(1,'status','单据状态','Number')]},
{'id':'2','name':'询价需求单详情查询','url':'https://openapi.xiekeyun.com/inquiryInnerOrder/getDetail.json','method':'POST','desc':'查询询价需求单详情','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('inquiryXkNo','询价需求单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','询价需求单详情','JSON对象'),(1,'inqXkNo','平台单据号','String'),(1,'billType','类型 1询价2招标','Number'),(1,'status','单据状态','Number'),(1,'publishName','发布人名称','String'),(1,'publishTime','发布时间','Number'),(1,'prodList','产品编码列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'purchaseQty','采购数量','Number'),(2,'newFlag','是否新产品','Number'),(2,'vquopList','供应商报价信息列表','Json对象数组'),(3,'status','报价状态','Number'),(3,'promiseDeliDays','承诺交期(天)','Number'),(3,'selectedType','中标类型','Number'),(3,'detailList','阶梯报价信息','Json对象数组'),(4,'thresholdStartQty','开始阈值','Number'),(4,'thresholdEndQty','结束阈值','Number'),(4,'quotePrice','单价','Number'),(1,'vendorList','询价企业列表','Json对象数组'),(2,'innerVendorName','询价企业名称','String')]},
{'id':'3','name':'询价招标单列表查询','url':'https://openapi.xiekeyun.com/inquiryOrder/getList.json','method':'POST','desc':'查询询价招标单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','询价招标单列表','JSON数组'),(1,'inqXkNo','询价招标平台单号','String'),(1,'status','单据状态','Number')]},
{'id':'4','name':'竞价单查询','url':'https://openapi.xiekeyun.com/bidOrder/getList.json','method':'POST','desc':'查询竞价单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','竞价单列表','JSON数组'),(1,'bidXkNo','竞价单平台单号','String'),(1,'status','单据状态','Number'),(1,'publishName','发布人姓名','String'),(1,'publishTime','发布时间','Number')]},
{'id':'5','name':'项目预询单查询','url':'https://openapi.xiekeyun.com/preProjectOrder/getList.json','method':'POST','desc':'查询项目预询单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','项目预询单列表','JSON数组'),(1,'ppbXkNo','预询单平台编码','String'),(1,'status','状态','Number')]},
{'id':'6','name':'项目招标单查询','url':'https://openapi.xiekeyun.com/projectOrder/getList.json','method':'POST','desc':'查询项目招标单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','项目招标单列表','JSON数组'),(1,'PbXkNo','项目招标单平台单号','String'),(1,'status','单据状态','Number')]},
{'id':'7','name':'核价单查询','url':'https://openapi.xiekeyun.com/pricingOrder/getList.json','method':'POST','desc':'查询核价单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','核价单列表','JSON数组'),(1,'prcXkNo','核价单平台单号','String'),(1,'prcErpNo','核价单ERP单据号','String'),(1,'status','单据状态','Number'),(1,'publishTime','发布时间','Number')]},
{'id':'8','name':'核价失效单查询','url':'https://openapi.xiekeyun.com/pricingInvalidOrder/getList.json','method':'POST','desc':'查询核价失效单列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','核价失效单列表','JSON数组'),(1,'priXkNo','核价失效单平台单号','String'),(1,'priErpNo','核价失效单ERP单据号','String'),(1,'status','单据状态','Number')]}]}

# 10. 委外物料管理
OS = {'name':'委外物料管理','interfaces':[
{'id':'1','name':'委外发料单详情查询','url':'https://openapi.xiekeyun.com/outsourceIssuanceOrder/getDetail.json','method':'POST','desc':'查询委外发料单详情','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('oioXkNo','委外发料单平台单号','选填','String'),('oioErpNo','委外发料单ERP单号','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','委外发料单详情','Json对象'),(1,'oioXkNo','委外发料单平台单号','String'),(1,'innerVendorCode','供应商编码','String'),(1,'innerVendorName','供应商名称','String'),(1,'status','状态','Number'),(1,'lineList','单身信息列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'issueQty','发料数量','Number'),(2,'waitIssueQty','应发料量','Number'),(2,'alreadyIssueQty','已发料量','Number'),(2,'storeCode','仓库编码','String'),(2,'storeName','仓库名称','String')]},
{'id':'2','name':'委外物料盘点单查询','url':'https://openapi.xiekeyun.com/outsourceInventoryOrder/getList.json','method':'POST','desc':'查询委外物料盘点单数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','盘点单数据','Json对象'),(1,'oioXkNo','盘点单平台单号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'status','盘点单状态','Number'),(1,'orderProductList','盘点明细','JSON对象数组'),(2,'prodCode','物料编码','String'),(2,'prodName','物料名称','String'),(2,'inventoryQty','库存数量','Number'),(2,'countedQty','盘点数量','Number')]}]}

# 11. 工装协同
TL = {'name':'工装协同','interfaces':[
{'id':'1','name':'工装信息变更单列表查询','url':'https://openapi.xiekeyun.com/toolsInfo/change/queryList.json','method':'POST','desc':'查询工装信息变更单列表','params':[
 ('startDate','更新时间开始','必填','Number'),('endDate','更新时间结束','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','变更单数据列表','JSON数组'),(1,'tcoXkNo','变更单平台唯一编码','String'),(1,'toolingsUniqueCode','工装平台唯一编码','String'),(1,'toolingsCode','工装编码','String'),(1,'toolingsName','工装名称','String'),(1,'status','状态','Number')]},
{'id':'2','name':'工装信息变更单详情查询','url':'https://openapi.xiekeyun.com/toolsInfo/change/detail.json','method':'POST','desc':'查询工装信息变更单详情','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('tcoXkNo','变更单平台唯一编码','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','变更单详情','JSON数组'),(1,'tcoXkNo','变更单平台唯一编码','String'),(1,'toolingsCode','工装编码','String'),(1,'toolingsName','工装名称','String'),(1,'status','状态','Number'),(1,'remark','备注','String'),(1,'changedRemark','变更后备注','String'),(1,'publishName','发布人名称','String'),(1,'publishTime','发布时间','Number')]},
{'id':'3','name':'工装档案查询','url':'https://openapi.xiekeyun.com/toolsInfo/getDetail.json','method':'POST','desc':'查询工装档案数据','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('toolingsUniqueCode','工装平台唯一编码','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','工装档案数据','JSON对象'),(1,'toolingsUniqueCode','工装平台唯一编码','String'),(1,'toolingsCode','工装编码','String'),(1,'toolingsName','工装名称','String'),(1,'status','状态','Number')]}]}

# 12. 附件更新查询
AT = {'name':'附件更新查询','interfaces':[
{'id':'1','name':'获取供应商新上传了附件的单据列表','url':'https://openapi.xiekeyun.com/orderHasUploadFile/queryByLastDate.json','method':'POST','desc':'查询某类型单据在时间范围内供应商有操作附件的单据编码列表','params':[
 ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('billType','单据类型 10采购订单','必填','Number'),('uploadType','附件上传者类型','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','单号数据列表','JSON数组'),(1,'billXkNo','单据平台编码','String'),(1,'billErpNo','单据ERP编码','String')]}]}

MODS = [VM, LC, IM, PO, RC, IK, SP, AP, IQ, OS, TL, AT]

def gen(path):
    wb=openpyxl.Workbook()
    ws=wb.active;ws.title='接口总览'
    write_overview(ws,MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn=api['name']
            if len(sn)>18:sn=sn[:18]+'..'
            nm=f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31:nm=nm[:31]
            write_subs(wb.create_sheet(title=nm),api)
    wb.save(path);print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\SRM_采购协同_接口梳理.xlsx')
