#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""OA工作流 + 免密登录 + 附件上传下载 - API接口梳理Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEEP_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
NEST_L1 = PatternFill(start_color='F4F7FC', end_color='F4F7FC', fill_type='solid')
NEST_L2 = PatternFill(start_color='EBEFF7', end_color='EBEFF7', fill_type='solid')

WHITE_B = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_F = Font(name='微软雅黑', size=10, bold=True)
TITLE_F = Font(name='微软雅黑', size=11, bold=True)
NORMAL_F = Font(name='微软雅黑', size=10)

THIN_B = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),
                top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_W = {'A':28,'B':52,'C':14,'D':20}
OV_C_W = {'A':10,'B':36,'C':60,'D':12,'E':50}

def apply_style(c, font=NORMAL_F, fill=None, align=ALIGN_LEFT, border=THIN_B):
    c.font = font; c.alignment = align; c.border = border
    if fill is not None: c.fill = fill

def write_overview(ws, modules):
    for k,v in OV_C_W.items(): ws.column_dimensions[k].width = v
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '系统工具类API - 接口总览'
    c.font = Font(name='微软雅黑',size=14,bold=True); c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        apply_style(ws.cell(row=2,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[2].height = 22
    row = 3
    for mod in modules:
        ws.merge_cells(f'A{row}:E{row}')
        apply_style(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
        for c_ in 'BCDE': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
        ws.row_dimensions[row].height = 22; row += 1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, LIGHT_GRAY,
                           ALIGN_CENTER if i in (1,4) else ALIGN_LEFT)
            row += 1
    ws.auto_filter.ref = f'A2:E{row-1}'

def write_subs(ws, api):
    for k,v in C_W.items(): ws.column_dimensions[k].width = v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        apply_style(ws.cell(row=idx,column=1,value=l), BOLD_F, WHITE_FILL, ALIGN_LEFT)
        ws.merge_cells(f'B{idx}:D{idx}')
        apply_style(ws.cell(row=idx,column=2,value=v), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        for c_ in 'CD': apply_style(ws[f'{c_}{idx}'], NORMAL_F, WHITE_FILL, ALIGN_LEFT)
    row = 5; ws.row_dimensions[4].height = 6
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='请求参数'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; row += 1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        apply_style(ws.cell(row=row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22; row += 1
    if api['params']:
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, WHITE_FILL,
                           ALIGN_CENTER if i in (3,4) else ALIGN_LEFT)
            row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}')
        apply_style(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        row += 1
    row += 2; ws.row_dimensions[row-1].height = 6
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='响应数据字段'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; hdr_row = row + 1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        apply_style(ws.cell(row=hdr_row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[hdr_row].height = 22; row = hdr_row + 1
    rl = []
    NEST_FILLS = {0: WHITE_FILL, 1: NEST_L1, 2: NEST_L2}
    for item in api['responses']:
        lvl, rname, rdesc, rtype = item[0], item[1], item[2] if len(item)>2 else '', item[3] if len(item)>3 else ''
        indent = '  ' * lvl
        row_fill = NEST_FILLS.get(lvl, WHITE_FILL)
        apply_style(ws.cell(row=row,column=1,value=indent+rname), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=2,value=rdesc), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=3,value=''), NORMAL_F, row_fill, ALIGN_CENTER)
        apply_style(ws.cell(row=row,column=4,value=rtype), NORMAL_F, row_fill, ALIGN_CENTER)
        rl.append((row, lvl)); row += 1
    if rl:
        max_lvl = max(l for _,l in rl)
        for lvl in range(max_lvl, 0, -1):
            i = 0
            while i < len(rl):
                r, l = rl[i]
                if l == lvl:
                    j = i + 1
                    while j < len(rl) and rl[j][1] >= lvl: j += 1
                    if j - 1 > i: ws.row_dimensions.group(r, rl[j-1][0], outline_level=lvl)
                    i = j
                else: i += 1
    return row

# =========================== DATA ===========================

# ---- 1. OA工作流 ----
OA = {
'name': 'OA工作流', 'interfaces': [
    {'id': '1', 'name': '查询已送审单据', 'url': 'https://openapi.xiekeyun.com/workflowBills/approvalList.json', 'method': 'POST',
     'desc': '查询需要OA审批的业务单据列表（单据状态=草稿且已送审），支持40+种单据类型',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billsType', '单据类型（1供应商对账单/2供应商费用单/3供应商发票单/4询价单/5潜在供应商/8检验单/9-ECN报告/10采购单/11核价单/12招标单/13付款申请单/14请购单/16项目招标/17核价失效单/18供应商竞价/19多物料竞价/20供应商准入评估报告/21索样通知单/22供应商信息变更/23供应商合同/15供应商退出通知/24采购方付款单/25内部商城引入单/26预付款单/27供应商通知/28物料停产报告/29品类管理单/30询价比价单/31招标比价单/32图纸变更通知/33采购变更单/34供应计划排程/35询价报价单/36招标报价单/37项目招标投标单/38询价需求单/39关务资料/40收货单/43入库单/44样品管理单/45物料申请单/46委外退料单/47委外申请单/48供应商申诉单/49供应商绩效评估报告/50供应商登记/53工序核价单/55交料通知单/56送货单申请）', '必填', 'Number'),
         ('startDate', '送审时间开始（毫秒时间戳，开区间）', '必填', 'Number'),
         ('endDate', '送审时间结束（毫秒时间戳，闭区间，最大120分钟）', '必填', 'Number'),
         ('checkUseType', '审批流使用类型（1平台审批/2OA审批）', '选填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'dataList', '已送审单据列表', 'Json数组'),
         (1, 'billsType', '单据类型编码', 'Number'),
         (1, 'innerVendorCode', '供应商编码（仅潜在供应商类型返回）', 'String'),
         (1, 'billXkNo', '平台单据号', 'String'),
         (1, 'status', '单据状态', 'Number'),
         (1, 'employeeName', '员工姓名', 'String'),
         (1, 'employeeMobile', '手机号', 'String'),
         (1, 'employeeNumber', '工号', 'String'),
         (1, 'erpAccount', 'ERP帐号', 'String'),
         (1, 'erpEmployeeCode', 'ERP员工编码', 'String'),
         (1, 'checkDesc', '送审说明', 'String'),
         (1, 'checkDateTime', '送审时间', 'Date'),
     ]},
    {'id': '2', 'name': 'OA回写审批信息', 'url': 'https://openapi.xiekeyun.com/workflowBills/publish.json', 'method': 'POST',
     'desc': '将OA审批过程及结果回写到携客云平台对应单据，支持直接发布单据',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billsType', '单据类型（同查询已送审单据的billsType定义）', '必填', 'Number'),
         ('billXkNo', '平台单据号', '必填', 'String'),
         ('status', '审批结果（1审批通过/2审批退回）', '必填', 'Number'),
         ('checkDesc', '审批意见', '选填', 'String'),
         ('checkUserName', '审批人姓名', '选填', 'String'),
         ('wfTypeCode', '流程类型编码', '选填', 'String'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
     ]},
    {'id': '3', 'name': '回写审批URL', 'url': 'https://openapi.xiekeyun.com/workflowBills/updateOAViewUrl.json', 'method': 'POST',
     'desc': '回写单据在OA系统中的审批查看URL，方便用户在携客云直接跳转OA查看进度',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billsType', '单据类型', '必填', 'Number'),
         ('billXkNo', '平台单据号', '必填', 'String'),
         ('oaViewUrl', 'OA系统中的审批查看URL', '必填', 'String'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
     ]},
    {'id': '4', 'name': '查询待审批单据列表', 'url': 'https://openapi.xiekeyun.com/workflowBills/getCheckInfoList.json', 'method': 'POST',
     'desc': '查询需要当前人员（erpCode对应员工）审批的单据列表',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billsType', '单据类型', '必填', 'Number'),
         ('crossEmployeeFlag', '是否跨员工查询（0否/1是）', '选填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'dataList', '待审批单据列表', 'Json数组'),
         (1, 'billXkNo', '平台单据号', 'String'),
         (1, 'billShowNo', '显示单号（ERP单号）', 'String'),
         (1, 'billsType', '单据类型', 'String'),
         (1, 'wfTypeCode', '流程类型编码', 'String'),
         (1, 'wfnName', '审批关卡名称', 'String'),
         (1, 'checkUserName', '审批人姓名', 'String'),
         (1, 'checkDesc', '审批说明', 'String'),
         (1, 'status', '审批状态', 'Number'),
         (1, 'sourceType', '来源类型', 'Number'),
         (1, 'applyCheckFlag', '是否申请加签', 'Number'),
         (1, 'recallFlag', '是否可撤回', 'Number'),
         (1, 'createTime', '创建时间', 'Number'),
         (1, 'updateTime', '更新时间', 'Number'),
     ]},
    {'id': '5', 'name': '查询审批记录列表', 'url': 'https://openapi.xiekeyun.com/workflowBills/getCheckRecordList.json', 'method': 'POST',
     'desc': '查询某单据在平台审批流中的完整审批记录',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billsType', '单据类型', '必填', 'Number'),
         ('billXkNo', '平台单据号', '必填', 'String'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'dataList', '审批记录列表', 'Json数组'),
         (1, 'billXkNo', '平台单据号', 'String'),
         (1, 'billShowNo', '显示单号', 'String'),
         (1, 'billsType', '单据类型', 'String'),
         (1, 'wfTypeCode', '流程类型编码', 'String'),
         (1, 'wfnName', '审批关卡名称', 'String'),
         (1, 'checkUserName', '审批人姓名', 'String'),
         (1, 'checkDesc', '审批意见', 'String'),
         (1, 'status', '审批状态（1通过/2退回/3加签）', 'Number'),
         (1, 'sourceType', '来源类型', 'Number'),
         (1, 'applyCheckFlag', '是否申请加签', 'Number'),
         (1, 'recallFlag', '是否可撤回', 'Number'),
         (1, 'checkDateTime', '审批时间', 'Number'),
     ]},
]}

# ---- 2. 免密登录 ----
SL = {
'name': '免密登录', 'interfaces': [
    {'id': '1', 'name': '获取免密登录URL', 'url': 'https://openapi.xiekeyun.com/system/ssoLogin.json', 'method': 'POST',
     'desc': '生成单个用户免密登录携客云的动态URL（有效期30秒）',
     'params': [
         ('erpCode', '请求者用户ERP帐号（用于平台鉴权检查）', '必填', 'String'),
         ('mobile', '携客云用户手机号（需已在平台注册）', '必填', 'String'),
         ('mobileArea', '手机区域码（+86/+852/+886/+853/+82/+84/+65/+81/+66/+1）', '必填', 'String'),
         ('isConstrain', '是否限制跳转到指定企业（0不限/1限制，默认0）', '选填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'data', '免密登录数据', 'Json对象'),
         (1, 'ssoUrl', '免密登录URL（有效期30秒）', 'String'),
         (1, 'employeeName', '员工姓名', 'String'),
     ]},
]}

# ---- 3. 附件上传下载 ----
FL = {
'name': '附件上传下载', 'interfaces': [
    {'id': '1', 'name': '获取附件下载地址', 'url': 'https://fileapi.xiekeyun.com/download/getUrls.json', 'method': 'POST',
     'desc': '获取某一单据中各附件的下载地址列表（downloadUrl有效期10分钟），然后用GET请求下载文件流',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('businessType', '单据类型（101采购单/102询价单/103检验单/104送货通知单/105询价需求单/106核价单/107-8D报告/108项目预询单/109项目招标单/110付款单/111发票通知单/112请购单/113比价单/114代签单/115客户询价招标单/117费用单/118竞价单/119多物料竞价单/120收货单/121入库单/122物流竞价单/127索样通知单/128变更单/129付款单电子回单/130交料通知单/131商机线索/133工序核价单/134客户订单供应商附件/135报价单/201物料图纸版本/202物料资质/203供应商合同/204供应商资质/206供应商信息变更/207供应商通知/209供应商准入评估/210供应商问卷/211物料停产报告/212违约索赔通知/213供应商ECN报告/218物料申请单/219潜在供应商/220预付款单/221付款申请单/222物料评估报告/223对账单/224改善报告）', '必填', 'Number'),
         ('businessKey', '单据编码（ERP单号或平台单号，根据businessType取不同值）', '必填', 'String'),
         ('businessKeyFrom', '单据编码来源（0平台单号/1-ERP单号）', '选填', 'Number'),
         ('downloadSorceType', '附件来源类型（1采购方/2供应商/3电子签章）', '选填', 'Number'),
         ('downloadWorkflowFile', '是否下载审批流附件（1是/0否）', '选填', 'Number'),
         ('downloadUnauditedCert', '是否下载未审核资质（物料资质时有效）', '选填', 'Number'),
         ('fileTypes', '附件类型过滤（如[.pdf,.xml]）', '选填', 'Array'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'data', '附件下载数据', 'Json对象'),
         (1, 'fileList', '单头附件列表', 'Json数组'),
         (2, 'fileName', '附件文件名称', 'String'),
         (2, 'downloadUrl', '附件下载地址（有效期10分钟）', 'String'),
         (2, 'downloadSorceType', '附件来源类型（1采购方/2供应商/3电子签章）', 'Number'),
         (2, 'innerVendorCode', '供应商内部编码（供应商上传时有值）', 'String'),
         (1, 'lineFileList', '单身附件列表', 'Json数组'),
         (2, 'lineNo', '项次号', 'String'),
         (2, 'billXkNo', '单据平台单号', 'String'),
         (2, 'productCode', '物料编码', 'String'),
         (2, 'fileName', '附件文件名称', 'String'),
         (2, 'downloadUrl', '附件下载地址（有效期10分钟）', 'String'),
         (1, 'workflowFileList', '审批流附件列表（开启平台审批且有审批附件时返回）', 'Json数组'),
         (2, 'fileName', '附件文件名称', 'String'),
         (2, 'downloadUrl', '附件下载地址（有效期10分钟）', 'String'),
         (2, 'workflowFileType', '审批流附件类型（1送审附件/2审批过程附件）', 'Number'),
     ]},
    {'id': '2', 'name': '获取附件上传地址', 'url': 'https://fileapi.xiekeyun.com/upload/getUrls.json', 'method': 'POST',
     'desc': '获取某一单据中各附件的上传地址（uploadUrl有效期10分钟），然后用POST multipart/form-data上传文件流',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('businessType', '单据类型（101采购单/102询价单/103检验单/104送货通知单/105询价需求单/106核价单/107-8D报告/108项目预询单/109项目招标单/110付款单/111发票单/112请购单/116客户报价单/117费用单/118竞价单/119多物料竞价单/124比价单/125样品管理单/127索样通知单/128采购变更单/130交料通知单/131商机线索/132工装档案/133工序核价单/134客户订单供应商附件/201物料图纸版本/202物料资质/203供应商合同/204供应商资质/205供应商绩效/207供应商通知/208图纸变更通知/212违约索赔通知/213供应商ECN报告/215预检申请单/216供应商问卷/217物料档案/220预付款单/221付款申请单/223对账单/224改善报告）', '必填', 'Number'),
         ('businessKey', '单据编码（根据businessType取不同值）', '必填', 'String'),
         ('subKey', '唯一值（图纸版本号/资质文件编码/合同编号等）', '选填', 'String'),
         ('innerVendorCode', '供应商编码（businessType=202/124时必填）', '选填', 'String'),
         ('autoCreateProdGraph', '是否自动创建图纸版本（0否/1是，businessType=201有效）', '选填', 'Number'),
         ('validFlag', '是否为最新有效版本（0否/1是，businessType=201有效）', '选填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1成功；9失败', 'Number'),
         (0, 'errorCode', '接口错误编码', 'String'),
         (0, 'errorMsg', '接口错误信息', 'String'),
         (0, 'data', '附件上传数据', 'Json对象'),
         (1, 'uploadUrl', '单头附件上传URL（有效期10分钟）', 'String'),
         (1, 'allowedMaxFileSize', '单头剩余最大上传文件累计大小（字节）', 'Number'),
         (1, 'lineUploadUrlList', '各单身附件上传URL列表', 'Json数组'),
         (2, 'lineNo', '项次号', 'String'),
         (2, 'productCode', '物料编码', 'String'),
         (2, 'allowedMaxFileSize', '单身剩余最大上传文件累计大小（字节）', 'Number'),
         (2, 'uploadUrl', '附件的上传地址（有效期10分钟）', 'String'),
         (2, 'innerUploadUrl', '内部附件的上传地址（仅本方可见）', 'String'),
         (2, 'innerAllowedMaxFileSize', '内部附件允许最大上传大小', 'Number'),
     ]},
]}

MODS = [OA, SL, FL]

def gen(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '接口总览'
    write_overview(ws, MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn = api['name']
            if len(sn)>18: sn = sn[:18]+'..'
            nm = f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31: nm = nm[:31]
            write_subs(wb.create_sheet(title=nm), api)
    wb.save(path); print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\6-9系统工具类_接口梳理.xlsx')
