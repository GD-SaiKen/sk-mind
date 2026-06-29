"""SRM采购协同 - 完整字段Excel生成器"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side

DEEP_BLUE=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
LIGHT_BLUE=PatternFill(start_color='D6E4F0',end_color='D6E4F0',fill_type='solid')
LIGHT_GRAY=PatternFill(start_color='F2F2F2',end_color='F2F2F2',fill_type='solid')
WHITE_B=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF')
BOLD_F=Font(name='微软雅黑',size=10,bold=True)
TITLE_F=Font(name='微软雅黑',size=11,bold=True)
NORMAL_F=Font(name='微软雅黑',size=10)
THIN_B=Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
WRAP=Alignment(wrap_text=True,vertical='center')
CENTER=Alignment(wrap_text=True,vertical='center',horizontal='center')
C_W={'A':28,'B':52,'C':14,'D':20}
OV_C_W={'A':10,'B':36,'C':60,'D':12,'E':50}

def ac(c,font=NORMAL_F,fill=None,align=WRAP,border=THIN_B):
    c.font=font;c.alignment=align;c.border=border
    if fill:c.fill=fill

def write_overview(ws,mods):
    for k,v in OV_C_W.items():ws.column_dimensions[k].width=v
    ws.merge_cells('A1:E1');c=ws['A1'];c.value='携客云SRM采购协同平台 - API接口总览';c.font=Font(name='微软雅黑',size=14,bold=True);c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        ac(ws.cell(row=2,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[2].height=22;row=3
    for mod in mods:
        ws.merge_cells(f'A{row}:E{row}')
        ac(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
        for c_ in 'BCDE':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
        ws.row_dimensions[row].height=22;row+=1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,LIGHT_GRAY,CENTER if i in(1,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    ws.auto_filter.ref=f'A2:E{row-1}'

def write_subs(ws,api):
    for k,v in C_W.items():ws.column_dimensions[k].width=v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        ac(ws.cell(row=idx,column=1,value=l),BOLD_F);ws.merge_cells(f'B{idx}:D{idx}')
        ac(ws.cell(row=idx,column=2,value=v))
        for c_ in 'CD':ac(ws[f'{c_}{idx}'])
        ws.row_dimensions[idx].height=22
    row=5;ws.row_dimensions[4].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='请求参数'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;row+=1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        ac(ws.cell(row=row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[row].height=22;row+=1
    if api['params']:
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,None,CENTER if i in(3,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    else:
        ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'),NORMAL_F)
        ws.row_dimensions[row].height=20;row+=1
    row+=2;ws.row_dimensions[row-1].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='响应数据字段'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;hdr_row=row+1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        ac(ws.cell(row=hdr_row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[hdr_row].height=22;row=hdr_row+1
    rl=[]
    for item in api['responses']:
        lvl=item[0];rname=item[1];rdesc=item[2]if len(item)>2 else'';rtype=item[3]if len(item)>3 else''
        indent='  '*lvl
        ac(ws.cell(row=row,column=1,value=indent+rname),NORMAL_F);ac(ws.cell(row=row,column=2,value=rdesc),NORMAL_F)
        ac(ws.cell(row=row,column=3,value=''),NORMAL_F,None,CENTER)
        ac(ws.cell(row=row,column=4,value=rtype),NORMAL_F,None,CENTER)
        rl.append((row,lvl));ws.row_dimensions[row].height=20;row+=1
    if rl:
        max_lvl=max(l for _,l in rl)
        for lvl in range(max_lvl,0,-1):
            i=0
            while i<len(rl):
                r,l=rl[i]
                if l==lvl:
                    j=i+1
                    while j<len(rl)and rl[j][1]>=lvl:j+=1
                    if j-1>i:ws.row_dimensions.group(r,rl[j-1][0],outline_level=lvl)
                    i=j
                else:i+=1
    return row

# ===================== COMPLETE RESPONSE DATA =====================
# ALL fields at ALL nesting levels extracted from source documents
# Format: (level, field_name, description, data_type)

# Since the full data is thousands of lines, we load it from the compiled module
import importlib.util
import sys

# The data module will be generated separately - for now use what we have inline
# to handle the critical missing fields, we'll expand the key interfaces

# ===================== MODULE DEFINITIONS =====================
# Using the same inline approach as before but with EXPANDED responses for ALL interfaces
# Based on agent extraction results

# Import the expanded data module
spec = importlib.util.spec_from_file_location("srm_data",
    "G:\\workspace\\GD-SaiiKen\\sk-mind\\.claude\\skills\\api-interface-excel\\srm_data.py")
srm_data = importlib.util.module_from_spec(spec)
sys.modules['srm_data'] = srm_data
spec.loader.exec_module(srm_data)

MODS = srm_data.MODS

def gen(path):
    wb=openpyxl.Workbook()
    ws=wb.active;ws.title='接口总览'
    write_overview(ws,MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn=api['name']
            if len(sn)>18:sn=sn[:18]+'..'
            nm=f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31:nm=nm[:31]
            write_subs(wb.create_sheet(title=nm),api)
    wb.save(path);print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\SRM_采购协同_接口梳理.xlsx')
