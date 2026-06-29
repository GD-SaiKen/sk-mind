#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""系统工作台 - API接口梳理Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEEP_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
NEST_L1 = PatternFill(start_color='F4F7FC', end_color='F4F7FC', fill_type='solid')
NEST_L2 = PatternFill(start_color='EBEFF7', end_color='EBEFF7', fill_type='solid')

WHITE_B = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_F = Font(name='微软雅黑', size=10, bold=True)
TITLE_F = Font(name='微软雅黑', size=11, bold=True)
NORMAL_F = Font(name='微软雅黑', size=10)

THIN_B = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),
                top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
C_W = {'A':28,'B':52,'C':14,'D':20}
OV_C_W = {'A':10,'B':36,'C':60,'D':12,'E':50}

def apply_style(c, font=NORMAL_F, fill=None, align=ALIGN_LEFT, border=THIN_B):
    c.font = font; c.alignment = align; c.border = border
    if fill is not None: c.fill = fill

def write_overview(ws, modules):
    for k,v in OV_C_W.items(): ws.column_dimensions[k].width = v
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '系统工作台 - API接口总览'
    c.font = Font(name='微软雅黑',size=14,bold=True); c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        apply_style(ws.cell(row=2,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[2].height = 22
    row = 3
    for mod in modules:
        ws.merge_cells(f'A{row}:E{row}')
        apply_style(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
        for c_ in 'BCDE': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
        ws.row_dimensions[row].height = 22; row += 1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, LIGHT_GRAY,
                           ALIGN_CENTER if i in (1,4) else ALIGN_LEFT)
            row += 1
    ws.auto_filter.ref = f'A2:E{row-1}'

def write_subs(ws, api):
    for k,v in C_W.items(): ws.column_dimensions[k].width = v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        apply_style(ws.cell(row=idx,column=1,value=l), BOLD_F, WHITE_FILL, ALIGN_LEFT)
        ws.merge_cells(f'B{idx}:D{idx}')
        apply_style(ws.cell(row=idx,column=2,value=v), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        for c_ in 'CD': apply_style(ws[f'{c_}{idx}'], NORMAL_F, WHITE_FILL, ALIGN_LEFT)
    row = 5; ws.row_dimensions[4].height = 6
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='请求参数'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; row += 1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        apply_style(ws.cell(row=row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22; row += 1
    if api['params']:
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, WHITE_FILL,
                           ALIGN_CENTER if i in (3,4) else ALIGN_LEFT)
            row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}')
        apply_style(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        row += 1
    row += 2; ws.row_dimensions[row-1].height = 6
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='响应数据字段'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; hdr_row = row + 1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        apply_style(ws.cell(row=hdr_row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[hdr_row].height = 22; row = hdr_row + 1
    rl = []
    NEST_FILLS = {0: WHITE_FILL, 1: NEST_L1, 2: NEST_L2}
    for item in api['responses']:
        lvl, rname, rdesc, rtype = item[0], item[1], item[2] if len(item)>2 else '', item[3] if len(item)>3 else ''
        indent = '  ' * lvl
        row_fill = NEST_FILLS.get(lvl, WHITE_FILL)
        apply_style(ws.cell(row=row,column=1,value=indent+rname), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=2,value=rdesc), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=3,value=''), NORMAL_F, row_fill, ALIGN_CENTER)
        apply_style(ws.cell(row=row,column=4,value=rtype), NORMAL_F, row_fill, ALIGN_CENTER)
        rl.append((row, lvl)); row += 1
    if rl:
        max_lvl = max(l for _,l in rl)
        for lvl in range(max_lvl, 0, -1):
            i = 0
            while i < len(rl):
                r, l = rl[i]
                if l == lvl:
                    j = i + 1
                    while j < len(rl) and rl[j][1] >= lvl: j += 1
                    if j - 1 > i: ws.row_dimensions.group(r, rl[j-1][0], outline_level=lvl)
                    i = j
                else: i += 1
    return row

# =========================== DATA ===========================

WB = {
'name': '系统工作台', 'interfaces': [
    {'id': '1', 'name': '待跟进待处理列表查询', 'url': 'https://openapi.xiekeyun.com/workbench/waitWorksList.json', 'method': 'POST',
     'desc': '查询工作台待跟进、待处理列表。billType单据类别编码见接口文档枚举表（10采购单/11变更单/12收货单/15检验单/22出货通知单/30排程/50询价/52对账单等40+种类型）',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('billType', '单据类别编码（10采购单/11变更单/12收货单/13入库单/15检验单/16交料通知单/17工装移转单/22出货通知单/23核价单/26延迟送货申请单/27-8D报告/29供应商/30排程/32预检申请/33物料准入报告/40合同/41资质/43供应商处罚通知单/45供应商信息变更/46项目预询单/49订单直运看板/50询价/52对账单/53发票单/54费用单/55竞价单/59比价单/60预付款单/61付款申请/62付款计划/64付款单/68金融凭证/69-ECN报告/71项目招标单/73供应商评价/74引入单/78工装档案/79供应商申诉单/82物料图纸变更/83改善报告/84多物料竞价/85提前付款申请/86物料资质/87潜在供应商/91进货折让单/92索样单/101供应商问卷/110供应商通知）', '必填', 'Number'),
         ('progressType', '进程类型（1:待处理 2:待跟进）', '必填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1:成功；9:失败', 'Number'),
         (0, 'errorCode', '接口错误编码（失败时有值）', 'String'),
         (0, 'errorMsg', '接口错误信息（失败时有值）', 'String'),
         (0, 'dataList', '待跟进/待处理单据列表', 'Json数组'),
         (1, 'billXkNo', '平台单号', 'String'),
         (1, 'billShowNo', '显示单号（对应ERP单号）', 'String'),
         (1, 'billType', '单据类型编码', 'String'),
         (1, 'billTypeSub', '单据操作类型编码', 'String'),
         (1, 'billStatus', '单据状态', 'Number'),
         (1, 'innerType', '交易对象类型（1:供应商 2:客户）', 'Number'),
         (1, 'innerCode', '交易对象编码', 'String'),
         (1, 'innerName', '交易对象名称', 'String'),
         (1, 'publishName', '发布人名称', 'String'),
         (1, 'publishTime', '发布时间（毫秒时间戳）', 'Number'),
     ]},
]}

MODS = [WB]

def gen(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '接口总览'
    write_overview(ws, MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn = api['name']
            if len(sn)>18: sn = sn[:18]+'..'
            nm = f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31: nm = nm[:31]
            write_subs(wb.create_sheet(title=nm), api)
    wb.save(path); print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\5系统工作台_接口梳理.xlsx')
