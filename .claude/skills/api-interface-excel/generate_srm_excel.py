#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
携客云SRM采购协同平台 - API接口梳理Excel生成（完整版）
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

DEEP_BLUE_FILL = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE_FILL = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY_FILL = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE_FONT = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_FONT = Font(name='微软雅黑', size=10, bold=True)
TITLE_FONT = Font(name='微软雅黑', size=11, bold=True)
NORMAL_FONT = Font(name='微软雅黑', size=10)
THIN_BORDER = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
WRAP_ALIGN = Alignment(wrap_text=True, vertical='center')
CENTER_ALIGN = Alignment(wrap_text=True, vertical='center', horizontal='center')
COL_WIDTHS = {'A': 28, 'B': 52, 'C': 14, 'D': 20}
OVERVIEW_COL_WIDTHS = {'A': 10, 'B': 36, 'C': 60, 'D': 12, 'E': 50}

def set_col_widths(ws):
    for col_letter, width in COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width

def write_overview(ws, modules):
    for col_letter, width in OVERVIEW_COL_WIDTHS.items():
        ws.column_dimensions[col_letter].width = width
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '携客云SRM采购协同平台 - API接口总览'; c.font = Font(name='微软雅黑',size=14,bold=True); c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30
    headers = ['序号','接口名称','接口地址','请求方式','接口描述']
    for i, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=i, value=h); cell.font = WHITE_FONT; cell.fill = DEEP_BLUE_FILL; cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER
    ws.row_dimensions[2].height = 22
    row = 3
    for module in modules:
        ws.merge_cells(f'A{row}:E{row}')
        cat_cell = ws.cell(row=row, column=1, value=f'▸ {module["name"]}'); cat_cell.font = TITLE_FONT; cat_cell.fill = LIGHT_BLUE_FILL; cat_cell.alignment = Alignment(vertical='center'); cat_cell.border = THIN_BORDER
        for c_ in 'BCDE': ws[f'{c_}{row}'].border = THIN_BORDER; ws[f'{c_}{row}'].fill = LIGHT_BLUE_FILL
        ws.row_dimensions[row].height = 22; row += 1
        for api in module['interfaces']:
            data = [api['id'], api['name'], api['url'], api['method'], api['desc']]
            for i, val in enumerate(data, 1):
                cell = ws.cell(row=row, column=i, value=val); cell.font = NORMAL_FONT; cell.fill = LIGHT_GRAY_FILL; cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if i in (1,4) else WRAP_ALIGN
            ws.row_dimensions[row].height = 20; row += 1
    ws.auto_filter.ref = f'A2:E{row-1}'

def write_subsheet(ws, api):
    set_col_widths(ws)
    # 基本信息
    info_rows = [('接口地址', api['url']), ('请求方式', api['method']), ('接口描述', api['desc'])]
    for idx, (label, value) in enumerate(info_rows, 1):
        a_cell = ws.cell(row=idx, column=1, value=label); a_cell.font = BOLD_FONT; a_cell.alignment = WRAP_ALIGN; a_cell.border = THIN_BORDER
        ws.merge_cells(f'B{idx}:D{idx}')
        v_cell = ws.cell(row=idx, column=2, value=value); v_cell.font = NORMAL_FONT; v_cell.alignment = WRAP_ALIGN; v_cell.border = THIN_BORDER
        for c_ in 'CD': ws[f'{c_}{idx}'].border = THIN_BORDER
        ws.row_dimensions[idx].height = 22
    row = 4; ws.row_dimensions[row].height = 6
    # 请求参数
    row = 5; ws.merge_cells(f'A{row}:D{row}')
    tc = ws.cell(row=row, column=1, value='请求参数'); tc.font = TITLE_FONT; tc.fill = LIGHT_BLUE_FILL; tc.alignment = Alignment(vertical='center'); tc.border = THIN_BORDER
    for c_ in 'BCD': ws[f'{c_}{row}'].border = THIN_BORDER; ws[f'{c_}{row}'].fill = LIGHT_BLUE_FILL
    ws.row_dimensions[row].height = 22; row += 1
    p_headers = ['参数名称','参数说明','是否必须','数据类型']
    for i, h in enumerate(p_headers, 1):
        cell = ws.cell(row=row, column=i, value=h); cell.font = WHITE_FONT; cell.fill = DEEP_BLUE_FILL; cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22; row += 1
    if api['params']:
        for p_name, p_desc, p_req, p_type in api['params']:
            for i, val in enumerate([p_name, p_desc, p_req, p_type], 1):
                cell = ws.cell(row=row, column=i, value=val); cell.font = NORMAL_FONT; cell.border = THIN_BORDER
                cell.alignment = CENTER_ALIGN if i in (3,4) else WRAP_ALIGN
            ws.row_dimensions[row].height = 20; row += 1
    else:
        cell = ws.cell(row=row, column=1, value='（该接口无业务请求参数）'); cell.font = NORMAL_FONT
        ws.merge_cells(f'A{row}:D{row}'); ws.row_dimensions[row].height = 20; row += 1
    row += 1; ws.row_dimensions[row].height = 6; row += 1
    # 响应数据字段
    ws.merge_cells(f'A{row}:D{row}')
    tc = ws.cell(row=row, column=1, value='响应数据字段'); tc.font = TITLE_FONT; tc.fill = LIGHT_BLUE_FILL; tc.alignment = Alignment(vertical='center'); tc.border = THIN_BORDER
    for c_ in 'BCD': ws[f'{c_}{row}'].border = THIN_BORDER; ws[f'{c_}{row}'].fill = LIGHT_BLUE_FILL
    ws.row_dimensions[row].height = 22; row += 1
    r_headers = ['字段名称','字段说明','','数据类型']
    for i, h in enumerate(r_headers, 1):
        cell = ws.cell(row=row, column=i, value=h); cell.font = WHITE_FONT; cell.fill = DEEP_BLUE_FILL; cell.alignment = CENTER_ALIGN; cell.border = THIN_BORDER
    ws.row_dimensions[row].height = 22; row += 1
    for r_item in api['responses']:
        if len(r_item) == 3:
            r_name, r_desc, r_type = r_item
            data_v = [r_name, r_desc, '', r_type]
        else:
            data_v = list(r_item) + [''] * (4 - len(r_item))
        for i, val in enumerate(data_v, 1):
            cell = ws.cell(row=row, column=i, value=val); cell.font = NORMAL_FONT; cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN if i in (3,4) else WRAP_ALIGN
        ws.row_dimensions[row].height = 20; row += 1

# ========== 模块1: 供应商管理 ==========
vendor_mgmt = {
 'name': '供应商管理',
 'interfaces': [
  {'id': '1', 'name': '查询已加入的供应商列表', 'url': 'https://openapi.xiekeyun.com/compVendor/confirmList.json', 'method': 'POST', 'desc': '查询邀约状态为"已加入"的供应商列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('handshakeTimeStart','握手成功查询开始时间','选填','Number'),('handshakeTimeEnd','握手成功查询结束时间','选填','Number'),('categoryCode01','供应商分类1编码','选填','String'),('categoryCode02','供应商分类2编码','选填','String'),('categoryCode03','供应商分类3编码','选填','String'),('performanceClassificationCode','绩效分类编码','选填','String')], 'responses': [
   ('result','返回结果状态 1:成功；9:失败','Number'),('errorCode','接口错误编码（失败时有值）','String'),('errorMsg','接口错误信息（失败时有值）','String'),('dataList','供应商档案集合','Json数组'),('vendorXkUniqueCode','供应商平台唯一编码','String'),('innerVendorCode','供应商编码','String'),('innerVendorName','供应商名称','String'),('employeeName','邀约供应商的员工姓名','String'),('handshakeTime','握手时间（时间戳）','Number'),('categoryCode01','供应商分类1编码','String'),('categoryName01','供应商分类1名称','String'),('categoryCode02','供应商分类2编码','String'),('categoryName02','供应商分类2名称','String'),('categoryCode03','供应商分类3编码','String'),('categoryName03','供应商分类3名称','String'),('performanceClassificationCode','绩效分类编码','String'),('performanceClassificationName','绩效分类名称','String')]},
  {'id': '2', 'name': '查询供应商绩效考评结果', 'url': 'https://openapi.xiekeyun.com/compVendor/assessmentScore.json', 'method': 'POST', 'desc': '查询供应商月度绩效考评结果列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型 1:月份2:季度3:年份','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型(上半年/下半年/全年)','选填','Number'),('monthTime','月份 01-12','选填','String'),('quarterTime','季度 1-4','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','供应商绩效得分列表','Json对象'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('categoryCode','分类编码','String'),('categoryName','分类名称','String'),('assessmentMonth','绩效对应的月份 yyyyMM','String'),('assessmentScore','绩效得分','Number'),('performanceScore','绩效评分','Number'),('deductionScore','绩效扣分','Number'),('gradeInfo','评级','String'),('detailList','绩效明细项','json数组'),('curSrcScore','本项达成率（%）','Number'),('curTargetScore','本项得分','Number'),('assementStandardItemCode','绩效考核项编码','String'),('assementStandardItemName','绩效考核项名称','String'),('vendorCategoryList','供应商分类列表','json数组'),('categoryType','分类类型(0001-0004)','String'),('categoryCode','分类编码','String'),('categoryName','分类名称','String')]},
  {'id': '3', 'name': '供应商档案详情', 'url': 'https://openapi.xiekeyun.com/compVendor/getDetail.json', 'method': 'POST', 'desc': '查询一个供应商详情明细数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码（ERP内）','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','供应商详情','Json对象'),('xkUniqueCode','内部供应商平台编码(32位)','String'),('vendorCode','供应商携客企业号','String'),('vendorName','企业名称','String'),('innerVendorCode','内部供应商编码（ERP）','String'),('innerVendorName','内部供应商名称','String'),('innerVendorAbbr','内部供应商简称','String'),('inviteStatus','邀约状态','number'),('handshakeTime','握手成功时间','Number'),('vendorStatus','供应商状态 0无效1有效3黑名单','number'),('buyerCode','采购员编码','String'),('buyerName','采购员名称','String'),('buyerMobile','采购员手机号','String'),('buyerEmail','采购员邮箱','String'),('lifeCycle','生命周期标识 0准入评估1试样2小批量3正式4退出','number'),('gradeInfo','供应商绩效等级','String'),('contacts','联系人明细列表','Json对象数组'),('vendorCategoryList','供应商分类/准入品类类别列表','Json对象数组'),('orgList','供应商组织关系列表','Json对象数组'),('sqeList','SQE列表','Json对象数组'),('companyEnglishName','公司英文名称','String'),('groupCompanyName','集团公司名称','String')]},
  {'id': '4', 'name': '查询供应商应用度报表', 'url': 'https://openapi.xiekeyun.com/compVendor/report.json', 'method': 'POST', 'desc': '查询供应商应用报表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','Number'),('year','年份','必填','Number'),('month','月份','必填','Number'),('type','查询类型 1:公司 2:组织','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','供应商应用度报表','Json对象'),('reportMonth','报表月份 yyyyMM','String'),('vendorTotalSum','供应商数量','Number'),('vendorHandShakeSum','已握手供应商数量','Number'),('reportList','供应商应用度报表列表','Json对象数组'),('companyCode','企业号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('currOrderSum','本月订单数','Number'),('orderHandRate','订单处理率','Number'),('onlineDeliveryRate','在线送货率','Number'),('enquiryJoinRate','询价参与率','Number'),('biddingJoinRate','竞价参与率','Number')]},
  {'id': '5', 'name': '查询集团法人绩效考评结果', 'url': 'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/list.json', 'method': 'POST', 'desc': '查询集团法人绩效考评结果列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','供应商绩效得分列表','Json对象'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('assessmentScore','绩效得分','Number'),('gradeInfo','评级','String'),('detailList','绩效明细项','json数组'),('curSrcScore','本项达成率（%）','Number'),('curTargetScore','本项得分','Number'),('assementStandardItemCode','绩效考核项编码','String'),('assementStandardItemName','绩效考核项名称','String'),('relationCompanyList','关联公司列表','json数组'),('relationCompanyCode','关联企业编码','String'),('relationCompanyName','关联企业名称','String')]},
  {'id': '6', 'name': '查询集团物料分类绩效考评结果', 'url': 'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/prodCategory/list.json', 'method': 'POST', 'desc': '查询集团物料分类绩效考评结果列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','供应商绩效得分列表','Json对象'),('prodCategoryCode','物料分类编码','String'),('prodCategoryName','物料分类名称','String'),('assessmentScore','绩效得分','Number'),('gradeInfo','评级','String'),('detailList','绩效明细项','json数组'),('vendorAsList','供应商列表','json数组'),('innerVendorCode','供应商编码','String'),('innerVendorName','供应商名称','String')]}]}

# ========== 模块2: 生命周期管理 ==========
lifecycle = {
 'name': '生命周期管理',
 'interfaces': [
  {'id': '1', 'name': '潜在供应商列表查询', 'url': 'https://openapi.xiekeyun.com/potentialVendor/getList.json', 'method': 'POST', 'desc': '查询企业潜在供应商列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('dateStart','资质审核完成时间-从','必填','Number'),('dateEnd','资质审核完成时间-至','必填','Number'),('createStart','创建时间-从','选填','Number'),('createEnd','创建时间-至','选填','Number'),('checkTimeStart','审核时间-从','选填','Number'),('checkTimeEnd','审核时间-至','选填','Number'),('applyStatus','申请状态','选填','Number数组'),('inviteStatus','邀约状态','选填','Number数组')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','潜在供应商数据列表','Json对象'),('innerVendorCode','内部供应商编码（ERP）','String'),('innerVendorName','内部供应商名称','String'),('certificateAuditFinishTime','资质审核完成时间','Number'),('admissionReportFlag','完成准入评估标识 0未完成1已完成','Number'),('applyStatus','申请状态 0草稿1申请中2拒绝3已确认4已转正式5转正中6不合格','Number'),('vendorStatus','供应商状态 0无效1有效2停用3黑名单','Number'),('inviteStatus','邀约状态','Number'),('checkTime','审核时间','Number'),('updateTime','最后更新时间','Number')]},
  {'id': '2', 'name': '潜在供应商详情查询', 'url': 'https://openapi.xiekeyun.com/potentialVendor/getDetail.json', 'method': 'POST', 'desc': '查询一个潜在供应商明细数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','潜在供应商编码','必填','String'),('approvalStatusFlag','是否返回审批状态 0不返回1返回','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','潜在供应商详情','Json对象'),('xkUniqueCode','内部供应商平台编码','String'),('innerVendorCode','内部供应商编码（ERP）','String'),('innerVendorName','内部供应商名称','String'),('innerVendorAbbr','内部供应商简称','String'),('applyStatus','审批状态','number'),('vendorStatus','供应商状态','number'),('inviteStatus','邀约状态','number'),('contacts','联系人明细列表','Json对象数组'),('purchaseOrgList','采购组织明细列表','Json对象数组'),('financeList','企业资料明细列表','Json对象数组'),('bankList','银行明细列表','Json对象数组'),('vendorCategoryList','供应商分类/准入品类类别列表','Json对象数组'),('mainCustomerList','主要客户','Json对象数组'),('mainProductList','主要供应产品','Json对象数组'),('keyManufacturingCapabilityList','关键制造能力','Json对象数组'),('companyEnglishName','公司英文名称','String'),('groupCompanyName','集团公司名称','String')]},
  {'id': '3', 'name': '样品管理列表查询', 'url': 'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getList.json', 'method': 'POST', 'desc': '查询样品管理列表信息', 'params': [
   ('startDate','创建时间开始','必填','Number'),('endDate','创建时间结束','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','状态 0待签收1待检验2待试产3待判定4已完成','选填','Integer[]')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','样品列表','Json数组'),('stkXkNo','样品单号','String'),('dnXkNo','送样单号','String'),('sampleDate','样品日期','Number'),('status','状态','Number'),('cableSampleTypeCode','索样类型编码','String'),('cableSampleTypeName','索样类型名称','String')]},
  {'id': '4', 'name': '样品单详情查询', 'url': 'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getDetailByStkXkNo.json', 'method': 'POST', 'desc': '查询样品单详情信息', 'params': [
   ('stkXkNo','样品单号','必填','String'),('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','样品单数据','Json对象'),('stkXkNo','样品单号','String'),('dnXkNo','送样单号','String'),('dnLineNo','送样单项次','Number'),('sampleType','样品类别 1一般物料2工装','Number'),('sampleDate','样品日期','Number'),('status','状态 0-待签收...4-已完成','Number'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('verifyUserName','检验人名称','String'),('pilotUserName','试产人名称','String'),('dnLine','送货明细列表','Json对象数组'),('verifies','检验列表','Json对象数组'),('produces','试产工艺列表','Json对象数组'),('cableSampleTypeCode','索样类型编码','String'),('cableSampleTypeName','索样类型名称','String')]},
  {'id': '5', 'name': 'PCN报告列表查询', 'url': 'https://openapi.xiekeyun.com/ecn/getList.json', 'method': 'POST', 'desc': '查询PCN报告列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','创建开始时间','必填','Number'),('endDate','创建结束时间','必填','Number'),('statusArray','状态过滤','选填','Number数组'),('changeTypeArray','变更类型过滤','选填','Number数组')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','PCN报告列表','JSON数组'),('reportXkNo','PCN报告平台编号','String'),('innerVendorCode','内部供应商编码','String'),('reportTitle','PCN报告名称','String'),('reportStatus','PCN报告状态','Number'),('changeDate','变更日期','Number'),('reportPublishTime','报告发布日期','Number'),('productSum','物料数量','Number'),('unfinishedWaitSum','未完成待交数量','Number'),('changeType','变更类型','Number')]},
  {'id': '6', 'name': 'PCN报告详情查询', 'url': 'https://openapi.xiekeyun.com/ecn/getDetail.json', 'method': 'POST', 'desc': '查询PCN报告详情数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('reportXkNo','PCN报告平台编号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','PCN报告详情','JSON对象'),('reportXkNo','PCN报告平台编号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('reportTitle','PCN报告名称','String'),('reportStatus','PCN报告状态','Number'),('changeType','变更类型','Number'),('reportPublishName','报告发布人名称','String'),('reportPublishEmail','报告发布人邮箱','String'),('contentList','PCN报告内容信息列表','JSON对象数组'),('productList','PCN报告物料信息列表','JSON对象数组'),('categoryCode','供应商分类编码','String'),('categoryName','供应商分类名称','String')]},
  {'id': '7', 'name': '供应商准入报告列表查询', 'url': 'https://openapi.xiekeyun.com/admissionReport/getList.json', 'method': 'POST', 'desc': '查询供应商准入报告列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('statusArray','状态过滤','选填','Number数组'),('workflowStatus','审批状态','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','供应商准入报告列表','JSON数组'),('admissionReportXkNo','供应商准入报告平台编码','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','供应商内部名称','String'),('assessmentType','评估类型 1品类准入2供应商准入','Number'),('assessmentWay','评估方式 0线上1现场','Number'),('assessmentDate','评估日期','Number'),('status','报告状态','Number'),('inspectionResult','评审结果 1合格2有条件使用3不合格','Number')]},
  {'id': '8', 'name': '供应商准入报告详情查询', 'url': 'https://openapi.xiekeyun.com/admissionReport/getDetail.json', 'method': 'POST', 'desc': '查询供应商准入报告详情数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('admissionReportXkNo','供应商准入报告平台编号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','供应商准入报告详情','JSON对象'),('admissionReportXkNo','报告平台编码','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('assessmentType','评估类型','Number'),('assessmentWay','评估方式','Number'),('status','报告状态','Number'),('inspectionResult','评审结果','Number'),('detailList','评审项明细信息列表','JSON对象数组'),('sumInfoData','汇总信息','JSON对象'),('vendorCategoryList','供应商分类','JSON对象数组'),('prodCategoryList','准入品类列表','JSON对象数组'),('sqeInfoList','SQE列表','JSON对象数组')]}]}

# ========== 模块3: 内部商城协同 ==========
internal_mall = {
 'name': '内部商城协同',
 'interfaces': [
  {'id': '1', 'name': '内部商城引入单列表查询', 'url': 'https://openapi.xiekeyun.com/introduceOrder/getList.json', 'method': 'POST', 'desc': '查询内部商城引入单列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','引入时间的开始时间','必填','Number'),('endDate','引入时间的结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','内部商城引入单列表','JSON数组'),('inductXkNo','内部商城引入单平台单据号','String'),('status','引入单状态 0草稿1待选品2待发布3已完成4已撤回5已作废','Number')]},
  {'id': '2', 'name': '内部商城引入单详情查询', 'url': 'https://openapi.xiekeyun.com/introduceOrder/getDetail.json', 'method': 'POST', 'desc': '查询内部商城引入单详情数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('inductXkNo','内部商城引入单平台单据号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','内部商城引入单详情','JSON对象'),('inductXkNo','平台单据号','String'),('status','引入单状态','Number'),('currencyCode','币别编码','String'),('currencyName','币别名称','String'),('sourceSystem','来源 0平台5商城选购','Number'),('publishTime','发布时间','Number'),('applyCompList','适用公司列表','JSON对象数组'),('lineList','明细信息列表','JSON对象数组'),('userInfo','引入用户信息','JSON对象')]}]}

# ========== 模块4: 采购订单协同 ==========
purchase_order = {
 'name': '采购订单协同',
 'interfaces': [
  {'id': '1', 'name': '请购单列表查询', 'url': 'https://openapi.xiekeyun.com/requisition/getList.json', 'method': 'POST', 'desc': '查找请购单列表信息', 'params': [
   ('startDate','请购时间开始','必填','Number'),('endDate','请购时间结束','必填','Number'),('dateType','查询时间类型','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','请购单状态','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','请购单列表','JSON数组'),('reqXkNo','请购单携客云平台单号','String'),('reqInnerNo','请购单ERP单据号','String'),('status','单据状态','Number')]},
  {'id': '2', 'name': '请购单详情查询', 'url': 'https://openapi.xiekeyun.com/requisition/getDetail.json', 'method': 'POST', 'desc': '查询请购单详情信息', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('reqXkNo','请购单平台编号','必填','String'),('reqErpNo','请购单ERP单号','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','请购单详情','JSON对象'),('reqXkNo','请购单平台单据号','String'),('reqInnerNo','请购单ERP单据号','String'),('status','单据状态','Number'),('departmentCode','请购部门编码','String'),('departmentName','请购部门名称','String'),('lineList','单身列表','Json对象数组'),('infoList','业务操作列表','Json对象数组'),('empList','协作人列表','Json对象数组'),('productCode','产品编码','String'),('productName','物料名称','String'),('reqQty','请购数量','Number'),('demandDate','需求日期','Number'),('estimatePrice','预估单价','Number'),('estimateAmount','预估金额','Number')]},
  {'id': '3', 'name': '查询请购未转采购预估金额', 'url': 'https://openapi.xiekeyun.com/requisition/sumRequisitionAmount.json', 'method': 'POST', 'desc': '根据WBS编码统计请购未转采购预估金额汇总', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('wbsCode','WBS编码','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','汇总数据列表','JSON对象'),('wbsCode','Wbs编码','String'),('estimateAmount','请购预估金额合计','Number'),('poAmount','未同步的采购单金额汇总','Number'),('syncPoAmount','已同步的采购单金额汇总','Number')]},
  {'id': '4', 'name': '采购订单列表查询(已答交)', 'url': 'https://openapi.xiekeyun.com/purchaseOrder/getList.json', 'method': 'POST', 'desc': '查询已答交的采购单号列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('status','订单状态','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','采购订单列表','JSON数组'),('poErpNo','采购单ERP单号','String'),('innerVendorCode','内部供应商编码','String'),('orderStatus','订单状态','Number'),('poCreateDate','订单推到平台的时间','Number'),('answerDate','答交时间','Number'),('vStatus','答交状态 1无差异2有差异','Number'),('diffConfirmStatus','差异确认状态','Number')]},
  {'id': '5', 'name': '采购订单详情查询(答交明细)', 'url': 'https://openapi.xiekeyun.com/purchaseOrder/getDetail.json', 'method': 'POST', 'desc': '查询采购单答交明细数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('poXkNo','采购订单平台编号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','采购订单详情','JSON对象'),('poXkNo','采购单平台单号','String'),('poErpNo','采购单ERP单号','String'),('purchaseType','采购类型','Number'),('innerVendorCode','供应商编码','String'),('innerVendorName','供应商名称','String'),('orderStatus','订单状态','Number'),('currencyCode','交易币种编码','String'),('currencyName','交易币种名称','String'),('lineList','产品明细列表','Json对象数组'),('empList','协作人列表','Json对象数组'),('afterSaleOrderList','售后单列表','Json对象数组'),('answer','答交数据','Json对象'),('productCode','物料编码','String'),('productName','产品名称','String'),('purchaseQty','采购数量','Number'),('price','无税单价','Number'),('taxPrice','含税单价','Number'),('expectedDate','交货日期','Number')]},
  {'id': '6', 'name': '采购变更单列表查询', 'url': 'https://openapi.xiekeyun.com/changeOrder/getList.json', 'method': 'POST', 'desc': '查询采购变更单列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','采购变更单列表','JSON数组'),('innerVendorCode','内部供应商编码','String'),('poErpNo','采购单ERP单号','String'),('pocXkNo','变更单平台单据号','String'),('changeSeq','变更版序','String'),('diffFlag','答交是否有差异','Number')]},
  {'id': '7', 'name': '采购变更单详情查询', 'url': 'https://openapi.xiekeyun.com/changeOrder/getDetail.json', 'method': 'POST', 'desc': '查询采购变更单详情', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('changeXkNo','变更单平台编号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','采购变更单详情','JSON对象'),('pocXkNo','变更单平台单据号','String'),('changeSeq','变更版序','Number'),('status','变更单状态','Number'),('poErpNo','采购单ERP单据号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('lineList','产品明细列表','Json对象数组'),('productCode','产品编码','String'),('purchaseQty','采购数量','Number'),('price','未税单价','Number'),('changedPurchaseQty','变更后采购数量','Number'),('changedPrice','变更后未税单价','Number'),('changedFlag','是否有变更','Number')]},
  {'id': '8', 'name': '交料通知单查询', 'url': 'https://openapi.xiekeyun.com/acNotifyOrder/getList.json', 'method': 'POST', 'desc': '查询交料通知单答交列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','交料通知单列表','JSON数组'),('acnXkNo','交料通知单平台单号','String'),('acnInnerNo','交料通知单ERP单号','String'),('innerVendorCode','内部供应商编码','String'),('status','通知单状态','Number'),('answerDate','答交时间','Number')]},
  {'id': '9', 'name': '索样通知单查询', 'url': 'https://openapi.xiekeyun.com/cableSampleOrder/getList.json', 'method': 'POST', 'desc': '查询索样通知单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','索样通知单列表','JSON数组'),('ncsoXkNo','平台单据号','String'),('ncsoErpNo','ERP单据号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('status','单据状态','Number'),('sampleSendingStatus','送样状态','Number')]}]}

# ========== 模块5: 采购收货协同 ==========
receiving = {
 'name': '采购收货协同',
 'interfaces': [
  {'id': '1', 'name': '送货单号列表查询', 'url': 'https://openapi.xiekeyun.com/delivery/getNoList.json', 'method': 'POST', 'desc': '按时间范围拉取新的送货单号列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','内部供应商编码','选填','String'),('profitCenterCode','收货部门编码','选填','String'),('status','状态过滤','选填','number数组'),('logisticsStatus','物流状态','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','送货单号数据列表','JSON数组'),('deliveryNo','送货单号','String'),('deliveryType','送货类型','Number'),('innerVendorCode','内部供应商编码','String'),('purchaseType','采购类型','Number'),('status','状态','String'),('logisticsStatus','物流状态','String'),('lastOperateTime','最后变更时间','String'),('buyerName','采购员名称','String'),('buyerCode','采购员编码','String'),('deliveryDate','送货日期','String')]},
  {'id': '2', 'name': '送货单详情查询', 'url': 'https://openapi.xiekeyun.com/delivery/getDetail.json', 'method': 'POST', 'desc': '查询一个送货单明细数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号或物流单号','必填','String'),('returnVerifyFlag','是否返回送检统计数量','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','送货单数据','Json对象'),('deliveryNo','送货单号','String'),('innerVendorCode','供应商编码','String'),('innerVendorName','供应商名称','String'),('purchaseType','采购类型','Number'),('deliveryType','送货类型','Number'),('deliveryDate','送货日期','Number'),('status','送货单状态','Number'),('logisticsStatus','物流状态','Number'),('logisticsNumber','物流单号','String'),('lineList','送货单身列表','Json对象数组'),('customsData','报关资料','Json对象'),('productCode','产品编码','String'),('productName','产品名称','String'),('deliveryQty','出货数量','Number'),('price','出货产品单价','Number')]},
  {'id': '3', 'name': '条码查询', 'url': 'https://openapi.xiekeyun.com/barcode/byDeliveryNo.json', 'method': 'POST', 'desc': '查询送货单包含的所有条码数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','条码数据','JsoN对象数组'),('packageList','条码数量信息','JsoN对象数组'),('dynamicDescList','条码规则描述','Json对象数组'),('productCode','产品编码','String'),('smallBarcode','小包条码','String'),('bigBarcode','大包条码','String'),('outerBarcode','外箱条码','String'),('includeQty','包含数量','Number'),('packLevel','包装层级','Number'),('dynamicData','动态字段值','MAP')]},
  {'id': '4', 'name': '送货预约看板查询', 'url': 'https://openapi.xiekeyun.com/reserves/dayInfo.json', 'method': 'POST', 'desc': '查询一天所有供应商送货预约数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('planArrivedDate','预计到厂时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','看板数据','JsoN对象数组'),('reserveRule','规则模板','Json对象'),('reservesList','预约点亮数据','Json对象数组')]},
  {'id': '5', 'name': '送货单供应商填写的检验数据查询', 'url': 'https://openapi.xiekeyun.com/delivery/getDeliveryVendorWriteStandard.json', 'method': 'POST', 'desc': '查询送货单供应商填写的检验数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String'),('standardDesc','检测说明','选填','String'),('rangStart','合格范围开始值','选填','String'),('rangEnd','合格范围结束值','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','供应商填写检验标准列表','JsoN数组'),('dnXkNo','送货单号','String'),('lineNo','送货单项次号','Number'),('productCode','产品编码','String'),('productName','产品名称','String'),('standardName','检验指标名称','String'),('standardValue','检测值','String')]},
  {'id': '6', 'name': '客供料代收单条码查询', 'url': 'https://openapi.xiekeyun.com/signforOrder/barcode/byDeliveryNo.json', 'method': 'POST', 'desc': '查询客供料代收单包含的所有条码数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','条码数据','JsoN对象数组'),('packageList','条码数量信息','JsoN对象数组'),('dynamicDescList','条码规则描述','Json对象数组')]},
  {'id': '7', 'name': '送货单待签收列表查询', 'url': 'https://openapi.xiekeyun.com/delivery/getWaitSignList.json', 'method': 'POST', 'desc': '按送货时间范围拉取待签收的送货单列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','待签收送货单数据列表','JSON数组'),('dnXkNo','送货单号','String'),('lineNo','送货单项次','Number'),('poErpNo','订单ERP单号','String'),('poLineNo','订单项次','String'),('deliveryQty','送货数量','Number'),('innerVendorCode','内部供应商编码','String'),('productCode','物料编码','String')]},
  {'id': '8', 'name': '送样通知单号列表查询', 'url': 'https://openapi.xiekeyun.com/sampleDelivery/getList.json', 'method': 'POST', 'desc': '按时间范围拉取新的送样通知单号列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','内部供应商编码','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','送样通知单号数据列表','JSON数组'),('deliveryNo','送样通知单号','String'),('innerVendorCode','内部供应商编码','String'),('status','状态','String'),('logisticsStatus','物流状态','String'),('lastOperateTime','最后变更时间','String')]},
  {'id': '9', 'name': '送样通知单详情查询', 'url': 'https://openapi.xiekeyun.com/sampleDelivery/getDetail.json', 'method': 'POST', 'desc': '查询一个送样通知单明细数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送样通知单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','送样通知单数据','Json对象'),('deliveryNo','送样通知单号','String'),('innerVendorCode','内部供应商编码','String'),('status','状态','String'),('deliveryDate','送货日期','Number'),('lineList','送样明细列表','Json对象数组')]}]}

# ========== 模块6: 来料检验协同 ==========
inspection = {
 'name': '来料检验协同',
 'interfaces': [
  {'id': '1', 'name': '检验单列表查询', 'url': 'https://openapi.xiekeyun.com/verifyOrder/getList.json', 'method': 'POST', 'desc': '查询检验单列表信息', 'params': [
   ('startDate','创建开始时间','必填','Number'),('endDate','创建结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','检验状态','选填','Integer[]'),('innerVendorCode','供应商编码','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','检验单列表','JSON数组'),('vfoXkNo','检验单携客云平台单号','String'),('sourceOrderType','来源单据 1送货单2收货单','Number'),('resultFlag','检验结果标识','Number'),('status','检验状态','Number')]},
  {'id': '2', 'name': '检验单详情查询', 'url': 'https://openapi.xiekeyun.com/verifyOrder/getDetail.json', 'method': 'POST', 'desc': '查询检验单详情信息', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('verifyXkNo','检验单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','检验单数据','Json对象'),('vfoXkNo','检验单据号','String'),('vfoInnerNo','检验单内部单据号','String'),('sourceOrderType','来源单据','Number'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('resultFlag','检验结果标识','Number'),('status','检验状态','Number'),('dnXkNo','送货单号','String'),('lineList','检测项次列表','Json对象数组'),('lineNo','检验单项次','Number'),('productCode','产品编码','String'),('productName','产品名称','String'),('verifyQty','送检数量','Number'),('allowQty','允收数量','Number'),('compromiseQty','让步收货数量','Number'),('refusedQty','拒收数量','Number'),('detailList','不良记录详情列表','Json对象数组'),('standardRecordList','检测项目列表','Json对象数组')]},
  {'id': '3', 'name': '8D报告列表查询', 'url': 'https://openapi.xiekeyun.com/improveReport/getList.json', 'method': 'POST', 'desc': '查询8D报告列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','8D报告列表','JSON数组'),('p8dXkNo','8D报告平台编号','String'),('innerVendorCode','内部供应商编码','String'),('reportType','报告类型','Number'),('status','8D报告状态','Number'),('reportResult','报告结果','Number'),('createTime','创建时间','Number'),('productCode','物料编码','String')]},
  {'id': '4', 'name': '8D报告详情查询', 'url': 'https://openapi.xiekeyun.com/improveReport/getDetail.json', 'method': 'POST', 'desc': '查询8D报告详情数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('irXkNo','8D报告编号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','8D报告详情','JSON对象'),('p8dXkNo','8D报告平台编号','String'),('p8dInnerNo','8D报告ERP单号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('reportType','报告类型','Number'),('status','8D报告状态','Number'),('reportResult','报告结果','Number'),('productCode','物料编码','String'),('productName','物料名称','String'),('badReason','不良原因','String'),('detailInfoList','8D报告项明细信息列表','JSON对象数组'),('reportTraceList','8D报告批次检验信息列表','JSON对象数组')]}]}

# ========== 模块7: 供应计划协同 ==========
supply_plan = {
 'name': '供应计划协同',
 'interfaces': [
  {'id': '1', 'name': '供应计划排程列表查询', 'url': 'https://openapi.xiekeyun.com/schedule/getList.json', 'method': 'POST', 'desc': '查询排程单号列表', 'params': [
   ('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('statusList','状态过滤','选填','Number数组')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','排程单号数据列表','JSON数组'),('batchNo','排程批次号','String'),('oldBatchNo','父排程单号','String'),('status','状态 1已发布2部分答交3有差异4全部答交5已失效','Number'),('publishName','发布人员工姓名','String'),('publishTime','发布时间','Number'),('schedulePlanType','计划类型','Number')]},
  {'id': '2', 'name': '供应计划排程详情查询', 'url': 'https://openapi.xiekeyun.com/schedule/getDetail.json', 'method': 'POST', 'desc': '查询排程详情信息', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('scheduleXkNo','排程单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','排程详情','Json对象'),('batchNo','排程号','String'),('status','排程状态','Number'),('publishName','发布人员工姓名','String'),('publishTime','发布时间','Number'),('itemList','排程明细列表','Json对象数组'),('prodCode','物料编码','String'),('prodName','物料名称','String'),('innerVendorCode','供应商编码','String'),('detailList','排程时间明细','Json数组'),('scheduleDate','排程日期','Number'),('scheduleQty','排程量','Number'),('replyQty','答交量','Number')]},
  {'id': '3', 'name': '生产计划排程查询', 'url': 'https://openapi.xiekeyun.com/productSchedule/getList.json', 'method': 'POST', 'desc': '查询生产计划排程数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','生产计划排程列表','JSON数组'),('batchNo','排程号','String'),('status','状态','Number'),('itemList','排程明细列表','Json对象数组')]},
  {'id': '4', 'name': '供应计划看板查询', 'url': 'https://openapi.xiekeyun.com/receiveboard/getList.json', 'method': 'POST', 'desc': '查询供应计划看板数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','数据列表','JSON数组'),('productCode','物料编码','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('itemList','天维度数据明细','JSON数组'),('boardDate','看板日期','Number'),('planQty','计划数量','Number'),('answerQty','答交数量','Number'),('deliveriedQty','累计已出货量','Number'),('receivedQty','累计签收数量','Number')]},
  {'id': '5', 'name': '直运看板查询', 'url': 'https://openapi.xiekeyun.com/directDeliveryBoard/getList.json', 'method': 'POST', 'desc': '查询直运看板数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','直运看板列表','JSON数组'),('poErpNo','采购单ERP单号','String'),('productCode','物料编码','String'),('productName','物料名称','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('directType','直送类型','Number'),('deliveryQty','送货数量','Number'),('status','状态','Number')]},
  {'id': '6', 'name': '供应商库存看板查询', 'url': 'https://openapi.xiekeyun.com/prodStockBoard/getList.json', 'method': 'POST', 'desc': '查询供应商库存看板数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','数据列表','JSON数组'),('prodCode','物料编码','String'),('prodName','物料名称','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('stockQty','库存数量','Number'),('planReceiveQty','预计收货','Number'),('itemList','天维度数据明细','JSON数组')]}]}

# ========== 模块8: 应付账款协同 ==========
ap = {
 'name': '应付账款协同',
 'interfaces': [
  {'id': '1', 'name': '对账单列表查询', 'url': 'https://openapi.xiekeyun.com/vendorStatementOrder/getList.json', 'method': 'POST', 'desc': '查询对账单列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','对账单单号数据列表','JSON数组'),('recoXkNo','对账单携客云平台单号','String'),('status','单据状态','Number'),('syncStatus','同步状态','Number'),('innerVendorCode','内部供应商编码','String'),('flowStatus','审批节点状态','Number'),('updateTime','最后修改时间','Number'),('esignStatus','签章状态','Number')]},
  {'id': '2', 'name': '对账单详情查询', 'url': 'https://openapi.xiekeyun.com/vendorStatementOrder/getDetail.json', 'method': 'POST', 'desc': '查询对账单详情数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('statementXkNo','对账单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','对账单单数据','Json对象'),('recoXkNo','对账单平台单号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('totalAmount','对账总金额','Number'),('totalTaxAmount','对账含税金额','Number'),('status','单据状态','Number'),('deliveryList','出货明细列表','Json对象数组'),('returnList','仓退明细列表','Json对象数组'),('costList','费用明细列表','Json对象数组'),('vmiDeliveryList','VMI入库明细','Json对象数组'),('prepayList','预付款单','Json对象数组'),('empList','协作人列表','Json对象数组'),('productCode','产品编码','String'),('deliveryQty','出货数量','Number'),('curReconAmount','本次对账未税金额','Number')]},
  {'id': '3', 'name': '发票单列表查询', 'url': 'https://openapi.xiekeyun.com/invoiceOrder/getList.json', 'method': 'POST', 'desc': '查询发票单列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','发票单号数据列表','JSON数组'),('invoiceNo','发票单携客云平台单号','String'),('innerVendorCode','内部供应商编码','String'),('status','发票状态','Number'),('publishTime','发布时间','Number'),('lastOperateTime','最后操作时间','Number'),('syncStatus','同步状态','Number')]},
  {'id': '4', 'name': '费用单列表查询', 'url': 'https://openapi.xiekeyun.com/costOrder/getList.json', 'method': 'POST', 'desc': '查询费用单列表数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','费用单号数据列表','JSON数组'),('costXkNo','费用单携客云平台单号','String'),('costErpNo','费用单ERP单号','String'),('innerVendorCode','内部供应商编码','String'),('status','状态','Number'),('statementStatus','对账状态','Number'),('invoiceStatus','开票状态','Number')]},
  {'id': '5', 'name': '付款申请单查询', 'url': 'https://openapi.xiekeyun.com/paymentApplyOrder/getList.json', 'method': 'POST', 'desc': '查询付款申请单数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','付款申请单号数据列表','JSON数组'),('pmaXkNo','付款申请单携客云平台单号','String'),('pmaInnerNo','付款申请单ERP单据号','String'),('status','单据状态','Number')]},
  {'id': '6', 'name': '预付款单查询', 'url': 'https://openapi.xiekeyun.com/prepaymentOrder/getList.json', 'method': 'POST', 'desc': '查询预付款单数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','预付款单数据列表','JSON数组'),('ppmXkNo','预付款单平台单号','String'),('innerVendorName','内部供应商名称','String'),('prepayTotalAmount','预付总金额','Number'),('status','状态','Number'),('publishName','发布人员工姓名','String'),('publishTime','发布时间','Number')]},
  {'id': '7', 'name': '付款单查询', 'url': 'https://openapi.xiekeyun.com/receipt/getList.json', 'method': 'POST', 'desc': '查询付款单数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','付款单号数据列表','JSON数组'),('rcpXkNo','付款单携客云平台单号','String'),('innerVendorCode','内部供应商编码','String'),('status','状态','Number'),('payStatus','支付状态','Number'),('syncStatus','同步状态','Number')]},
  {'id': '8', 'name': '折让单查询', 'url': 'https://openapi.xiekeyun.com/vendorStatementDiscountOrder/getList.json', 'method': 'POST', 'desc': '查询折让单数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','折让单数据列表','JSON数组'),('disXkNo','折让单携客云平台单号','String'),('status','单据状态','Number')]},
  {'id': '9', 'name': '付款计划查询', 'url': 'https://openapi.xiekeyun.com/paymentPlanOrder/getList.json', 'method': 'POST', 'desc': '查询付款计划数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','付款计划数据列表','JSON数组'),('pmpXkNo','付款计划携客云平台单号','String'),('innerVendorCode','内部供应商编码','String'),('innerVendorName','内部供应商名称','String'),('sourceType','来源类型','Number'),('planTaxAmount','排定付款金额','Number'),('planDate','排定付款日期','Number'),('payStatus','付款状态','Number')]}]}

# ========== 模块9: 询价招标竞价 ==========
inquiry = {
 'name': '询价招标竞价',
 'interfaces': [
  {'id': '1', 'name': '询价需求单列表查询', 'url': 'https://openapi.xiekeyun.com/inquiryInnerOrder/getList.json', 'method': 'POST', 'desc': '查询询价需求单列表信息', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','询价需求单列表','JSON数组'),('inqXkNo','询价需求单携客云平台单号','String'),('status','单据状态','Number')]},
  {'id': '2', 'name': '询价需求单详情查询', 'url': 'https://openapi.xiekeyun.com/inquiryInnerOrder/getDetail.json', 'method': 'POST', 'desc': '查询询价需求单详情信息', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('inquiryXkNo','询价需求单号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','询价需求单详情','JSON对象'),('inqXkNo','询价需求单平台单据号','String'),('billType','类型 1询价2招标','Number'),('status','单据状态','Number'),('publishName','发布人名称','String'),('prodList','产品编码列表','Json对象数组'),('vendorList','询价企业列表','Json对象数组'),('productCode','物料编码','String'),('productName','物料名称','String'),('purchaseQty','采购数量','Number'),('vquopList','供应商报价信息列表','Json对象数组')]},
  {'id': '3', 'name': '询价招标单列表查询', 'url': 'https://openapi.xiekeyun.com/inquiryOrder/getList.json', 'method': 'POST', 'desc': '查询询价招标单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','询价招标单列表','JSON数组'),('inqXkNo','询价招标携客云平台单号','String'),('status','单据状态','Number'),('updateTime','单据更新时间','Number')]},
  {'id': '4', 'name': '竞价单查询', 'url': 'https://openapi.xiekeyun.com/bidOrder/getList.json', 'method': 'POST', 'desc': '查询竞价单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','竞价单列表','JSON数组'),('bidXkNo','竞价单携客云平台单号','String'),('biddingOrderType','竞价单类型','Number'),('status','单据状态','Number'),('publishName','发布人姓名','String'),('publishTime','发布时间','Number')]},
  {'id': '5', 'name': '项目预询单查询', 'url': 'https://openapi.xiekeyun.com/preProjectOrder/getList.json', 'method': 'POST', 'desc': '查询项目预询单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','项目预询单列表','JSON数组'),('ppbXkNo','预询单平台唯一编码','String'),('status','模板状态','Number'),('updateTime','单据更新时间','Number')]},
  {'id': '6', 'name': '项目招标单查询', 'url': 'https://openapi.xiekeyun.com/projectOrder/getList.json', 'method': 'POST', 'desc': '查询项目招标单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','项目招标单列表','JSON数组'),('PbXkNo','项目招标单携客云平台单号','String'),('status','单据状态','Number'),('updateTime','单据更新时间','Number')]},
  {'id': '7', 'name': '核价单查询', 'url': 'https://openapi.xiekeyun.com/pricingOrder/getList.json', 'method': 'POST', 'desc': '查询核价单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','核价单列表','JSON数组'),('prcXkNo','核价单携客云平台单号','String'),('prcErpNo','核价单ERP单据号','String'),('status','单据状态','Number'),('syncStatus','同步状态','Number'),('publishTime','发布时间','Number'),('checkTime','审核时间','Number')]},
  {'id': '8', 'name': '核价失效单查询', 'url': 'https://openapi.xiekeyun.com/pricingInvalidOrder/getList.json', 'method': 'POST', 'desc': '查询核价失效单列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','核价失效单列表','JSON数组'),('priXkNo','核价失效单携客云平台单号','String'),('priErpNo','核价失效单ERP单据号','String'),('status','单据状态','Number')]}]}

# ========== 模块10: 委外物料管理 ==========
outsource = {
 'name': '委外物料管理',
 'interfaces': [
  {'id': '1', 'name': '委外发料单详情查询', 'url': 'https://openapi.xiekeyun.com/outsourceIssuanceOrder/getDetail.json', 'method': 'POST', 'desc': '查询委外发料单详情', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('oioXkNo','委外发料单平台单号','选填','String'),('oioErpNo','委外发料单ERP单号','选填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','委外发料单详情','Json对象'),('oioXkNo','委外发料单平台单号','String'),('oioErpNo','委外发料单ERP单号','String'),('innerVendorCode','供应商编码','String'),('innerVendorName','供应商名称','String'),('status','状态','Number'),('lineList','单身信息列表','Json对象数组'),('productCode','产品编码','String'),('productName','产品名称','String'),('issuanceQty','发料数量','Number')]},
  {'id': '2', 'name': '委外物料盘点单查询', 'url': 'https://openapi.xiekeyun.com/outsourceInventoryOrder/getList.json', 'method': 'POST', 'desc': '查询委外物料盘点单数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','盘点单列表','JSON数组')]}]}

# ========== 模块11: 工装协同 ==========
tooling = {
 'name': '工装协同',
 'interfaces': [
  {'id': '1', 'name': '工装信息变更单列表查询', 'url': 'https://openapi.xiekeyun.com/toolsInfo/change/queryList.json', 'method': 'POST', 'desc': '查询工装信息变更单列表', 'params': [
   ('startDate','单据更新时间范围的开始时间','必填','Number'),('endDate','单据更新时间范围的结束时间','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','工装信息变更单数据列表','JSON数组'),('tcoXkNo','工装信息变更单平台唯一编码','String'),('toolingsUniqueCode','工装信息平台唯一编码','String'),('toolingsCode','工装编码','String'),('toolingsName','工装名称','String'),('status','状态','Number')]},
  {'id': '2', 'name': '工装信息变更单详情查询', 'url': 'https://openapi.xiekeyun.com/toolsInfo/change/detail.json', 'method': 'POST', 'desc': '查询工装信息变更单详情', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('tcoXkNo','工装信息变更单平台唯一编码','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','工装信息变更单详情','JSON数组'),('tcoXkNo','变更单平台唯一编码','String'),('toolingsCode','工装编码','String'),('toolingsName','工装名称','String'),('status','状态','Number'),('remark','备注','String'),('changedRemark','变更后备注','String'),('publishName','发布人名称','String'),('publishTime','发布时间','Number')]},
  {'id': '3', 'name': '工装档案查询', 'url': 'https://openapi.xiekeyun.com/toolsInfo/getDetail.json', 'method': 'POST', 'desc': '查询工装档案数据', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('toolingsUniqueCode','工装信息平台唯一编码','必填','String')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('data','工装档案数据','JSON对象'),('toolingsUniqueCode','工装信息平台唯一编码','String'),('toolingsCode','工装编码','String'),('toolingsName','工装名称','String'),('toolingsTypeName','工装类型名称','String'),('status','状态','Number')]}]}

# ========== 模块12: 附件更新查询 ==========
attachment = {
 'name': '附件更新查询',
 'interfaces': [
  {'id': '1', 'name': '获取供应商新上传了附件的单据列表', 'url': 'https://openapi.xiekeyun.com/orderHasUploadFile/queryByLastDate.json', 'method': 'POST', 'desc': '查询某类型单据在时间范围内供应商有操作附件的单据编码列表', 'params': [
   ('erpCode','请求者用户ERP帐号','必填','String'),('startDate','时间范围的开始时间','必填','Number'),('endDate','时间范围的结束时间','必填','Number'),('billType','单据类型 10:采购订单','必填','Number'),('uploadType','附件上传者类型 默认2','选填','Number')], 'responses': [
   ('result','返回结果状态','Number'),('errorCode','接口错误编码','String'),('errorMsg','接口错误信息','String'),('dataList','单据编码列表','JSON数组')]}]}

all_modules = [vendor_mgmt, lifecycle, internal_mall, purchase_order, receiving, inspection, supply_plan, ap, inquiry, outsource, tooling, attachment]

def generate_excel(output_path):
    wb = openpyxl.Workbook()
    ws_overview = wb.active; ws_overview.title = '接口总览'
    write_overview(ws_overview, all_modules)
    for module in all_modules:
        for api in module['interfaces']:
            short_name = api['name']
            if len(short_name) > 18: short_name = short_name[:18] + '..'
            sheet_name = f"{module['name']}_{api['id']}_{short_name}"
            if len(sheet_name) > 31: sheet_name = sheet_name[:31]
            ws = wb.create_sheet(title=sheet_name)
            write_subsheet(ws, api)
    wb.save(output_path)
    print(f'Excel generated: {output_path}')

if __name__ == '__main__':
    output = 'G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\SRM_采购协同_接口梳理.xlsx'
    generate_excel(output)
