#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""SRM采购协同 - 完整嵌套字段Excel"""
import openpyxl
from openpyxl.styles import Font,PatternFill,Alignment,Border,Side

DEEP_BLUE=PatternFill(start_color='1F4E79',end_color='1F4E79',fill_type='solid')
LIGHT_BLUE=PatternFill(start_color='D6E4F0',end_color='D6E4F0',fill_type='solid')
LIGHT_GRAY=PatternFill(start_color='F2F2F2',end_color='F2F2F2',fill_type='solid')
WHITE_B=Font(name='微软雅黑',size=10,bold=True,color='FFFFFF')
BOLD_F=Font(name='微软雅黑',size=10,bold=True)
TITLE_F=Font(name='微软雅黑',size=11,bold=True)
NORMAL_F=Font(name='微软雅黑',size=10)
THIN_B=Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))
WRAP=Alignment(wrap_text=True,vertical='center')
CENTER=Alignment(wrap_text=True,vertical='center',horizontal='center')
C_W={'A':28,'B':52,'C':14,'D':20}
OV_C_W={'A':10,'B':36,'C':60,'D':12,'E':50}

def ac(c,font=NORMAL_F,fill=None,align=WRAP,border=THIN_B):
    c.font=font;c.alignment=align;c.border=border
    if fill:c.fill=fill

def write_overview(ws,mods):
    for k,v in OV_C_W.items():ws.column_dimensions[k].width=v
    ws.merge_cells('A1:E1');c=ws['A1'];c.value='携客云SRM采购协同平台 - API接口总览';c.font=Font(name='微软雅黑',size=14,bold=True);c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        ac(ws.cell(row=2,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[2].height=22;row=3
    for mod in mods:
        ws.merge_cells(f'A{row}:E{row}')
        ac(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
        for c_ in 'BCDE':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
        ws.row_dimensions[row].height=22;row+=1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,LIGHT_GRAY,CENTER if i in(1,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    ws.auto_filter.ref=f'A2:E{row-1}'

def write_subs(ws,api):
    for k,v in C_W.items():ws.column_dimensions[k].width=v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        ac(ws.cell(row=idx,column=1,value=l),BOLD_F);ws.merge_cells(f'B{idx}:D{idx}')
        ac(ws.cell(row=idx,column=2,value=v))
        for c_ in 'CD':ac(ws[f'{c_}{idx}'])
        ws.row_dimensions[idx].height=22
    row=5;ws.row_dimensions[4].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='请求参数'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;row+=1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        ac(ws.cell(row=row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[row].height=22;row+=1
    if api.get('params'):
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                ac(ws.cell(row=row,column=i,value=val),NORMAL_F,None,CENTER if i in(3,4)else WRAP)
            ws.row_dimensions[row].height=20;row+=1
    else:
        ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'),NORMAL_F)
        ws.row_dimensions[row].height=20;row+=1
    row+=2;ws.row_dimensions[row-1].height=6
    ws.merge_cells(f'A{row}:D{row}');ac(ws.cell(row=row,column=1,value='响应数据字段'),TITLE_F,LIGHT_BLUE,Alignment(vertical='center',wrap_text=True))
    for c_ in 'BCD':ac(ws[f'{c_}{row}'],NORMAL_F,LIGHT_BLUE,WRAP)
    ws.row_dimensions[row].height=22;hdr_row=row+1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        ac(ws.cell(row=hdr_row,column=i,value=h),WHITE_B,DEEP_BLUE,CENTER)
    ws.row_dimensions[hdr_row].height=22;row=hdr_row+1
    rl=[]
    for item in api['responses']:
        lvl=item[0];rname=('  '*lvl)+item[1]if lvl>0 else item[1];rdesc=item[2]if len(item)>2 else'';rtype=item[3]if len(item)>3 else''
        ac(ws.cell(row=row,column=1,value=rname),NORMAL_F);ac(ws.cell(row=row,column=2,value=rdesc),NORMAL_F)
        ac(ws.cell(row=row,column=3,value=''),NORMAL_F,None,CENTER)
        ac(ws.cell(row=row,column=4,value=rtype),NORMAL_F,None,CENTER)
        rl.append((row,lvl));ws.row_dimensions[row].height=20;row+=1
    if rl:
        max_lvl=max(l for _,l in rl)
        for lvl in range(max_lvl,0,-1):
            i=0
            while i<len(rl):
                r,l=rl[i]
                if l==lvl:
                    j=i+1
                    while j<len(rl)and rl[j][1]>=lvl:j+=1
                    if j-1>i:ws.row_dimensions.group(r,rl[j-1][0],outline_level=lvl)
                    i=j
                else:i+=1

# ============== ALL MODULES ==============

# M1: 供应商管理
VM = {
'name':'供应商管理','interfaces':[
{'id':'1','name':'查询已加入的供应商列表','url':'https://openapi.xiekeyun.com/compVendor/confirmList.json','method':'POST','desc':'查询邀约状态为已加入的供应商列表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','选填','String'),('handshakeTimeStart','握手查询开始时间','选填','Number'),('handshakeTimeEnd','握手查询结束时间','选填','Number'),('categoryCode01','供应商分类1编码','选填','String'),('categoryCode02','供应商分类2编码','选填','String'),('categoryCode03','供应商分类3编码','选填','String'),('performanceClassificationCode','绩效分类编码','选填','String')],
 'responses':[(0,'result','返回结果状态 1成功9失败','Number'),(0,'errorCode','接口错误编码(失败时有值)','String'),(0,'errorMsg','接口错误信息(失败时有值)','String'),(0,'dataList','供应商档案集合','Json数组'),(1,'vendorXkUniqueCode','供应商平台唯一编码','String'),(1,'innerVendorCode','供应商编码','String'),(1,'innerVendorName','供应商名称','String'),(1,'employeeName','邀约供应商的员工姓名','String'),(1,'handshakeTime','握手时间(时间戳)','Number'),(1,'categoryCode01','供应商分类1编码','String'),(1,'categoryName01','供应商分类1名称','String'),(1,'categoryCode02','供应商分类2编码','String'),(1,'categoryName02','供应商分类2名称','String'),(1,'categoryCode03','供应商分类3编码','String'),(1,'categoryName03','供应商分类3名称','String'),(1,'performanceClassificationCode','绩效分类编码','String'),(1,'performanceClassificationName','绩效分类名称','String')]},
{'id':'2','name':'查询供应商绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore.json','method':'POST','desc':'查询供应商月度绩效考评结果列表',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表(限长50)','必填','数组'),('timeType','绩效时间类型 1月份2季度3年份','必填','Number'),('yearTime','年份 2022,2023...','必填','Number'),('yearType','年份类型 1上半年2下半年3全年','选填','Number'),('monthTime','月份 01-12','选填','String'),('quarterTime','季度 1-4','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'categoryCode','分类编码','String'),(1,'categoryName','分类名称','String'),(1,'assessmentMonth','绩效月份 yyyyMM','String'),(1,'assessmentScore','绩效得分','Number'),(1,'performanceScore','绩效评分','Number'),(1,'deductionScore','绩效扣分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'vendorCategoryList','供应商分类列表','json数组'),(2,'categoryType','分类类型(0001-0004)','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String')]},
{'id':'3','name':'供应商档案详情','url':'https://openapi.xiekeyun.com/compVendor/getDetail.json','method':'POST','desc':'查询一个供应商详情明细数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码(ERP内)','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商详情','Json对象'),(1,'xkUniqueCode','供应商平台编码(32位)','String'),(1,'vendorCode','企业号','String'),(1,'vendorName','企业名称','String'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'innerVendorAbbr','内部供应商简称','String'),(1,'inviteStatus','邀约状态 0未发起1已发起2已确认3失败','Number'),(1,'handshakeTime','握手成功时间(时间戳)','Number'),(1,'vendorStatus','供应商状态 0无效1有效3黑名单','Number'),(1,'buyerName','采购员名称','String'),(1,'lifeCycle','生命周期 0准入评估1试样2小批量3正式4退出','Number'),(1,'gradeInfo','供应商绩效等级 A/B/C/D','String'),(1,'contacts','联系人明细列表','Json对象数组'),(2,'fullName','姓名','String'),(2,'departmentName','部门名称','String'),(2,'positionName','职位名称','String'),(2,'phoneNo','联系电话','String'),(2,'mobile','手机号','String'),(2,'email','邮箱','String'),(2,'mainFlag','是否主要联系人 0否1是','Number'),(1,'vendorCategoryList','供应商分类/准入品类列表','Json对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(2,'status','状态 1有效2备用3暂停4退出','Number'),(1,'orgList','供应商组织关系列表','Json对象数组'),(2,'innerVendorCode','内部供应商编码(ERP)','String'),(2,'isOverseasDirect','是否境外直送 0否1是','Number'),(2,'innerTradePartnerCode','贸易伙伴供应商编码','String'),(2,'innerTradePartnerName','贸易伙伴供应商名称','String'),(2,'purchaseOrgCode','采购组织编码','String'),(2,'purchaseOrgName','采购组织名称','String'),(2,'currencyCode','交易币种编码','String'),(2,'currencyName','交易币种名称','String'),(2,'payWayCode','付款条件编码','String'),(2,'payWayName','付款条件名称','String'),(2,'status','有效否 0无效1有效','Number'),(2,'priceConditionList','价格条件列表','Json对象数组'),(3,'profitCenterCode','工厂编码','String'),(3,'profitCenterName','工厂名称','String'),(3,'erpPriceConditionCode','价格条件编码','String'),(3,'erpPriceConditionName','价格条件名称','String'),(3,'currencyCode','交易币种编码','String'),(3,'currencyName','交易币种名称','String'),(3,'payWayCode','付款条件编码','String'),(3,'payWayName','付款条件名称','String'),(1,'sqeList','SQE列表','Json对象数组'),(2,'sqeCode','SQE编码','String'),(2,'sqeName','SQE名称','String'),(1,'companyEnglishName','公司英文名称','String'),(1,'groupCompanyName','集团公司名称','String')]},
{'id':'4','name':'查询供应商应用度报表','url':'https://openapi.xiekeyun.com/compVendor/report.json','method':'POST','desc':'查询供应商应用报表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','Number'),('year','年份','必填','Number'),('month','月份','必填','Number'),('type','查询类型 1公司2组织','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商应用度报表','Json对象'),(1,'reportMonth','报表月份 yyyyMM','String'),(1,'vendorTotalSum','供应商数量','Number'),(1,'vendorHandShakeSum','已握手供应商数量','Number'),(1,'reportList','供应商应用度报表列表','Json对象数组'),(2,'companyCode','企业号','String'),(2,'innerVendorCode','内部供应商编码','String'),(2,'innerVendorName','内部供应商名称','String'),(2,'currOrderSum','本月订单数','Number'),(2,'orderHandRate','订单处理率','Number'),(2,'onlineDeliveryRate','在线送货率','Number'),(2,'barPrintRate','条码打印率','Number'),(2,'onlineReconciliationRate','在线对账率','Number'),(2,'enquiryJoinRate','询价参与率','Number'),(2,'biddingJoinRate','竞价参与率','Number')]},
{'id':'5','name':'查询集团法人绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/list.json','method':'POST','desc':'查询集团法人绩效考评结果列表',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表(限长50)','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'assessmentScore','绩效得分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'relationCompanyList','关联公司列表','json数组'),(2,'relationCompanyCode','关联企业编码','String'),(2,'relationCompanyName','关联企业名称','String'),(2,'assessmentScore','绩效总得分','Number'),(2,'gradeInfo','评级','String'),(2,'relAsDetailList','关联公司绩效明细项','json数组'),(3,'curSrcScore','本项达成率(%)','Number'),(3,'curTargetScore','本项得分','Number'),(3,'assementStandardItemCode','绩效考核项编码','String'),(3,'assementStandardItemName','绩效考核项名称','String')]},
{'id':'6','name':'查询集团物料分类绩效考评结果','url':'https://openapi.xiekeyun.com/compVendor/assessmentScore/relation/company/prodCategory/list.json','method':'POST','desc':'查询集团物料分类绩效考评结果列表',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','供应商编码','必填','String'),('innerVendorCodeList','供应商编码列表(限长50)','必填','数组'),('timeType','绩效时间类型','必填','Number'),('yearTime','年份','必填','Number'),('yearType','年份类型','选填','Number'),('monthTime','月份','选填','String'),('quarterTime','季度','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商绩效得分列表','Json对象'),(1,'prodCategoryCode','物料分类编码','String'),(1,'prodCategoryName','物料分类名称','String'),(1,'assessmentScore','绩效得分','Number'),(1,'gradeInfo','评级','String'),(1,'detailList','绩效明细项','json数组'),(2,'curSrcScore','本项达成率(%)','Number'),(2,'curTargetScore','本项得分','Number'),(2,'assementStandardItemCode','绩效考核项编码','String'),(2,'assementStandardItemName','绩效考核项名称','String'),(1,'vendorAsList','供应商列表','json数组'),(2,'innerVendorCode','供应商编码','String'),(2,'innerVendorName','供应商名称','String'),(2,'assessmentScore','绩效总得分','Number'),(2,'gradeInfo','评级','String'),(2,'vendorDetailList','供应商绩效明细项','json数组'),(3,'curSrcScore','本项达成率(%)','Number'),(3,'curTargetScore','本项得分','Number'),(3,'assementStandardItemCode','绩效考核项编码','String'),(3,'assementStandardItemName','绩效考核项名称','String')]}]}

# M2: 生命周期管理
LC = {
'name':'生命周期管理','interfaces':[
{'id':'1','name':'潜在供应商列表查询','url':'https://openapi.xiekeyun.com/potentialVendor/getList.json','method':'POST','desc':'查询企业潜在供应商列表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('dateStart','资质审核完成时间-从','选填','Number'),('dateEnd','资质审核完成时间-至','选填','Number'),('createStart','创建时间-从','选填','Number'),('createEnd','创建时间-至','选填','Number'),('checkTimeStart','审核时间-从','选填','Number'),('checkTimeEnd','审核时间-至','选填','Number'),('turnFormalTimeStart','转正时间-从','选填','Number'),('turnFormalTimeEnd','转正时间-至','选填','Number'),('applyStatus','申请状态 0草稿1申请中2拒绝3已确认4已转正式...','选填','Number数组'),('inviteStatus','邀约状态 0未发起1已发起2已确认3失败4待采购审核5退回','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','潜在供应商数据列表','Json对象'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'certificateAuditFinishTime','资质审核完成时间','Number'),(1,'admissionReportFlag','完成准入评估标识 0未完成1已完成','Number'),(1,'applyStatus','申请状态','Number'),(1,'vendorStatus','供应商状态 0无效1有效2停用3黑名单','Number'),(1,'inviteStatus','邀约状态','Number'),(1,'checkTime','审核时间','Number'),(1,'updateTime','最后更新时间','Number')]},
{'id':'2','name':'潜在供应商详情查询','url':'https://openapi.xiekeyun.com/potentialVendor/getDetail.json','method':'POST','desc':'查询一个潜在供应商明细数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('innerVendorCode','潜在供应商编码','必填','String'),('approvalStatusFlag','是否返回审批状态 0不返回1返回','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','潜在供应商详情','Json对象'),(1,'xkUniqueCode','内部供应商平台编码','String'),(1,'innerVendorCode','内部供应商编码(ERP)','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'innerVendorAbbr','内部供应商简称','String'),(1,'applyStatus','审批状态 3已确认4已转正式','number'),(1,'vendorStatus','供应商状态','number'),(1,'inviteStatus','邀约状态','number'),(1,'handshakeTime','握手成功时间','Number'),(1,'creditCode','统一社会信用代码','String'),(1,'paymentDaysCode','账期编码','String'),(1,'paymentDaysName','账期名称','String'),(1,'contacts','联系人明细列表','Json对象数组'),(2,'fullName','姓名','String'),(2,'departmentName','部门名称','String'),(2,'phoneNo','联系电话','String'),(2,'mobile','手机号','String'),(2,'email','邮箱','String'),(2,'mainFlag','是否主要联系人','Number'),(1,'purchaseOrgList','采购组织明细列表','Json对象数组'),(2,'purchaseOrgCode','采购组织编码','String'),(2,'purchaseOrgName','采购组织名称','String'),(2,'currencyCode','交易币种编码','String'),(2,'currencyName','交易币种名称','String'),(1,'financeList','企业资料明细列表','Json对象数组'),(2,'erpCompanyCode','公司代码编码','String'),(2,'erpCompanyName','公司代码名称','String'),(2,'payWayCode','付款条件编码','String'),(2,'currencyCode','交易币种编码','String'),(1,'bankList','银行明细列表','Json对象数组'),(2,'bankCardNo','银行帐号','String'),(2,'bankTypeName','银行类别名称','String'),(2,'unionBankNo','联行号','String'),(1,'vendorCategoryList','供应商分类/准入品类列表','Json对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(1,'mainProductList','主要供应产品','Json对象数组'),(2,'prodName','产品名称','String'),(2,'prodCapacity','产品产能','String'),(1,'keyManufacturingCapabilityList','关键制造能力','Json对象数组'),(2,'equipmentName','设备名称','String'),(2,'equipmentQty','设备数量','Number'),(1,'companyEnglishName','公司英文名称','String'),(1,'groupCompanyName','集团公司名称','String')]},
{'id':'3','name':'样品管理列表查询','url':'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getList.json','method':'POST','desc':'查询样品管理列表信息',
 'params':[('startDate','创建时间开始','必填','Number'),('endDate','创建时间结束','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','状态 0待签收1待检验2待试产3待判定4已完成','选填','Integer[]')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','样品列表','Json数组'),(1,'stkXkNo','样品单号','String'),(1,'dnXkNo','送样单号','String'),(1,'sampleDate','样品日期','Number'),(1,'status','状态 0待签收1待检验2待试产3待判定4已完成','Number'),(1,'cableSampleTypeCode','索样类型编码','String'),(1,'cableSampleTypeName','索样类型名称','String')]},
{'id':'4','name':'样品单详情查询','url':'https://openapi.xiekeyun.com/vendorSampleTrackOrder/getDetailByStkXkNo.json','method':'POST','desc':'查询样品单详情信息',
 'params':[('stkXkNo','样品单号','必填','String'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','样品单数据','Json对象'),(1,'stkXkNo','样品单号','String'),(1,'dnXkNo','送样单号','String'),(1,'dnLineNo','送样单项次','Number'),(1,'sampleType','样品类别 1一般物料2工装','Number'),(1,'sampleDate','样品日期','Number'),(1,'trialOrderNo','试产工单号','String'),(1,'trialDate','试产日期','Number'),(1,'sampleResultFlag','样品是否合格 0不合格1合格','Number'),(1,'status','状态 0待签收1待检验2待试产3待判定4已完成','Number'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'verifyUserName','检验人名称','String'),(1,'pilotUserName','试产人名称','String'),(1,'verifyTime','检验完成时间','Number'),(1,'pilotTime','试产完成时间','Number'),(1,'cableSampleTypeCode','索样类型编码','String'),(1,'cableSampleTypeName','索样类型名称','String'),(1,'dnLine','送货明细列表','Json对象数组'),(2,'dnXkNo','送样单号','String'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'productScale','产品规格','String'),(2,'deliveryDate','答交日期','Number'),(2,'deliveryQty','出货数量','Number'),(1,'verifies','检验列表','Json对象数组'),(2,'verifyType','检验类别 1检验记录2试产检验','Number'),(2,'standardName','检验指标名称','String'),(2,'standardResultFlag','指标是否合格 0不合格1合格','Number'),(2,'resultDesc','检验结果','String'),(1,'produces','试产工艺列表','Json对象数组'),(2,'paramName','工艺参数名称','String'),(2,'paramValue','工艺参数取值','String')]},
{'id':'5','name':'PCN报告列表查询','url':'https://openapi.xiekeyun.com/ecn/getList.json','method':'POST','desc':'查询PCN报告列表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('startDate','创建开始时间','必填','Number'),('endDate','创建结束时间','必填','Number'),('statusArray','状态 0草稿1待SQM2待IQC4已审核5撤回6退回7作废8结案9待RD','选填','Number数组'),('changeTypeArray','变更类型 1立即2自然3批量4指定日期','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','PCN报告列表','JSON数组'),(1,'reportXkNo','PCN报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'reportTitle','PCN报告名称','String'),(1,'reportStatus','PCN报告状态','Number'),(1,'changeDate','变更日期(时间戳)','Number'),(1,'reportPublishTime','报告发布日期','Number'),(1,'productSum','物料数量','Number'),(1,'unfinishedWaitSum','未完成待交数量','Number'),(1,'changeType','变更类型','Number')]},
{'id':'6','name':'PCN报告详情查询','url':'https://openapi.xiekeyun.com/ecn/getDetail.json','method':'POST','desc':'查询PCN报告详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('reportXkNo','PCN报告平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','PCN报告详情','JSON对象'),(1,'reportXkNo','PCN报告平台编号','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'reportTitle','PCN报告名称','String'),(1,'reportStatus','PCN报告状态 0草稿1待SQM2待IQC4已审核5撤回6退回7作废8结案9待RD','Number'),(1,'changeType','变更类型 1立即2自然3批量4指定日期5变更结束日期','Number'),(1,'changeDate','变更日期','Number'),(1,'changeDateEnd','变更结束日期','Number'),(1,'categoryCode','供应商分类编码','String'),(1,'categoryName','供应商分类名称','String'),(1,'reportPublishName','报告发布人名称','String'),(1,'reportPublishEmail','报告发布人邮箱','String'),(1,'reportPublishTime','报告发布日期','Number'),(1,'productSum','物料数量','Number'),(1,'unfinishedWaitSum','未完成待交数量','Number'),(1,'contentList','报告内容信息列表','JSON对象数组'),(2,'changeType','变更类型','String'),(2,'changeOption','变更项','String'),(2,'changeBefore','变更前','String'),(2,'changeAfter','变更后','String'),(2,'changeReason','变更原因','String'),(2,'alternativeProduct','推荐替代物料','String'),(2,'lastReceiveDate','最后接单日期','Number'),(2,'lastDeliveryDate','最后出货日期','Number'),(1,'productList','报告物料信息列表','JSON对象数组'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'productScale','物料规格','String'),(2,'brandName','品牌名称','String'),(2,'manufacturerProdCode','制造商料号','String'),(2,'unfinishedOrderNum','未完成订单数量','Number'),(2,'unfinishedWaitNum','未完成订单待交数量','Number'),(1,'remarks','备注说明','String')]},
{'id':'7','name':'供应商准入报告列表查询','url':'https://openapi.xiekeyun.com/admissionReport/getList.json','method':'POST','desc':'查询供应商准入报告列表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间','必填','Number'),('endDate','结束时间','必填','Number'),('statusArray','状态 0待发出1待回复2待评审3退回4待审批5已完成6作废','选填','Number数组'),('workflowStatus','审批状态 0待送审1审批中2审批完成','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','供应商准入报告列表','JSON数组'),(1,'admissionReportXkNo','准入报告平台编码','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','供应商内部名称','String'),(1,'assessmentType','评估类型 1品类准入2供应商准入','Number'),(1,'assessmentWay','评估方式 0线上1现场','Number'),(1,'assessmentDate','评估日期','Number'),(1,'status','报告状态','Number'),(1,'publishName','发布用户名称','String'),(1,'publishTime','发布时间','Number'),(1,'inspectionResult','评审结果 1合格2有条件3不合格','Number')]},
{'id':'8','name':'供应商准入报告详情查询','url':'https://openapi.xiekeyun.com/admissionReport/getDetail.json','method':'POST','desc':'查询供应商准入报告详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('admissionReportXkNo','供应商准入报告平台编号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','供应商准入报告详情','JSON对象'),(1,'admissionReportXkNo','准入报告平台编码','String'),(1,'innerVendorCode','内部供应商编码','String'),(1,'innerVendorName','内部供应商名称','String'),(1,'assessmentType','评估类型 1品类准入2供应商准入3年度复评4QSA稽核','Number'),(1,'assessmentWay','评估方式 0线上1现场','Number'),(1,'assessmentDate','评估开始日期','Number'),(1,'assessmentEndDate','评估结束日期','Number'),(1,'status','报告状态','Number'),(1,'inspectionResult','评审结果','Number'),(1,'publishName','发布用户名称','String'),(1,'publishTime','发布时间','Number'),(1,'improveDate','改善日期','Number'),(1,'validTime','有效到期时间','Number'),(1,'lifeCycle','生命周期 1试样2小批量3正式','String'),(1,'grade','评分等级','String'),(1,'vendorCategoryCode','供应商ERP分类编码','String'),(1,'vendorCategoryName','供应商ERP分类名称','String'),(1,'detailList','评审项明细信息列表','JSON对象数组'),(2,'inspectionItemCode','评审项编码','String'),(2,'itemName','评审项目名称','String'),(2,'scoreDescription','评审说明','String'),(2,'ownScore','自评分','Number'),(2,'inspectionScore','评审得分','Number'),(2,'itemScore','评审项总分','Number'),(2,'scoreEmployeeName','评分人员工姓名','String'),(2,'scoreEmployeeNumber','评分人工号','String'),(1,'sumInfoData','汇总信息','JSON对象'),(2,'totalItemScore','总分','String'),(2,'totalInspectionScore','评审得分','String'),(2,'totalScoreRate','总得分','String'),(2,'resultScore','结果得分','String'),(1,'vendorCategoryList','供应商分类','JSON对象数组'),(2,'categoryType','分类类型','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(1,'prodCategoryList','准入品类列表','JSON对象数组'),(2,'categoryType','分类类型(物料ERP分类)','String'),(2,'categoryCode','分类编码','String'),(2,'categoryName','分类名称','String'),(2,'status','审核状态 0不同意1同意','Number'),(1,'sqeInfoList','SQE列表','JSON对象数组'),(2,'sqeCode','SQE员工编码','String'),(2,'sqeName','SQE员工名称','String')]}]}

# M3: 内部商城协同
IM = {
'name':'内部商城协同','interfaces':[
{'id':'1','name':'内部商城引入单列表查询','url':'https://openapi.xiekeyun.com/introduceOrder/getList.json','method':'POST','desc':'查询内部商城引入单列表数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('startDate','引入时间开始','必填','Number'),('endDate','引入时间结束','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','内部商城引入单列表','JSON数组'),(1,'inductXkNo','平台单据号','String'),(1,'status','状态 0草稿1待选品2待发布3已完成4已撤回5已作废','Number')]},
{'id':'2','name':'内部商城引入单详情查询','url':'https://openapi.xiekeyun.com/introduceOrder/getDetail.json','method':'POST','desc':'查询内部商城引入单详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('inductXkNo','引入单平台单据号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','引入单详情','JSON对象'),(1,'inductXkNo','平台单据号','String'),(1,'profitCenterCode','利润中心编码','String'),(1,'profitCenterName','利润中心名称','String'),(1,'status','引入单状态','Number'),(1,'currencyCode','币别编码','String'),(1,'currencyName','币别名称','String'),(1,'sourceSystem','来源 0平台5商城选购','Number'),(1,'publishTime','发布时间','Number'),(1,'introduceTime','引入时间','Number'),(1,'applyCompList','适用公司列表','JSON对象数组'),(2,'applyCompanyName','适用公司名称','String'),(1,'lineList','明细信息列表','JSON对象数组'),(2,'lineNo','项次','Number'),(2,'detailLineNo','详情项次','Number'),(2,'innerVendorCode','内部供应商编码','String'),(2,'innerVendorName','内部供应商名称','String'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'productScale','物料规格','String'),(2,'price','未税价','Number'),(2,'taxPrice','含税价','Number'),(2,'taxCode','交易税种编码','String'),(2,'taxRate','交易税率','Number'),(2,'unitCode','单位编码','String'),(2,'unitName','单位名称','String'),(2,'deliveryTime','承诺交期','Number'),(2,'mroProductName','商城产品名称','String'),(2,'mroBrandName','品牌名称','String'),(2,'productStatus','状态 0过期1正常','Number'),(1,'userInfo','引入用户信息','JSON对象'),(2,'employeeName','员工姓名','String'),(2,'erpAccount','ERP账号','String'),(2,'employeeNumber','工号','String'),(2,'departmentName','部门','String'),(2,'mobile','手机号','String')]}]}

print('Modules loaded successfully')
MODS = [VM, LC, IM]

def gen(path):
    wb=openpyxl.Workbook()
    ws=wb.active;ws.title='接口总览'
    write_overview(ws,MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn=api['name']
            if len(sn)>18:sn=sn[:18]+'..'
            nm=f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31:nm=nm[:31]
            write_subs(wb.create_sheet(title=nm),api)
    wb.save(path);print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\SRM_采购协同_接口梳理.xlsx')
