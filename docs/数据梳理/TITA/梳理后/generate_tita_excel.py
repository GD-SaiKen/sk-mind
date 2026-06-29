"""
TITA 查询接口梳理 Excel 生成器
读取 TITA 查询接口清单 → 生成规范Excel（总览表+子表）
"""
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── 统一样式常量 ──
DEEP_BLUE   = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE  = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
WHITE_FILL  = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
LIGHT_GRAY  = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
NEST_L1     = PatternFill(start_color='F4F7FC', end_color='F4F7FC', fill_type='solid')
NEST_L2     = PatternFill(start_color='EBEFF7', end_color='EBEFF7', fill_type='solid')

WHITE_BOLD  = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_10     = Font(name='微软雅黑', size=10, bold=True)
BOLD_11     = Font(name='微软雅黑', size=11, bold=True)
NORMAL_10   = Font(name='微软雅黑', size=10)
TITLE_14    = Font(name='微软雅黑', size=14, bold=True)

THIN_BORDER = Border(
    left=Side(style='thin', color='B0B0B0'),
    right=Side(style='thin', color='B0B0B0'),
    top=Side(style='thin', color='B0B0B0'),
    bottom=Side(style='thin', color='B0B0B0'),
)
NO_BORDER = Border()

ALIGN_LEFT   = Alignment(horizontal='left',   vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

OUT = "docs/数据梳理/TITA/梳理后/TITA查询接口梳理.xlsx"
BASE_URL = "https://oapi.tita.com"

# ── 接口数据定义 ──

INTERFACES = [
    # ===== CRM API (8个) =====
    {
        "category": "CRM API",
        "module": "查询商机",
        "url": "/crm/api/{tenantId}/business/search",
        "method": "GET",
        "desc": "查询商机列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "errorMsg", "type": "string", "desc": "错误信息"},
            {"name": "data.total", "type": "int", "desc": "总条数"},
            {"name": "data.records", "type": "list", "desc": "数据列表"},
            {"name": "  records[].id", "type": "string", "desc": "主键id"},
            {"name": "  records[].name", "type": "string", "desc": "商机名称"},
            {"name": "  records[].principalUserId", "type": "string", "desc": "负责人id"},
            {"name": "  records[].amount", "type": "double", "desc": "预计金额"},
            {"name": "  records[].stage", "type": "string", "desc": "商机阶段"},
            {"name": "  records[].customerId", "type": "string", "desc": "客户id"},
            {"name": "  records[].competitors", "type": "string", "desc": "对手名称"},
            {"name": "  records[].finishDealTime", "type": "string", "desc": "预计成交时间"},
            {"name": "  records[].remark", "type": "string", "desc": "备注"},
            {"name": "  records[].createdBy", "type": "string", "desc": "创建人id"},
            {"name": "  records[].createdTime", "type": "string", "desc": "创建时间"},
        ]
    },
    {
        "category": "CRM API",
        "module": "查询合同",
        "url": "/crm/api/{tenantId}/contract/search",
        "method": "GET",
        "desc": "查询合同列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "data.total", "type": "int", "desc": "总条数"},
            {"name": "data.records", "type": "list", "desc": "数据列表"},
            {"name": "  records[].id", "type": "string", "desc": "主键id"},
            {"name": "  records[].code", "type": "string", "desc": "合同编码"},
            {"name": "  records[].name", "type": "string", "desc": "合同名称"},
            {"name": "  records[].customerId", "type": "string", "desc": "关联客户id"},
            {"name": "  records[].businessId", "type": "string", "desc": "关联商机id"},
            {"name": "  records[].amount", "type": "double", "desc": "合同金额"},
            {"name": "  records[].signTime", "type": "string", "desc": "签订时间"},
            {"name": "  records[].startTime", "type": "string", "desc": "开始时间"},
            {"name": "  records[].endTime", "type": "string", "desc": "结束时间"},
            {"name": "  records[].contractType", "type": "string", "desc": "合同类型"},
            {"name": "  records[].remark", "type": "string", "desc": "备注"},
        ]
    },
    {
        "category": "CRM API",
        "module": "查询客户",
        "url": "/crm/api/{tenantId}/customer/search",
        "method": "GET",
        "desc": "查询客户列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "data.records[].id", "type": "string", "desc": "主键id"},
            {"name": "data.records[].Name", "type": "string", "desc": "客户名称"},
            {"name": "data.records[].Region", "type": "string", "desc": "地区"},
            {"name": "data.records[].Industry", "type": "string", "desc": "行业"},
            {"name": "data.records[].Status", "type": "string", "desc": "状态"},
            {"name": "data.records[].Stage", "type": "string", "desc": "阶段"},
            {"name": "data.records[].CustomerSource", "type": "string", "desc": "客户来源"},
            {"name": "data.records[].EmployeeTotal", "type": "string", "desc": "员工总数"},
            {"name": "data.records[].Telephone", "type": "string", "desc": "电话"},
            {"name": "data.records[].Address", "type": "string", "desc": "地址"},
            {"name": "data.records[].Owner", "type": "string", "desc": "负责人"},
            {"name": "data.records[].CreatedBy", "type": "string", "desc": "创建人"},
            {"name": "data.records[].CreatedTime", "type": "string", "desc": "创建时间"},
        ]
    },
    {
        "category": "CRM API",
        "module": "查询发票",
        "url": "/crm/api/{tenantId}/invoice/search",
        "method": "GET",
        "desc": "查询发票列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "data.total", "type": "int", "desc": "总条数"},
            {"name": "data.records", "type": "list", "desc": "数据列表(通用CRM search结构)"},
        ]
    },
    {
        "category": "CRM API",
        "module": "查询回款",
        "url": "/crm/api/{tenantId}/receivables/search",
        "method": "GET",
        "desc": "查询回款列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "data.total", "type": "int", "desc": "总条数"},
            {"name": "data.records", "type": "list", "desc": "数据列表(通用CRM search结构)"},
        ]
    },
    {
        "category": "CRM API",
        "module": "查询线索",
        "url": "/crm/api/{tenantId}/salesleads/search",
        "method": "GET",
        "desc": "查询销售线索列表",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "data.total", "type": "int", "desc": "总条数"},
            {"name": "data.records", "type": "list", "desc": "数据列表(通用CRM search结构)"},
        ]
    },
    {
        "category": "CRM API",
        "module": "字典项获取",
        "url": "/crm/api/{tenant_id}/datasource/get",
        "method": "GET",
        "desc": "获取字典项列表（如地区、行业、客户来源等）",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "code", "type": "string", "required": "是", "desc": "字典编码(如Crm.SalesStage, Crm.Industry等)"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data", "type": "array", "desc": "字典列表(code, name等)"},
        ]
    },
    {
        "category": "CRM API",
        "module": "CRM通用查询",
        "url": "/crm/api/{tenantId}/search",
        "method": "POST",
        "desc": "通用CRM查询（支持客户/联系人/合同/发票/回款/产品/线索/产品类别）",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页条数"},
            {"name": "metaObjectCode", "type": "string", "required": "是", "desc": "对象编码(Crm.Customer/Crm.Contract/Crm.Invoice/Crm.Receivables/Crm.Product/Crm.SalesLeads/Crm.Contacts等)"},
            {"name": "items", "type": "object", "required": "否", "desc": "条件列表"},
            {"name": "  items[].name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "  items[].value", "type": "string", "required": "否", "desc": "字段值"},
            {"name": "  items[].operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "  items[].conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "errorMsg", "type": "string", "desc": "错误信息"},
            {"name": "data[].fields", "type": "array", "desc": "字段键值对数组"},
            {"name": "  fields[].name", "type": "string", "desc": "字段名称"},
            {"name": "  fields[].value", "type": "string", "desc": "字段值"},
        ]
    },

    # ===== 业务API - 面谈 (2个) =====
    {
        "category": "业务API - 面谈",
        "module": "获取面谈列表",
        "url": "/api/{tenant_id}/interview/search",
        "method": "GET",
        "desc": "获取指定时间段内面谈列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "面谈条数"},
            {"name": "data.interviews", "type": "List<InterviewModel>", "desc": "面谈列表"},
            {"name": "  interviews[].interview_id", "type": "string", "desc": "面谈标识"},
            {"name": "  interviews[].name", "type": "string", "desc": "面谈标题"},
            {"name": "  interviews[].interview_date", "type": "string", "desc": "面谈时间"},
            {"name": "  interviews[].user", "type": "UserModel", "desc": "面谈人(userId,name,email,mobile,departmentName,wx_userId)"},
            {"name": "  interviews[].toUser", "type": "UserModel", "desc": "被面谈人"},
        ]
    },
    {
        "category": "业务API - 面谈",
        "module": "获取面谈详情",
        "url": "/api/{tenant_id}/interview/getInterviewRecords",
        "method": "POST",
        "desc": "获取指定面谈信息详情",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "interview_ids", "type": "List<string>", "required": "是", "desc": "面谈标识列表"},
        ],
        "response": [
            {"name": "data[].id", "type": "string", "desc": "面谈标识"},
            {"name": "data[].name", "type": "string", "desc": "面谈标题"},
            {"name": "data[].interviewTime", "type": "string", "desc": "面谈时间"},
            {"name": "data[].content", "type": "string", "desc": "面谈内容(富文本JSON)"},
            {"name": "data[].user", "type": "UserModel", "desc": "面谈人"},
            {"name": "data[].toUser", "type": "UserModel", "desc": "被面谈人"},
        ]
    },

    # ===== 业务API - 总结 (2个) =====
    {
        "category": "业务API - 总结",
        "module": "获取总结列表",
        "url": "/api/{tenant_id}/daily/search",
        "method": "GET",
        "desc": "获取指定时间段内创建的总结",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "daily_type", "type": "int", "required": "否", "desc": "总结类型(8日报 26周报 27月报)"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
            {"name": "date_type", "type": "int", "required": "是", "desc": "0周期时间 1创建时间"},
            {"name": "is_content", "type": "bool", "required": "否", "desc": "是否包含总结内容"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "数量"},
            {"name": "data.dailys", "type": "List<PlanDailyModel>", "desc": "总结列表"},
            {"name": "  dailys[].daily_id", "type": "int", "desc": "ID"},
            {"name": "  dailys[].daily_name", "type": "string", "desc": "名称"},
            {"name": "  dailys[].daily_date", "type": "string", "desc": "总结所属日期"},
            {"name": "  dailys[].daily_type", "type": "int", "desc": "总结类型"},
            {"name": "  dailys[].submit_type", "type": "int", "desc": "1按时 2补交"},
            {"name": "    dailys[].summary", "type": "string", "desc": "总结内容"},
            {"name": "    dailys[].work_plan", "type": "string", "desc": "下一步计划"},
            {"name": "  dailys[].create_user", "type": "UserModel", "desc": "创建人"},
        ]
    },
    {
        "category": "业务API - 总结",
        "module": "获取总结ID集合",
        "url": "/api/{tenant_id}/daily/search/ids",
        "method": "GET",
        "desc": "获取指定时间段内创建的总结ID集合",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "数量"},
            {"name": "data.ids", "type": "List", "desc": "ID列表"},
        ]
    },

    # ===== 业务API - 部门(1个) =====
    {
        "category": "业务API - 部门",
        "module": "获取所有部门",
        "url": "/api/{tenant_id}/department/get/all",
        "method": "GET",
        "desc": "获取所有部门",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识(URL中)"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data[].id", "type": "int", "desc": "部门ID"},
            {"name": "data[].name", "type": "string", "desc": "部门名称"},
            {"name": "data[].code", "type": "string", "desc": "部门编码"},
            {"name": "data[].parent_id", "type": "int", "desc": "父部门"},
            {"name": "data[].order", "type": "int", "desc": "排序"},
            {"name": "data[].leader_users", "type": "List<UserModel>", "desc": "部门负责人"},
            {"name": "data[].hrbp_users", "type": "List<UserModel>", "desc": "部门HRBP"},
        ]
    },

    # ===== 业务API - 群组 (2个) =====
    {
        "category": "业务API - 群组",
        "module": "获取所有群组",
        "url": "/api/{tenant_id}/group/search",
        "method": "GET",
        "desc": "获取租户所有群组",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "群组条数"},
            {"name": "data.groups", "type": "List<GroupModel>", "desc": "群组列表"},
            {"name": "  groups[].id", "type": "int", "desc": "群组ID"},
            {"name": "  groups[].name", "type": "string", "desc": "群组名称"},
            {"name": "  groups[].create_date", "type": "string", "desc": "创建时间"},
            {"name": "  groups[].create_user", "type": "object", "desc": "创建人信息"},
        ]
    },
    {
        "category": "业务API - 群组",
        "module": "获取群组下人员",
        "url": "/api/{tenant_id}/group/search/user",
        "method": "GET",
        "desc": "获取群组下人员列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户ID"},
            {"name": "group_id", "type": "int", "required": "是", "desc": "群组ID"},
        ],
        "response": [
            {"name": "data[].userId", "type": "int", "desc": "人员Id"},
            {"name": "data[].name", "type": "string", "desc": "人员名称"},
            {"name": "data[].email", "type": "string", "desc": "邮箱"},
            {"name": "data[].mobile", "type": "string", "desc": "手机号"},
            {"name": "data[].wx_userId", "type": "string", "desc": "企业微信ID"},
            {"name": "data[].departmentName", "type": "string", "desc": "部门名称"},
            {"name": "data[].groupName", "type": "string", "desc": "群组名称"},
        ]
    },

    # ===== 业务API - 消息/待办 (2个) =====
    {
        "category": "业务API - 消息/待办",
        "module": "按时间获取待办",
        "url": "/api/v1/{tenantId}/todo/get",
        "method": "POST",
        "desc": "按时间获取待办列表",
        "params": [
            {"name": "tenantId", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "status", "type": "int", "required": "否", "desc": "状态(1未读 2已读 默认0全部)"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "每页数量，单页最大100条"},
            {"name": "begin_time", "type": "long", "required": "是", "desc": "开始时间(毫秒时间戳)"},
            {"name": "last_time", "type": "long", "required": "是", "desc": "结束时间(毫秒时间戳)"},
        ],
        "response": [
            {"name": "total", "type": "int", "desc": "数量"},
            {"name": "messages", "type": "List<Todo>", "desc": "消息(Todo对象)集合"},
            {"name": "  Todo.fromUser", "type": "UserInfo", "desc": "消息发送人"},
            {"name": "  Todo.title", "type": "string", "desc": "消息标题"},
            {"name": "  Todo.SubTitle", "type": "string", "desc": "消息子标题"},
            {"name": "  Todo.content", "type": "string", "desc": "消息内容(含富文本)"},
            {"name": "  Todo.objUrl", "type": "string", "desc": "地址"},
            {"name": "  Todo.todo_id", "type": "string", "desc": "待办标识"},
            {"name": "  Todo.createDate", "type": "string", "desc": "创建时间"},
        ]
    },
    {
        "category": "业务API - 消息/待办",
        "module": "按时间获取消息",
        "url": "/api/v1/{tenantId}/message/get",
        "method": "POST",
        "desc": "获取消息列表",
        "params": [
            {"name": "tenantId", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "status", "type": "int", "required": "是", "desc": "状态(1未读 2已读 默认0全部)"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "每页数量，单页最大100条"},
            {"name": "begin_time", "type": "long", "required": "是", "desc": "开始时间(毫秒时间戳)"},
            {"name": "last_time", "type": "long", "required": "是", "desc": "结束时间(毫秒时间戳)"},
        ],
        "response": [
            {"name": "total", "type": "int", "desc": "数量"},
            {"name": "messages", "type": "List<Message>", "desc": "消息(Message对象)集合"},
            {"name": "  Message.fromUser", "type": "UserInfo", "desc": "消息发送人"},
            {"name": "  Message.title", "type": "string", "desc": "消息标题"},
            {"name": "  Message.content", "type": "string", "desc": "消息内容(含富文本)"},
            {"name": "  Message.url", "type": "string", "desc": "地址"},
            {"name": "  Message.messageId", "type": "string", "desc": "消息标识"},
            {"name": "  Message.createDate", "type": "string", "desc": "创建时间"},
        ]
    },

    # ===== 业务API - 里程碑 (1个) =====
    {
        "category": "业务API - 里程碑",
        "module": "获取项目里程碑",
        "url": "/api/{tenant_id}/milestone/search",
        "method": "POST",
        "desc": "获取项目里程碑",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "work_id", "type": "int", "required": "是", "desc": "里程碑归属项目id"},
            {"name": "relation_id", "type": "string", "required": "否", "desc": "第三方ID"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "message", "type": "string", "desc": "错误信息"},
            {"name": "data.total", "type": "int", "desc": "项目条数"},
            {"name": "data.milestones", "type": "List<MilestoneModel>", "desc": "里程碑列表"},
            {"name": "  milestones[].id", "type": "int", "desc": "里程碑ID"},
            {"name": "  milestones[].name", "type": "string", "desc": "里程碑名称"},
            {"name": "  milestones[].work_id", "type": "int", "desc": "归属项目ID"},
            {"name": "  milestones[].relation_id", "type": "string", "desc": "第三方ID"},
            {"name": "  milestones[].principal_user", "type": "UserModel", "desc": "负责人"},
            {"name": "  milestones[].start_date", "type": "string", "desc": "开始时间"},
            {"name": "  milestones[].end_date", "type": "string", "desc": "截止时间"},
            {"name": "  milestones[].status", "type": "int", "desc": "状态(1进行中 2完成)"},
            {"name": "  milestones[].description", "type": "string", "desc": "描述"},
        ]
    },

    # ===== 业务API - OKR (3个) =====
    {
        "category": "业务API - OKR",
        "module": "获取目标(OKR)",
        "url": "/api/{tenant_id}/okr/search",
        "method": "GET",
        "desc": "获取指定时间段内创建的目标",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
            {"name": "fetch_kr", "type": "bool", "required": "否", "desc": "是否查询目标下的关键成果"},
            {"name": "fetch_progress", "type": "bool", "required": "否", "desc": "是否查询目标下的进展"},
            {"name": "fetch_krprogress", "type": "bool", "required": "否", "desc": "是否查询关键成果的进展"},
            {"name": "date_type", "type": "int", "required": "否", "desc": "1创建时间 2更新时间 3开始时间"},
            {"name": "okr_ids", "type": "string", "required": "否", "desc": "OKRId(多个逗号分割)"},
            {"name": "kr_ids", "type": "string", "required": "否", "desc": "KRId(多个逗号分割)"},
            {"name": "cycle_Type", "type": "int", "required": "否", "desc": "周期(0自定义 1年度 2季度 3月度 4双月 5半年度)"},
            {"name": "year_Num", "type": "int", "required": "否", "desc": "年度"},
            {"name": "yqm_Num", "type": "int", "required": "否", "desc": "月份扩展"},
            {"name": "fetch_relation", "type": "bool", "required": "否", "desc": "是否展示关联信息"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "目标条数"},
            {"name": "data.okrs", "type": "List<OkrModel>", "desc": "目标列表"},
            {"name": "  okrs[].okr_id", "type": "int", "desc": "目标Id"},
            {"name": "  okrs[].okr_name", "type": "string", "desc": "目标名称"},
            {"name": "  okrs[].label_name", "type": "string", "desc": "目标分类名称"},
            {"name": "  okrs[].create_user", "type": "UserModel", "desc": "创建人"},
            {"name": "  okrs[].principal_user", "type": "UserModel", "desc": "负责人"},
            {"name": "  okrs[].participants", "type": "List<UserModel>", "desc": "参与人"},
            {"name": "  okrs[].followers", "type": "List<UserModel>", "desc": "关注人"},
            {"name": "  okrs[].progress", "type": "int", "desc": "进度"},
            {"name": "  okrs[].okr_type", "type": "int", "desc": "目标类型(1个人 2部门 3公司 4团队)"},
            {"name": "  okrs[].cycle_type", "type": "int", "desc": "周期类型"},
            {"name": "  okrs[].cycle_name", "type": "string", "desc": "周期名称"},
            {"name": "  okrs[].start_date", "type": "string", "desc": "开始时间"},
            {"name": "  okrs[].end_date", "type": "string", "desc": "截止时间"},
            {"name": "  okrs[].status", "type": "int", "desc": "状态(1进行中 2完成)"},
            {"name": "  okrs[].evaluate_score", "type": "decimal", "desc": "评分"},
            {"name": "  okrs[].risk_level", "type": "string", "desc": "OKR状态(2有风险 3已延期 6正常)"},
            {"name": "  okrs[].krs", "type": "List<KrModel>", "desc": "关键成果"},
            {"name": "  okrs[].okr_progress", "type": "List<ProgressModel>", "desc": "进展"},
            {"name": "  okrs[].parent_okr", "type": "List<RelationOkrModel>", "desc": "对齐的目标"},
            {"name": "  okrs[].relation_models", "type": "List<RelationModel>", "desc": "关联信息"},
        ]
    },
    {
        "category": "业务API - OKR",
        "module": "按负责人/部门筛选OKR",
        "url": "/api/{tenant_id}/okr/searchByUser",
        "method": "POST",
        "desc": "获取指定时间段内创建的目标（可根据OKR负责人、部门编码筛选查询）",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
            {"name": "account_type", "type": "int", "required": "否", "desc": "账号类型(0手机号或邮箱 1企业微信 4工号 5TitaUserId)"},
            {"name": "accounts", "type": "List<string>", "required": "否", "desc": "OKR负责人"},
            {"name": "department_codes", "type": "List<string>", "required": "否", "desc": "部门编码"},
            {"name": "status", "type": "int", "required": "否", "desc": "状态(1进行中 2完成)"},
            {"name": "risk_level", "type": "int", "required": "否", "desc": "OKR状态(1正常推进 2有风险 3已延期 4完成 5延期完成 6正常)"},
        ],
        "response": [
            {"name": "同 /okr/search 响应结构", "type": "", "desc": "返回字段同获取目标(OKR)接口"},
        ]
    },
    {
        "category": "业务API - OKR",
        "module": "获取OKR分类",
        "url": "/api/{tenant_id}/okr/label/get",
        "method": "GET",
        "desc": "获取OKR分类",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "message", "type": "string", "desc": "错误信息"},
            {"name": "data.total", "type": "int", "desc": "标签分类条数"},
            {"name": "data.lables", "type": "List<OkrLabelModel>", "desc": "目标分类列表"},
            {"name": "  lables[].id", "type": "int", "desc": "目标分类Id"},
            {"name": "  lables[].name", "type": "string", "desc": "目标分类名称"},
            {"name": "  lables[].status", "type": "string", "desc": "停启用状态"},
        ]
    },

    # ===== 业务API - 绩效 (4个) =====
    {
        "category": "业务API - 绩效",
        "module": "获取绩效考核列表",
        "url": "/api/{tenant_id}/performance/search",
        "method": "GET",
        "desc": "获取指定时间段内绩效考核列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.assess_forms", "type": "List<AssessFormModel>", "desc": "考核列表"},
            {"name": "  assess_forms[].assess_formId", "type": "string", "desc": "考核Id"},
            {"name": "  assess_forms[].assess_templateId", "type": "string", "desc": "考核模板标识"},
            {"name": "  assess_forms[].assess_template_name", "type": "string", "desc": "考核模板名称"},
            {"name": "  assess_forms[].assess_activityId", "type": "string", "desc": "考核活动Id"},
            {"name": "  assess_forms[].name", "type": "string", "desc": "考核名称"},
            {"name": "  assess_forms[].user", "type": "UserModel", "desc": "被考核人"},
            {"name": "  assess_forms[].current_nodeName", "type": "string", "desc": "当前考核流程节点"},
            {"name": "  assess_forms[].start_date", "type": "string", "desc": "开始时间"},
            {"name": "  assess_forms[].end_date", "type": "string", "desc": "截止时间"},
            {"name": "  assess_forms[].score", "type": "decimal", "desc": "考核分数"},
            {"name": "  assess_forms[].rank", "type": "string", "desc": "考核等级"},
            {"name": "  assess_forms[].is_end", "type": "bool", "desc": "是否结束"},
            {"name": "  assess_forms[].cycle_type", "type": "int", "desc": "周期类型"},
            {"name": "  assess_forms[].cycle_name", "type": "string", "desc": "考核周期"},
            {"name": "  assess_forms[].nodes", "type": "List<NodeModel>", "desc": "流程节点集合"},
        ]
    },
    {
        "category": "业务API - 绩效",
        "module": "获取绩效评价详情",
        "url": "/api/{tenant_id}/performance/getevaluations",
        "method": "POST",
        "desc": "获取指定绩效表单评价详情",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "assessFormIds", "type": "List<string>", "required": "是", "desc": "表单标识列表"},
        ],
        "response": [
            {"name": "data[].id", "type": "string", "desc": "表单标识"},
            {"name": "data[].assessTemplateId", "type": "string", "desc": "表单使用模板标识"},
            {"name": "data[].assessTemplateName", "type": "string", "desc": "表单使用模板名称"},
            {"name": "data[].user", "type": "UserModel", "desc": "用户"},
            {"name": "data[].name", "type": "string", "desc": "考核名称"},
            {"name": "data[].currentNodeName", "type": "string", "desc": "当前考核流程节点"},
            {"name": "data[].createDate", "type": "date", "desc": "创建时间"},
            {"name": "data[].isEnd", "type": "bool", "desc": "流程是否结束"},
            {"name": "data[].score", "type": "decimal", "desc": "考核分数"},
            {"name": "data[].rank", "type": "string", "desc": "考核等级"},
            {"name": "data[].modules", "type": "List<AssessModuleModel>", "desc": "考核维度"},
            {"name": "  modules[].id", "type": "string", "desc": "维度标识"},
            {"name": "  modules[].name", "type": "string", "desc": "维度名称"},
            {"name": "  modules[].type", "type": "string", "desc": "维度类型"},
            {"name": "  modules[].weight", "type": "decimal", "desc": "维度权重"},
            {"name": "  modules[].assessItems", "type": "List<AssessItemModel>", "desc": "考核指标"},
            {"name": "    assessItems[].id", "type": "string", "desc": "指标标识"},
            {"name": "    assessItems[].name", "type": "string", "desc": "指标名称"},
            {"name": "    assessItems[].targetValue", "type": "string", "desc": "目标值"},
            {"name": "    assessItems[].actualValue", "type": "string", "desc": "完成值"},
            {"name": "    assessItems[].weight", "type": "decimal", "desc": "指标权重"},
            {"name": "    assessItems[].scoreType", "type": "string", "desc": "评分类型(Manual手动 Auto自动)"},
            {"name": "data[].evaluationTotals", "type": "List<AssessFormEvaluationTotalModel>", "desc": "考核总评"},
            {"name": "  evaluationTotals[].score", "type": "string", "desc": "评分"},
            {"name": "  evaluationTotals[].rank", "type": "string", "desc": "评级"},
            {"name": "  evaluationTotals[].evaluateContents", "type": "List", "desc": "高级评语"},
            {"name": "data[].subjectResults", "type": "List", "desc": "多主题评价结果"},
            {"name": "data[].is_Appeal", "type": "bool", "desc": "是否存在绩效校准"},
        ]
    },
    {
        "category": "业务API - 绩效",
        "module": "考核活动列表查询",
        "url": "/api/{tenant_id}/performance/activity/list",
        "method": "POST",
        "desc": "考核活动列表查询",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "beginTime", "type": "string", "required": "否", "desc": "根据活动创建时间查询(闭区间)"},
            {"name": "endTime", "type": "string", "required": "否", "desc": "根据活动创建时间查询(开区间)"},
            {"name": "activityName", "type": "string", "required": "否", "desc": "活动名称，模糊查询"},
            {"name": "pageNum", "type": "int", "required": "是", "desc": "页号"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "页码"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.activities", "type": "List", "desc": "活动列表"},
            {"name": "  activities[].id", "type": "string", "desc": "活动id"},
            {"name": "  activities[].name", "type": "string", "desc": "活动名称"},
            {"name": "  activities[].departmentId", "type": "int", "desc": "所属部门"},
            {"name": "  activities[].assessType", "type": "int", "desc": "考核类型(0正式考核 1试用期考核)"},
            {"name": "  activities[].cycleType", "type": "int", "desc": "周期类型"},
            {"name": "  activities[].cycleName", "type": "string", "desc": "周期名称"},
            {"name": "  activities[].startDate", "type": "string", "desc": "活动开始时间"},
            {"name": "  activities[].endDate", "type": "string", "desc": "活动结束时间"},
            {"name": "  activities[].status", "type": "int", "desc": "活动状态(1进行中 2结束 3未开始)"},
            {"name": "  activities[].createTime", "type": "string", "desc": "活动创建时间"},
            {"name": "  activities[].userTotal", "type": "int", "desc": "被考核人数"},
        ]
    },
    {
        "category": "业务API - 绩效",
        "module": "考核结果列表查询",
        "url": "/api/{tenant_id}/performance/getAssessFormResults",
        "method": "POST",
        "desc": "考核结果列表查询",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "activityId", "type": "string", "required": "是", "desc": "考核活动Id"},
            {"name": "pageNum", "type": "int", "required": "是", "desc": "页号"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "页码"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.assessForms", "type": "List", "desc": "考核表单列表"},
            {"name": "  assessForms[].assessFormId", "type": "string", "desc": "考核id"},
            {"name": "  assessForms[].name", "type": "string", "desc": "活动名称"},
            {"name": "  assessForms[].user", "type": "UserModel", "desc": "被考核人"},
            {"name": "  assessForms[].score", "type": "decimal", "desc": "考核得分"},
            {"name": "  assessForms[].rank", "type": "string", "desc": "等级"},
            {"name": "  assessForms[].factor", "type": "decimal", "desc": "绩效系数"},
            {"name": "  assessForms[].isEnd", "type": "bool", "desc": "考核是否结束"},
            {"name": "  assessForms[].cycleName", "type": "string", "desc": "周期名称"},
        ]
    },

    # ===== 业务API - 职级/职务/标签 (3个) =====
    {
        "category": "业务API - 职级/职务/标签",
        "module": "获取人员职务列表",
        "url": "/api/{tenant_id}/position/search",
        "method": "GET",
        "desc": "获取人员职务列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.positions", "type": "List<PositionModel>", "desc": "职务列表"},
            {"name": "  positions[].id", "type": "int", "desc": "职务标识"},
            {"name": "  positions[].code", "type": "string", "desc": "职务编码"},
            {"name": "  positions[].name", "type": "string", "desc": "职务名称"},
            {"name": "  positions[].description", "type": "string", "desc": "职务描述"},
        ]
    },
    {
        "category": "业务API - 职级/职务/标签",
        "module": "获取人员职级列表",
        "url": "/api/{tenant_id}/rank/search",
        "method": "GET",
        "desc": "获取人员职级列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.ranks", "type": "List<RankModel>", "desc": "职级列表"},
            {"name": "  ranks[].rank_id", "type": "int", "desc": "职级标识"},
            {"name": "  ranks[].rank_name", "type": "string", "desc": "职级名称"},
        ]
    },
    {
        "category": "业务API - 职级/职务/标签",
        "module": "获取人员标签列表",
        "url": "/api/{tenant_id}/label/search",
        "method": "GET",
        "desc": "获取人员标签列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.labels", "type": "List<LabelModel>", "desc": "标签列表"},
            {"name": "  labels[].label_id", "type": "int", "desc": "标签标识"},
            {"name": "  labels[].label_name", "type": "string", "desc": "标签名称"},
        ]
    },

    # ===== 业务API - 复盘 (1个) =====
    {
        "category": "业务API - 复盘",
        "module": "获取复盘",
        "url": "/api/{tenant_id}/review/search",
        "method": "GET",
        "desc": "获取复盘列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
            {"name": "review_type", "type": "string", "required": "否", "desc": "复盘类型逗号隔开(87单OKR复盘 88周期OKR复盘 89单项目复盘 90自定义复盘)"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "数量"},
            {"name": "data.reviews", "type": "List<ReviewModel>", "desc": "复盘列表"},
            {"name": "  reviews[].review_id", "type": "string", "desc": "复盘ID"},
            {"name": "  reviews[].review_name", "type": "string", "desc": "复盘名称"},
            {"name": "  reviews[].review_type", "type": "int", "desc": "复盘类型"},
            {"name": "  reviews[].create_user", "type": "UserModel", "desc": "发起人"},
            {"name": "  reviews[].create_date", "type": "string", "desc": "创建时间"},
            {"name": "  reviews[].obj_id", "type": "string", "desc": "复盘关联的对象ID"},
            {"name": "  reviews[].obj_type", "type": "int", "desc": "复盘关联的对象类型"},
            {"name": "  reviews[].details", "type": "List<ReviewModuleModel>", "desc": "复盘组件详情"},
        ]
    },

    # ===== 业务API - 员工 (3个) =====
    {
        "category": "业务API - 员工",
        "module": "获取用户列表",
        "url": "/api/{tenant_id}/staff/search",
        "method": "GET",
        "desc": "获取用户列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "content", "type": "string", "required": "否", "desc": "关键词"},
            {"name": "department_code", "type": "string", "required": "否", "desc": "所属部门编码"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "人员信息条数"},
            {"name": "data.users", "type": "List<UserDetailModel>", "desc": "人员列表"},
            {"name": "  users[].userId", "type": "int", "desc": "用户Id"},
            {"name": "  users[].name", "type": "string", "desc": "用户姓名"},
            {"name": "  users[].email", "type": "string", "desc": "邮箱"},
            {"name": "  users[].mobile", "type": "string", "desc": "手机号"},
            {"name": "  users[].position", "type": "string", "desc": "职务"},
            {"name": "  users[].staff_code", "type": "string", "desc": "人员编码/工号"},
            {"name": "  users[].department_id", "type": "int", "desc": "部门Id"},
            {"name": "  users[].department_name", "type": "string", "desc": "部门名称"},
            {"name": "  users[].department_code", "type": "string", "desc": "部门编码"},
            {"name": "  users[].sex", "type": "string", "desc": "性别"},
            {"name": "  users[].jobState", "type": "int", "desc": "人员状态(1在职 0离职)"},
            {"name": "  users[].rank_name", "type": "string", "desc": "职级"},
            {"name": "  users[].labels", "type": "List<string>", "desc": "标签集合"},
            {"name": "  users[].lineManger_name", "type": "string", "desc": "上级人员名称"},
            {"name": "  users[].lineManger_userId", "type": "string", "desc": "上级人员UserId"},
            {"name": "  users[].birthday", "type": "string", "desc": "生日"},
            {"name": "  users[].employed_Date", "type": "string", "desc": "入职时间"},
        ]
    },
    {
        "category": "业务API - 员工",
        "module": "获取工号为空的用户列表",
        "url": "/api/{tenant_id}/staff/emptyStaffCode",
        "method": "GET",
        "desc": "获取工号为空的用户列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
        ],
        "response": [
            {"name": "同 获取用户列表 响应结构", "type": "", "desc": "返回字段同获取用户列表接口"},
        ]
    },
    {
        "category": "业务API - 员工",
        "module": "获取用户详情",
        "url": "/api/{tenant_id}/staff/get",
        "method": "GET",
        "desc": "获取用户详情（根据账号查询）",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "acccount_name", "type": "string", "required": "是", "desc": "账号(邮箱或手机号)"},
            {"name": "account_type", "type": "int", "required": "否", "desc": "账号类型(0手机号或邮箱 1企业微信 4工号 5TitaUserId)"},
        ],
        "response": [
            {"name": "data.userId", "type": "int", "desc": "用户Id"},
            {"name": "data.name", "type": "string", "desc": "用户姓名"},
            {"name": "data.email", "type": "string", "desc": "邮箱"},
            {"name": "data.mobile", "type": "string", "desc": "手机号"},
            {"name": "data.position", "type": "string", "desc": "职务"},
            {"name": "data.staff_code", "type": "string", "desc": "人员编码/工号"},
            {"name": "data.department_id", "type": "int", "desc": "部门Id"},
            {"name": "data.department_name", "type": "string", "desc": "部门名称"},
            {"name": "data.department_code", "type": "string", "desc": "部门编码"},
            {"name": "data.sex", "type": "string", "desc": "性别"},
            {"name": "data.jobState", "type": "int", "desc": "人员状态(1在职 0离职)"},
            {"name": "data.rank_name", "type": "string", "desc": "职级"},
            {"name": "data.labels", "type": "List<string>", "desc": "标签集合"},
            {"name": "data.lineManger_name", "type": "string", "desc": "上级人员名称"},
        ]
    },

    # ===== 业务API - 360评估 (1个) =====
    {
        "category": "业务API - 360评估",
        "module": "获取360评估结果列表",
        "url": "/api/{tenant_id}/survey/getSurveyByCycle",
        "method": "GET",
        "desc": "获取指定时间段内360评估结果列表",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "总数"},
            {"name": "data.surveyInfos", "type": "List<SurveyModel>", "desc": "考核列表"},
            {"name": "  surveyInfos[].surveyName", "type": "string", "desc": "评估活动名称"},
            {"name": "  surveyInfos[].surveyYear", "type": "string", "desc": "评估年度"},
            {"name": "  surveyInfos[].surveyCycle", "type": "string", "desc": "评估周期"},
            {"name": "  surveyInfos[].startTime", "type": "string", "desc": "活动开始时间"},
            {"name": "  surveyInfos[].endTime", "type": "string", "desc": "活动结束时间"},
            {"name": "  surveyInfos[].userModel", "type": "UserModel", "desc": "被评人信息"},
            {"name": "  surveyInfos[].score", "type": "decimal", "desc": "评估分数"},
        ]
    },

    # ===== 业务API - 任务 (3个) =====
    {
        "category": "业务API - 任务",
        "module": "获取指定时间段内创建的任务",
        "url": "/api/{tenant_id}/task/search",
        "method": "POST",
        "desc": "获取指定时间段内创建的任务",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码(默认1)"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量(默认20，最大500)"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
            {"name": "obj_ids", "type": "List<string>", "required": "否", "desc": "所属对象ID(目标ID,项目ID)"},
            {"name": "obj_type", "type": "int", "required": "否", "desc": "所属对象类型(项目4 目标62)"},
            {"name": "is_getQuantifyValues", "type": "bool", "required": "否", "desc": "是否返回任务定量信息"},
            {"name": "task_ids", "type": "string", "required": "否", "desc": "任务ID(多个逗号分割)"},
            {"name": "is_getEvalute", "type": "bool", "required": "否", "desc": "是否返回任务评价信息"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "任务条数"},
            {"name": "data.tasks", "type": "List<TaskModel>", "desc": "任务列表"},
            {"name": "  tasks[].task_id", "type": "string", "desc": "任务ID"},
            {"name": "  tasks[].task_name", "type": "string", "desc": "任务名称"},
            {"name": "  tasks[].description", "type": "string", "desc": "描述"},
            {"name": "  tasks[].progress", "type": "int", "desc": "任务进度"},
            {"name": "  tasks[].relation_id", "type": "string", "desc": "第三方Id"},
            {"name": "  tasks[].principal_user", "type": "UserModel", "desc": "负责人"},
            {"name": "  tasks[].create_user", "type": "UserModel", "desc": "发起人"},
            {"name": "  tasks[].participators", "type": "List<UserModel>", "desc": "参与人"},
            {"name": "  tasks[].status", "type": "int", "desc": "状态(1进行中 2完成 3延迟 4取消 5未接受 6暂停)"},
            {"name": "  tasks[].visibility", "type": "int", "desc": "是否私密(1公开 2私密)"},
            {"name": "  tasks[].start_date", "type": "string", "desc": "开始时间"},
            {"name": "  tasks[].end_date", "type": "string", "desc": "截止时间"},
            {"name": "  tasks[].target_value", "type": "decimal", "desc": "目标值"},
            {"name": "  tasks[].actual_value", "type": "decimal", "desc": "实际值"},
            {"name": "  tasks[].estimate_hours", "type": "decimal", "desc": "预估工时"},
            {"name": "  tasks[].worktime_hours", "type": "double", "desc": "实际工时"},
            {"name": "  tasks[].priority", "type": "int", "desc": "等级"},
            {"name": "  tasks[].evaluate_model", "type": "List<EvaluateModel>", "desc": "任务评价列表"},
        ]
    },
    {
        "category": "业务API - 任务",
        "module": "获取任务进展",
        "url": "/api/{tenant_id}/task/progressDescription/get",
        "method": "GET",
        "desc": "获取任务进展",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "task_Id", "type": "string", "required": "是", "desc": "任务ID"},
            {"name": "start_date", "type": "string", "required": "否", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "否", "desc": "截止时间"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "数量"},
            {"name": "data.evolves", "type": "List<EvolveModel>", "desc": "进展信息"},
            {"name": "  evolves[].progress_id", "type": "string", "desc": "进展标识"},
            {"name": "  evolves[].create_user", "type": "UserModel", "desc": "进展创建人信息"},
            {"name": "  evolves[].progress_description", "type": "string", "desc": "进展描述"},
            {"name": "  evolves[].create_date", "type": "string", "desc": "进展创建时间"},
            {"name": "  evolves[].start_date", "type": "string", "desc": "进展开始时间"},
            {"name": "  evolves[].end_date", "type": "string", "desc": "进展结束时间"},
        ]
    },
    {
        "category": "业务API - 任务",
        "module": "获取任务工时",
        "url": "/api/{tenant_id}/task/worktime/get",
        "method": "GET",
        "desc": "获取任务工时",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "task_Id", "type": "string", "required": "是", "desc": "任务ID"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "数量"},
            {"name": "data.work_Times", "type": "List<WorkTimeModel>", "desc": "工时信息"},
            {"name": "  work_Times[].create_user", "type": "UserModel", "desc": "创建人信息"},
            {"name": "  work_Times[].description", "type": "string", "desc": "工时描述"},
            {"name": "  work_Times[].hours", "type": "decimal", "desc": "工时"},
            {"name": "  work_Times[].time", "type": "string", "desc": "时间"},
            {"name": "  work_Times[].progress_id", "type": "string", "desc": "进展标识"},
        ]
    },

    # ===== 业务API - 待办 (1个) =====
    {
        "category": "业务API - 待办",
        "module": "获取待办",
        "url": "/api/v1/{tenantId}/{userId}/todo",
        "method": "GET",
        "desc": "获取用户的待办列表",
        "params": [
            {"name": "tenantId", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "userId", "type": "int", "required": "是", "desc": "用户标识"},
            {"name": "status", "type": "int", "required": "是", "desc": "状态(1未处理 2已处理)"},
            {"name": "pageNum", "type": "int", "required": "是", "desc": "页码(未读状态不用传值)"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页数量"},
        ],
        "response": [
            {"name": "total", "type": "int", "desc": "数量"},
            {"name": "todos", "type": "List<Todo>", "desc": "待办(Todo对象)集合"},
            {"name": "  Todo.fromUser", "type": "object", "desc": "待办发起人(UserInfo对象)"},
            {"name": "  Todo.title", "type": "string", "desc": "标题"},
            {"name": "  Todo.content", "type": "string", "desc": "内容"},
            {"name": "  Todo.url", "type": "string", "desc": "跳转地址"},
            {"name": "  Todo.status", "type": "int", "desc": "状态(1未处理 2已处理)"},
            {"name": "  Todo.appName", "type": "string", "desc": "产品线名称"},
            {"name": "  Todo.createDate", "type": "string", "desc": "创建时间"},
        ]
    },

    # ===== 业务API - 项目 (2个) =====
    {
        "category": "业务API - 项目",
        "module": "获取指定时间段内创建的项目",
        "url": "/api/{tenant_id}/work/search",
        "method": "GET",
        "desc": "获取指定时间段内创建的项目",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "是", "desc": "页码"},
            {"name": "page_size", "type": "int", "required": "是", "desc": "单页数量"},
            {"name": "start_date", "type": "string", "required": "是", "desc": "开始时间"},
            {"name": "end_date", "type": "string", "required": "是", "desc": "截止时间"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "项目条数"},
            {"name": "data.works", "type": "List<WorkModel>", "desc": "项目列表"},
            {"name": "  works[].work_name", "type": "string", "desc": "项目名称"},
            {"name": "  works[].progress", "type": "int", "desc": "进度"},
            {"name": "  works[].relation_id", "type": "string", "desc": "第三方Id"},
            {"name": "  works[].principal_user", "type": "UserModel", "desc": "负责人"},
            {"name": "  works[].start_date", "type": "string", "desc": "开始时间"},
            {"name": "  works[].end_date", "type": "string", "desc": "截止时间"},
            {"name": "  works[].status", "type": "int", "desc": "状态(1进行中 2完成 7暂停中 8已取消)"},
            {"name": "  works[].visibility", "type": "int", "desc": "是否私密(1公开 2私密)"},
            {"name": "  works[].description", "type": "string", "desc": "描述"},
            {"name": "  works[].participants", "type": "List<UserModel>", "desc": "参与人"},
            {"name": "  works[].followers", "type": "List<UserModel>", "desc": "关注人"},
            {"name": "  works[].lable_name", "type": "string", "desc": "分类名称"},
            {"name": "  works[].department_name", "type": "string", "desc": "所属部门名称"},
            {"name": "  works[].workForm_name", "type": "string", "desc": "项目类型"},
            {"name": "  works[].relation_okrIds", "type": "List<string>", "desc": "关联OKRid"},
            {"name": "  works[].work_sets", "type": "List", "desc": "所属项目集列表(work_set_id, work_set_name)"},
        ]
    },
    {
        "category": "业务API - 项目",
        "module": "获取项目进展",
        "url": "/api/{tenant_id}/work/progress/get",
        "method": "POST",
        "desc": "获取项目进展",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "work_id", "type": "int", "required": "是", "desc": "项目ID"},
        ],
        "response": [
            {"name": "data.total", "type": "int", "desc": "项目进展条数"},
            {"name": "data.progress", "type": "List<WorkProgressModel>", "desc": "项目进展列表"},
            {"name": "  progress[].content", "type": "string", "desc": "项目进展内容"},
            {"name": "  progress[].progress_description", "type": "string", "desc": "项目进展说明"},
            {"name": "  progress[].old_progress", "type": "string", "desc": "旧的进度"},
            {"name": "  progress[].progress", "type": "string", "desc": "新的进度"},
            {"name": "  progress[].create_date", "type": "string", "desc": "进展添加日期"},
            {"name": "  progress[].create_user", "type": "UserModel", "desc": "进展添加人"},
        ]
    },

    # ===== 业务API - 项目集 (1个) =====
    {
        "category": "业务API - 项目集",
        "module": "搜索项目集",
        "url": "/api/{tenant_id}/workset/search",
        "method": "GET",
        "desc": "搜索项目集（GET方式）",
        "params": [
            {"name": "tenant_id", "type": "int", "required": "是", "desc": "租户标识"},
            {"name": "page_num", "type": "int", "required": "否", "desc": "页码，默认1"},
            {"name": "page_size", "type": "int", "required": "否", "desc": "每页数量，默认20，最大500"},
            {"name": "work_set_id", "type": "string", "required": "否", "desc": "项目集Id(查询指定项目集时传入)"},
            {"name": "key_words", "type": "string", "required": "否", "desc": "搜索关键字"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码"},
            {"name": "data.total", "type": "int", "desc": "项目集条数"},
            {"name": "data.work_sets", "type": "List<WorkSetModel>", "desc": "项目集列表"},
            {"name": "  work_sets[].work_set_id", "type": "string", "desc": "项目集Id"},
            {"name": "  work_sets[].work_set_name", "type": "string", "desc": "项目集名称"},
            {"name": "  work_sets[].parent_id", "type": "string", "desc": "父级项目集Id"},
            {"name": "  work_sets[].has_child", "type": "bool", "desc": "是否有子项目集"},
            {"name": "  work_sets[].work_count", "type": "int", "desc": "关联项目数量"},
            {"name": "  work_sets[].user_count", "type": "int", "desc": "人员数量"},
            {"name": "  work_sets[].manager_ids", "type": "List<int>", "desc": "管理员用户Id列表"},
            {"name": "  work_sets[].participator_ids", "type": "List<int>", "desc": "成员用户Id列表"},
            {"name": "  work_sets[].created_time", "type": "string", "desc": "创建时间"},
        ]
    },

    # ===== PaaS API (1个) =====
    {
        "category": "PaaS API",
        "module": "PaaS通用查询接口",
        "url": "/general/api/paas/{tenantId}/search",
        "method": "POST",
        "desc": "PaaS通用查询接口",
        "params": [
            {"name": "pageNum", "type": "int", "required": "是", "desc": "第几页"},
            {"name": "pageSize", "type": "int", "required": "是", "desc": "每页几条数据"},
            {"name": "metaObjectCode", "type": "string", "required": "是", "desc": "对象编码"},
            {"name": "appName", "type": "string", "required": "是", "desc": "应用名称(如Test, Sign)"},
            {"name": "items", "type": "object[]", "required": "否", "desc": "条件列表"},
            {"name": "  items[].name", "type": "string", "required": "否", "desc": "字段名称"},
            {"name": "  items[].value", "type": "string", "required": "否", "desc": "字段具体值"},
            {"name": "  items[].operatorType", "type": "int", "required": "否", "desc": "1等于 2≥ 3≤ 4> 5< 6范围 7为空 8不为空 9≠ 10开头等于 11结尾等于 12包含 14模糊"},
            {"name": "  items[].conditionType", "type": "int", "required": "否", "desc": "1且 2或"},
        ],
        "response": [
            {"name": "code", "type": "int", "desc": "状态码(0成功)"},
            {"name": "errorMsg", "type": "string", "desc": "错误信息"},
            {"name": "data", "type": "object[]", "desc": "数据列表"},
            {"name": "  data[].fields", "type": "array", "desc": "字段键值对数组"},
            {"name": "    fields[].name", "type": "string", "desc": "字段名称"},
            {"name": "    fields[].value", "type": "string", "desc": "字段值"},
        ]
    },
]


# ──────────────────────── 生成 Excel ────────────────────────

def get_nest_level(name):
    """根据字段名的缩进空格数返回嵌套层级"""
    stripped = name.lstrip()
    if len(name) - len(stripped) >= 4:
        return 2
    elif len(name) - len(stripped) >= 2:
        return 1
    return 0

def get_nest_fill(level):
    if level == 2:
        return NEST_L2
    elif level == 1:
        return NEST_L1
    return WHITE_FILL

def write_section(ws, start_row, title, data, is_response=False):
    """写入请求参数或响应数据字段分区"""
    # 分区标题行
    ws.merge_cells(f'A{start_row}:D{start_row}')
    cell = ws.cell(row=start_row, column=1, value=title)
    cell.font = BOLD_11
    cell.fill = LIGHT_BLUE
    cell.alignment = ALIGN_LEFT
    cell.border = THIN_BORDER
    for c in range(2, 5):
        ws.cell(row=start_row, column=c).border = THIN_BORDER
        ws.cell(row=start_row, column=c).fill = LIGHT_BLUE
    ws.row_dimensions[start_row].height = 22

    # 表头行
    hdr_row = start_row + 1
    if is_response:
        hdrs = ['字段名称', '字段说明', '', '数据类型']
    else:
        hdrs = ['参数名称', '参数说明', '是否必须', '数据类型']

    for col, h in enumerate(hdrs, 1):
        cell = ws.cell(row=hdr_row, column=col, value=h)
        cell.font = WHITE_BOLD
        cell.fill = DEEP_BLUE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws.row_dimensions[hdr_row].height = 22

    # 数据行
    if not data:
        dr = hdr_row + 1
        ws.merge_cells(f'A{dr}:D{dr}')
        cell = ws.cell(row=dr, column=1, value='（该接口无业务请求参数）' if not is_response else '（暂无响应字段说明）')
        cell.font = NORMAL_10
        cell.fill = WHITE_FILL
        cell.alignment = ALIGN_LEFT
        cell.border = THIN_BORDER
        for c in range(2, 5):
            ws.cell(row=dr, column=c).border = THIN_BORDER
            ws.cell(row=dr, column=c).fill = WHITE_FILL
    else:
        for idx, d in enumerate(data):
            dr = hdr_row + 1 + idx
            name = d.get('name', '')
            desc = d.get('desc', '')
            dtype = d.get('type', '')

            ws.cell(row=dr, column=1, value=name).font = NORMAL_10
            ws.cell(row=dr, column=2, value=desc).font = NORMAL_10
            if is_response:
                ws.cell(row=dr, column=3, value='').font = NORMAL_10
            else:
                ws.cell(row=dr, column=3, value=d.get('required', '')).font = NORMAL_10
            ws.cell(row=dr, column=4, value=dtype).font = NORMAL_10

            # 响应字段按缩进深度着色
            if is_response:
                level = get_nest_level(name)
                fill = get_nest_fill(level)
            else:
                fill = WHITE_FILL

            for c in range(1, 5):
                ws.cell(row=dr, column=c).fill = fill
                ws.cell(row=dr, column=c).border = THIN_BORDER
                if c in (1, 2):
                    ws.cell(row=dr, column=c).alignment = ALIGN_LEFT
                else:
                    ws.cell(row=dr, column=c).alignment = ALIGN_CENTER

    return hdr_row + len(data) + 1  # 返回下一个可用行


def create_sub_sheet(ws, iface):
    """创建单个接口的子表"""
    # 列宽
    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 52
    ws.column_dimensions['C'].width = 14
    ws.column_dimensions['D'].width = 20

    # 基本信息区 (Row 1-3)
    url_full = f"{BASE_URL}{iface.get('url', '')}"
    info = [
        ('接口地址', url_full),
        ('请求方式', iface.get('method', '')),
        ('接口描述', iface.get('desc', '')),
    ]

    for r, (label, value) in enumerate(info, 1):
        cell_a = ws.cell(row=r, column=1, value=label)
        cell_a.font = BOLD_10
        cell_a.fill = WHITE_FILL
        cell_a.alignment = ALIGN_LEFT
        cell_a.border = THIN_BORDER

        ws.merge_cells(f'B{r}:D{r}')
        cell_b = ws.cell(row=r, column=2, value=value)
        cell_b.font = NORMAL_10
        cell_b.fill = WHITE_FILL
        cell_b.alignment = ALIGN_LEFT
        cell_b.border = THIN_BORDER
        for c in range(3, 5):
            ws.cell(row=r, column=c).border = THIN_BORDER
            ws.cell(row=r, column=c).fill = WHITE_FILL

    # 空行分隔 (Row 4)
    sep_row = 4
    ws.row_dimensions[sep_row].height = 6
    for c in range(1, 5):
        ws.cell(row=sep_row, column=c).fill = WHITE_FILL
        ws.cell(row=sep_row, column=c).border = NO_BORDER

    # 请求参数区
    params = iface.get('params', [])
    next_row = write_section(ws, 5, '请求参数', params, is_response=False)

    # 空行分隔
    sep2 = next_row
    ws.row_dimensions[sep2].height = 6
    for c in range(1, 5):
        ws.cell(row=sep2, column=c).fill = WHITE_FILL
        ws.cell(row=sep2, column=c).border = NO_BORDER

    # 响应数据字段区
    response = iface.get('response', [])
    write_section(ws, sep2 + 1, '响应数据字段', response, is_response=True)


def generate_excel():
    wb = Workbook()
    wb.remove(wb.active)

    # ── 总览表 ──
    ws_ov = wb.create_sheet('接口总览')
    ws_ov.sheet_properties.tabColor = '1F4E79'

    # 标题行
    ws_ov.merge_cells('A1:E1')
    ws_ov['A1'] = 'TITA 查询接口梳理'
    ws_ov['A1'].font = TITLE_14
    ws_ov['A1'].alignment = ALIGN_CENTER
    ws_ov['A1'].fill = WHITE_FILL
    ws_ov.row_dimensions[1].height = 30

    # 表头
    headers_ov = ['序号', '接口名称', '接口地址', '请求方式', '接口描述']
    for col, h in enumerate(headers_ov, 1):
        cell = ws_ov.cell(row=2, column=col, value=h)
        cell.font = WHITE_BOLD
        cell.fill = DEEP_BLUE
        cell.alignment = ALIGN_CENTER
        cell.border = THIN_BORDER
    ws_ov.row_dimensions[2].height = 22

    ws_ov.column_dimensions['A'].width = 10
    ws_ov.column_dimensions['B'].width = 36
    ws_ov.column_dimensions['C'].width = 60
    ws_ov.column_dimensions['D'].width = 12
    ws_ov.column_dimensions['E'].width = 50

    # 按分类分组
    cat_order = []
    seen = set()
    for iface in INTERFACES:
        cat = iface['category']
        if cat not in seen:
            seen.add(cat)
            cat_order.append(cat)

    grouped = {}
    for iface in INTERFACES:
        cat = iface['category']
        if cat not in grouped:
            grouped[cat] = []
        grouped[cat].append(iface)

    row = 3

    for cat in cat_order:
        items = grouped[cat]

        # 分类标题行
        ws_ov.merge_cells(f'A{row}:E{row}')
        cell = ws_ov.cell(row=row, column=1, value=f'▸ {cat}')
        cell.font = BOLD_11
        cell.fill = LIGHT_BLUE
        cell.alignment = Alignment(horizontal='left', vertical='center')
        cell.border = THIN_BORDER
        for c in range(2, 6):
            ws_ov.cell(row=row, column=c).border = THIN_BORDER
            ws_ov.cell(row=row, column=c).fill = LIGHT_BLUE
        ws_ov.row_dimensions[row].height = 22
        row += 1

        for idx, item in enumerate(items, 1):
            name = item['module']
            url = f"{BASE_URL}{item['url']}"
            method = item['method']
            desc = item['desc']

            ws_ov.cell(row=row, column=1, value=idx).font = NORMAL_10
            ws_ov.cell(row=row, column=2, value=name).font = NORMAL_10
            ws_ov.cell(row=row, column=3, value=url).font = NORMAL_10
            ws_ov.cell(row=row, column=4, value=method).font = NORMAL_10
            ws_ov.cell(row=row, column=5, value=desc[:200] if desc else '').font = NORMAL_10

            for c in range(1, 6):
                ws_ov.cell(row=row, column=c).fill = LIGHT_GRAY
                ws_ov.cell(row=row, column=c).border = THIN_BORDER
                if c in (1, 4):
                    ws_ov.cell(row=row, column=c).alignment = ALIGN_CENTER
                else:
                    ws_ov.cell(row=row, column=c).alignment = ALIGN_LEFT

            # ── 子表 ──
            sheet_label = f'{idx}.{name[:25]}'
            sheet_label = re.sub(r'[\\\*\?\/\[\]:]', '', sheet_label)[:31]
            # 保证唯一性
            if sheet_label in wb.sheetnames:
                sheet_label = f'{idx}_{name[:20]}'
                sheet_label = re.sub(r'[\\\*\?\/\[\]:]', '', sheet_label)[:31]

            ws_sub = wb.create_sheet(sheet_label)
            create_sub_sheet(ws_sub, item)

            row += 1

    # 保存
    wb.save(OUT)
    total = len(INTERFACES)
    print(f"Excel saved: {OUT}")
    print(f"Total interfaces: {total}")
    print(f"Categories: {cat_order}")


if __name__ == '__main__':
    generate_excel()
