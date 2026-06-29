#!/usr/bin/env python3
"""Step 4: Generate Business Object → Data Table Mapping Excel (产物4)

Maps platform BusinessObjects to ERP tables and fields.
Based on: solution doc §5.7 制造业行业语义模板 + actual ERP schema
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent.parent / "梳理后"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Styling ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
TITLE_FONT = Font(name="微软雅黑", size=12, bold=True, color="1F4E79")
BODY_FONT = Font(name="微软雅黑", size=9)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
OBJ_FILL = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")

# ═══════════════════════════════════════════════════
# 业务对象 → 数据表映射定义
# ═══════════════════════════════════════════════════

BUSINESS_OBJECTS = [
    # ── 主数据对象 ──
    {
        "object": "Material / 物料",
        "type": "主数据对象",
        "category": "库存管理",
        "mappings": [
            ("物料编码",       "E10_ERP", "ITEM",              "ITEM_NO",               "人工确认"),
            ("物料名称",       "E10_ERP", "ITEM",              "ITEM_NAME",             "人工确认"),
            ("规格",          "E10_ERP", "ITEM",              "ITEM_SPEC",             "人工确认"),
            ("单位",          "E10_ERP", "ITEM",              "ITEM_UNIT",             "人工确认"),
            ("物料分类",       "E10_ERP", "ITEM",              "ITEM_CATEGORY",         "系统推断"),
            ("物料状态",       "E10_ERP", "ITEM",              "ITEM_STATUS",           "系统推断"),
            ("品号财务信息",    "E10_ERP", "ITEM_FINANCIAL",    "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号采购信息",    "E10_ERP", "ITEM_PURCHASE",     "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号销售信息",    "E10_ERP", "ITEM_SALES",        "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号工厂信息",    "E10_ERP", "ITEM_PLANT",        "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号仓库库存",    "E10_ERP", "ITEM_WAREHOUSE",    "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号供应商信息",  "E10_ERP", "ITEM_SUPPLIER",     "ITEM_ID → ITEM_NO",     "系统推断"),
            ("工艺路线",       "E10_ERP", "ITEM_ROUTING",       "ITEM_ID → ITEM_NO",     "系统推断"),
            ("品号条码",       "E10_ERP", "ITEM_BARCODE",       "ITEM_ID → ITEM_NO",     "系统推断"),
        ],
    },
    {
        "object": "Supplier / 供应商",
        "type": "主数据对象",
        "category": "采购供应",
        "mappings": [
            ("供应商编码",      "E10_ERP", "SUPPLIER",          "SUPPLIER_CODE",         "系统推断"),
            ("供应商名称",      "E10_ERP", "SUPPLIER",          "SUPPLIER_NAME",         "系统推断"),
            ("供应商分类",      "E10_ERP", "SUPPLIER",          "SUPPLIER_CATEGORY",     "系统推断"),
            ("供应商状态",      "E10_ERP", "SUPPLIER",          "SUPPLIER_STATUS",       "系统推断"),
            ("供应商联系人",    "E10_ERP", "SUPPLIER_CONTACT",   "CONTACT_NAME",          "人工确认"),
            ("供应商地址",      "E10_ERP", "SUPPLIER_ADDRESS",   "ADDRESS",               "人工确认"),
            ("供应商银行",      "E10_ERP", "SUPPLIER_BANK",      "BANK_ACCOUNT",          "系统推断"),
            ("供应商财务信息",  "E10_ERP", "SUPPLIER_FI",        "FI_CATEGORY",           "系统推断"),
            ("供应商采购信息",  "E10_ERP", "SUPPLIER_PURCHASE",  "PURCHASE_TYPE",         "系统推断"),
        ],
    },
    {
        "object": "Customer / 客户",
        "type": "主数据对象",
        "category": "客户关系",
        "mappings": [
            ("客户编码",       "E10_ERP", "CUSTOMER",           "CUSTOMER_NO",           "人工确认"),
            ("客户名称",       "E10_ERP", "CUSTOMER",           "CUSTOMER_NAME",         "人工确认"),
            ("客户分类",       "E10_ERP", "CUSTOMER",           "CUSTOMER_CATEGORY",     "系统推断"),
            ("信用额度",       "E10_ERP", "CUSTOMER",           "CREDIT_LIMIT",          "系统推断"),
            ("客户联系人",     "E10_ERP", "CUSTOMER_CONTACT",    "CONTACT_NAME",          "系统推断"),
            ("客户地址",       "E10_ERP", "CUSTOMER_ADDRESS",    "ADDRESS",               "系统推断"),
        ],
    },
    {
        "object": "Account / 会计科目",
        "type": "主数据对象",
        "category": "财务核算",
        "mappings": [
            ("科目编码",       "E10_ERP", "ACCOUNT_CODE",        "ACCOUNT_NO",            "人工确认"),
            ("科目名称",       "E10_ERP", "ACCOUNT_CODE",        "ACCOUNT_NAME",          "人工确认"),
            ("科目分类",       "E10_ERP", "ACCOUNT_CATEGORY",    "CATEGORY_NAME",         "系统推断"),
            ("科目余额",       "E10_ERP", "ACCOUNT_BALANCE",     "BALANCE",               "系统推断"),
            ("科目组",         "E10_ERP", "ACCOUNT_GROUP",       "GROUP_NAME",            "系统推断"),
        ],
    },

    # ── 结构对象 ──
    {
        "object": "BOM / 物料清单",
        "type": "结构对象",
        "category": "生产制造",
        "mappings": [
            ("BOM编码",        "E10_ERP", "BOM",                 "BOM_NO",                "系统推断"),
            ("父件编码",       "E10_ERP", "BOM",                 "PARENT_ITEM_ID",        "系统推断"),
            ("子件编码",       "E10_ERP", "BOM_D",               "CHILD_ITEM_ID",         "系统推断"),
            ("用量",          "E10_ERP", "BOM_D",               "QUANTITY",              "系统推断"),
            ("损耗率",         "E10_ERP", "BOM_D",               "LOSS_RATE",             "系统推断"),
            ("BOM版本",       "E10_ERP", "BOM",                 "VERSION",               "系统推断"),
            ("生效日期",       "E10_ERP", "BOM",                 "EFFECTIVE_DATE",        "系统推断"),
        ],
    },

    # ── 交易对象 ──
    {
        "object": "PurchaseOrder / 采购订单",
        "type": "交易对象",
        "category": "采购供应",
        "mappings": [
            ("采购订单号",      "E10_ERP", "PURCHASE_ORDER",      "DOC_NO",                "人工确认"),
            ("订单日期",       "E10_ERP", "PURCHASE_ORDER",      "DOC_DATE",              "人工确认"),
            ("供应商",         "E10_ERP", "PURCHASE_ORDER",      "SUPPLIER_ID",           "人工确认"),
            ("物料编码",       "E10_ERP", "PURCHASE_ORDER_D",    "ITEM_ID",               "人工确认"),
            ("采购数量",       "E10_ERP", "PURCHASE_ORDER_D",    "ORDER_QTY",             "人工确认"),
            ("单价",          "E10_ERP", "PURCHASE_ORDER_D",    "UNIT_PRICE",            "系统推断"),
            ("金额",          "E10_ERP", "PURCHASE_ORDER_D",    "AMOUNT",                "系统推断"),
            ("交货日期",       "E10_ERP", "PURCHASE_ORDER_D",    "DELIVERY_DATE",         "系统推断"),
            ("订单状态",       "E10_ERP", "PURCHASE_ORDER",      "STATUS",                "系统推断"),
            ("收货信息",       "E10_ERP", "PURCHASE_RECEIPT",    "RECEIPT_QTY",           "系统推断"),
            ("到货检验",       "E10_ERP", "PO_ARRIVAL_INSPECTION","INSPECTION_QTY",        "系统推断"),
            ("采购合同",       "E10_ERP", "PURCHASE_CONTRACT",   "CONTRACT_NO",           "系统推断"),
            ("采购变更",       "E10_ERP", "PO_CHANGE",           "CHANGE_DATE",           "系统推断"),
        ],
    },
    {
        "object": "ProductionOrder / 制造工单",
        "type": "过程对象",
        "category": "生产制造",
        "mappings": [
            ("工单号",         "E10_ERP", "MO",                  "MO_NO",                 "人工确认"),
            ("工单类型",       "E10_ERP", "MO",                  "MO_TYPE",               "人工确认"),
            ("物料编码",       "E10_ERP", "MO",                  "ITEM_ID",               "人工确认"),
            ("计划数量",       "E10_ERP", "MO",                  "PLAN_QTY",              "人工确认"),
            ("完工数量",       "E10_ERP", "MO",                  "COMPLETED_QTY",         "人工确认"),
            ("工单状态",       "E10_ERP", "MO",                  "STATUS",                "人工确认"),
            ("计划开工",       "E10_ERP", "MO",                  "PLAN_START_DATE",       "系统推断"),
            ("计划完工",       "E10_ERP", "MO",                  "PLAN_END_DATE",         "系统推断"),
            ("实际开工",       "E10_ERP", "MO",                  "ACTUAL_START_DATE",     "系统推断"),
            ("实际完工",       "E10_ERP", "MO",                  "ACTUAL_END_DATE",       "系统推断"),
            ("工序明细",       "E10_ERP", "MO_ROUTING",          "ROUTING_SEQ",           "系统推断"),
            ("投料明细",       "E10_ERP", "MO_MATERIAL",         "MATERIAL_ID",           "系统推断"),
            ("工单工艺",       "E10_ERP", "ITEM_ROUTING",        "ITEM_ID → MO.ITEM_ID",  "系统推断"),
            ("工作中心",       "E10_ERP", "WORK_CENTER",         "WORK_CENTER_NO",        "系统推断"),
        ],
    },
    {
        "object": "SalesOrder / 销售订单",
        "type": "交易对象",
        "category": "销售分销",
        "mappings": [
            ("销售订单号",      "E10_ERP", "SALE",                "DOC_NO",                "人工确认"),
            ("订单日期",       "E10_ERP", "SALE",                "DOC_DATE",              "人工确认"),
            ("客户",          "E10_ERP", "SALE",                "CUSTOMER_ID",           "人工确认"),
            ("物料编码",       "E10_ERP", "SALE_D",              "ITEM_ID",               "系统推断"),
            ("销售数量",       "E10_ERP", "SALE_D",              "ORDER_QTY",             "系统推断"),
            ("单价",          "E10_ERP", "SALE_D",              "UNIT_PRICE",            "系统推断"),
            ("金额",          "E10_ERP", "SALE_D",              "AMOUNT",                "系统推断"),
            ("交货日期",       "E10_ERP", "SALE_D",              "DELIVERY_DATE",         "系统推断"),
            ("订单状态",       "E10_ERP", "SALE",                "STATUS",                "系统推断"),
            ("出货指令",       "E10_ERP", "SHIPPING",            "SHIP_QTY",              "系统推断"),
        ],
    },

    # ── 资源对象 ──
    {
        "object": "Inventory / 库存",
        "type": "资源对象",
        "category": "库存管理",
        "mappings": [
            ("物料编码",       "E10_ERP", "ITEM_WAREHOUSE",      "ITEM_ID",               "人工确认"),
            ("仓库",          "E10_ERP", "ITEM_WAREHOUSE",      "WAREHOUSE_ID",          "人工确认"),
            ("库存数量",       "E10_ERP", "ITEM_WAREHOUSE",      "ON_HAND_QTY",           "人工确认"),
            ("可用数量",       "E10_ERP", "ITEM_WAREHOUSE",      "AVAILABLE_QTY",         "系统推断"),
            ("在途数量",       "E10_ERP", "ITEM_WAREHOUSE",      "IN_TRANSIT_QTY",        "系统推断"),
            ("库存成本",       "E10_ERP", "INV_STOCK",           "UNIT_COST",             "系统推断"),
            ("库存价值",       "E10_ERP", "INV_STOCK",           "TOTAL_VALUE",           "系统推断"),
            ("批号",          "E10_ERP", "ITEM_LOT",            "LOT_NO",                "系统推断"),
            ("条码",          "E10_ERP", "ITEM_BARCODE",         "BARCODE",               "系统推断"),
            ("序列号",        "E10_ERP", "SN",                   "SERIAL_NO",             "系统推断"),
            ("货位",          "E10_ERP", "FIL",                  "LOCATION",              "系统推断"),
        ],
    },
    {
        "object": "WorkCenter / 工作中心",
        "type": "资源对象",
        "category": "生产制造",
        "mappings": [
            ("工作中心编码",    "E10_ERP", "WORK_CENTER",         "WORK_CENTER_NO",        "人工确认"),
            ("工作中心名称",    "E10_ERP", "WORK_CENTER",         "WORK_CENTER_NAME",      "人工确认"),
            ("产能",          "E10_ERP", "WORK_CENTER",         "CAPACITY",              "系统推断"),
            ("班次",          "E10_ERP", "WORK_SHIFT",           "SHIFT_NAME",            "系统推断"),
            ("效率",          "E10_ERP", "WORK_CENTER",         "EFFICIENCY",            "系统推断"),
        ],
    },
    {
        "object": "Equipment / 设备",
        "type": "资源对象",
        "category": "资产设备",
        "mappings": [
            ("设备编码",       "E10_ERP", "EQT",                  "EQT_NO",                "系统推断"),
            ("设备名称",       "E10_ERP", "EQT",                  "EQT_NAME",              "系统推断"),
            ("设备类型",       "E10_ERP", "EQT",                  "EQT_TYPE",              "系统推断"),
            ("设备状态",       "E10_ERP", "EQT",                  "STATUS",                "系统推断"),
            ("维护记录",       "E10_ERP", "MAINTAIN",             "MAINTENANCE_DATE",      "系统推断"),
            ("设备台账",       "E10_ERP", "EQUIPMENT",            "EQUIPMENT_NO",          "系统推断"),
        ],
    },
    {
        "object": "Asset / 固定资产",
        "type": "资源对象",
        "category": "资产设备",
        "mappings": [
            ("资产编码",       "E10_ERP", "ASSET",                "ASSET_NO",              "系统推断"),
            ("资产名称",       "E10_ERP", "ASSET",                "ASSET_NAME",            "系统推断"),
            ("资产分类",       "E10_ERP", "ASSET",                "ASSET_CATEGORY",        "系统推断"),
            ("原值",          "E10_ERP", "ASSET",                "ORIGINAL_VALUE",        "系统推断"),
            ("累计折旧",       "E10_ERP", "ASSET",                "ACCUMULATED_DEPR",      "系统推断"),
            ("净值",          "E10_ERP", "ASSET",                "NET_VALUE",             "系统推断"),
            ("折旧方法",       "E10_ERP", "ASSET",                "DEPR_METHOD",           "系统推断"),
            ("资产卡片",       "E10_ERP", "ASSET_CARD",           "CARD_NO",               "系统推断"),
        ],
    },

    # ── 财务对象 ──
    {
        "object": "AR / 应收账款",
        "type": "财务对象",
        "category": "财务核算",
        "mappings": [
            ("应收单号",       "E10_ERP", "AR",                   "DOC_NO",                "人工确认"),
            ("客户",          "E10_ERP", "AR",                   "CUSTOMER_ID",           "人工确认"),
            ("应收金额",       "E10_ERP", "AR",                   "AMOUNT",                "人工确认"),
            ("已收金额",       "E10_ERP", "AR",                   "RECEIVED_AMOUNT",       "系统推断"),
            ("余额",          "E10_ERP", "AR",                   "BALANCE",               "系统推断"),
            ("到期日",        "E10_ERP", "AR",                   "DUE_DATE",              "系统推断"),
            ("收款记录",       "E10_ERP", "PAYMENT",              "PAYMENT_AMOUNT",        "系统推断"),
        ],
    },
    {
        "object": "AP / 应付账款",
        "type": "财务对象",
        "category": "财务核算",
        "mappings": [
            ("应付单号",       "E10_ERP", "AP",                   "DOC_NO",                "人工确认"),
            ("供应商",         "E10_ERP", "AP",                   "SUPPLIER_ID",           "人工确认"),
            ("应付金额",       "E10_ERP", "AP",                   "AMOUNT",                "人工确认"),
            ("已付金额",       "E10_ERP", "AP",                   "PAID_AMOUNT",           "系统推断"),
            ("余额",          "E10_ERP", "AP",                   "BALANCE",               "系统推断"),
            ("到期日",        "E10_ERP", "AP",                   "DUE_DATE",              "系统推断"),
            ("付款记录",       "E10_ERP", "PAYMENT",              "PAYMENT_AMOUNT",        "系统推断"),
        ],
    },
    {
        "object": "GL / 总账",
        "type": "财务对象",
        "category": "财务核算",
        "mappings": [
            ("凭证号",        "E10_ERP", "VOUCHER",              "VOUCHER_NO",            "人工确认"),
            ("凭证日期",       "E10_ERP", "VOUCHER",              "VOUCHER_DATE",          "人工确认"),
            ("科目",          "E10_ERP", "JOURNAL",              "ACCOUNT_ID",            "人工确认"),
            ("借方金额",       "E10_ERP", "JOURNAL",              "DEBIT_AMOUNT",          "人工确认"),
            ("贷方金额",       "E10_ERP", "JOURNAL",              "CREDIT_AMOUNT",         "人工确认"),
            ("摘要",          "E10_ERP", "JOURNAL",              "DESCRIPTION",           "系统推断"),
            ("科目余额",       "E10_ERP", "GL",                   "BALANCE",               "系统推断"),
        ],
    },
    {
        "object": "Cost / 成本核算",
        "type": "财务对象",
        "category": "财务核算",
        "mappings": [
            ("物料编码",       "E10_ERP", "COST",                 "ITEM_ID",               "人工确认"),
            ("材料成本",       "E10_ERP", "COST",                 "MATERIAL_COST",         "系统推断"),
            ("人工成本",       "E10_ERP", "COST",                 "LABOR_COST",            "系统推断"),
            ("制造费用",       "E10_ERP", "COST",                 "OVERHEAD_COST",         "系统推断"),
            ("单位成本",       "E10_ERP", "COST",                 "UNIT_COST",             "系统推断"),
            ("库存成本流转",    "E10_ERP", "INV_STOCK",           "COST_FLOW",             "系统推断"),
        ],
    },

    # ── 质量对象 ──
    {
        "object": "Inspection / 检验管理",
        "type": "过程对象",
        "category": "质量管理",
        "mappings": [
            ("检验单号",       "E10_ERP", "INSPECTION",           "DOC_NO",                "人工确认"),
            ("检验类型",       "E10_ERP", "INSPECTION",           "INSPECTION_TYPE",       "系统推断"),
            ("检验标准",       "E10_ERP", "INSPECTION",           "INSPECTION_STANDARD",   "系统推断"),
            ("抽样方案",       "E10_ERP", "INSPECTION",           "SAMPLING_PLAN",         "系统推断"),
            ("检验结果",       "E10_ERP", "INSPECTION",           "RESULT",                "系统推断"),
            ("合格数量",       "E10_ERP", "INSPECTION",           "ACCEPTABLE_QTY",        "系统推断"),
            ("不合格数量",     "E10_ERP", "INSPECTION",           "UNQUALIFIED_QTY",       "系统推断"),
            ("来料检验",       "E10_ERP", "PO_ARRIVAL_INSPECTION","INSPECTION_QTY",        "系统推断"),
            ("不良品记录",     "E10_ERP", "DEFECTIVE",            "DEFECT_QTY",            "系统推断"),
            ("验收确认",       "E10_ERP", "ACCEPTANCE",           "ACCEPT_QTY",            "系统推断"),
        ],
    },

    # ── 事件对象 ──
    {
        "object": "Shipment / 发货",
        "type": "事件对象",
        "category": "销售分销",
        "mappings": [
            ("发货单号",       "E10_ERP", "SHIPPING",             "DOC_NO",                "人工确认"),
            ("销售订单",       "E10_ERP", "SHIPPING",             "SALE_ID",               "系统推断"),
            ("客户",          "E10_ERP", "SHIPPING",             "CUSTOMER_ID",           "系统推断"),
            ("发货数量",       "E10_ERP", "SHIPPING",             "SHIP_QTY",              "系统推断"),
            ("发货日期",       "E10_ERP", "SHIPPING",             "SHIP_DATE",             "系统推断"),
            ("物流跟踪",       "E10_ERP", "SHIPPING",             "TRACKING_NO",           "系统推断"),
        ],
    },
]

# ═══════════════════════════════════════════════════
# 构建 Excel
# ═══════════════════════════════════════════════════

def build():
    wb = openpyxl.Workbook()

    # ── Sheet 1: 对象映射总览 ──
    ws = wb.active
    ws.title = "对象映射总览"
    _write_mapping_sheet(ws)

    # ── Sheet 2: 对象清单 ──
    ws2 = wb.create_sheet("业务对象清单")
    _write_object_list(ws2)

    # ── Sheet 3: 统计汇总 ──
    ws3 = wb.create_sheet("统计汇总")
    _write_stats(ws3)

    out_path = OUTPUT_DIR / "ERP业务对象映射表.xlsx"
    wb.save(str(out_path))
    print(f"✅ 已生成: {out_path}")
    print(f"   业务对象: {len(BUSINESS_OBJECTS)}个, 映射关系: {sum(len(obj['mappings']) for obj in BUSINESS_OBJECTS)}条")


def _write_mapping_sheet(ws):
    """完整的对象→属性→表→字段映射"""
    ws.merge_cells("A1:H1")
    ws["A1"] = "ERP 业务对象 → 数据表映射 (DataMapping)"
    ws["A1"].font = TITLE_FONT

    headers = ["业务对象", "对象类型", "业务域", "对象属性", "来源系统", "来源表", "来源字段", "映射可信度"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    widths = [20, 12, 12, 14, 10, 24, 28, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.freeze_panes = "A3"

    row = 3
    current_object = None
    for obj in BUSINESS_OBJECTS:
        obj_name = obj["object"]
        obj_type = obj["type"]
        obj_cat = obj["category"]

        for attr_name, source_sys, source_tbl, source_col, confidence in obj["mappings"]:
            # Write object info only for first row of each object
            vals = [
                obj_name if current_object != obj_name else "",
                obj_type if current_object != obj_name else "",
                obj_cat if current_object != obj_name else "",
                attr_name, source_sys, source_tbl, source_col, confidence,
            ]
            for c, v in enumerate(vals, 1):
                cell = ws.cell(row=row, column=c, value=v)
                cell.font = BODY_FONT
                cell.border = THIN_BORDER
                if c in (4, 6, 7):
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            # Highlight object header rows
            if current_object != obj_name:
                for c in range(1, 9):
                    ws.cell(row=row, column=c).fill = OBJ_FILL
                current_object = obj_name

            row += 1


def _write_object_list(ws):
    """业务对象清单（不含字段级映射）"""
    ws.merge_cells("A1:F1")
    ws["A1"] = "业务对象清单"
    ws["A1"].font = TITLE_FONT

    headers = ["业务对象", "对象类型", "业务域", "映射属性数", "关联核心模块", "说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

    widths = [22, 14, 12, 12, 36, 40]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # Object → modules mapping
    obj_modules = {
        "Material / 物料":         "ITEM, ITEM_PLANT, ITEM_WAREHOUSE, ITEM_SUPPLIER, ITEM_PURCHASE, ITEM_SALES, ITEM_ROUTING, ITEM_BARCODE, ITEM_FINANCIAL",
        "Supplier / 供应商":       "SUPPLIER, SUPPLIER_CONTACT, SUPPLIER_ADDRESS, SUPPLIER_BANK, SUPPLIER_FI, SUPPLIER_PURCHASE",
        "Customer / 客户":         "CUSTOMER, CUSTOMER_CONTACT, CUSTOMER_ADDRESS",
        "Account / 会计科目":      "ACCOUNT_CODE, ACCOUNT_CATEGORY, ACCOUNT_BALANCE, ACCOUNT_GROUP",
        "BOM / 物料清单":          "BOM, BOM_D",
        "PurchaseOrder / 采购订单":"PURCHASE_ORDER, PURCHASE_ORDER_D, PURCHASE_RECEIPT, PO_ARRIVAL_INSPECTION, PURCHASE_CONTRACT, PO_CHANGE",
        "ProductionOrder / 制造工单":"MO, MO_ROUTING, MO_MATERIAL, ITEM_ROUTING, WORK_CENTER",
        "SalesOrder / 销售订单":   "SALE, SALE_D, SHIPPING",
        "Inventory / 库存":        "ITEM_WAREHOUSE, INV_STOCK, ITEM_LOT, ITEM_BARCODE, SN, FIL",
        "WorkCenter / 工作中心":   "WORK_CENTER, WORK_SHIFT",
        "Equipment / 设备":        "EQT, MAINTAIN, EQUIPMENT",
        "Asset / 固定资产":        "ASSET, ASSET_CARD",
        "AR / 应收账款":           "AR, PAYMENT",
        "AP / 应付账款":           "AP, PAYMENT",
        "GL / 总账":              "VOUCHER, JOURNAL, GL",
        "Cost / 成本核算":         "COST, INV_STOCK",
        "Inspection / 检验管理":   "INSPECTION, PO_ARRIVAL_INSPECTION, DEFECTIVE, ACCEPTANCE",
        "Shipment / 发货":         "SHIPPING",
    }

    obj_descriptions = {
        "Material / 物料":         "制造企业核心主数据，所有业务单据的基础参照对象",
        "Supplier / 供应商":       "采购供应源头，连接采购订单、应付账款",
        "Customer / 客户":         "销售需求源头，连接销售订单、应收账款",
        "Account / 会计科目":      "财务核算基础维度，贯穿总账、应收应付、成本",
        "BOM / 物料清单":          "产品结构定义，连接物料与生产工单",
        "PurchaseOrder / 采购订单":"采购业务核心交易单据，从下单到收货入库",
        "ProductionOrder / 制造工单":"生产执行核心单据，从排产到完工入库",
        "SalesOrder / 销售订单":   "销售业务核心交易单据，从接单到发货",
        "Inventory / 库存":        "供需缓冲，连接采购收货和生产消耗",
        "WorkCenter / 工作中心":   "产能资源，关联工单排程和工序执行",
        "Equipment / 设备":        "生产设备资产，关联维护和产能",
        "Asset / 固定资产":        "企业固定资产全生命周期管理",
        "AR / 应收账款":           "销售端财务闭环：发货→应收→收款",
        "AP / 应付账款":           "采购端财务闭环：收货→应付→付款",
        "GL / 总账":              "财务核算中枢：凭证→科目余额→报表",
        "Cost / 成本核算":         "成本归集与分配，支撑盈利分析",
        "Inspection / 检验管理":   "质量管控：来料检、过程检、不良品处理",
        "Shipment / 发货":         "物流执行：拣货、包装、发运、跟踪",
    }

    for i, obj in enumerate(BUSINESS_OBJECTS, 3):
        obj_name = obj["object"]
        vals = [
            obj_name,
            obj["type"],
            obj["category"],
            len(obj["mappings"]),
            obj_modules.get(obj_name, ""),
            obj_descriptions.get(obj_name, ""),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER
            if c in (5, 6):
                cell.alignment = Alignment(wrap_text=True, vertical="top")


def _write_stats(ws):
    """统计汇总"""
    ws.merge_cells("A1:D1")
    ws["A1"] = "映射统计汇总"
    ws["A1"].font = TITLE_FONT

    headers = ["统计项", "数值", "说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = THIN_BORDER

    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 50

    total_mappings = sum(len(obj["mappings"]) for obj in BUSINESS_OBJECTS)
    human_confirmed = sum(1 for obj in BUSINESS_OBJECTS for _, _, _, _, conf in obj["mappings"] if conf == "人工确认")
    system_inferred = total_mappings - human_confirmed

    stats = [
        ("业务对象总数", len(BUSINESS_OBJECTS), "覆盖制造价值链核心对象"),
        ("映射关系总数", total_mappings, "对象属性 → 来源表.来源字段"),
        ("人工确认", human_confirmed, "基于ERP培训手册和业务知识确认的字段"),
        ("系统推断", system_inferred, "基于命名规则和外键关系推断的字段，需人工复核"),
        ("主数据对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "主数据对象"), "Material, Supplier, Customer, Account"),
        ("交易对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "交易对象"), "PurchaseOrder, SalesOrder"),
        ("过程对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "过程对象"), "ProductionOrder, Inspection"),
        ("资源对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "资源对象"), "Inventory, WorkCenter, Equipment, Asset"),
        ("财务对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "财务对象"), "AR, AP, GL, Cost"),
        ("结构对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "结构对象"), "BOM"),
        ("事件对象", sum(1 for obj in BUSINESS_OBJECTS if obj["type"] == "事件对象"), "Shipment"),
    ]

    for i, (name, val, desc) in enumerate(stats, 3):
        for c, v in enumerate([name, val, desc], 1):
            cell = ws.cell(row=i, column=c, value=v)
            cell.font = BODY_FONT
            cell.border = THIN_BORDER


if __name__ == "__main__":
    build()
