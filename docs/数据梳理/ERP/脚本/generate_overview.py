# -*- coding: utf-8 -*-
"""
生成 ERP 核心数据表总览目录 Excel（产物2）
1 Excel × 9 sheet（总览统计 + 8 业务域）
"""
import os, re, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

BASE = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
CORE_DIR = os.path.join(BASE, "核心模块")
MODULES_DIR = os.path.join(BASE, "全量模块")
OUT_DIR = os.path.join(BASE, "梳理后")
os.makedirs(OUT_DIR, exist_ok=True)

# ── 样式定义 ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
TITLE_FONT  = Font(name="微软雅黑", size=11, bold=True, color="1F4E79")
BODY_FONT   = Font(name="微软雅黑", size=9)
P0_FILL     = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")  # 浅橙
P2_FILL     = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")  # 浅灰
THIN_BORDER = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT   = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════════
# 8 个业务域定义（MOC 分类 + ITEM 手工归入库存管理）
# ═══════════════════════════════════════════════════

DOMAINS = {
    "01财务核算": {
        "desc": "总账、应收应付、成本、预算、税务、凭证",
        "全量模块": {
            "ACCOUNT":   ("会计科目", 10, 885),
            "ADJUST":    ("调整凭证", 4, 177),
            "AP":        ("应付账款", 15, 692),
            "AR":        ("应收账款", 22, 1012),
            "BANK":      ("银行管理", 9, 442),
            "BUDGET":    ("预算管理", 9, 409),
            "CASHFLOW":  ("现金流量", 8, 351),
            "CFS":       ("财务结转", 23, 999),
            "CNNM":      ("大陆税务", 2, 22),
            "COST":      ("成本核算", 17, 748),
            "CP":        ("商业票据", 4, 197),
            "EXP":       ("费用分摊", 8, 466),
            "EXPENSES":  ("费用管理", 6, 287),
            "GL":        ("总账", 4, 44),
            "INNER":     ("内部交易", 12, 775),
            "INV":       ("库存成本", 8, 356),
            "JOURNAL":   ("日记账", 5, 273),
            "OTHER":     ("其他收支", 8, 413),
            "PAYMENT":   ("付款管理", 19, 975),
            "PREPAID":   ("预付账款", 3, 165),
            "SIMULATE":  ("成本模拟", 6, 299),
            "TAX":       ("税务管理", 13, 588),
            "TRANS":     ("交易行", 3, 121),
            "TW":        ("台湾发票", 17, 1144),
            "TWNM":      ("台湾税务", 0, 0),
            "VOUCHER":   ("凭证管理", 5, 266),
            "WRITEOFFS": ("核销管理", 0, 0),
        },
    },
    "02销售分销": {
        "desc": "销售、门店、柜台、报价、出货、退货、促销",
        "全量模块": {
            "BEHALF":     ("代销业务", 5, 212),
            "BILL":       ("账单促销", 14, 507),
            "CALL":       ("叫货管理", 4, 178),
            "CONTRACT":   ("合同管理", 5, 318),
            "COUNTER":    ("柜台交易", 20, 1055),
            "CST":        ("报关管理", 6, 171),
            "DELIVERY":   ("交货条款", 2, 114),
            "DIRECT":     ("直运业务", 2, 99),
            "DISCOUNT":   ("折扣表", 4, 171),
            "DISTRIBUTION": ("分销配送", 9, 440),
            "EXPORT":     ("出口管理", 8, 381),
            "GOOD":       ("商品促销", 14, 513),
            "GROUP":      ("组合促销", 14, 493),
            "LETTER":     ("信用证", 4, 207),
            "ORDER":      ("订货管理", 2, 94),
            "PACKING":    ("包装清单", 4, 186),
            "POINT":      ("积分促销", 4, 157),
            "PRICE":      ("价格策略", 9, 346),
            "PROMOTION":  ("促销管理", 14, 472),
            "QUOTATION":  ("报价管理", 5, 242),
            "RETURN":     ("退货管理", 6, 266),
            "SALE":       ("销售单据", 23, 1019),
            "SALES":      ("销售中心", 23, 1776),
            "SELLING":    ("销售限制", 5, 195),
            "SHIPPING":   ("出货指令", 3, 111),
            "SHOP":       ("门店管理", 18, 787),
            "SHOPPE":     ("门店费用", 19, 997),
            "SO":         ("销售变更", 6, 591),
        },
    },
    "03库存管理": {
        "desc": "物料、条码、调拨、序列号、盘点、寄售",
        "全量模块": {
            "AJST":       ("库存调整核算", 0, 0),
            "BC":         ("条码管理", 32, 1276),
            "BORROW":     ("借入借出", 4, 273),
            "CHECK":      ("盘点调整", 2, 98),
            "CONSIGN":    ("寄售管理", 5, 356),
            "COUNTING":   ("盘点计划", 4, 220),
            "DAMAGE":     ("损坏管理", 3, 134),
            "FIL":        ("货位管理", 6, 248),
            "ISSUE":      ("发料出库", 16, 703),
            "ITEM":       ("物料商品", 53, 2609),   # ← 手工归入库存管理
            "KIT":        ("套件管理", 4, 196),
            "LOST":       ("报损管理", 3, 131),
            "MATERIAL":   ("物料调拨", 3, 144),
            "PARTS":      ("备件管理", 3, 150),
            "SN":         ("序列号", 11, 525),
            "STOCK":      ("库存请购", 2, 104),
            "TRANSFER":   ("调拨管理", 8, 478),
            "VMI":        ("供应商寄售", 5, 254),
        },
    },
    "04生产制造": {
        "desc": "工单、BOM、工程变更、在制、工作中心",
        "全量模块": {
            "APS":        ("高级排程计划", 84, 4599),
            "BOM":        ("物料清单", 11, 481),
            "CHANGE":     ("变更执行", 8, 436),
            "ECN":        ("工程变更", 12, 612),
            "FORECAST":   ("需求预测", 4, 278),
            "MO":         ("制造工单", 42, 2212),
            "OUT":        ("委外加工", 7, 369),
            "PLAN":       ("计划管理", 9, 440),
            "PRE":        ("预置数据", 6, 241),
            "PROJECT":    ("项目管理", 2, 83),
            "RESOURCE":   ("资源管理", 5, 184),
            "SCRAP":      ("报废管理", 2, 123),
            "SF":         ("数据采集", 10, 412),
            "SUGGESTION": ("建议计划", 6, 366),
            "WORK":       ("工作中心", 12, 434),
        },
    },
    "05采购供应": {
        "desc": "采购、供应商、请购、进口、询价",
        "全量模块": {
            "IMPORT":     ("进口管理", 10, 455),
            "NEGOTIATE":  ("议价管理", 4, 170),
            "PO":         ("采购订单", 15, 1009),
            "PURCHASE":   ("采购管理", 22, 1763),
            "REQUISITION":("请购管理", 3, 174),
            "SUPPLIER":   ("供应商", 8, 361),
            "SUPPLY":     ("供应中心", 6, 284),
        },
    },
    "06客户关系": {
        "desc": "客户、会员、卡券、服务、维修",
        "全量模块": {
            "BP":         ("业务伙伴", 4, 158),
            "CARD":       ("卡券管理", 29, 1249),
            "CRM":        ("客户关系", 8, 297),
            "CUSTOMER":   ("客户管理", 24, 1142),
            "MEMBER":     ("会员管理", 44, 1694),
            "REPAIR":     ("维修管理", 15, 640),
            "REPLACE":    ("替换管理", 7, 283),
            "SERVICE":    ("售后服务", 30, 1407),
            "TICKET":     ("票券管理", 18, 697),
        },
    },
    "07资产设备": {
        "desc": "固定资产、设备、维护、模具",
        "全量模块": {
            "ASSET":      ("固定资产", 39, 2459),
            "EQT":        ("设备管理", 34, 1497),
            "EQUIPMENT":  ("设备台账", 3, 159),
            "MAINTAIN":   ("设备维护", 8, 395),
        },
    },
    "08质量管理": {
        "desc": "检验、QC、不良品、验收",
        "全量模块": {
            "ACCEPTANCE": ("验收管理", 7, 283),
            "DEFECTIVE":  ("不良品", 4, 163),
            "INSP":       ("检验单据", 6, 219),
            "INSPECTION": ("检验管理", 7, 305),
        },
    },
}

# ═══════════════════════════════════════════════════
# P0 级别判定规则（精确的表级规则）
# ═══════════════════════════════════════════════════

# P0 核心模块：整个模块的表都是 P0（制造业价值链不可缺的主表）
P0_MODULES = {
    "ACCOUNT", "AP", "AR",          # 财务核心
    "GL", "VOUCHER", "COST",        # 财务核心
    "JOURNAL", "PAYMENT",           # 财务核心
    "SALE", "SALES", "SHIPPING",    # 销售核心
    "ITEM", "INV",                  # 库存核心（ITEM归入库存管理）
    "MO", "BOM", "WORK",            # 生产核心
    "PO", "PURCHASE", "SUPPLIER",   # 采购核心
    "CUSTOMER",                     # 客户核心（CRM域的CUSTOMER也是P0）
    "ASSET", "EQT",                 # 资产核心
    "INSPECTION", "ACCEPTANCE",     # 质量核心
}

# P0 模块内仍需排除的表（日志/历史/模板/临时）
P0_EXCLUDE_PATTERNS = [
    r"_LOG$", r"_HISTORY$", r"_AUDIT$",   # 日志/历史/审计
    r"_TEMPLATE$", r"_TEMP$", r"_TMP$",   # 模板/临时
    r"_BAK$", r"_BACKUP$",                # 备份
    r"_SNAPSHOT$",                         # 快照
]

# P1 模块（辅助模块，表默认 P1，部分关键表可提为 P0）
P1_MODULES = {
    # 财务
    "BANK", "BUDGET", "CASHFLOW", "TAX", "CFS",
    "EXP", "EXPENSES", "INNER", "SIMULATE", "OTHER", "PREPAID",
    # 销售
    "QUOTATION", "RETURN", "CONTRACT", "SHOP",
    "DISTRIBUTION", "PRICE", "SO",
    # 库存
    "BC", "SN", "TRANSFER", "ISSUE", "STOCK",
    "CONSIGN", "VMI", "BORROW", "COUNTING",
    # 生产
    "APS", "ECN", "SF", "PLAN", "OUT",
    "CHANGE", "FORECAST", "SCRAP",
    # 采购
    "REQUISITION", "SUPPLY", "IMPORT",
    # CRM
    "MEMBER", "SERVICE", "CARD", "TICKET", "REPAIR",
    # 资产
    "MAINTAIN", "EQUIPMENT",
    # 质量
    "DEFECTIVE", "INSP",
}

# P1 模块中特别提升为 P0 的关键表（精确表名匹配）
P1_UPGRADE_TO_P0 = {
    # 库存管理
    "ITEM", "ITEM_PLANT", "ITEM_WAREHOUSE", "ITEM_SUPPLIER",
    "ITEM_COST", "ITEM_BARCODE", "ITEM_FINANCIAL",
    "ITEM_PURCHASE", "ITEM_SALES", "ITEM_ROUTING",
    "ITEM_GROUP", "ITEM_ROUTING_D", "ITEM_ROUTING_WORK_CENTER",
    "ITEM_LOT", "INV_STOCK", "INV_TRANSACTION",
    # 生产制造
    "MO", "MO_D", "MO_ROUTING", "MO_MATERIAL",
    "BOM", "BOM_D", "BOM_MATERIAL",
    "ECN", "ECN_D",
    "WORK_CENTER", "WORK_SHIFT",
    "SF_DATA", "SF_RESULT",
    # 采购
    "PO", "PO_D",
    "REQUISITION", "REQUISITION_D",
    "SUPPLIER", "SUPPLIER_CONTACT",
    # 财务辅助
    "CASHFLOW_ITEM", "CASHFLOW_STATEMENT",
    "TAX_RATE", "TAX_CODE",
    "BANK_ACCOUNT", "BANK_STATEMENT",
    # CRM
    "CUSTOMER", "CUSTOMER_CONTACT", "CUSTOMER_ADDRESS",
    "MEMBER", "MEMBER_CARD",
    # 资产
    "ASSET", "ASSET_CARD", "ASSET_DEPRECIATION",
    "EQT", "EQT_MAINTENANCE",
    # 质量
    "INSPECTION_ORDER", "INSPECTION_RESULT",
    "DEFECTIVE_RECORD",
}

# 所有其他模块 → P2
P2_MODULES = {
    "ADJUST", "TRANS", "CP", "CNNM", "TWNM", "WRITEOFFS", "TW",
    "COUNTER", "SHOPPE", "BILL", "GOOD", "PROMOTION", "GROUP",
    "BEHALF", "CALL", "LETTER", "DISCOUNT", "POINT", "PACKING",
    "DIRECT", "DELIVERY", "ORDER", "CST", "EXPORT", "SELLING",
    "AJST", "FIL", "KIT", "PARTS", "MATERIAL", "DAMAGE", "LOST", "CHECK",
    "PRE", "PROJECT", "RESOURCE", "SUGGESTION",
    "NEGOTIATE",
    "BP", "CRM", "REPLACE",
}

# ═══════════════════════════════════════════════════
# 业务含义摘要（模块级，P0模块更具体）
# ═══════════════════════════════════════════════════
MODULE_SUMMARY = {
    # P0 模块 — 制造价值链主线
    "ACCOUNT":    "会计科目基础档案：科目编码、名称、分类、初始余额",
    "AP":         "应付账款：供应商发票、应付立账、付款核销",
    "AR":         "应收账款：客户发票、应收立账、收款核销",
    "GL":         "总账：凭证过账、科目余额、期末结转",
    "VOUCHER":    "凭证管理：会计凭证录入、审核、过账",
    "COST":       "成本核算：材料成本、人工成本、制造费用分摊",
    "JOURNAL":    "日记账：日常会计分录、凭证分录明细",
    "PAYMENT":    "付款管理：付款单、付款审批、银行付款",
    "SALE":       "销售单据：销售订单、发货通知、销售出库、销售发票",
    "SALES":      "销售中心：销售预测、销售计划、销售统计",
    "SHIPPING":   "出货指令：拣货、包装、发货、物流跟踪",
    "ITEM":       "物料主数据：品号、规格、单位、BOM、工艺路线、成本",
    "INV":        "库存成本：库存价值核算、成本流转、月末加权平均",
    "MO":         "制造工单：生产订单、工序、投料、报工、完工入库",
    "BOM":        "物料清单：产品结构、父子件关系、用量、损耗率",
    "WORK":       "工作中心：设备组、产能、班次、效率",
    "PO":         "采购订单：采购下单、审批、收货、入库、对账",
    "PURCHASE":   "采购管理：采购策略、货源、价格、配额",
    "SUPPLIER":   "供应商主数据：供应商编码、名称、分类、资质、联系人",
    "INSPECTION": "来料检验：检验标准、抽样方案、检验记录、判定",
    "ACCEPTANCE": "验收管理：采购收货验收、质量确认、合格入库",
    "ASSET":      "固定资产：资产卡片、折旧、盘点、处置",
    "EQT":        "设备管理：设备台账、维修记录、保养计划、备件",
    # P1 模块
    "BANK":       "银行管理：银行账户、银行对账、银行流水",
    "BUDGET":     "预算管理：预算编制、预算控制、预算执行分析",
    "CASHFLOW":   "现金流量：现金流项目、现金流量表",
    "TAX":        "税务管理：税率、税务计算、税务申报数据",
    "CFS":        "财务结转：月末/年末结转、利润结转",
    "INNER":      "内部交易：内部往来、内部购销、内部结算",
    "CUSTOMER":   "客户主数据：客户编码、名称、分类、信用额度、地址",
    "QUOTATION":  "报价管理：销售报价单、报价审批、报价转订单",
    "RETURN":     "退货管理：销售退货、退货入库、退款处理",
    "CONTRACT":   "合同管理：销售合同、合同条款、合同执行跟踪",
    "DISTRIBUTION":"分销配送：分销渠道、配送计划、配送执行",
    "PRICE":      "价格策略：定价规则、折扣策略、促销价格",
    "SHOP":       "门店管理：门店档案、门店库存、门店销售",
    "BC":         "条码管理：条码生成、条码打印、条码扫描",
    "SN":         "序列号：序列号生成、序列号追溯、序列号库存",
    "TRANSFER":   "调拨管理：库存调拨、调拨出库、调拨入库",
    "ISSUE":      "发料出库：生产领料、委外发料、其他出库",
    "STOCK":      "库存请购：安全库存、再订货点、自动请购",
    "CONSIGN":    "寄售管理：寄售库存、寄售结算",
    "VMI":        "供应商寄售：VMI库存、VMI结算",
    "APS":        "高级排程：产能规划、排程计算、排程甘特图",
    "ECN":        "工程变更：ECN申请、审批、生效、版本管理",
    "SF":         "数据采集：车间数据采集、设备数据采集",
    "PLAN":       "计划管理：MPS主生产计划、MRP物料需求计划",
    "MEMBER":     "会员管理：会员档案、积分、等级、消费记录",
    "SERVICE":    "售后服务：服务工单、维修、保养、投诉",
    "CARD":       "卡券管理：会员卡、优惠券、储值卡",
    "MAINTAIN":   "设备维护：预防性维护、故障维修、保养计划",
    "DEFECTIVE":  "不良品：不良品记录、不良原因分析、不良品处理",
}

# ═══════════════════════════════════════════════════
# 读取模块表名（优先 core/，其次 modules/）
# ═══════════════════════════════════════════════════
def read_module_tables(module_code):
    """读取模块的表名列表，返回 [(table_name, cn_name), ...]"""
    core_file = os.path.join(CORE_DIR, f"{module_code} (.*).md")
    # search in core/
    for fname in os.listdir(CORE_DIR):
        if fname.startswith(f"{module_code} (") and fname.endswith(".md"):
            return _parse_module_md(os.path.join(CORE_DIR, fname))
    # search in modules/
    for fname in os.listdir(MODULES_DIR):
        if fname.startswith(f"{module_code} (") and fname.endswith(".md"):
            return _parse_module_md(os.path.join(MODULES_DIR, fname))
    return []

def _parse_module_md(filepath):
    """从 .md 文件中提取表名和中文名"""
    tables = []
    seen = set()
    try:
        with open(filepath, "r", encoding="utf-8") as f:
            content = f.read()
    except:
        return tables

    # 方式1: 匹配 ### TABLE_NAME (中文名)  —— 三级标题（modules/ 文件的主格式）
    heading_pattern = r'^###\s+([A-Z][A-Z0-9_]*(?:_[A-Z0-9]+)*)\s*(?:\(([^)]+)\))?'
    for m in re.finditer(heading_pattern, content, re.MULTILINE):
        table_name = m.group(1)
        cn_name = m.group(2)
        # 排除非表名的匹配（如模块标题、导航标题）
        if len(table_name) < 3 or table_name in ("ERP", "E10", "MOC", "Phase", "Tags",
            "Tables", "Related", "Table", "Reference", "Next", "Core", "Filtered"):
            continue
        # 排除明显不是表名的（纯数字、全小写单词等）
        if table_name.islower() and "_" not in table_name:
            continue
        key = table_name
        if key not in seen:
            seen.add(key)
            tables.append((table_name, cn_name.strip() if cn_name else None))

    # 方式2: 匹配 **TABLE_NAME (中文名)**  —— 加粗列表（core/ 文件的主格式）
    bold_pattern = r'\*{0,2}\*{2}([A-Z][A-Z0-9_]*(?:_[A-Z0-9]+)*)\s*(?:\(([^)]+)\))?\*{2}'
    for m in re.finditer(bold_pattern, content):
        table_name = m.group(1)
        cn_name = m.group(2)
        if len(table_name) < 3 or table_name in ("ERP", "E10", "MOC", "Phase", "Tags"):
            continue
        # 优先保留 bold 格式中的中文名（可能比 heading 更完整）
        key = table_name
        if key not in seen:
            seen.add(key)
            tables.append((table_name, cn_name.strip() if cn_name else None))
        elif cn_name and cn_name.strip():
            # 已有此表但无中文名，用 bold 格式的中文名补上
            for i, (t, c) in enumerate(tables):
                if t == key and not c:
                    tables[i] = (t, cn_name.strip())
                    break

    return tables

# ═══════════════════════════════════════════════════
# 判定 P0/P1/P2 + Agent可用
# ═══════════════════════════════════════════════════
def classify_table(module_code, table_name):
    """返回 (level, agent_available, business_summary)"""
    level = "P1"
    agent = "✓"

    # P0 模块：主表 P0，排除日志/历史/模板
    if module_code in P0_MODULES:
        level = "P0"
        for pat in P0_EXCLUDE_PATTERNS:
            if re.search(pat, table_name):
                level = "P1"
                agent = "✗"
                break

    # P1 模块：默认 P1，关键表提升为 P0
    elif module_code in P1_MODULES:
        level = "P1"
        if table_name in P1_UPGRADE_TO_P0:
            level = "P0"

    # P2 模块：默认 P2
    elif module_code in P2_MODULES:
        level = "P2"

    # Agent 可用判定（与级别无关，独立判断）
    # 排除日志/历史/审计/临时/配置/备份表
    _agent_exclude = [
        r"_LOG$", r"_HISTORY$", r"_AUDIT$",
        r"_TEMPLATE$", r"_TEMP$", r"_TMP$",
        r"_BAK$", r"_BACKUP$",
        r"_CONFIG$", r"_SETTING$", r"_PARAM$",
        r"_SNAPSHOT$", r"_TRACE$",
    ]
    for pat in _agent_exclude:
        if re.search(pat, table_name):
            agent = "✗"
            break

    # 业务含义摘要
    summary = MODULE_SUMMARY.get(module_code, f"{module_code}模块相关数据表")

    return level, agent, summary

# ═══════════════════════════════════════════════════
# 构建 Excel
# ═══════════════════════════════════════════════════

def build_excel():
    wb = openpyxl.Workbook()
    # 删除默认 sheet
    wb.remove(wb.active)

    # ── Sheet 0: 总览统计 ──
    ws_overview = wb.create_sheet("总览统计", 0)
    _write_overview(ws_overview)

    # ── Sheet 1-8: 各业务域 ──
    all_domain_stats = {}
    for domain_key, domain_info in DOMAINS.items():
        ws = wb.create_sheet(domain_key)
        stats = _write_domain_sheet(ws, domain_key, domain_info)
        all_domain_stats[domain_key] = stats

    # 更新总览统计（填入实际梳理数量）
    _update_overview(ws_overview, all_domain_stats)

    # 保存
    out_path = os.path.join(OUT_DIR, "ERP核心数据表总览.xlsx")
    wb.save(out_path)
    print(f"✅ 已生成: {out_path}")
    total_modules = sum(s["全量模块"] for s in all_domain_stats.values())
    total_tables = sum(s["tables"] for s in all_domain_stats.values())
    total_cols = sum(s["columns"] for s in all_domain_stats.values())
    print(f"   Sheets: {wb.sheetnames}")
    print(f"   域数: {len(all_domain_stats)}, 模块总数: {total_modules}, 表总数: {total_tables}, 列总数: {total_cols}")

def _write_overview(ws):
    """写入总览统计 sheet"""
    ws.merge_cells("A1:G1")
    ws["A1"] = "ERP 核心数据表总览目录"
    ws["A1"].font = Font(name="微软雅黑", size=14, bold=True, color="1F4E79")
    ws["A1"].alignment = CENTER

    ws.merge_cells("A3:G3")
    ws["A3"] = f"数据库: E10_3.0.0.2_CHS | 服务器: .\\SQLEXPRESS | 梳理范围: 8个业务域"
    ws["A3"].font = Font(name="微软雅黑", size=9, color="666666")

    headers = ["序号", "业务域", "模块数", "总表数", "总列数", "已梳理表数", "说明"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=5, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 列宽
    widths = [6, 16, 8, 8, 8, 10, 45]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    ws.sheet_properties.tabColor = "1F4E79"

def _update_overview(ws, stats):
    """更新总览统计（填充数据行）"""
    row = 6
    for domain_key in DOMAINS:
        s = stats[domain_key]
        p0_count = s.get("p0", 0)
        p1_count = s.get("p1", 0)
        note = f"P0模块: {s['p0_modules']}个, P0表: {p0_count}张, P1表: {p1_count}张"

        values = [int(domain_key[:2]), domain_key, s["全量模块"], s["tables"], s["columns"], s["scanned"], note]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = CENTER if c <= 6 else LEFT
            cell.border = THIN_BORDER
        row += 1

    # 合计行
    total_modules = sum(s["全量模块"] for s in stats.values())
    total_tables = sum(s["tables"] for s in stats.values())
    total_cols = sum(s["columns"] for s in stats.values())
    total_scanned = sum(s["scanned"] for s in stats.values())
    values = [None, "合计", total_modules, total_tables, total_cols, total_scanned, "8个业务域合计（不含系统管理、单据管理、其他模块）"]
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = Font(name="微软雅黑", size=9, bold=True)
        cell.alignment = CENTER if c <= 6 else LEFT
        cell.border = THIN_BORDER

def _write_domain_sheet(ws, domain_key, domain_info):
    """写入一个业务域 sheet，返回统计信息"""
    desc = domain_info["desc"]
    modules = domain_info["全量模块"]

    # 标题行
    ws.merge_cells("A1:I1")
    ws["A1"] = f"{domain_key} — {desc}  |  模块: {len(modules)}个  |  表: {sum(m[1] for m in modules.values())}张"
    ws["A1"].font = TITLE_FONT
    ws["A1"].alignment = LEFT

    # 表头
    headers = ["模块编码", "模块中文名", "表名", "表中文名", "重要级别", "字段数(估)", "业务含义摘要", "Agent可用", "备注"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 列宽
    widths = [12, 14, 30, 26, 9, 10, 48, 9, 20]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # 冻结表头
    ws.freeze_panes = "A3"

    # 数据行
    row = 3
    stats = {"全量模块": len(modules), "tables": 0, "columns": 0, "scanned": 0, "p0": 0, "p1": 0, "p0_modules": 0}

    for module_code, (cn_name, table_count, col_count) in modules.items():
        stats["tables"] += table_count
        stats["columns"] += col_count

        # 统计P0模块数
        if module_code in P0_MODULES:
            stats["p0_modules"] += 1

        # 读取该模块的表名
        table_list = read_module_tables(module_code)

        if not table_list:
            # 无法读取详情，仅输出模块级信息
            level = "P0" if module_code in P0_MODULES else ("P2" if module_code in P2_MODULES else "P1")
            summary = MODULE_SUMMARY.get(module_code, f"{cn_name}模块相关数据表")
            values = [module_code, cn_name, f"[{table_count}张表，待细化]", None, level, None, summary, "—", "模块详情待补充"]
            _write_row(ws, row, values, level)
            row += 1
            stats["scanned"] += 1
            if level == "P0": stats["p0"] += 1
            elif level == "P1": stats["p1"] += 1
            continue

        # 逐表输出
        for table_name, table_cn in table_list:
            level, agent, summary = classify_table(module_code, table_name)
            # 中文名缺失时用英文名+标记
            display_cn = table_cn if table_cn else f"{table_name}(待译)"
            values = [module_code, cn_name, table_name, display_cn, level, None, summary, agent, None]
            _write_row(ws, row, values, level)
            row += 1
            stats["scanned"] += 1
            if level == "P0": stats["p0"] += 1
            elif level == "P1": stats["p1"] += 1

    # Sheet tab 颜色
    colors = ["2E75B6", "ED7D31", "70AD47", "FFC000", "9B59B6", "E74C3C", "1ABC9C", "E67E22"]
    domain_idx = int(domain_key[:2]) - 1
    if domain_idx < len(colors):
        ws.sheet_properties.tabColor = colors[domain_idx]

    return stats

def _write_row(ws, row, values, level):
    """写入一行数据，根据 level 着色"""
    for c, v in enumerate(values, 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.font = BODY_FONT
        cell.alignment = CENTER if c in (1, 5, 8) else LEFT
        cell.border = THIN_BORDER

        # P0 行浅橙背景
        if level == "P0":
            cell.fill = P0_FILL
        # P2 行浅灰背景
        elif level == "P2":
            cell.fill = P2_FILL

if __name__ == "__main__":
    build_excel()
