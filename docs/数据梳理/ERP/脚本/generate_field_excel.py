#!/usr/bin/env python3
"""Step 3: Generate field-level Excel files per domain for P0 tables.

Reads the overview Excel (产物2) to get P0 tables per domain,
then reads core/ and modules/ markdown files to extract field definitions,
and generates one Excel per domain with:
  - Overview sheet: all tables in domain with P0/P1/P2
  - Sub-sheets: one per P0 table with field details
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
import os
from pathlib import Path

BASE = Path(__file__).parent.parent  # ERP/
CORE_DIR = BASE / "核心模块"
MODULES_DIR = BASE / "全量模块"
OUTPUT_DIR = BASE / "梳理后"
OVERVIEW_XLSX = OUTPUT_DIR / "ERP核心数据表总览.xlsx"

# ── Styling ──────────────────────────────────────────────
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9)
TITLE_FONT = Font(name="微软雅黑", size=11, bold=True)
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
P0_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")  # light yellow
P1_FILL = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # light green

SENSITIVE_KEYWORDS = ["金额", "单价", "价格", "税额", "税率", "工资", "薪水", "手机", "电话",
                       "AMOUNT", "PRICE", "TAX", "SALARY", "PHONE", "MOBILE"]

OUTPUT_FILES = {
    "01财务核算": "01财务核算_字段梳理.xlsx",
    "02销售分销": "02销售分销_字段梳理.xlsx",
    "03库存管理": "03库存管理_字段梳理.xlsx",
    "04生产制造": "04生产制造_字段梳理.xlsx",
    "05采购供应": "05采购供应_字段梳理.xlsx",
    "06客户关系": "06客户关系_字段梳理.xlsx",
    "07资产设备": "07资产设备_字段梳理.xlsx",
    "08质量管理": "08质量管理_字段梳理.xlsx",
}

# ── Parsing ──────────────────────────────────────────────

def find_module_file(module_code):
    """Find the markdown file for a module, checking core/ then modules/."""
    for d in [CORE_DIR, MODULES_DIR]:
        for f in d.iterdir():
            if f.suffix == ".md" and f.stem.startswith(module_code):
                return f
    return None

def parse_md_tables(filepath):
    """Parse a module .md file and return {table_name: [list of field dicts]}."""
    if not filepath or not filepath.exists():
        return {}
    text = filepath.read_text(encoding="utf-8")
    tables = {}
    current_table = None
    in_table = False
    header_cols = []

    for line in text.split("\n"):
        line = line.strip()
        # Detect table section: "### TABLE_NAME (中文名)" or "### TABLE_NAME"
        if line.startswith("### ") and not line.startswith("#### "):
            if in_table and current_table and header_cols:
                pass  # table ended naturally
            in_table = False
            current_table = None
            header_cols = []
            name = line[4:].strip()
            # Extract table name before any parenthetical
            m = re.match(r"(\S+)", name)
            if m:
                current_table = m.group(1)
                tables[current_table] = []

        # Detect table header row: "| # | Column | Description | Type | Required | PK |"
        elif line.startswith("| # |") or line.startswith("|# |"):
            parts = [p.strip() for p in line.split("|")]
            parts = [p for p in parts if p]
            if len(parts) >= 5:
                header_cols = parts
                in_table = True
                continue

        # Data rows
        elif in_table and current_table and line.startswith("|") and header_cols:
            parts = [p.strip() for p in line.split("|")]
            parts = [p for p in parts if p]
            if len(parts) >= 4 and parts[0].isdigit():
                field = {}
                for i, h in enumerate(header_cols):
                    if i < len(parts):
                        field[h.lower()] = parts[i]
                if "column" in field:
                    tables[current_table].append(field)
                elif "description" in field and len(parts) >= 2:
                    # Some files have slightly different header
                    field["column"] = parts[1] if len(parts) > 1 else ""
                    field["description"] = parts[2] if len(parts) > 2 else ""
                    field["type"] = parts[3] if len(parts) > 3 else ""
                    field["required"] = parts[4] if len(parts) > 4 else ""
                    field["pk"] = parts[5] if len(parts) > 5 else ""
                    tables[current_table].append(field)

        # End of table section
        elif (line.startswith("---") or line.startswith("## ")) and in_table:
            in_table = False

    return tables

def is_sensitive(desc, col_name):
    """Check if a field looks sensitive based on description or name."""
    text = f"{desc} {col_name}".upper()
    for kw in SENSITIVE_KEYWORDS:
        if kw.upper() in text:
            return "✓"
    return ""

def is_pk_marker(val):
    """Check if PK column has 'Y'."""
    return "✓" if val and val.strip().upper() == "Y" else ""

def is_required_marker(val):
    return "✓" if val and val.strip().upper() == "Y" else ""

def normalize_type(t):
    """Normalize data type string."""
    if not t:
        return ""
    t = t.strip()
    # Map common types
    type_map = {
        "string(40)": "varchar(40)", "string(120)": "varchar(120)",
        "string(510)": "varchar(510)", "string(400)": "varchar(400)",
        "string(72)": "varchar(72)", "string(144)": "varchar(144)",
        "string(80)": "varchar(80)", "string(240)": "varchar(240)",
        "string(8000)": "varchar(8000)", "string(4)": "varchar(4)",
        "string(2)": "varchar(2)", "string": "varchar",
        "number(16,6)": "decimal(16,6)", "number(23,8)": "decimal(23,8)",
        "number(18,9)": "decimal(18,9)", "number(5,4)": "decimal(5,4)",
        "number(0/1)": "bit", "number": "numeric",
        "binary": "varbinary",
    }
    return type_map.get(t, t)

def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = THIN_BORDER

def style_body_cell(ws, row, col, wrap=False):
    cell = ws.cell(row=row, column=col)
    cell.font = BODY_FONT
    cell.border = THIN_BORDER
    if wrap:
        cell.alignment = Alignment(wrap_text=True, vertical="top")
    return cell

# ── Main ─────────────────────────────────────────────────

def generate_domain_excel(domain_name, sheet_key, output_name):
    """Generate one domain Excel."""
    print(f"\n{'='*60}")
    print(f"处理: {domain_name} → {output_name}")

    # Read overview to get P0 tables
    wb = openpyxl.load_workbook(OVERVIEW_XLSX)
    if sheet_key not in wb.sheetnames:
        print(f"  ⚠ Sheet '{sheet_key}' not found, skipping")
        wb.close()
        return

    ws_overview = wb[sheet_key]

    # Collect all tables and P0 tables
    all_rows = []
    p0_tables = []  # (module_code, module_cn, table_name, table_cn)
    for row in ws_overview.iter_rows(min_row=3, values_only=True):
        if row[0] and row[0] != "模块编码":
            mod_code, mod_cn, tbl_name, tbl_cn, level, _, desc, agent, remark = row[:9]
            all_rows.append((mod_code, mod_cn, tbl_name, tbl_cn, level, desc, agent, remark))
            if level == "P0" and tbl_name and tbl_name != "[0张表，待细化]":
                p0_tables.append((mod_code, mod_cn, tbl_name, tbl_cn or tbl_name))
    wb.close()

    print(f"  总表: {len(all_rows)}, P0表: {len(p0_tables)}")

    # Create output workbook
    out = openpyxl.Workbook()
    out.remove(out.active)

    # ── Overview sheet ──
    ws_ov = out.create_sheet("总览")
    # Title row
    ws_ov.merge_cells("A1:H1")
    ws_ov.cell(row=1, column=1, value=f"{domain_name} — 字段梳理总览  |  表: {len(all_rows)}张  |  P0表: {len(p0_tables)}张")
    ws_ov.cell(row=1, column=1).font = TITLE_FONT

    # Header
    ov_headers = ["模块编码", "模块中文名", "表名", "表中文名", "重要级别", "业务含义摘要", "Agent可用", "备注"]
    for i, h in enumerate(ov_headers, 1):
        ws_ov.cell(row=2, column=i, value=h)
    style_header_row(ws_ov, 2, len(ov_headers))

    # Data
    for r, row_data in enumerate(all_rows, 3):
        mod_code, mod_cn, tbl_name, tbl_cn, level, desc, agent, remark = row_data
        vals = [mod_code, mod_cn, tbl_name, tbl_cn, level, desc, agent, remark]
        for c, v in enumerate(vals, 1):
            style_body_cell(ws_ov, r, c, wrap=(c == 6))
            ws_ov.cell(row=r, column=c, value=v)
        # Row fill
        fill = P0_FILL if level == "P0" else (P1_FILL if level == "P1" else None)
        if fill:
            for c in range(1, len(vals) + 1):
                ws_ov.cell(row=r, column=c).fill = fill

    # Column widths
    ws_ov.column_dimensions["A"].width = 14
    ws_ov.column_dimensions["B"].width = 16
    ws_ov.column_dimensions["C"].width = 28
    ws_ov.column_dimensions["D"].width = 24
    ws_ov.column_dimensions["E"].width = 10
    ws_ov.column_dimensions["F"].width = 36
    ws_ov.column_dimensions["G"].width = 10
    ws_ov.column_dimensions["H"].width = 20

    # ── P0 table sub-sheets ──
    field_headers = ["字段名", "字段中文名", "数据类型", "主键", "外键", "业务含义", "枚举/取值范围", "敏感标记", "可空", "质量备注"]

    # Cache module files
    module_cache = {}

    for idx, (mod_code, mod_cn, tbl_name, tbl_cn) in enumerate(p0_tables):
        # Build sheet name: TABLE_NAME (中文名), max 31 chars
        cn_part = (tbl_cn or "").strip()
        # 如果是 "(待译)" 占位，只用英文名做标签
        if cn_part and "(待译)" not in cn_part:
            full = f"{tbl_name}({cn_part})"
            if len(full) <= 31:
                sheet_name = full
            else:
                max_cn = 31 - len(tbl_name) - 3
                if max_cn >= 2:
                    sheet_name = f"{tbl_name}({cn_part[:max_cn]})"
                else:
                    sheet_name = tbl_name[:31]
        else:
            sheet_name = tbl_name[:31]
        sheet_name = re.sub(r'[\\/*?\[\]:]', '_', sheet_name)

        # De-duplicate if sheet already exists
        if sheet_name in out.sheetnames:
            fallback = f"{mod_code}_{tbl_name}"
            sheet_name = fallback[:31]
            sheet_name = re.sub(r'[\\/*?\[\]:]', '_', sheet_name)
        if sheet_name in out.sheetnames:
            sheet_name = f"{mod_code}_{tbl_name}_{idx}"[:31]
            sheet_name = re.sub(r'[\\/*?\[\]:]', '_', sheet_name)

        ws = out.create_sheet(sheet_name)

        # Title
        ws.merge_cells("A1:J1")
        title_cn = tbl_cn if tbl_cn and "(待译)" not in str(tbl_cn) else ""
        title_text = f"{tbl_name} ({title_cn})  |  模块: {mod_code} {mod_cn}  |  P0核心表" if title_cn else f"{tbl_name}  |  模块: {mod_code} {mod_cn}  |  P0核心表"
        ws.cell(row=1, column=1, value=title_text)
        ws.cell(row=1, column=1).font = TITLE_FONT

        # Header
        for i, h in enumerate(field_headers, 1):
            ws.cell(row=2, column=i, value=h)
        style_header_row(ws, 2, len(field_headers))

        # Load module file
        if mod_code not in module_cache:
            fp = find_module_file(mod_code)
            if fp:
                module_cache[mod_code] = parse_md_tables(fp)
            else:
                module_cache[mod_code] = {}

        fields = module_cache[mod_code].get(tbl_name, [])

        if not fields:
            # Table not found in module file
            style_body_cell(ws, 3, 1)
            ws.cell(row=3, column=1, value="(字段详情待补充 — 未在模块文件中找到)")
            ws.merge_cells("A3:J3")
        else:
            for r_offset, f in enumerate(fields):
                r = 3 + r_offset
                col_name = f.get("column", "")
                desc = f.get("description", "")
                dtype = normalize_type(f.get("type", ""))
                pk = is_pk_marker(f.get("pk", ""))
                required = is_required_marker(f.get("required", ""))
                sensitive = is_sensitive(desc, col_name)

                vals = [col_name, desc, dtype, pk, "", desc, "", sensitive, required, ""]
                for c, v in enumerate(vals, 1):
                    style_body_cell(ws, r, c, wrap=(c in [2, 6, 7, 10]))
                    ws.cell(row=r, column=c, value=v)

                # Highlight PK rows
                if pk:
                    for c in range(1, len(vals) + 1):
                        ws.cell(row=r, column=c).font = Font(name="微软雅黑", size=9, bold=True)

        # Column widths
        widths = [22, 18, 14, 6, 8, 28, 16, 8, 6, 16]
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # Save
    outpath = OUTPUT_DIR / output_name
    out.save(str(outpath))
    print(f"  ✅ 已生成: {outpath}")
    print(f"     Sheets: 总览 + {len(p0_tables)} P0表子sheets, 共{len(out.sheetnames)} sheets")


def main():
    print("=" * 60)
    print("Step 3: 生成字段梳理 Excel（8个业务域）")
    print("=" * 60)

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    for domain_name, filename in OUTPUT_FILES.items():
        generate_domain_excel(domain_name, domain_name, filename)

    print(f"\n{'='*60}")
    print("全部完成！8 份字段梳理 Excel 已生成")
    print(f"输出目录: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
