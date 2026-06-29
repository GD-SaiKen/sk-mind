#!/usr/bin/env python3
"""
Step 5: 生成 ERP 业务关系图谱边 Excel（产物5）
BusinessGraphEdge — 业务对象间的关系边
"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from pathlib import Path

OUTPUT_DIR = Path(__file__).parent.parent / "梳理后"
OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

# ── Styling ──
HEADER_FILL = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
HEADER_FONT = Font(name="微软雅黑", size=10, bold=True, color="FFFFFF")
BODY_FONT = Font(name="微软雅黑", size=9)
TITLE_FONT = Font(name="微软雅黑", size=11, bold=True)
HIGH_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
MED_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
LOW_FILL = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid")
THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)

# ═══════════════════════════════════════════════════
# 业务关系图谱边定义
# ═══════════════════════════════════════════════════

EDGES = [
    # ── 采购供应域 ──
    {
        "from": "Supplier",
        "relation": "supplies",
        "to": "Material",
        "join": "ITEM_SUPPLIER.ITEM_ID → ITEM.ITEM_ID / ITEM_SUPPLIER.SUPPLIER_ID → SUPPLIER.SUPPLIER_ID",
        "source": "FK: ITEM_SUPPLIER",
        "confidence": "high",
    },
    {
        "from": "PurchaseOrder",
        "relation": "references",
        "to": "Supplier",
        "join": "PO.SUPPLIER_ID → SUPPLIER.SUPPLIER_ID",
        "source": "FK: PURCHASE_ORDER.SUPPLIER_ID",
        "confidence": "high",
    },
    {
        "from": "PurchaseOrder",
        "relation": "contains",
        "to": "Material",
        "join": "PURCHASE_ORDER_D.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: PURCHASE_ORDER_D.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "PurchaseOrder",
        "relation": "fulfills",
        "to": "RequisitionOrder",
        "join": "REQUISITION_D.REQUISITION_ID → REQUISITION.REQUISITION_ID / PURCHASE_ORDER_D sourced from REQUISITION",
        "source": "FK: REQUISITION → PO linkage",
        "confidence": "medium",
    },
    {
        "from": "RequisitionOrder",
        "relation": "requests",
        "to": "Material",
        "join": "REQUISITION_D.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: REQUISITION_D.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "PurchaseOrder",
        "relation": "generates",
        "to": "AP",
        "join": "AP invoice sourced from PURCHASE_RECEIPT → PURCHASE_ORDER",
        "source": "商务流程: 采购收货→应付立账",
        "confidence": "medium",
    },
    {
        "from": "PurchaseReceipt",
        "relation": "received_from",
        "to": "PurchaseOrder",
        "join": "PURCHASE_RECEIPT.SOURCE_ID → PURCHASE_ORDER.PURCHASE_ORDER_ID",
        "source": "FK: PURCHASE_RECEIPT.SOURCE_ID",
        "confidence": "high",
    },
    {
        "from": "PurchaseReceipt",
        "relation": "updates",
        "to": "Inventory",
        "join": "PURCHASE_RECEIPT → ITEM_WAREHOUSE (收货入库增加库存)",
        "source": "商务流程: 采购收货→库存增加",
        "confidence": "medium",
    },

    # ── 生产制造域 ──
    {
        "from": "BOM",
        "relation": "composes",
        "to": "Material",
        "join": "BOM.ITEM_ID (parent) → ITEM.ITEM_ID / BOM_D.ITEM_ID (child) → ITEM.ITEM_ID",
        "source": "FK: BOM(BOM_D).ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "ProductionOrder",
        "relation": "produces",
        "to": "Material",
        "join": "MO.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: MO.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "ProductionOrder",
        "relation": "uses",
        "to": "BOM",
        "join": "MO_MATERIAL.BOM_ID → BOM.BOM_ID",
        "source": "FK: MO_MATERIAL references BOM",
        "confidence": "medium",
    },
    {
        "from": "ProductionOrder",
        "relation": "executed_at",
        "to": "WorkCenter",
        "join": "MO_ROUTING.WORK_CENTER_ID → WORK.WORK_CENTER_ID",
        "source": "FK: MO_ROUTING.WORK_CENTER_ID",
        "confidence": "high",
    },
    {
        "from": "ProductionOrder",
        "relation": "consumes",
        "to": "Material",
        "join": "MO_MATERIAL.ITEM_ID → ITEM.ITEM_ID (投料)",
        "source": "FK: MO_MATERIAL.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "ProductionOrder",
        "relation": "outputs_to",
        "to": "Inventory",
        "join": "工单完工入库 → ITEM_WAREHOUSE (增加成品库存)",
        "source": "商务流程: 完工入库",
        "confidence": "medium",
    },
    {
        "from": "Material",
        "relation": "has_routing",
        "to": "WorkCenter",
        "join": "ITEM_ROUTING_WORK_CENTER.WORK_CENTER_ID → WORK.WORK_CENTER_ID",
        "source": "FK: ITEM_ROUTING_WORK_CENTER",
        "confidence": "high",
    },

    # ── 销售分销域 ──
    {
        "from": "Customer",
        "relation": "places",
        "to": "SalesOrder",
        "join": "SALE.CUSTOMER_ID → CUSTOMER.CUSTOMER_ID",
        "source": "FK: SALE.CUSTOMER_ID",
        "confidence": "high",
    },
    {
        "from": "SalesOrder",
        "relation": "references",
        "to": "Material",
        "join": "SALE_D.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: SALE_D.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "SalesOrder",
        "relation": "ships_via",
        "to": "Shipment",
        "join": "SHIPPING.SOURCE_ID → SALE.SALE_ID",
        "source": "FK: SHIPPING.SOURCE_ID",
        "confidence": "high",
    },
    {
        "from": "Shipment",
        "relation": "depletes",
        "to": "Inventory",
        "join": "SHIPPING → ITEM_WAREHOUSE (发货出库减少库存)",
        "source": "商务流程: 销售出库",
        "confidence": "medium",
    },
    {
        "from": "SalesOrder",
        "relation": "generates",
        "to": "AR",
        "join": "AR invoice sourced from SALE → 应收立账",
        "source": "商务流程: 销售出库→应收立账",
        "confidence": "medium",
    },
    {
        "from": "SalesOrder",
        "relation": "may_return",
        "to": "ReturnOrder",
        "join": "RETURN.SOURCE_ID → SALE.SALE_ID",
        "source": "FK: RETURN.SOURCE_ID",
        "confidence": "high",
    },

    # ── 库存管理域 ──
    {
        "from": "Inventory",
        "relation": "tracks",
        "to": "Material",
        "join": "ITEM_WAREHOUSE.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: ITEM_WAREHOUSE.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "Inventory",
        "relation": "valued_by",
        "to": "Cost",
        "join": "INV_STOCK → COST (库存成本核算)",
        "source": "商务流程: 库存成本核算",
        "confidence": "medium",
    },
    {
        "from": "TransferOrder",
        "relation": "moves",
        "to": "Inventory",
        "join": "TRANSFER → ITEM_WAREHOUSE (库存调拨)",
        "source": "商务流程: 库存调拨",
        "confidence": "high",
    },

    # ── 财务核算域 ──
    {
        "from": "AR",
        "relation": "settled_by",
        "to": "Payment",
        "join": "PAYMENT → AR (收款核销应收账款)",
        "source": "商务流程: 收款核销",
        "confidence": "medium",
    },
    {
        "from": "AP",
        "relation": "settled_by",
        "to": "Payment",
        "join": "PAYMENT → AP (付款核销应付账款)",
        "source": "商务流程: 付款核销",
        "confidence": "medium",
    },
    {
        "from": "AR",
        "relation": "posts_to",
        "to": "GL",
        "join": "AR → VOUCHER → GL (应收凭证过账总账)",
        "source": "商务流程: 凭证过账",
        "confidence": "medium",
    },
    {
        "from": "AP",
        "relation": "posts_to",
        "to": "GL",
        "join": "AP → VOUCHER → GL (应付凭证过账总账)",
        "source": "商务流程: 凭证过账",
        "confidence": "medium",
    },
    {
        "from": "Cost",
        "relation": "allocates_to",
        "to": "Material",
        "join": "ITEM_COST.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: ITEM_COST.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "Cost",
        "relation": "posts_to",
        "to": "GL",
        "join": "COST → VOUCHER → GL (成本凭证过账)",
        "source": "商务流程: 成本核算→凭证过账",
        "confidence": "medium",
    },
    {
        "from": "Payment",
        "relation": "via",
        "to": "BankAccount",
        "join": "PAYMENT.BANK_ID → BANK_ACCOUNT",
        "source": "FK: PAYMENT.BANK_ID",
        "confidence": "high",
    },
    {
        "from": "GL",
        "relation": "uses",
        "to": "Account",
        "join": "VOUCHER_D.ACCOUNT_ID → ACCOUNT.ACCOUNT_ID",
        "source": "FK: VOUCHER_D.ACCOUNT_ID",
        "confidence": "high",
    },

    # ── 质量管理域 ──
    {
        "from": "InspectionOrder",
        "relation": "inspects",
        "to": "PurchaseReceipt",
        "join": "PO_ARRIVAL_INSPECTION.SOURCE_ID → PURCHASE_RECEIPT",
        "source": "FK: PO_ARRIVAL_INSPECTION → 采购收货",
        "confidence": "high",
    },
    {
        "from": "InspectionOrder",
        "relation": "inspects",
        "to": "Material",
        "join": "PO_ARRIVAL_INSPECTION.ITEM_ID → ITEM.ITEM_ID",
        "source": "FK: PO_ARRIVAL_INSPECTION.ITEM_ID",
        "confidence": "high",
    },
    {
        "from": "InspectionOrder",
        "relation": "may_create",
        "to": "DefectiveRecord",
        "join": "检验不合格 → DEFECTIVE_RECORD",
        "source": "商务流程: 来料检验→不良品记录",
        "confidence": "medium",
    },
    {
        "from": "InspectionOrder",
        "relation": "judged_by",
        "to": "Acceptance",
        "join": "ACCEPTANCE sourced from inspection result",
        "source": "商务流程: 检验判定→验收",
        "confidence": "medium",
    },

    # ── 资产设备域 ──
    {
        "from": "Asset",
        "relation": "depreciates_to",
        "to": "Cost",
        "join": "ASSET_DEPRECIATION → COST (折旧计入成本)",
        "source": "商务流程: 折旧分摊",
        "confidence": "medium",
    },
    {
        "from": "Equipment",
        "relation": "maintained_by",
        "to": "Maintenance",
        "join": "EQT_MAINTENANCE.EQUIPMENT_ID → EQT.EQT_ID",
        "source": "FK: EQT_MAINTENANCE.EQUIPMENT_ID",
        "confidence": "high",
    },
    {
        "from": "Equipment",
        "relation": "located_at",
        "to": "WorkCenter",
        "join": "WORK_CENTER → EQT (设备归属工作中心)",
        "source": "商务流程: 设备归属",
        "confidence": "medium",
    },

    # ── 客户关系域 ──
    {
        "from": "Customer",
        "relation": "has",
        "to": "Member",
        "join": "MEMBER.CUSTOMER_ID → CUSTOMER.CUSTOMER_ID",
        "source": "FK: MEMBER.CUSTOMER_ID",
        "confidence": "high",
    },
    {
        "from": "Member",
        "relation": "owns",
        "to": "Card",
        "join": "MEMBER_CARD.MEMBER_ID → MEMBER.MEMBER_ID",
        "source": "FK: MEMBER_CARD",
        "confidence": "high",
    },
    {
        "from": "Customer",
        "relation": "requests",
        "to": "ServiceOrder",
        "join": "SERVICE.CUSTOMER_ID → CUSTOMER.CUSTOMER_ID",
        "source": "FK: SERVICE.CUSTOMER_ID",
        "confidence": "high",
    },

    # ── 跨域关键关系 ──
    {
        "from": "SalesOrder",
        "relation": "triggers",
        "to": "ProductionOrder",
        "join": "MO.SOURCE_ID → SALE.SALE_ID (销售订单驱动生产)",
        "source": "商务流程: 销产联动 (MTO模式)",
        "confidence": "medium",
    },
    {
        "from": "SalesOrder",
        "relation": "triggers",
        "to": "PurchaseOrder",
        "join": "销售订单→MRP→采购建议→采购订单",
        "source": "商务流程: MRP运算",
        "confidence": "low",
    },
    {
        "from": "ProductionOrder",
        "relation": "triggers",
        "to": "PurchaseOrder",
        "join": "工单缺料→MRP→采购建议→采购订单",
        "source": "商务流程: MRP运算",
        "confidence": "low",
    },
    {
        "from": "PurchaseOrder",
        "relation": "inspected_by",
        "to": "InspectionOrder",
        "join": "PO_ARRIVAL_INSPECTION → PURCHASE_ORDER (到货检验)",
        "source": "商务流程: 采购到货检验",
        "confidence": "high",
    },
]

# ═══════════════════════════════════════════════════
# 构建 Excel
# ═══════════════════════════════════════════════════

def build_excel():
    wb = openpyxl.Workbook()

    # ── Sheet 1: 关系边列表 ──
    ws = wb.active
    ws.title = "关系边列表"

    # Title
    ws.merge_cells("A1:H1")
    ws["A1"] = "ERP 业务关系图谱边 (BusinessGraphEdge)  |  数据库: E10_3.0.0.2_CHS  |  关系边: {}条".format(len(EDGES))
    ws["A1"].font = TITLE_FONT

    # Header
    headers = ["序号", "起点对象\n(from_entity)", "关系名\n(relation_name)", "终点对象\n(to_entity)", "关联条件\n(join_condition)", "来源\n(source)", "可信度\n(confidence)", "业务域"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    ws.freeze_panes = "A3"

    # Domain assignment
    domain_map = {
        "Supplier": "采购供应", "PurchaseOrder": "采购供应", "RequisitionOrder": "采购供应",
        "PurchaseReceipt": "采购供应", "AP": "财务核算", "Payment": "财务核算",
        "GL": "财务核算", "Account": "财务核算", "Cost": "财务核算", "BankAccount": "财务核算",
        "Material": "库存管理", "BOM": "生产制造", "ProductionOrder": "生产制造",
        "WorkCenter": "生产制造", "Inventory": "库存管理", "TransferOrder": "库存管理",
        "Customer": "客户关系", "SalesOrder": "销售分销", "Shipment": "销售分销",
        "ReturnOrder": "销售分销", "AR": "财务核算",
        "InspectionOrder": "质量管理", "DefectiveRecord": "质量管理",
        "Acceptance": "质量管理",
        "Asset": "资产设备", "Equipment": "资产设备", "Maintenance": "资产设备",
        "Member": "客户关系", "Card": "客户关系", "ServiceOrder": "客户关系",
    }

    # Data rows
    for i, edge in enumerate(EDGES):
        row = 3 + i
        domain = domain_map.get(edge["from"], "—")

        values = [
            i + 1,
            edge["from"],
            edge["relation"],
            edge["to"],
            edge["join"],
            edge["source"],
            edge["confidence"],
            domain,
        ]
        for c, v in enumerate(values, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = CENTER if c in (1, 7, 8) else LEFT
            cell.border = THIN_BORDER

        # Confidence coloring
        conf = edge["confidence"]
        if conf == "high":
            fill = HIGH_FILL
        elif conf == "medium":
            fill = MED_FILL
        else:
            fill = LOW_FILL
        ws.cell(row=row, column=7).fill = fill

    # Column widths
    widths = [6, 18, 16, 18, 52, 32, 10, 12]
    for c, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(c)].width = w

    # ── Sheet 2: 按业务域汇总 ──
    ws2 = wb.create_sheet("按域汇总")

    ws2.merge_cells("A1:D1")
    ws2["A1"] = "业务关系图谱边 — 按业务域汇总"
    ws2["A1"].font = TITLE_FONT

    domains = ["采购供应", "生产制造", "销售分销", "库存管理", "财务核算", "质量管理", "资产设备", "客户关系", "跨域"]
    h2 = ["业务域", "关系边数", "高可信度", "中可信度", "低可信度", "典型关系"]
    for c, h in enumerate(h2, 1):
        cell = ws2.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, domain in enumerate(domains):
        row = 3 + i
        if domain == "跨域":
            domain_edges = [e for e in EDGES if e["from"] in ("SalesOrder", "ProductionOrder", "PurchaseOrder")
                           and e["confidence"] in ("medium", "low")
                           and e["to"] in ("ProductionOrder", "PurchaseOrder", "InspectionOrder")]
            # Actually just count cross-domain edges
            domain_edges = [
                e for e in EDGES
                if domain_map.get(e["from"]) != domain_map.get(e["to"])
                or e["confidence"] == "low"
            ]
            # Simpler: use the last 4 edges which are explicitly cross-domain
            domain_edges = EDGES[-5:-1]  # SalesOrder→ProductionOrder, SalesOrder→PurchaseOrder, ProductionOrder→PurchaseOrder, PurchaseOrder→InspectionOrder
        else:
            domain_edges = [e for e in EDGES if domain_map.get(e["from"]) == domain
                           and not (e["from"] in ("SalesOrder", "ProductionOrder") and e["to"] in ("ProductionOrder", "PurchaseOrder"))]
            # Exclude cross-domain from domain counts (just keep same-domain)
            domain_edges = [e for e in domain_edges
                           if domain_map.get(e.get("to", "")) == domain_map.get(e["from"])]

        high = sum(1 for e in domain_edges if e["confidence"] == "high")
        med = sum(1 for e in domain_edges if e["confidence"] == "medium")
        low = sum(1 for e in domain_edges if e["confidence"] == "low")
        examples = ", ".join([f'{e["from"]}→{e["to"]}' for e in domain_edges[:3]])

        # Fix for cross-domain
        if domain == "跨域":
            cross_edges = EDGES[-4:]  # last 4 edges
            high = sum(1 for e in cross_edges if e["confidence"] == "high")
            med = sum(1 for e in cross_edges if e["confidence"] == "medium")
            low = sum(1 for e in cross_edges if e["confidence"] == "low")
            examples = "SalesOrder→ProductionOrder, SalesOrder→PurchaseOrder, ProductionOrder→PurchaseOrder..."

        values = [domain, len(domain_edges) if domain != "跨域" else 4, high, med, low, examples]
        for c, v in enumerate(values, 1):
            cell = ws2.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = CENTER if c <= 5 else LEFT
            cell.border = THIN_BORDER

    widths2 = [14, 10, 10, 10, 10, 60]
    for c, w in enumerate(widths2, 1):
        ws2.column_dimensions[get_column_letter(c)].width = w

    # ── Sheet 3: 制造价值链路径 ──
    ws3 = wb.create_sheet("制造价值链路径")

    ws3.merge_cells("A1:E1")
    ws3["A1"] = "制造价值链 — 关键数据流路径"
    ws3["A1"].font = TITLE_FONT

    paths = [
        ("供应→库存", "Supplier → PurchaseOrder → PurchaseReceipt → Inventory",
         "供应商供货→采购下单→到货入库→库存增加", "采购供应→库存管理"),
        ("需求→供应", "Customer → SalesOrder → (MRP) → PurchaseOrder → Supplier",
         "客户下单→销售订单→MRP运算→采购建议→供应商", "客户关系→销售分销→采购供应"),
        ("需求→生产", "Customer → SalesOrder → (MTO) → ProductionOrder",
         "客户下单→销售订单→按单生产→生产工单", "客户关系→销售分销→生产制造"),
        ("库存→生产", "Inventory → ProductionOrder → BOM → Material",
         "库存物料→生产工单→BOM展开→子件物料", "库存管理→生产制造"),
        ("生产→库存", "ProductionOrder → Inventory (完工入库)",
         "生产完工→成品入库→库存增加", "生产制造→库存管理"),
        ("供应→质量", "PurchaseOrder → PurchaseReceipt → InspectionOrder → Acceptance",
         "采购到货→收货→来料检验→验收判定", "采购供应→质量管理"),
        ("销售→财务", "SalesOrder → Shipment → AR → Payment → GL",
         "销售出库→发货→应收立账→收款核销→总账", "销售分销→财务核算"),
        ("采购→财务", "PurchaseOrder → PurchaseReceipt → AP → Payment → GL",
         "采购收货→应付立账→付款核销→总账", "采购供应→财务核算"),
        ("生产→财务", "ProductionOrder → Cost → GL",
         "生产工单→成本核算→凭证过账→总账", "生产制造→财务核算"),
        ("资产→财务", "Asset → Depreciation → Cost → GL",
         "固定资产→折旧计提→成本→总账", "资产设备→财务核算"),
    ]

    h3 = ["路径名称", "对象路径", "业务含义", "涉及业务域"]
    for c, h in enumerate(h3, 1):
        cell = ws3.cell(row=2, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    for i, (name, path, meaning, domains) in enumerate(paths):
        row = 3 + i
        for c, v in enumerate([name, path, meaning, domains], 1):
            cell = ws3.cell(row=row, column=c, value=v)
            cell.font = BODY_FONT
            cell.alignment = LEFT
            cell.border = THIN_BORDER

    widths3 = [14, 60, 50, 28]
    for c, w in enumerate(widths3, 1):
        ws3.column_dimensions[get_column_letter(c)].width = w

    # Save
    out_path = OUTPUT_DIR / "ERP业务关系图谱边.xlsx"
    wb.save(str(out_path))
    print(f"✅ 已生成: {out_path}")
    print(f"   关系边: {len(EDGES)}条, Sheets: {wb.sheetnames}")


if __name__ == "__main__":
    build_excel()
