#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""携选工品 - API接口梳理Excel"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Fills ──
DEEP_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
NEST_L1 = PatternFill(start_color='F4F7FC', end_color='F4F7FC', fill_type='solid')
NEST_L2 = PatternFill(start_color='EBEFF7', end_color='EBEFF7', fill_type='solid')

# ── Fonts ──
WHITE_B = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_F = Font(name='微软雅黑', size=10, bold=True)
TITLE_F = Font(name='微软雅黑', size=11, bold=True)
NORMAL_F = Font(name='微软雅黑', size=10)

# ── Border ──
THIN_B = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),
                top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))

# ── Alignments ──
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)

C_W = {'A':28,'B':52,'C':14,'D':20}
OV_C_W = {'A':10,'B':36,'C':60,'D':12,'E':50}

def apply_style(c, font=NORMAL_F, fill=None, align=ALIGN_LEFT, border=THIN_B):
    c.font = font; c.alignment = align; c.border = border
    if fill is not None: c.fill = fill

def write_overview(ws, modules):
    for k,v in OV_C_W.items(): ws.column_dimensions[k].width = v
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '携选工品 - API接口总览'
    c.font = Font(name='微软雅黑',size=14,bold=True); c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        apply_style(ws.cell(row=2,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[2].height = 22
    row = 3
    for mod in modules:
        ws.merge_cells(f'A{row}:E{row}')
        apply_style(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
        for c_ in 'BCDE': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
        ws.row_dimensions[row].height = 22; row += 1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, LIGHT_GRAY,
                           ALIGN_CENTER if i in (1,4) else ALIGN_LEFT)
            row += 1
    ws.auto_filter.ref = f'A2:E{row-1}'

def write_subs(ws, api):
    for k,v in C_W.items(): ws.column_dimensions[k].width = v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        apply_style(ws.cell(row=idx,column=1,value=l), BOLD_F, WHITE_FILL, ALIGN_LEFT)
        ws.merge_cells(f'B{idx}:D{idx}')
        apply_style(ws.cell(row=idx,column=2,value=v), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        for c_ in 'CD': apply_style(ws[f'{c_}{idx}'], NORMAL_F, WHITE_FILL, ALIGN_LEFT)
    row = 5; ws.row_dimensions[4].height = 6
    # 请求参数
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='请求参数'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; row += 1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        apply_style(ws.cell(row=row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22; row += 1
    if api['params']:
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, WHITE_FILL,
                           ALIGN_CENTER if i in (3,4) else ALIGN_LEFT)
            row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}')
        apply_style(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        row += 1
    row += 2; ws.row_dimensions[row-1].height = 6
    # 响应数据字段
    ws.merge_cells(f'A{row}:D{row}')
    apply_style(ws.cell(row=row,column=1,value='响应数据字段'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; hdr_row = row + 1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        apply_style(ws.cell(row=hdr_row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[hdr_row].height = 22; row = hdr_row + 1
    rl = []
    NEST_FILLS = {0: WHITE_FILL, 1: NEST_L1, 2: NEST_L2}
    for item in api['responses']:
        lvl, rname, rdesc, rtype = item[0], item[1], item[2] if len(item)>2 else '', item[3] if len(item)>3 else ''
        indent = '  ' * lvl
        row_fill = NEST_FILLS.get(lvl, WHITE_FILL)
        apply_style(ws.cell(row=row,column=1,value=indent+rname), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=2,value=rdesc), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=3,value=''), NORMAL_F, row_fill, ALIGN_CENTER)
        apply_style(ws.cell(row=row,column=4,value=rtype), NORMAL_F, row_fill, ALIGN_CENTER)
        rl.append((row, lvl)); row += 1
    if rl:
        max_lvl = max(l for _,l in rl)
        for lvl in range(max_lvl, 0, -1):
            i = 0
            while i < len(rl):
                r, l = rl[i]
                if l == lvl:
                    j = i + 1
                    while j < len(rl) and rl[j][1] >= lvl: j += 1
                    if j - 1 > i: ws.row_dimensions.group(r, rl[j-1][0], outline_level=lvl)
                    i = j
                else: i += 1
    return row

# =========================== DATA ===========================

MRO = {
'name': '携选工品', 'interfaces': [
    {'id': '1', 'name': '商品列表查询', 'url': 'https://openapi.xiekeyun.com/mro/sku/list', 'method': 'POST',
     'desc': '查询携选工品商品列表数据',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('pageNum', '页码', '必填', 'Number'),
         ('pageSize', '页面数量', '必填', 'Number'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1:成功；9:失败', 'Number'),
         (0, 'errorCode', '接口错误编码（失败时有值）', 'String'),
         (0, 'errorMsg', '接口错误信息（失败时有值）', 'String'),
         (0, 'data', '商品列表数据', 'Json对象数组'),
         (1, 'skuCode', 'sku编码', 'String'),
         (1, 'skuName', 'sku名字', 'String'),
     ]},
    {'id': '2', 'name': '商品详情查询', 'url': 'https://openapi.xiekeyun.com/mro/sku/getDetail', 'method': 'POST',
     'desc': '查询单个商品详细信息，包含售卖价格、单位、产地、主图、规格参数等',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
         ('skuCode', 'Sku编码', '必填', 'String'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1:成功；9:失败', 'Number'),
         (0, 'errorCode', '接口错误编码（失败时有值）', 'String'),
         (0, 'errorMsg', '接口错误信息（失败时有值）', 'String'),
         (0, 'data', '商品详情数据', 'Json对象'),
         (1, 'brandCode', '品牌编号', 'String'),
         (1, 'brandName', '品牌名称', 'String'),
         (1, 'skuCode', 'sku编码', 'String'),
         (1, 'skuName', 'sku名称', 'String'),
         (1, 'categoryCode', '商品品类编码', 'String'),
         (1, 'categoryName', '商品品类名称', 'String'),
         (1, 'unitCode', '单位编码', 'String'),
         (1, 'unitName', '单位名称', 'String'),
         (1, 'qty', 'sku数量', 'Decimal'),
         (1, 'minOrderQty', '最小起订量', 'Decimal'),
         (1, 'salePrice', '销售价', 'Decimal'),
         (1, 'oriPrice', '原始价', 'Decimal'),
         (1, 'totalAmount', '总金额', 'Decimal'),
         (1, 'status', '上架状态 1:已上架 2:已下架', 'Integer'),
         (1, 'saleStatus', '销售状态 1:销售中 2:已售罄 3:可预订', 'Integer'),
         (1, 'saleTime', '可销售时间', 'String'),
         (1, 'deliveryTime', '承诺交期/几天内交货', 'Integer'),
         (1, 'orderStockType', '订单库存类型 1:订单 2:现货/交货方式', 'Integer'),
         (1, 'taxRate', '税率', 'Decimal'),
         (1, 'orderQtyMultiple', '订单倍数/订购批量', 'Decimal'),
         (1, 'productionPlace', '产地', 'String'),
         (1, 'imageUrl', '主图地址', 'String'),
         (1, 'isPromotion', '是否促销 1:无 2:是', 'Integer'),
         (1, 'invoiceType', '发票类型 0:普票 1:专票', 'Number'),
         (1, 'productType', '商品类型 0:全新 1:二手', 'Number'),
         (1, 'productModelNo', '产品型号', 'String'),
         (1, 'packageSpec', '包装规格', 'String'),
         (1, 'isCustomMade', '是否支持定制', 'Number'),
         (1, 'remark', '商品备注', 'String'),
         (1, 'msg', '购买说明', 'String'),
         (1, 'imageList', '图片列表', 'Json数组'),
         (2, 'imageUrl', '图片地址', 'String'),
         (2, 'isDefault', '是否默认 1:否 2:是', 'Integer'),
         (2, 'sort', '排序', 'Integer'),
         (1, 'attributeList', '属性列表', 'Json数组'),
         (2, 'attributeCode', '属性编码', 'String'),
         (2, 'attributeName', '属性名称', 'String'),
         (2, 'remark', '备注', 'String'),
     ]},
]}

FRT = {
'name': '携选工品运费', 'interfaces': [
    {'id': '1', 'name': '运费模板列表查询', 'url': 'https://openapi.xiekeyun.com/mro/freight/template/list', 'method': 'POST',
     'desc': '查询携选工品运费模板列表数据。运费从最小梯度开始计算，超过所有梯度免运费',
     'params': [
         ('erpCode', '请求者用户ERP帐号', '必填', 'String'),
     ],
     'responses': [
         (0, 'result', '返回结果状态 1:成功；9:失败', 'Number'),
         (0, 'errorCode', '接口错误编码（失败时有值）', 'String'),
         (0, 'errorMsg', '接口错误信息（失败时有值）', 'String'),
         (0, 'data', '运费模板数据', 'Json对象'),
         (1, 'list', '运费模板列表', 'Array'),
         (2, 'tempId', '运费模板ID', 'String'),
         (2, 'enable', '是否生效中（只有一个会生效）', 'Boolean'),
         (2, 'list', '运费梯度列表', 'Array'),
         (3, 'lineNo', '梯度项次', 'Number'),
         (3, 'cost', '所需运费', 'BigDecimal'),
         (3, 'value', '阈值（总商品金额大于此值进入下一梯度）', 'BigDecimal'),
     ]},
]}

MODS = [MRO, FRT]

def gen(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '接口总览'
    write_overview(ws, MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn = api['name']
            if len(sn)>18: sn = sn[:18]+'..'
            nm = f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31: nm = nm[:31]
            write_subs(wb.create_sheet(title=nm), api)
    wb.save(path); print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\4携选工品_接口梳理.xlsx')
