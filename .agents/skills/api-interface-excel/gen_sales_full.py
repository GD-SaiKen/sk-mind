#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""携客云SRM销售协同平台 - API接口梳理Excel（完整嵌套 + 分组模式）"""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

# ── Fills ──
DEEP_BLUE = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
LIGHT_BLUE = PatternFill(start_color='D6E4F0', end_color='D6E4F0', fill_type='solid')
LIGHT_GRAY = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
WHITE_FILL = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
NEST_L1 = PatternFill(start_color='F4F7FC', end_color='F4F7FC', fill_type='solid')
NEST_L2 = PatternFill(start_color='EBEFF7', end_color='EBEFF7', fill_type='solid')

# ── Fonts ──
WHITE_B = Font(name='微软雅黑', size=10, bold=True, color='FFFFFF')
BOLD_F = Font(name='微软雅黑', size=10, bold=True)
TITLE_F = Font(name='微软雅黑', size=11, bold=True)
NORMAL_F = Font(name='微软雅黑', size=10)

# ── Border ──
THIN_B = Border(left=Side(style='thin',color='B0B0B0'),right=Side(style='thin',color='B0B0B0'),top=Side(style='thin',color='B0B0B0'),bottom=Side(style='thin',color='B0B0B0'))

# ── Alignments ──
ALIGN_LEFT = Alignment(horizontal='left', vertical='center', wrap_text=True)
ALIGN_CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALIGN_VC = Alignment(vertical='center', wrap_text=True)

# ── Column Widths ──
C_W = {'A':28,'B':52,'C':14,'D':20}
OV_C_W = {'A':10,'B':36,'C':60,'D':12,'E':50}

def apply_style(c, font=NORMAL_F, fill=None, align=ALIGN_LEFT, border=THIN_B):
    c.font = font; c.alignment = align; c.border = border
    if fill is not None: c.fill = fill

def write_overview(ws, modules):
    for k,v in OV_C_W.items(): ws.column_dimensions[k].width = v
    ws.merge_cells('A1:E1')
    c = ws['A1']; c.value = '携客云SRM销售协同平台 - API接口总览'; c.font = Font(name='微软雅黑',size=14,bold=True); c.alignment = Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height = 30
    for i,h in enumerate(['序号','接口名称','接口地址','请求方式','接口描述'],1):
        cell = ws.cell(row=2,column=i,value=h); apply_style(cell, WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[2].height = 22
    row = 3
    for mod in modules:
        ws.merge_cells(f'A{row}:E{row}')
        apply_style(ws.cell(row=row,column=1,value=f'▸ {mod["name"]}'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
        for c_ in 'BCDE': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
        ws.row_dimensions[row].height = 22; row += 1
        for api in mod['interfaces']:
            for i,val in enumerate([api['id'],api['name'],api['url'],api['method'],api['desc']],1):
                cell = ws.cell(row=row,column=i,value=val)
                apply_style(cell, NORMAL_F, LIGHT_GRAY, ALIGN_CENTER if i in (1,4) else ALIGN_LEFT)
            row += 1
    ws.auto_filter.ref = f'A2:E{row-1}'

def write_subs(ws, api):
    for k,v in C_W.items(): ws.column_dimensions[k].width = v
    for idx,(l,v) in enumerate([('接口地址',api['url']),('请求方式',api['method']),('接口描述',api['desc'])],1):
        apply_style(ws.cell(row=idx,column=1,value=l), BOLD_F, WHITE_FILL, ALIGN_LEFT)
        ws.merge_cells(f'B{idx}:D{idx}')
        apply_style(ws.cell(row=idx,column=2,value=v), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        for c_ in 'CD': apply_style(ws[f'{c_}{idx}'], NORMAL_F, WHITE_FILL, ALIGN_LEFT)
    row = 5; ws.row_dimensions[4].height = 6
    # 请求参数标题
    ws.merge_cells(f'A{row}:D{row}'); apply_style(ws.cell(row=row,column=1,value='请求参数'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; row += 1
    for i,h in enumerate(['参数名称','参数说明','是否必须','数据类型'],1):
        apply_style(ws.cell(row=row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[row].height = 22; row += 1
    if api['params']:
        for pn,pd,pr,pt in api['params']:
            for i,val in enumerate([pn,pd,pr,pt],1):
                apply_style(ws.cell(row=row,column=i,value=val), NORMAL_F, WHITE_FILL, ALIGN_CENTER if i in (3,4) else ALIGN_LEFT)
            row += 1
    else:
        ws.merge_cells(f'A{row}:D{row}'); apply_style(ws.cell(row=row,column=1,value='（该接口无业务请求参数）'), NORMAL_F, WHITE_FILL, ALIGN_LEFT)
        row += 1
    row += 2; ws.row_dimensions[row-1].height = 6
    # 响应数据字段标题
    ws.merge_cells(f'A{row}:D{row}'); apply_style(ws.cell(row=row,column=1,value='响应数据字段'), TITLE_F, LIGHT_BLUE, ALIGN_LEFT)
    for c_ in 'BCD': apply_style(ws[f'{c_}{row}'], NORMAL_F, LIGHT_BLUE, ALIGN_LEFT)
    ws.row_dimensions[row].height = 22; hdr_row = row + 1
    for i,h in enumerate(['字段名称','字段说明','','数据类型'],1):
        apply_style(ws.cell(row=hdr_row,column=i,value=h), WHITE_B, DEEP_BLUE, ALIGN_CENTER)
    ws.row_dimensions[hdr_row].height = 22; row = hdr_row + 1
    # Write response fields with levels
    rl = []
    NEST_FILLS = {0: WHITE_FILL, 1: NEST_L1, 2: NEST_L2}
    for item in api['responses']:
        lvl, rname, rdesc, rtype = item[0], item[1], item[2] if len(item)>2 else '', item[3] if len(item)>3 else ''
        indent = '  ' * lvl
        row_fill = NEST_FILLS.get(lvl, WHITE_FILL)
        apply_style(ws.cell(row=row,column=1,value=indent+rname), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=2,value=rdesc), NORMAL_F, row_fill, ALIGN_LEFT)
        apply_style(ws.cell(row=row,column=3,value=''), NORMAL_F, row_fill, ALIGN_CENTER)
        apply_style(ws.cell(row=row,column=4,value=rtype), NORMAL_F, row_fill, ALIGN_CENTER)
        rl.append((row, lvl)); row += 1
    if rl:
        max_lvl = max(l for _,l in rl)
        for lvl in range(max_lvl, 0, -1):
            i = 0
            while i < len(rl):
                r, l = rl[i]
                if l == lvl:
                    j = i + 1
                    while j < len(rl) and rl[j][1] >= lvl: j += 1
                    if j - 1 > i: ws.row_dimensions.group(r, rl[j-1][0], outline_level=lvl)
                    i = j
                else: i += 1
    return row

# =========================== DATA ===========================
# response format: (nest_level, field_name, description, data_type)

CO = {
'name':'客户订单协同','interfaces':[
{'id':'1','name':'客户订单列表查询','url':'https://openapi.xiekeyun.com/customerOrder/orderList.json','method':'POST','desc':'获取客户订单列表数据，支持按发布时间/确认时间/状态筛选',
 'params':[('orderStatus','订单状态 1待答交2差异待确认3退回待答交4变更确认中5已确认6已结案7已冻结8已留置10撤回答交11采购方撤回12作废13供应商方拒绝','选填','Number'),('publishStartTime','发布开始时间(毫秒)','必填','Number'),('publishEndTime','发布结束时间(毫秒)','必填','Number'),('confirmStartTime','确认时间起始(毫秒)','必填','Number'),('confirmEndTime','确认时间截止(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('customerXkUniqueCode','客户档案平台唯一编码','选填','String')],
 'responses':[(0,'result','返回结果状态 1成功9失败','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','客户订单列表','JSON数组'),(1,'coXkNo','客户订单平台单号','String'),(1,'coInnerNo','客户订单内部单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'orderStatus','订单状态','Number'),(1,'publishTime','发布时间(毫秒)','Number'),(1,'confirmTime','确认时间(毫秒)','Number')]},
{'id':'2','name':'客户订单详情查询','url':'https://openapi.xiekeyun.com/customerOrder/getDetail.json','method':'POST','desc':'按平台单号获取客户订单详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('coXkNo','客户订单平台单号','必填','String')],
 'responses':[(0,'result','返回结果状态 1成功9失败','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','客户订单详情','Json对象'),(1,'coXkNo','客户订单平台单号','String'),(1,'coInnerNo','客户订单内部单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'customerName','客户企业名称','String'),(1,'orderStatus','订单状态','Number'),(1,'currencyCode','交易币种编码','String'),(1,'currencyName','交易币种名称','String'),(1,'payWayCode','付款条件编码','String'),(1,'payWayName','付款条件名称','String'),(1,'totalAmount','订单总金额','Number'),(1,'remark','备注','String'),(1,'publishTime','发布时间','Number'),(1,'confirmTime','确认时间','Number'),(1,'lineList','客户订单单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'productScale','产品规格','String'),(2,'customerProductCode','客户物料编码','String'),(2,'customerProductName','客户物料名称','String'),(2,'unitCode','单位编码','String'),(2,'unitName','单位名称','String'),(2,'orderQty','订单数量','Number'),(2,'price','单价','Number'),(2,'taxPrice','含税单价','Number'),(2,'amount','未税金额','Number'),(2,'taxAmount','含税金额','Number'),(2,'expectedDate','预计交期(时间戳)','Number'),(2,'remark','备注','String'),(2,'receiveAddressCode','收货部门编码','String'),(2,'receiveAddressName','收货部门名称','String'),(2,'purchaseDesc','采购说明','String'),(2,'extendN01','扩展字段1','String'),(2,'extendN02','扩展字段2','String')]},
{'id':'3','name':'客户订单条码查询','url':'https://openapi.xiekeyun.com/customerOrder/getBarcodeList.json','method':'POST','desc':'获取客户订单生产条码数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('coXkNo','客户订单平台单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','条码数据列表','JSON数组'),(1,'barcode','条码号','String'),(1,'productCode','产品编码','String'),(1,'productName','产品名称','String'),(1,'productScale','产品规格','String'),(1,'packLevel','包装层级标记','Number'),(1,'lineNo','项次号','Number')]},
{'id':'4','name':'客户订单进度列表查询','url':'https://openapi.xiekeyun.com/customerOrder/getProgressList.json','method':'POST','desc':'获取客户订单执行进度查询数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('startDate','开始时间(毫秒)','必填','Number'),('endDate','结束时间(毫秒)','必填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','订单进度列表','JSON数组'),(1,'coXkNo','客户订单平台单号','String'),(1,'productCode','产品编码','String'),(1,'productName','产品名称','String'),(1,'orderQty','订单数量','Number'),(1,'deliveryQty','已出货数量','Number'),(1,'remainQty','待出货数量','Number'),(1,'progressRate','进度百分比','Number')]},
{'id':'5','name':'客户变更单列表查询','url':'https://openapi.xiekeyun.com/customerChangeOrder/orderList.json','method':'POST','desc':'获取客户变更单列表数据',
 'params':[('startDate','单据发布时间开始(毫秒)','必填','Number'),('endDate','单据发布时间结束(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('customerXkUniqueCode','客户档案平台唯一编码','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','客户变更单列表','JSON数组'),(1,'ccXkNo','客户变更单平台单号','String'),(1,'coXkNo','关联客户订单平台单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'status','状态1待确认2已确认3退回','Number'),(1,'publishTime','发布时间','Number')]},
{'id':'6','name':'客户变更单详情查询','url':'https://openapi.xiekeyun.com/customerChangeOrder/getDetail.json','method':'POST','desc':'按平台单号获取客户变更单详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('ccXkNo','客户变更单平台单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','客户变更单详情','Json对象'),(1,'ccXkNo','客户变更单平台单号','String'),(1,'coXkNo','关联客户订单平台单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'status','状态','Number'),(1,'remark','变更说明','String'),(1,'publishTime','发布时间','Number'),(1,'changeOrderlineList','变更单单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'oldQty','原数量','Number'),(2,'newQty','新数量','Number'),(2,'changeOrderLine','变更单项次明细','Json对象数组'),(3,'lineNo','变更单单身项次号','Number'),(3,'changeType','变更类型','Number'),(3,'oldValue','原值','String'),(3,'newValue','新值','String')]}]}

DS = {
'name':'供应出货协同','interfaces':[
{'id':'1','name':'送货单号列表查询','url':'https://openapi.xiekeyun.com/customer/delivery/getNoList.json','method':'POST','desc':'按时间范围拉取新的送货单号列表',
 'params':[('startDate','开始时间(毫秒)','必填','Number'),('endDate','结束时间(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerCustomerCode','内部客户编码','选填','String'),('status','状态数组例[1,3]','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','送货单号数据列表','JSON数组'),(1,'deliveryNo','送货单号','String'),(1,'deliveryType','送货类型1普通2补货3备品4样品5VMI','Number'),(1,'innerCustomerCode','内部客户编码','String'),(1,'status','状态1待签收3已完成5申请中6退回7已同意9部分同意','String'),(1,'logisticsStatus','物流状态0未发出1已发货2已到货','String'),(1,'publishTime','发布时间(毫秒)','String')]},
{'id':'2','name':'送货单详情查询','url':'https://openapi.xiekeyun.com/customer/delivery/getDetail.json','method':'POST','desc':'查询一个送货单明细数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','送货单数据','Json对象'),(1,'deliveryNo','送货单号','String'),(1,'deliveryType','送货类型1普通2补货3备品4样品5VMI','Number'),(1,'innerCustomerCode','客户编码','String'),(1,'deliveryDate','送货日期(时间戳)','Number'),(1,'planArrivedDate','预计送达日期(时间戳)','Number'),(1,'contactPersion','送货联系人','String'),(1,'contactMobile','送货人联系电话','String'),(1,'addressInfo','收货明细地址','String'),(1,'remark','送货通知备注','String'),(1,'status','送货单状态1待签收2签收中3已完成4撤回','Number'),(1,'directDeliveryFlag','直运标识1是0否','Number'),(1,'logisticsStatus','物流状态0未发出1已发货2已到货','Number'),(1,'publishTime','最后一次发布时间','Number'),(1,'updateTime','最后一次更新时间','Number'),(1,'profitCenterCode','利润中心编码','String'),(1,'grossWeight','整单毛重','Number'),(1,'netWeight','整单净重','Number'),(1,'totalPackingQty','总包装数','Number'),(1,'transportTypeCode','运输方式编码','String'),(1,'transportTypeName','运输方式名称','String'),(1,'customsData','报关资料','Json对象'),(2,'dnXkNo','送货通知单据号','String'),(2,'billLadingXkNo','报单号码','String'),(2,'customsTypeCode','报关类型编码','String'),(2,'customsTypeName','报关类型名称','String'),(2,'destinationPortTypeCode','目的港编码','String'),(2,'overseasDeliveryFlag','海外送货0否1是','Number'),(1,'fieldConfigs','字段配置列表','Json对象数组'),(2,'billType','单据类型10采购单22送货单','String'),(2,'fieldLevel','字段层级','Number'),(2,'fieldCode','字段编码','String'),(2,'fieldName','字段名称','String'),(1,'lineList','送货单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'poErpNo','ERP采购单号','String'),(2,'purchaseType','采购单类型1一般2委外','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'productScale','产品规格','String'),(2,'deliveryUnitCode','出货单位编码','String'),(2,'deliveryQty','出货数量','Number'),(2,'price','出货产品单价','Number'),(2,'taxPrice','出货产品含税单价','Number'),(2,'srcBillType','原始单据类型1订单2排程3按退货4按备品','Number'),(2,'status','状态1待签收2签收中3已完成','Number'),(2,'receiveQty','收货量','Number'),(2,'returnQty','退货量','Number'),(2,'storageQty','入库量','Number'),(2,'buyerName','采购员','String'),(2,'expectedDate','订单交期(时间戳)','Number'),(2,'lineExtendList','送货单身扩展信息','Json对象数组'),(3,'toolingsUniqueCode','工装信息平台唯一编码','String'),(3,'toolingsCode','工装编码','String'),(3,'toolingsName','工装名称','String'),(3,'unitNetWeight','单位净重','Number')]},
{'id':'3','name':'送货单条码查询','url':'https://openapi.xiekeyun.com/customer/barcode/byDeliveryNo.json','method':'POST','desc':'查询一个送货单包含的所有条码数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('deliveryNo','送货单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dynamicDescList','小包条码规则描述','Json对象数组'),(0,'packageList','包装列表(version>=1.1)','Json对象数组'),(0,'dataList','条码数据列表','JSON数组'),(1,'barcode','条码号','String'),(1,'productCode','产品编码','String'),(1,'productName','产品名称','String'),(1,'productScale','产品规格','String'),(1,'qty','数量','Number'),(1,'packLevel','包装层级','Number')]},
{'id':'4','name':'客供料代收单单号列表','url':'https://openapi.xiekeyun.com/customer/delivery/getCollectionNoList.json','method':'POST','desc':'按时间范围拉取新的客供料代收单单号列表',
 'params':[('startDate','开始时间(毫秒)','必填','Number'),('endDate','结束时间(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','客供料代收单号列表','JSON数组'),(1,'collectionNo','客供料代收单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'status','状态','Number'),(1,'publishTime','发布时间','Number')]},
{'id':'5','name':'客供料代收单详情','url':'https://openapi.xiekeyun.com/customer/delivery/getCollectionDetail.json','method':'POST','desc':'查询一个客供料代收单明细数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('collectionNo','客供料代收单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','客供料代收单详情','Json对象'),(1,'collectionNo','客供料代收单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'status','状态','Number'),(1,'publishTime','发布时间','Number'),(1,'lineList','客供料代收单单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'collectionQty','代收数量','Number')]}]}

SP = {
'name':'客户计划协同','interfaces':[
{'id':'1','name':'客户排程列表查询','url':'https://openapi.xiekeyun.com/customer/schedule/getList.json','method':'POST','desc':'按时间范围拉取新的客户排程列表',
 'params':[('startDate','单据发布时间开始(毫秒)','必填','Number'),('endDate','单据发布时间结束(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerCustomerCodes','内部客户编码列表','选填','String数组'),('status','状态数组[1,2,3,4,5,6,7,8,9,10]','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','排程单号数据列表','JSON数组'),(1,'batchNo','排程批次号','String'),(1,'scheduleNo','排程单号','String'),(1,'refBatchNo','来源排程批次号','String'),(1,'status','状态1已发布2答交已保存3差异待确认4差异已确认5全部答交6退回7部分答交有差异8部分答交无差异9已失效10作废','Number'),(1,'schedulePlanType','排程计划类型1交货计划2预测计划','Number'),(1,'oldBatchType','1更新计划2调整计划3增量计划4排程变更','String'),(1,'oldBatchNo','历史版本批次号','String'),(1,'publishName','发布人','String'),(1,'publishTime','发布时间','Number')]},
{'id':'2','name':'客户排程详情查询','url':'https://openapi.xiekeyun.com/customer/schedule/getDetail.json','method':'POST','desc':'查询一个客户排程详情(含答交数据)',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('batchNo','排程批次号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','排程数据','Json对象'),(1,'batchNo','排程号','String'),(1,'refBatchNo','来源排程单号','String'),(1,'status','排程状态1已发布2部分答交无差异3有差异未确认4全部答交5已失效','Number'),(1,'publishName','发布人员工姓名','String'),(1,'publishTime','发布时间(毫秒)','Number'),(1,'oldBatchNo','老版本批次号','String'),(1,'oldBatchType','0非老版本1更新2计划调整','Number'),(1,'schDesc','计划说明','String'),(1,'itemList','排程明细列表','Json对象数组'),(2,'itemId','排程物料id(平台唯一编码)','String'),(2,'scheduleNo','排程供应商维度批次号','String'),(2,'prodCode','物料编码','String'),(2,'prodName','物料名称','String'),(2,'prodScale','物料规格','String'),(2,'innerVendorCode','供应商编码','String'),(2,'innerVendorName','供应商名称','String'),(2,'poErpNo','订单Erp编码','String'),(2,'lineNo','订单项次号','String'),(2,'profitCenterCode','利润中心编码','String'),(2,'profitCenterName','利润中心名称','String'),(2,'receiveAddressCode','收货地点编码','String'),(2,'receiveAddressName','收货地点名称','String'),(2,'status','物料状态1已发布2答交已保存3答交发布有差异4答交发布无差异5差异已确认6退回7已失效8作废','Number'),(2,'urgentFlag','是否急料0否1是','Number'),(2,'waitDeliveryQty','订单待出数量','Number'),(2,'stockQty','库存量','Number'),(2,'replyTime','答交时间(毫秒)','Number'),(2,'brandName','品牌名称','String'),(2,'remark','备注','String'),(2,'detailList','排程时间明细','Json数组'),(3,'detailId','排程单身明细主键列','String'),(3,'sourceType','来源1采购方2销售方','String'),(3,'scheduleDate','排程日期(unix毫秒)','Number'),(3,'subList','排程时间明细','Json数组'),(4,'dtlSubId','排程单身明细sub主键列','String'),(4,'sourceType','来源1采购方2销售方','String'),(4,'detailDate','排程日期(unix毫秒)','Number'),(4,'scheduleQty','排程量','Number'),(4,'replyQty','答交量','Number'),(4,'prodPlanQty','生产计划量','Number')]},
{'id':'3','name':'排程发布结果查询','url':'https://openapi.xiekeyun.com/customer/schedule/getPublishResult.json','method':'POST','desc':'查询排程发布处理结果',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('batchNo','排程批次号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','发布结果','Json对象'),(1,'batchNo','排程批次号','String'),(1,'publishStatus','发布状态1成功9失败','Number'),(1,'publishTime','发布时间','Number'),(1,'errorDesc','错误描述','String')]}]}

VM = {
'name':'VMI库存看板','interfaces':[
{'id':'1','name':'VMI库存看板列表查询','url':'https://openapi.xiekeyun.com/vmiVendorBoard/queryList.json','method':'POST','desc':'VMI库存看板列表查询',
 'params':[('startDate','单据更新时间开始(毫秒)','必填','Number'),('endDate','单据更新时间结束(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('customerKey','内部客户编码或名称','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','VMI库存看板数据列表','JSON数组'),(1,'cProdCode','客户物料编码','String'),(1,'cProdName','客户物料名称','String'),(1,'cProdScale','客户物料规格','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'plantCode','工厂编码','String'),(1,'plantName','工厂名称','String'),(1,'currentInventoryQty','当前库存量','Number'),(1,'safeInventoryQty','安全库存量','Number'),(1,'highestInventoryQty','最高库存量','Number'),(1,'status','状态0无效1有效','Number'),(1,'storeCode','仓库编码','String'),(1,'storeName','仓库名称','String'),(1,'locationCode','库位编码','String'),(1,'locationName','库位名称','String'),(1,'erpCategoryCode','ERP中分类编码','String'),(1,'erpCategoryName','ERP中分类名称','String'),(1,'erpStorageDate','入库日期','Number'),(1,'prodFeature','产品特征','String'),(1,'storeUnitCode','库存单位编码','String'),(1,'storeUnitName','库存单位名称','String'),(1,'batchNo','批次号','String'),(1,'lateDeliveryDate','最近送货时间','Number'),(1,'lateDeliveryQty','最近送货数量','Number'),(1,'dnXkNo','送货单号','String')]}]}

AR = {
'name':'应收账款协同','interfaces':[
{'id':'1','name':'对账单列表查询','url':'https://openapi.xiekeyun.com/customer/statement/getList.json','method':'POST','desc':'查找对账单列表信息',
 'params':[('startDate','发布时间开始(毫秒)','必填','Number'),('endDate','发布时间结束(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','状态数组[1,5,6,7]','选填','Number数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','对账单单号数据列表','JSON数组'),(1,'recoXkNo','对账单携客云平台单号','String'),(1,'status','1待采购方确认5已确认6数量待确认7待供应方确认','Number'),(1,'innerCustomerCode','内部客户编码(ERP)','String'),(1,'innerCustomerName','内部客户名称(ERP)','String'),(1,'publishTime','发布时间','Number'),(1,'lastOperateTime','最后一次更新时间','Number')]},
{'id':'2','name':'对账单详情查询','url':'https://openapi.xiekeyun.com/customer/statement/getDetail.json','method':'POST','desc':'查询对账单详情信息',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('recoXkNo','对账单携客云平台单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','对账单单数据','Json对象'),(1,'recoXkNo','对账单平台单号','String'),(1,'recoInnerNo','对账单内部单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'recoStartTime','对账周期开始时间','Number'),(1,'recoEndTime','对账周期结束时间','Number'),(1,'openAccountBankCode','开户行编码','String'),(1,'openAccountBankName','开户行名称','String'),(1,'bankCardNo','银行账号','String'),(1,'vendorCreditCode','纳税人识别号','String'),(1,'currencyCode','交易币种编码','String'),(1,'currencyName','交易币种名称','String'),(1,'payWayCode','付款条件编码','String'),(1,'payWayName','付款条件名称','String'),(1,'purchaseType','采购类型1一般2委外3多角','Number'),(1,'remark','采购方备注','String'),(1,'vendorRemark','供应商备注','String'),(1,'status','单据状态','Number'),(1,'totalAmount','总对账金额','Number'),(1,'totalTaxAmount','总对账含税金额','Number'),(1,'totalConfirmAmount','总确认对账金额','Number'),(1,'totalConfirmTaxAmount','总确认对账含税金额','Number'),(1,'alreadyInvoiceAmount','已开票金额','Number'),(1,'alreadyInvoiceTaxAmount','已开票含税金额','Number'),(1,'invoiceStatus','开票状态0未开票1部分开票2全部开票','Number'),(1,'confirmEmpName','确认人','String'),(1,'confirmErpEmpNumber','确认员工erp编号','String'),(1,'publishName','发布人员工姓名','String'),(1,'publishTime','发布时间','Number'),(1,'lastOperateTime','最后一次更新时间','Number'),(1,'openAccountBankAddress','开户行地址','String'),(1,'deliveryList','出货明细列表','Json对象数组'),(2,'lineNo','项次','String'),(2,'dnXkNo','送货通知单平台单号','String'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'productScale','产品规格','String'),(2,'deliveryQty','出货数量','Number'),(2,'deliveryAmount','出货未税金额','Number'),(2,'deliveryTaxAmount','出货含税金额','Number'),(2,'storageQty','入库数量','Number'),(2,'curReconQty','本次对账数量','Number'),(2,'curReconAmount','本次对账未税金额','Number'),(2,'curReconTaxAmount','本次对账含税金额','Number'),(2,'confirmStatus','状态0待采购方确认1采购方本次不对2采购方已确认','Number'),(1,'returnList','仓退明细列表','Json对象数组'),(2,'lineNo','项次','Number'),(2,'rtoErpNo','退货单ERP单号','String'),(2,'productCode','产品编码','String'),(2,'productName','产品名称','String'),(2,'returnQty','仓退数量','Number'),(2,'returnAmount','仓退未税金额','Number'),(2,'curReconAmount','本次对账未税金额','Number'),(2,'confirmStatus','确认状态','Number'),(1,'costList','应收折让明细列表','Json对象数组'),(2,'lineNo','项次','Number'),(2,'csoXkNo','费用单平台单号','String'),(2,'costAmount','费用单未税金额','Number'),(2,'curReconAmount','本次对账未税金额','Number'),(2,'confirmStatus','状态','Number'),(1,'prepayList','预付款明细列表','Json对象数组'),(2,'lineNo','项次','Number'),(2,'ppmXkNo','预付款单平台单号','String'),(2,'prepayAmount','预付款单总金额','Number'),(2,'curReconAmount','本次对账抵扣金额','Number'),(2,'confirmStatus','状态','Number'),(1,'vmiDeliveryList','VMI入库明细','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'billQty','入库数量','Number'),(2,'billAmount','单据未税金额','Number')]},
{'id':'3','name':'发票单列表查询','url':'https://openapi.xiekeyun.com/customer/invoice/getList.json','method':'POST','desc':'查找发票单列表信息',
 'params':[('startDate','发布时间开始(毫秒)','必填','Number'),('endDate','发布时间结束(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('status','状态','选填','Number')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','发票单号列表','JSON数组'),(1,'invXkNo','发票平台单号','String'),(1,'status','状态','Number'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'publishTime','发布时间','Number')]},
{'id':'4','name':'发票单详情查询','url':'https://openapi.xiekeyun.com/customer/invoice/getDetail.json','method':'POST','desc':'查询发票单详情信息',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('invXkNo','发票平台单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','发票单详情','Json对象'),(1,'invXkNo','发票平台单号','String'),(1,'invInnerNo','发票内部单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'invoiceNo','发票号','String'),(1,'invoiceDate','开票日期','Number'),(1,'totalTaxAmount','含税总金额','Number'),(1,'status','状态','Number'),(1,'publishTime','发布时间','Number'),(1,'lineList','发票单身列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'invoicingQty','开票数量','Number'),(2,'totalAmount','未税金额','Number'),(2,'totalTaxAmountAdjust','含税金额','Number')]}]}

IQ = {
'name':'客户询价招标','interfaces':[
{'id':'1','name':'客户物料询价招标列表查询','url':'https://openapi.xiekeyun.com/customer/inquiryOrder/getList.json','method':'POST','desc':'获取客户询价招标单列表(version=1按发布时间,version=2按更新时间)',
 'params':[('startDate','发布开始时间(毫秒)','必填','Number'),('endDate','发布结束时间(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','询价单号列表','Json对象数组'),(1,'inqXkNo','询价单号','String'),(1,'status','单据状态1未开始2报价中3已取消4比价中5已完成6已结束','Number')]},
{'id':'2','name':'客户物料询价招标详情查询','url':'https://openapi.xiekeyun.com/customer/inquiryOrder/getDetail.json','method':'POST','desc':'获取客户询价招标单详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('inqXkNo','询价招标单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','询价招标单数据','Json对象'),(1,'inqXkNo','询价招标单号','String'),(1,'inqInnerNo','内部询价招标单号','String'),(1,'innerCustomerCode','内部客户编码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'billTitle','标题','String'),(1,'billType','类型1询价2招标','Number'),(1,'startDate','开始时间','Number'),(1,'endDate','结束时间','Number'),(1,'payWayName','付款条件名称','String'),(1,'quoteCurrencyCode','报价币别编码','String'),(1,'quoteCurrencyName','报价币别名称','String'),(1,'priceContainTaxFlag','价格是否含税0不限1含税2未税','Number'),(1,'status','1未开始2报价中3已取消4比价中5已完成6已结束7比价结束','Number'),(1,'publishName','发布人名称','String'),(1,'publishTime','发布时间','Number'),(1,'taxRate','税率','Number'),(1,'taxCode','税别编码','String'),(1,'taxName','税别名称','String'),(1,'transportTypeName','运输方式名称','String'),(1,'deliveryPoint','交货地点','String'),(1,'inquiryTypeName','询价类别名称','String'),(1,'deliveryTypeName','交货方式名称','String'),(1,'invoiceType','发票类型1专票2普票','Number'),(1,'oppProductName','商机产品名称','String'),(1,'forecastCycleType','预测周期类型1每月2每季3每年4总量','Number'),(1,'requireQty','预计需求数量','Number'),(1,'productList','物料列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'productScale','物料规格','String'),(2,'unitCode','单位编码','String'),(2,'unitName','单位名称','String'),(2,'purchaseQty','采购数量','Number'),(2,'expectedPrice','期望价格','Number'),(2,'upperLimitPrice','价格上限','Number'),(2,'newFlag','是否为新产品0否1是','Number'),(2,'description','相关说明','String'),(2,'pricingUnitCode','计价单位编码','String'),(2,'pricingUnitName','计价单位名称','String'),(2,'pricingQuantity','计价数量','Number'),(2,'brandName','品牌名称','String'),(2,'demandDate','需求日期','Number'),(2,'remark','备注','String'),(2,'costFieldList','单身成本字段','Json对象数组'),(3,'code','编码','String'),(3,'name','名称','String'),(2,'dynamicFieldList','单身自定义字段','Json对象数组'),(3,'code','编码','String'),(3,'name','名称','String'),(3,'value','值','String')]},
{'id':'3','name':'报价单列表查询','url':'https://openapi.xiekeyun.com/customer/inquiryOrder/getQuoteList.json','method':'POST','desc':'获取客户询价招标报价列表',
 'params':[('startDate','客户报价单发布开始时间(毫秒)','必填','Number'),('endDate','客户报价单发布结束时间(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String'),('innerCustomerCode','内部客户编码','选填','String'),('inqXkNo','询价单号','选填','String'),('status','状态1已发布2发布退回3未入选4部分入选5全部入选6已作废','选填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','报价单列表','Json对象数组'),(1,'quoXkNo','平台报价单号','String'),(1,'quoInnerNo','平台内部报价单号','String'),(1,'inqXkNo','询价单号','String'),(1,'quoteCode','报价码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'innerCustomerCode','内部客户编号','String'),(1,'quoteName','报价人姓名','String'),(1,'quoteTime','报价日期(时间戳)','Number'),(1,'description','报价说明','String'),(1,'rejectReason','退回原因','String'),(1,'createTime','创建日期(时间戳)','Number'),(1,'status','状态1已发布2发布退回3未入选4部分入选5全部入选6已作废','Number')]},
{'id':'4','name':'报价单详情查询','url':'https://openapi.xiekeyun.com/customer/inquiryOrder/getQuoteDetail.json','method':'POST','desc':'获取客户询价招标单报价详情',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('quoXkNo','报价单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','报价单详情','Json对象'),(1,'quoXkNo','平台报价单号','String'),(1,'quoInnerNo','平台内部报价单号','String'),(1,'inqXkNo','询价单号','String'),(1,'quoteCode','报价码','String'),(1,'innerCustomerName','内部客户名称','String'),(1,'innerCustomerCode','内部客户编号','String'),(1,'quoteName','报价人姓名','String'),(1,'quoteTime','报价日期','Number'),(1,'description','报价说明','String'),(1,'status','状态','Number'),(1,'products','报价物料列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'productScale','物料规格','String'),(2,'purchaseQty','采购数量','Number'),(2,'unitCode','单位编码','String'),(2,'unitName','单位名称','String'),(2,'quotePrice','报价单价','Number'),(2,'selectedType','入选状态','Number'),(2,'quoteSelectedFlag','报价入选标识','Number')]},
{'id':'5','name':'竞价单报价列表查询','url':'https://openapi.xiekeyun.com/customer/biddingOrder/getQuoteList.json','method':'POST','desc':'获取竞价单报价列表信息',
 'params':[('startDate','开始时间(毫秒)','必填','Number'),('endDate','结束时间(毫秒)','必填','Number'),('erpCode','请求者用户ERP帐号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','竞价报价列表','Json对象数组'),(1,'bidXkNo','竞价单号','String'),(1,'quoXkNo','报价单号','String'),(1,'status','状态','Number'),(1,'quoteName','报价人','String'),(1,'quoteTime','报价时间','Number')]},
{'id':'6','name':'竞价单报价详情查询','url':'https://openapi.xiekeyun.com/customer/biddingOrder/getQuoteDetail.json','method':'POST','desc':'获取客户竞价单报价详情数据',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('quoXkNo','报价单号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','竞价报价详情','Json对象'),(1,'bidXkNo','竞价单号','String'),(1,'quoXkNo','报价单号','String'),(1,'quoteName','报价人','String'),(1,'quoteTime','报价时间','Number'),(1,'status','状态','Number'),(1,'products','报价物料列表','Json对象数组'),(2,'lineNo','项次号','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'quotePrice','报价单价','Number'),(2,'selectedType','入选状态','Number')]}]}

OP = {
'name':'商机管理','interfaces':[
{'id':'1','name':'商机线索单详情查询','url':'https://openapi.xiekeyun.com/oppOrder/getDetail.json','method':'POST','desc':'商机线索单详情信息',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('oppInnerNo','商机线索内部单据号','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'data','商机线索单数据','Json对象'),(1,'oppXkNo','商机线索平台单据号','String'),(1,'oppInnerNo','商机内部单据号','String'),(1,'oppTitle','商机名称','String'),(1,'customerType','客户类型0潜在1正式2录入','Number'),(1,'customerInnerCode','客户内部编码','String'),(1,'customerInnerName','客户内部名称','String'),(1,'oppTypeCode','商机类型编码','String'),(1,'oppTypeName','商机类型名称','String'),(1,'oppProductName','商机产品','String'),(1,'productionCode','生产企业','String'),(1,'prodCategoryName','所选产品分类字符串名称','String'),(1,'profitCenterCode','工厂编码','String'),(1,'profitCenterName','工厂名称','String'),(1,'forecastCycleType','预测周期','Number'),(1,'requireQty','预测数量','Number'),(1,'productionTime','量产时间','Number'),(1,'productionCycle','生产周期','Number'),(1,'expectedPrice','期望价格','Number'),(1,'deadline','截止时间','Number'),(1,'developEmpName','研发人员姓名','String'),(1,'developErpAccount','研发负责人ERP帐号','String'),(1,'buyerEmpName','采购员姓名','String'),(1,'buyerErpAccount','采购员员工档案ERP帐号','String'),(1,'currencyCode','币别编码','String'),(1,'currencyName','币别名称','String'),(1,'oppStatusCode','商机状态编码','String'),(1,'oppStatusName','商机状态名称','String'),(1,'exchangeRate','币别到本位币汇率','Number'),(1,'selectedFlag','商机结案标识0否1是','Number'),(1,'inqStatus','询价状态0部分1完成','Number'),(1,'returnDesc','退回原因说明','String'),(1,'createName','创建人姓名','String'),(1,'publishName','发布人姓名','String'),(1,'publishTime','发布时间','Number'),(1,'description','备注说明','String'),(1,'innerInqXkNo','询价需求单','String'),(1,'quoteSum','询价成本合计','Number'),(1,'totalBaseCurrencyPrice','本位币别合计','Number'),(1,'totalOppPrice','商机币别合计','Number'),(1,'dynamicFields','Excel物料清单自定义列字符串','String'),(1,'oppProdGroupList','物料群组询价成本列表','Json对象数组'),(2,'productGroup','物料群组','String'),(2,'quoteSum','询价成本','Number'),(2,'totalBaseCurrencyPrice','本位币别合计','Number'),(2,'totalOppPrice','商机币别合计','Number'),(1,'lineList','成本中心列表','Json对象数组'),(2,'productGroup','物料群组','String'),(2,'productCategoryCode','物料分类编码','String'),(2,'productCategoryName','物料分类名称','String'),(2,'lineNo','项次','Number'),(2,'productCode','物料编码','String'),(2,'productName','物料名称','String'),(2,'productScale','物料规格','String'),(2,'unitName','单位','String'),(2,'brandName','品牌','String'),(2,'description','说明','String'),(2,'demandQty','需求数量','Number'),(2,'unitDosage','单位用量','Number'),(2,'quotePrice','最新报价单价','Number'),(2,'quoteCurrencyCode','最新报价单价币别编码','String'),(2,'quoteCurrencyName','最新报价单价币别名称','String'),(2,'unitCost','单位用量成本','Number'),(2,'purchasePrice','现有价格','Number'),(2,'currencyCode','采购单币别编码','String'),(2,'currencyName','采购单币别名称','String'),(2,'innerInqXkNo','询价需求单','String'),(2,'status','状态6作废','Number'),(2,'newFlag','是否新物料0否1是','Number'),(2,'alternativeGroup','替代组','String'),(2,'alternativeRelation','替代关系标识S或M','String'),(2,'dynamicFields','Excel物料清单自定义列字符串','String'),(2,'purchaseQuota','采购配额','Number'),(2,'fileList','商机导入附件URL对象集合','Json对象数组'),(3,'fileName','文件名称','String'),(3,'fileUrl','文件url','String'),(3,'fileSize','文件大小(字节)','Number')]},
{'id':'2','name':'产品分类树查询','url':'https://openapi.xiekeyun.com/oppOrder/getProdTrees.json','method':'POST','desc':'返回产品分类树信息',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('categoryTypes','分类所属类型["1001","1002","1003"]','必填','String数组')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','产品分类数据列表','Json对象数组'),(1,'type','分类所属类型1001/1002/1003','String'),(1,'data','分类对象','Json对象数组'),(2,'code','分类编码','String'),(2,'name','分类名称','String'),(2,'selectedFlag','分类是否指定默认0否1是','Number'),(2,'children','分类子节点','Json对象数组'),(3,'code','分类编码','String'),(3,'name','分类名称','String'),(3,'selectedFlag','分类是否指定默认','Number'),(3,'children','子节点(递归)','Json对象数组')]},
{'id':'3','name':'产品分类工厂关系查询','url':'https://openapi.xiekeyun.com/oppOrder/getCategoryFactoryRel.json','method':'POST','desc':'返回产品分类工厂关系信息',
 'params':[('erpCode','请求者用户ERP帐号','必填','String'),('categoryType','分类所属类型','必填','String'),('categoryCode','分类编码','必填','String')],
 'responses':[(0,'result','返回结果状态','Number'),(0,'errorCode','接口错误编码','String'),(0,'errorMsg','接口错误信息','String'),(0,'dataList','产品分类工厂关系数据','Json对象数组'),(1,'categoryType','分类所属类型','String'),(1,'categoryCode','分类编码','String'),(1,'profitCenterCode','利润中心编码','String'),(1,'profitCenterName','利润中心名称','String'),(1,'developEmpName','研发人员姓名','String'),(1,'developEmpNumber','研发人员工号','String'),(1,'developRoleCode','研发岗位编码','String'),(1,'developRoleName','研发岗位名称','String'),(1,'developErpAccount','研发负责人ERP帐号','String'),(1,'developErpEmployeeCode','研发负责人ERP员工编码','String'),(1,'buyerEmpName','采购人员姓名','String'),(1,'buyerEmpNumber','采购人员工号','String'),(1,'buyerRoleCode','采购岗位编码','String'),(1,'buyerRoleName','采购岗位名称','String'),(1,'buyerErpAccount','采购负责人ERP帐号','String'),(1,'buyerErpEmployeeCode','采购负责人ERP员工编码','String')]}]}

MODS = [CO, DS, SP, VM, AR, IQ, OP]

def gen(path):
    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = '接口总览'
    write_overview(ws, MODS)
    for mod in MODS:
        for api in mod['interfaces']:
            sn = api['name']
            if len(sn)>18: sn = sn[:18]+'..'
            nm = f"{mod['name']}_{api['id']}_{sn}"
            if len(nm)>31: nm = nm[:31]
            write_subs(wb.create_sheet(title=nm), api)
    wb.save(path); print(f'OK: {path}')

gen('G:\\workspace\\GD-SaiiKen\\sk-mind\\docs\\数据梳理\\SRM\\梳理后\\3销售协同_接口梳理.xlsx')
